"""
Advanced TCO Calculation Commands - LLM-PRD - 03 Implementation

Management commands for advanced TCO calculation with:
- Comprehensive TCO calculations with 12 categories
- Industry benchmark comparisons
- Excel export with formulas and charts
- Sensitivity analysis and reporting
- Batch calculations and comparisons

Usage:
    python manage.py calculate-tco --vendor-product-id 1 --users 1000
    python manage.py batch-tco --vendor-ids 1,2,3 --users 1000
    python manage.py export-tco --tco-id 1
    python manage.py tco-benchmarks
"""

import logging

import click
from flask.cli import with_appcontext

from app import db
from app.models.vendor.vendor_organization import TCOCalculation, VendorProduct
from app.services.advanced_tco_engine import AdvancedTCOEngine

logger = logging.getLogger(__name__)


@click.command()
@click.option("--vendor-product-id", required=True, type=int, help="Vendor product ID")
@click.option("--users", required=True, type=int, help="Number of users")
@click.option("--years", default=5, type=int, help="TCO period in years")
@click.option(
    "--deployment",
    default="cloud",
    type=click.Choice(["cloud", "on-premise", "hybrid"]),
    help="Deployment model",
)
@click.option(
    "--org-size",
    default="medium",
    type=click.Choice(["small", "medium", "large", "enterprise"]),
    help="Organization size",
)
@click.option("--industry", default="manufacturing", type=str, help="Industry sector")
@click.option("--sensitivity", is_flag=True, help="Include sensitivity analysis")
@click.option("--export", is_flag=True, help="Export to Excel")
@with_appcontext
def calculate_tco(
    vendor_product_id, users, years, deployment, org_size, industry, sensitivity, export
):
    """Calculate comprehensive TCO for a vendor product."""

    click.echo("🧮 Advanced TCO Calculation")
    click.echo("=" * 40)

    try:
        # Initialize TCO engine
        engine = AdvancedTCOEngine()

        click.echo(f"📊 Calculating TCO for vendor product {vendor_product_id}")
        click.echo(f"👥 Users: {users:,}")
        click.echo(f"⏱️  Period: {years} years")
        click.echo(f"☁️  Deployment: {deployment}")
        click.echo(f"🏢 Organization: {org_size}")
        click.echo(f"🏭 Industry: {industry}")
        click.echo(f"📈 Sensitivity: {'Yes' if sensitivity else 'No'}")
        click.echo(f"📤 Export: {'Yes' if export else 'No'}")
        click.echo()

        # Calculate TCO
        results = engine.calculate_comprehensive_tco(
            vendor_product_id=vendor_product_id,
            user_count=users,
            tco_period_years=years,
            deployment_model=deployment,
            organization_size=org_size,
            industry=industry,
            include_sensitivity_analysis=sensitivity,
        )

        # Display results
        vendor_info = results["vendor_product"]
        cost_summary = results["cost_breakdown"]["summary"]
        comparative = results["comparative_metrics"]

        click.echo("✅ TCO Calculation Results:")
        click.echo(f"   Product: {vendor_info['name']} ({vendor_info['vendor_name']})")
        click.echo(f"   Total TCO: ${cost_summary['total_tco']:,.2f}")
        click.echo(f"   Annual Average: ${cost_summary['annual_average']:,.2f}")
        click.echo(f"   Per User Annual: ${cost_summary['per_user_annual']:,.2f}")
        click.echo(f"   Per User Total: ${cost_summary['per_user_total']:,.2f}")
        click.echo(f"   Confidence: {results['confidence_level'].title()}")

        if comparative:
            vs_industry = comparative["cost_comparison"]["total_tco_vs_benchmark"]
            click.echo(f"   vs Industry: {vs_industry:+.1f}%")

        # Cost breakdown
        click.echo("\n💰 Cost Breakdown:")
        costs = results["cost_breakdown"]["costs"]
        for category, amount in costs.items():
            percentage = (amount / cost_summary["total_tco"]) * 100
            click.echo(
                f"   {category.replace('_', ' ').title()}: ${amount:,.2f} ({percentage:.1f}%)"
            )

        # Sensitivity analysis
        if sensitivity and results.get("sensitivity_analysis"):
            sens = results["sensitivity_analysis"]
            click.echo(f"\n📊 Sensitivity Analysis:")
            click.echo(f"   Base TCO: ${sens.base_tco:,.2f}")
            click.echo(f"   Best Case: ${sens.best_case_tco:,.2f}")
            click.echo(f"   Worst Case: ${sens.worst_case_tco:,.2f}")
            click.echo(f"   Confidence: {sens.confidence_level.title()}")

        # Export to Excel
        if export:
            click.echo("\n📤 Exporting to Excel...")
            try:
                # Get the saved TCO calculation ID
                tco_calc = (
                    db.session.query(TCOCalculation)
                    .filter_by(
                        vendor_product_id=vendor_product_id,
                        user_count=users,
                        tco_period_years=years,
                    )
                    .first()
                )

                if tco_calc:
                    export_result = engine.export_tco_to_excel(
                        tco_calculation_id=tco_calc.id,
                        include_charts=True,
                        include_sensitivity=sensitivity,
                        include_pivot_tables=True,
                    )

                    click.echo(f"   ✅ Excel file created: {export_result['filename']}")
                    click.echo(f"   📄 Sheets: {export_result['sheets_created']}")
                    click.echo(
                        f"   📊 Charts: {'Yes' if export_result['includes_charts'] else 'No'}"
                    )
                    click.echo(
                        f"   📈 Sensitivity: {'Yes' if export_result['includes_sensitivity'] else 'No'}"
                    )
                    click.echo(
                        f"   📋 Pivot Tables: {'Yes' if export_result['includes_pivot_tables'] else 'No'}"
                    )

                    # Save file
                    import base64

                    excel_data = base64.b64decode(export_result["excel_data"])
                    with open(export_result["filename"], "wb") as f:
                        f.write(excel_data)

                    click.echo(f"   💾 Saved to: {export_result['filename']}")
                else:
                    click.echo("   ❌ No TCO calculation found to export")

            except ImportError as e:
                click.echo(f"   ❌ Export failed: {e}")
                click.echo("   💡 Install openpyxl: pip install openpyxl")
            except Exception as e:
                click.echo(f"   ❌ Export failed: {e}")

        click.echo("\n🎉 TCO calculation completed successfully!")
        return 0

    except Exception as e:
        click.echo(f"❌ TCO calculation failed: {e}")
        return 1


@click.command()
@click.option("--vendor-ids", required=True, help="Comma-separated vendor product IDs")
@click.option("--users", required=True, type=int, help="Number of users")
@click.option("--years", default=5, type=int, help="TCO period in years")
@click.option(
    "--deployment",
    default="cloud",
    type=click.Choice(["cloud", "on-premise", "hybrid"]),
    help="Deployment model",
)
@click.option(
    "--org-size",
    default="medium",
    type=click.Choice(["small", "medium", "large", "enterprise"]),
    help="Organization size",
)
@click.option("--industry", default="manufacturing", type=str, help="Industry sector")
@click.option("--export", is_flag=True, help="Export comparison to Excel")
@with_appcontext
def batch_tco(vendor_ids, users, years, deployment, org_size, industry, export):
    """Calculate TCO for multiple vendor products."""

    click.echo("🔄 Batch TCO Calculation")
    click.echo("=" * 40)

    try:
        # Parse vendor IDs
        vendor_product_ids = [int(id.strip()) for id in vendor_ids.split(",")]

        click.echo(f"📊 Calculating TCO for {len(vendor_product_ids)} vendor products")
        click.echo(f"👥 Users: {users:,}")
        click.echo(f"⏱️  Period: {years} years")
        click.echo(f"☁️  Deployment: {deployment}")
        click.echo(f"🏢 Organization: {org_size}")
        click.echo(f"🏭 Industry: {industry}")
        click.echo(f"📤 Export: {'Yes' if export else 'No'}")
        click.echo()

        # Initialize TCO engine
        engine = AdvancedTCOEngine()

        results = []
        successful = 0
        failed = 0

        for vendor_product_id in vendor_product_ids:
            try:
                click.echo(f"📈 Processing vendor product {vendor_product_id}...")

                result = engine.calculate_comprehensive_tco(
                    vendor_product_id=vendor_product_id,
                    user_count=users,
                    tco_period_years=years,
                    deployment_model=deployment,
                    organization_size=org_size,
                    industry=industry,
                    include_sensitivity_analysis=False,  # Skip for batch
                )

                results.append(
                    {
                        "vendor_product_id": vendor_product_id,
                        "vendor_product_name": result["vendor_product"]["name"],
                        "vendor_name": result["vendor_product"]["vendor_name"],
                        "total_tco": result["cost_breakdown"]["summary"]["total_tco"],
                        "per_user_annual": result["cost_breakdown"]["summary"]["per_user_annual"],
                        "vs_industry_percentage": result["comparative_metrics"]["cost_comparison"][
                            "total_tco_vs_benchmark"
                        ],
                        "confidence_level": result["confidence_level"],
                    }
                )

                successful += 1
                click.echo(
                    f"   ✅ {result['vendor_product']['name']}: ${result['cost_breakdown']['summary']['total_tco']:,.2f}"
                )

            except Exception as e:
                failed += 1
                click.echo(f"   ❌ Vendor product {vendor_product_id}: {e}")

        # Sort by total TCO
        results.sort(key=lambda x: x["total_tco"])

        # Display comparison
        click.echo(f"\n📊 TCO Comparison Results:")
        click.echo(f"   Successful: {successful}")
        click.echo(f"   Failed: {failed}")
        click.echo()

        if results:
            click.echo("🏆 Ranking by Total TCO:")
            for i, result in enumerate(results, 1):
                click.echo(f"   {i}. {result['vendor_product_name']} ({result['vendor_name']})")
                click.echo(f"      TCO: ${result['total_tco']:,.2f}")
                click.echo(f"      Per User: ${result['per_user_annual']:,.2f}")
                click.echo(f"      vs Industry: {result['vs_industry_percentage']:+.1f}%")
                click.echo()

        # Export comparison
        if export and results:
            click.echo("📤 Exporting comparison to Excel...")
            try:
                # Create comparison Excel file
                import base64
                import io

                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill

                wb = Workbook()
                wb.remove(wb.active)
                sheet = wb.create_sheet("TCO Comparison")

                # Title
                sheet["A1"] = "TCO Comparison Analysis"
                sheet["A1"].font = Font(size=16, bold=True)
                sheet["A1"].alignment = Alignment(horizontal="center")
                sheet.merge_cells("A1:G1")

                # Headers
                headers = [
                    "Rank",
                    "Product",
                    "Vendor",
                    "Total TCO",
                    "Per User Annual",
                    "vs Industry",
                    "Confidence",
                ]
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )

                # Data
                row = 4
                for i, result in enumerate(results, 1):
                    sheet.cell(row=row, column=1, value=i)
                    sheet.cell(row=row, column=2, value=result["vendor_product_name"])
                    sheet.cell(row=row, column=3, value=result["vendor_name"])
                    sheet.cell(row=row, column=4, value=result["total_tco"])
                    sheet.cell(row=row, column=4).number_format = "$#,##0"
                    sheet.cell(row=row, column=5, value=result["per_user_annual"])
                    sheet.cell(row=row, column=5).number_format = "$#,##0"
                    sheet.cell(row=row, column=6, value=f"{result['vs_industry_percentage']:+.1f}%")
                    sheet.cell(row=row, column=7, value=result["confidence_level"].title())
                    row += 1

                # Summary
                row += 2
                sheet[f"A{row}"] = "Summary"
                sheet[f"A{row}"].font = Font(size=14, bold=True)
                row += 1

                tco_values = [r["total_tco"] for r in results]
                sheet[f"A{row}"] = "Lowest TCO:"
                sheet[f"B{row}"] = min(tco_values)
                sheet[f"B{row}"].number_format = "$#,##0"
                row += 1

                sheet[f"A{row}"] = "Highest TCO:"
                sheet[f"B{row}"] = max(tco_values)
                sheet[f"B{row}"].number_format = "$#,##0"
                row += 1

                sheet[f"A{row}"] = "Average TCO:"
                sheet[f"B{row}"] = sum(tco_values) / len(tco_values)
                sheet[f"B{row}"].number_format = "$#,##0"
                row += 1

                sheet[f"A{row}"] = "TCO Range:"
                sheet[f"B{row}"] = max(tco_values) - min(tco_values)
                sheet[f"B{row}"].number_format = "$#,##0"

                # Save file
                filename = f"tco_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                wb.save(filename)

                click.echo(f"   ✅ Comparison exported to: {filename}")

            except ImportError as e:
                click.echo(f"   ❌ Export failed: {e}")
                click.echo("   💡 Install openpyxl: pip install openpyxl")
            except Exception as e:
                click.echo(f"   ❌ Export failed: {e}")

        click.echo(f"\n🎉 Batch TCO calculation completed!")
        return 0

    except Exception as e:
        click.echo(f"❌ Batch TCO calculation failed: {e}")
        return 1


@click.command()
@click.option("--tco-id", required=True, type=int, help="TCO calculation ID")
@click.option("--charts", is_flag=True, default=True, help="Include charts")
@click.option("--sensitivity", is_flag=True, default=True, help="Include sensitivity analysis")
@click.option("--pivot", is_flag=True, default=True, help="Include pivot tables")
@with_appcontext
def export_tco(tco_id, charts, sensitivity, pivot):
    """Export TCO calculation to Excel."""

    click.echo("📤 TCO Excel Export")
    click.echo("=" * 40)

    try:
        # Initialize TCO engine
        engine = AdvancedTCOEngine()

        click.echo(f"📊 Exporting TCO calculation {tco_id}")
        click.echo(f"📈 Charts: {'Yes' if charts else 'No'}")
        click.echo(f"📊 Sensitivity: {'Yes' if sensitivity else 'No'}")
        click.echo(f"📋 Pivot Tables: {'Yes' if pivot else 'No'}")
        click.echo()

        # Export TCO
        export_result = engine.export_tco_to_excel(
            tco_calculation_id=tco_id,
            include_charts=charts,
            include_sensitivity=sensitivity,
            include_pivot_tables=pivot,
        )

        # Save file
        import base64

        excel_data = base64.b64decode(export_result["excel_data"])
        with open(export_result["filename"], "wb") as f:
            f.write(excel_data)

        click.echo("✅ Export completed successfully!")
        click.echo(f"📄 Filename: {export_result['filename']}")
        click.echo(f"📊 File size: {export_result['file_size']:,} bytes")
        click.echo(f"📋 Sheets created: {export_result['sheets_created']}")
        click.echo(f"📈 Charts included: {'Yes' if export_result['includes_charts'] else 'No'}")
        click.echo(
            f"📊 Sensitivity analysis: {'Yes' if export_result['includes_sensitivity'] else 'No'}"
        )
        click.echo(f"📋 Pivot tables: {'Yes' if export_result['includes_pivot_tables'] else 'No'}")
        click.echo(f"💾 Saved to: {export_result['filename']}")

        return 0

    except ImportError as e:
        click.echo(f"❌ Export failed: {e}")
        click.echo("💡 Install openpyxl: pip install openpyxl")
        return 1
    except Exception as e:
        click.echo(f"❌ Export failed: {e}")
        return 1


@click.command()
@click.option("--industry", help="Filter by industry")
@click.option("--org-size", help="Filter by organization size")
@with_appcontext
def tco_benchmarks(industry, org_size):
    """Display TCO industry benchmarks."""

    click.echo("📊 TCO Industry Benchmarks")
    click.echo("=" * 40)

    try:
        # Initialize TCO engine
        engine = AdvancedTCOEngine()

        click.echo(f"📈 Available Benchmarks:")
        click.echo(f"   Industries: {len(engine.INDUSTRY_BENCHMARKS)}")
        click.echo(f"   Organization Sizes: 3 per industry")
        click.echo()

        # Display benchmarks
        for ind_name, ind_benchmarks in engine.INDUSTRY_BENCHMARKS.items():
            if industry and ind_name != industry:
                continue

            click.echo(f"🏭 {ind_name.title()}:")
            for org_size, benchmark in ind_benchmarks.items():
                if org_size and org_size != org_size:
                    continue

                click.echo(f"   📏 {org_size.title()}:")
                click.echo(f"      Median TCO/User: ${benchmark.median_tco_per_user:,.2f}")
                click.echo(f"      Implementation: {benchmark.implementation_months} months")
                click.echo(f"      Cost Distribution:")
                for category, percentage in benchmark.cost_distribution.items():
                    click.echo(f"        {category}: {percentage}%")
                click.echo()

        return 0

    except Exception as e:
        click.echo(f"❌ Failed to get benchmarks: {e}")
        return 1


@click.command()
@click.option("--limit", default=10, type=int, help="Number of recent calculations")
@with_appcontext
def tco_history(limit):
    """Display recent TCO calculations."""

    click.echo("📚 TCO Calculation History")
    click.echo("=" * 40)

    try:
        # Get recent calculations
        calculations = (
            db.session.query(TCOCalculation)
            .order_by(TCOCalculation.created_at.desc())
            .limit(limit)
            .all()
        )

        if not calculations:
            click.echo("📝 No TCO calculations found.")
            return 0

        click.echo(f"📊 Recent {len(calculations)} calculations:")
        click.echo()

        for calc in calculations:
            vendor_product = (
                db.session.query(VendorProduct).filter_by(id=calc.vendor_product_id).first()
            )
            product_name = vendor_product.name if vendor_product else "Unknown"
            vendor_name = (
                vendor_product.vendor_organization.name
                if vendor_product and vendor_product.vendor_organization
                else "Unknown"
            )

            click.echo(f"📈 {product_name} ({vendor_name})")
            click.echo(f"   Users: {calc.user_count:,}")
            click.echo(f"   Period: {calc.tco_period_years} years")
            click.echo(f"   Total TCO: ${calc.total_tco:,.2f}")
            click.echo(f"   Per User: ${calc.per_user_annual:,.2f}")
            click.echo(f"   Confidence: {calc.confidence_level.title()}")
            click.echo(f"   Created: {calc.created_at.strftime('%Y-%m-%d %H:%M')}")
            click.echo()

        return 0

    except Exception as e:
        click.echo(f"❌ Failed to get history: {e}")
        return 1


# Register commands with Flask CLI
def register_commands(cli):
    """Register TCO calculation commands with Flask CLI."""
    cli.add_command(calculate_tco)
    cli.add_command(batch_tco)
    cli.add_command(export_tco)
    cli.add_command(tco_benchmarks)
    cli.add_command(tco_history)
