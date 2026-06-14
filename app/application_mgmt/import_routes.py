"""
Import Routes for Application Management

Handles application data import (CSV, Excel, JSON, manual) and import history.
"""

import csv
import io
import json
import os
import unicodedata
from datetime import datetime

from flask import (
    current_app,
    flash,
    jsonify,
    redirect,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required

from .. import db
from ..models.application_portfolio import ApplicationComponent
from . import application_mgmt
from .routes import (
    INTEGER_RANGE_FIELDS,
    _link_application_to_apqc_by_ids,
    _link_application_to_capabilities,
    _link_application_to_processes,
    _organize_apqc_links_by_row,
    _validate_and_clean_import_row,
    parse_integer_from_range,
)
import logging
logger = logging.getLogger(__name__)


@application_mgmt.route("/applications/import", methods=["POST"])
@login_required
def application_import():
    """Import Application Components from CSV or JSON"""
    if "file" not in request.files:
        flash("No file uploaded", "error")
        return redirect(url_for("unified_applications.application_list"))

    file = request.files["file"]

    if file.filename == "":
        flash("No file selected", "error")
        return redirect(url_for("unified_applications.application_list"))

    # Get import mode: 'skip', 'update', or 'duplicate' (default to skip for safety)
    import_mode = request.form.get("import_mode", "skip")

    try:
        filename = (file.filename or "").lower()

        def _clean(value):
            if isinstance(value, str):
                value = value.strip()
                return value or None
            return value

        def _parse_user_count(raw_value):
            if raw_value in (None, ""):
                return None

            if isinstance(raw_value, bool):
                return int(raw_value)

            if isinstance(raw_value, (int, float)):
                try:
                    return int(raw_value)
                except (TypeError, ValueError):
                    return None

            cleaned = _clean(raw_value)
            if cleaned is None:
                return None

            sanitized = cleaned.replace(",", "")
            if sanitized.isdigit():
                return int(sanitized)

            matches = re.findall(r"\d+", sanitized)
            if len(matches) == 1:
                return int(matches[0])

            return None

        if filename.endswith(".csv"):
            # Parse CSV
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)

            # Skip leading blank lines that break DictReader header detection
            while True:
                pos = stream.tell()
                line = stream.readline()
                if not line:
                    stream.seek(0)
                    break
                if line.strip():
                    stream.seek(pos)
                    break

            reader = csv.DictReader(stream)

            if not reader.fieldnames:
                raise ValueError("CSV file is missing headers.")

            # Check for Name column using flexible matching
            name_column_found = any(
                alias in reader.fieldnames
                for alias in IMPORT_COLUMN_ALIASES.get("name", ["Name"])
            )
            if not name_column_found:
                raise ValueError(
                    "CSV file must include a 'Name' (or 'Application Name', 'App Name') column."
                )

            count = 0
            skipped = 0
            updated = 0
            failed = 0
            non_numeric_user_counts = []
            row_errors = []
            # Track names processed in THIS import to detect duplicates within the same CSV file
            processed_names = set()
            # Track capability and process linking statistics
            capabilities_linked = 0
            processes_linked = 0
            capabilities_not_found = []
            processes_not_found = []

            # Prefetch all existing applications to avoid N+1 query in import loop
            existing_apps_list = ApplicationComponent.query.all()
            existing_apps_by_name = {
                app.name.lower(): app for app in existing_apps_list if app.name
            }

            for index, row in enumerate(
                reader, start=2
            ):  # start=2 accounts for header row
                if not any(
                    (value or "").strip() for value in row.values() if value is not None
                ):
                    continue  # skip empty rows

                try:
                    # Use flexible column mapping and validation
                    cleaned = _validate_and_clean_import_row(row, _clean)
                    name = cleaned["name"]

                    if not name:
                        row_errors.append((index, "'Name' is required"))
                        failed += 1
                        continue

                    # Normalize name for duplicate checking (case-insensitive)
                    name_lower = name.lower()

                    # Check if this name was already processed in THIS import (within same CSV)
                    if name_lower in processed_names:
                        if import_mode == "skip":
                            skipped += 1
                            continue
                        elif import_mode != "duplicate":
                            # For 'update' mode, we already updated it earlier in this import
                            skipped += 1
                            continue
                        # For 'duplicate' mode, allow it to continue and create duplicate

                    # Parse user count
                    user_count = _parse_user_count(cleaned["user_count_raw"])
                    if user_count is None and cleaned["user_count_raw"]:
                        non_numeric_user_counts.append(
                            (index, str(cleaned["user_count_raw"]))
                        )

                    # Check if application exists (case-insensitive match)
                    # Use prefetched lookup instead of query to avoid N+1
                    existing_app = existing_apps_by_name.get(name.lower())

                    # Debug
                    current_app.logger.debug(
                        f"Row {index}: name='{name}', existing={existing_app is not None}, mode={import_mode}"
                    )

                    if existing_app and import_mode == "skip":
                        skipped += 1
                        processed_names.add(name_lower)
                        continue
                    elif existing_app and import_mode == "update":
                        # Update existing record with validated/truncated values
                        existing_app.component_type = cleaned["component_type"]
                        existing_app.application_category = cleaned[
                            "application_category"
                        ]
                        existing_app.technology_stack = cleaned["technology_stack"]
                        existing_app.version = cleaned["version"]
                        existing_app.deployment_status = cleaned["deployment_status"]
                        existing_app.business_domain = cleaned["business_domain"]
                        existing_app.business_owner = cleaned["business_owner"]
                        existing_app.development_team = cleaned["development_team"]
                        existing_app.user_count = user_count
                        existing_app.business_criticality = cleaned[
                            "business_criticality"
                        ]

                        # Link to capabilities (merge - adds missing, keeps existing)
                        if cleaned["capabilities"]:
                            cap_result = _link_application_to_capabilities(
                                existing_app, cleaned["capabilities"]
                            )
                            capabilities_linked += cap_result["linked"]
                            capabilities_not_found.extend(cap_result["not_found"])

                        # Link to business processes (merge - adds missing, keeps existing)
                        if cleaned["business_process"]:
                            proc_result = _link_application_to_processes(
                                existing_app, cleaned["business_process"]
                            )
                            processes_linked += proc_result["linked"]
                            processes_not_found.extend(proc_result["not_found"])

                        updated += 1
                        processed_names.add(name_lower)
                    else:
                        # Create new record (duplicate mode or doesn't exist)
                        app = ApplicationComponent(
                            name=name,
                            component_type=cleaned["component_type"],
                            application_category=cleaned["application_category"],
                            technology_stack=cleaned["technology_stack"],
                            version=cleaned["version"],
                            deployment_status=cleaned["deployment_status"],
                            business_domain=cleaned["business_domain"],
                            business_owner=cleaned["business_owner"],
                            development_team=cleaned["development_team"],
                            user_count=user_count,
                            business_criticality=cleaned["business_criticality"],
                        )
                        db.session.add(app)
                        db.session.flush()  # Flush to get app.id for linking

                        # Link to capabilities
                        if cleaned["capabilities"]:
                            cap_result = _link_application_to_capabilities(
                                app, cleaned["capabilities"]
                            )
                            capabilities_linked += cap_result["linked"]
                            capabilities_not_found.extend(cap_result["not_found"])

                        # Link to business processes
                        if cleaned["business_process"]:
                            proc_result = _link_application_to_processes(
                                app, cleaned["business_process"]
                            )
                            processes_linked += proc_result["linked"]
                            processes_not_found.extend(proc_result["not_found"])

                        count += 1
                        processed_names.add(name_lower)

                except Exception as row_error:
                    current_app.logger.error(
                        f"Error processing row {index}: {str(row_error)}"
                    )
                    row_errors.append((index, str(row_error)))
                    failed += 1
                    # Continue processing next row instead of failing entire import
                    continue

            db.session.commit()

            # Build success message
            msg_parts = []
            if count > 0:
                msg_parts.append(f"{count} created")
            if updated > 0:
                msg_parts.append(f"{updated} updated")
            if skipped > 0:
                msg_parts.append(f"{skipped} skipped")
            if failed > 0:
                msg_parts.append(f"{failed} failed")

            if msg_parts:
                flash(
                    f"Import complete: {', '.join(msg_parts)}",
                    "success" if failed == 0 else "warning",
                )
            else:
                flash("Import complete: No rows processed", "info")

            # Report capability and process linking results
            if capabilities_linked > 0 or processes_linked > 0:
                link_parts = []
                if capabilities_linked > 0:
                    link_parts.append(f"{capabilities_linked} capability mappings")
                if processes_linked > 0:
                    link_parts.append(f"{processes_linked} process mappings")
                flash(f"Linked: {', '.join(link_parts)}", "info")

            # Report capabilities/processes not found
            if capabilities_not_found:
                unique_caps = list(set(capabilities_not_found))[:5]
                preview = ", ".join(unique_caps)
                if len(set(capabilities_not_found)) > 5:
                    preview += ", ..."
                flash(f"Capabilities not found in database: {preview}", "warning")

            if processes_not_found:
                unique_procs = list(set(processes_not_found))[:5]
                preview = ", ".join(unique_procs)
                if len(set(processes_not_found)) > 5:
                    preview += ", ..."
                flash(f"Business processes not found in database: {preview}", "warning")

            if non_numeric_user_counts:
                preview = ", ".join(
                    f"row {row_num} ('{value}')"
                    for row_num, value in non_numeric_user_counts[:5]
                )
                if len(non_numeric_user_counts) > 5:
                    preview += ", ..."
                flash(
                    "Some user counts were non-numeric and were left blank: " + preview,
                    "warning",
                )

            # Report row-level errors
            if row_errors:
                preview = ", ".join(
                    f"row {row_num}: {error[:50]}" for row_num, error in row_errors[:3]
                )
                if len(row_errors) > 3:
                    preview += f", ... ({len(row_errors)} total errors)"
                flash(f"Row errors: {preview}", "error")

        elif filename.endswith(".json"):
            # Parse JSON
            data = json.load(file.stream)

            if not isinstance(data, list):
                raise ValueError("JSON payload must be a list of application objects.")

            count = 0
            skipped = 0
            updated = 0
            failed = 0
            row_errors = []
            # Track names processed in THIS import to detect duplicates within the same JSON file
            processed_names = set()
            # Track capability and process linking statistics
            capabilities_linked = 0
            processes_linked = 0
            capabilities_not_found = []
            processes_not_found = []

            # Prefetch all existing applications to avoid N+1 query in import loop
            existing_apps_list = ApplicationComponent.query.all()
            existing_apps_by_name = {
                app.name.lower(): app for app in existing_apps_list if app.name
            }

            for index, item in enumerate(data, start=1):
                try:
                    # Use flexible column mapping and validation
                    cleaned = _validate_and_clean_import_row(item, _clean)
                    name = cleaned["name"]

                    if not name:
                        row_errors.append((index, "'name' is required"))
                        failed += 1
                        continue

                    # Normalize name for duplicate checking (case-insensitive)
                    name_lower = name.lower()

                    # Check if this name was already processed in THIS import (within same JSON)
                    if name_lower in processed_names:
                        if import_mode == "skip":
                            skipped += 1
                            continue
                        elif import_mode != "duplicate":
                            # For 'update' mode, we already updated it earlier in this import
                            skipped += 1
                            continue
                        # For 'duplicate' mode, allow it to continue and create duplicate

                    # Check if application exists (case-insensitive match)
                    # Use prefetched lookup instead of query to avoid N+1
                    existing_app = existing_apps_by_name.get(name.lower())

                    if existing_app and import_mode == "skip":
                        skipped += 1
                        processed_names.add(name_lower)
                        continue
                    elif existing_app and import_mode == "update":
                        # Update existing record with validated/truncated values
                        existing_app.component_type = cleaned["component_type"]
                        existing_app.application_category = cleaned[
                            "application_category"
                        ]
                        existing_app.technology_stack = cleaned["technology_stack"]
                        existing_app.version = cleaned["version"]
                        existing_app.deployment_status = cleaned["deployment_status"]
                        existing_app.business_domain = cleaned["business_domain"]
                        existing_app.business_owner = cleaned["business_owner"]
                        existing_app.development_team = cleaned["development_team"]
                        existing_app.user_count = _parse_user_count(
                            cleaned["user_count_raw"]
                        )
                        existing_app.business_criticality = cleaned[
                            "business_criticality"
                        ]

                        # Link to capabilities (merge - adds missing, keeps existing)
                        capabilities_str = cleaned["capabilities"] or _clean(
                            item.get("business_functions")
                        )
                        if capabilities_str:
                            cap_result = _link_application_to_capabilities(
                                existing_app, capabilities_str
                            )
                            capabilities_linked += cap_result["linked"]
                            capabilities_not_found.extend(cap_result["not_found"])

                        # Link to business processes (merge - adds missing, keeps existing)
                        processes_str = cleaned["business_process"] or _clean(
                            item.get("business_processes")
                        )
                        if processes_str:
                            proc_result = _link_application_to_processes(
                                existing_app, processes_str
                            )
                            processes_linked += proc_result["linked"]
                            processes_not_found.extend(proc_result["not_found"])

                        updated += 1
                        processed_names.add(name_lower)
                    else:
                        # Create new record (duplicate mode or doesn't exist)
                        app = ApplicationComponent(
                            name=name,
                            component_type=cleaned["component_type"],
                            application_category=cleaned["application_category"],
                            technology_stack=cleaned["technology_stack"],
                            version=cleaned["version"],
                            deployment_status=cleaned["deployment_status"],
                            business_domain=cleaned["business_domain"],
                            business_owner=cleaned["business_owner"],
                            development_team=cleaned["development_team"],
                            user_count=_parse_user_count(cleaned["user_count_raw"]),
                            business_criticality=cleaned["business_criticality"],
                        )
                        db.session.add(app)
                        db.session.flush()  # Flush to get app.id for linking

                        # Link to capabilities
                        capabilities_str = cleaned["capabilities"] or _clean(
                            item.get("business_functions")
                        )
                        if capabilities_str:
                            cap_result = _link_application_to_capabilities(
                                app, capabilities_str
                            )
                            capabilities_linked += cap_result["linked"]
                            capabilities_not_found.extend(cap_result["not_found"])

                        # Link to business processes
                        processes_str = cleaned["business_process"] or _clean(
                            item.get("business_processes")
                        )
                        if processes_str:
                            proc_result = _link_application_to_processes(
                                app, processes_str
                            )
                            processes_linked += proc_result["linked"]
                            processes_not_found.extend(proc_result["not_found"])

                        count += 1
                        processed_names.add(name_lower)

                except Exception as row_error:
                    current_app.logger.error(
                        f"Error processing record {index}: {str(row_error)}"
                    )
                    row_errors.append((index, str(row_error)))
                    failed += 1
                    continue

            db.session.commit()

            # Build success message
            msg_parts = []
            if count > 0:
                msg_parts.append(f"{count} created")
            if updated > 0:
                msg_parts.append(f"{updated} updated")
            if skipped > 0:
                msg_parts.append(f"{skipped} skipped")
            if failed > 0:
                msg_parts.append(f"{failed} failed")

            if msg_parts:
                flash(
                    f"Import complete: {', '.join(msg_parts)}",
                    "success" if failed == 0 else "warning",
                )
            else:
                flash("Import complete: No records processed", "info")

            # Report capability and process linking results
            if capabilities_linked > 0 or processes_linked > 0:
                link_parts = []
                if capabilities_linked > 0:
                    link_parts.append(f"{capabilities_linked} capability mappings")
                if processes_linked > 0:
                    link_parts.append(f"{processes_linked} process mappings")
                flash(f"Linked: {', '.join(link_parts)}", "info")

            # Report capabilities/processes not found
            if capabilities_not_found:
                unique_caps = list(set(capabilities_not_found))[:5]
                preview = ", ".join(unique_caps)
                if len(set(capabilities_not_found)) > 5:
                    preview += ", ..."
                flash(f"Capabilities not found in database: {preview}", "warning")

            if processes_not_found:
                unique_procs = list(set(processes_not_found))[:5]
                preview = ", ".join(unique_procs)
                if len(set(processes_not_found)) > 5:
                    preview += ", ..."
                flash(f"Business processes not found in database: {preview}", "warning")

            # Report row-level errors
            if row_errors:
                preview = ", ".join(
                    f"record {row_num}: {error[:50]}"
                    for row_num, error in row_errors[:3]
                )
                if len(row_errors) > 3:
                    preview += f", ... ({len(row_errors)} total errors)"
                flash(f"Record errors: {preview}", "error")

        else:
            flash("Invalid file format. Please upload CSV or JSON", "error")

    except ValueError as exc:
        db.session.rollback()
        flash(f"Invalid import data: {exc}", "error")
    except Exception as e:
        db.session.rollback()
        flash("Error importing file. Please try again.", "error")

    return redirect(url_for("unified_applications.application_list"))



@application_mgmt.route("/applications/import-template/<format>")
@login_required
def application_import_template(format):
    """Provide downloadable templates for application import."""
    headers = [
        "Name",
        "Type",
        "Category",
        "Technology Stack",
        "Version",
        "Status",
        "Business Domain",
        "Owner",
        "Team",
        "Users",
        "Criticality",
        "Capabilities",
        "Business Process",
    ]

    sample_row = [
        "Customer Portal",
        "Web Application",
        "Enterprise Application",
        "React, Flask, PostgreSQL",
        "1.0",
        "production",
        "Digital Channels",
        "Jordan Lee",
        "Experience Team",
        "2500",
        "High",
        "Customer Management, Order Management",
        "Order-to-Cash",
    ]

    if format.lower() == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerow(sample_row)
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name="application_import_template.csv",
        )

    if format.lower() == "json":
        template = [
            {
                "name": sample_row[0],
                "component_type": sample_row[1],
                "application_category": sample_row[2],
                "technology_stack": "React, Flask, PostgreSQL",
                "version": sample_row[4],
                "deployment_status": sample_row[5],
                "business_domain": sample_row[6],
                "business_owner": sample_row[7],
                "development_team": sample_row[8],
                "user_count": int(sample_row[9]),
                "business_criticality": sample_row[10],
                "capabilities": sample_row[11],
                "business_process": sample_row[12],
            }
        ]
        payload = json.dumps(template, indent=2)
        return send_file(
            io.BytesIO(payload.encode("utf-8")),
            mimetype="application/json",
            as_attachment=True,
            download_name="application_import_template.json",
        )

    flash("Invalid template format requested", "error")
    return redirect(url_for("unified_applications.application_list"))



@application_mgmt.route("/applications/import-fields", methods=["GET"])
@login_required
def get_import_fields():
    """
    Return all valid ApplicationComponent model fields for import mapping.
    This allows the frontend to dynamically show available fields based on the actual model.
    """
    from sqlalchemy import inspect

    # Get model columns via SQLAlchemy inspection
    mapper = inspect(ApplicationComponent)
    columns = mapper.columns

    # Define field categories for better UX organization
    field_categories = {
        "identity": ["name", "description", "application_code"],
        "classification": [
            "component_type",
            "application_type",
            "application_category",
            "deployment_model",
            "deployment_status",
            "criticality",
            "business_criticality",
        ],
        "capabilities_processes": [
            "archimate_capability",  # For "Application Capability (Archimate)" column
            "functionality_capabilities",  # For "Functionality (Capabilities)" column - contains APQC PCF
        ],
        "business": [
            "business_domain",
            "business_purpose",
            "business_functions",
            "user_base_size",
            "user_types",
        ],
        "strategic": [
            "strategic_importance",
            "business_value",
            "competitive_advantage",
            "differentiation_level",
        ],
        "technology": [
            "technology_stack",
            "programming_languages",
            "database_platforms",
            "integration_methods",
            "api_available",
            "api_documentation",
            "version",
        ],
        "financial": [
            "total_cost_of_ownership",
            "license_cost",
            "maintenance_cost",
            "infrastructure_cost",
            "support_cost",
            "implementation_cost",
            "roi_score",
        ],
        "vendor": [
            "vendor_name",
            "package_vendor",  # For "Package Vendor" column
            "package_name",  # For "Package Name" column - used for vendor inference
            "vendor_type",
            "contract_type",
            "contract_expiry_date",
            "support_level",
        ],
        "lifecycle": [
            "lifecycle_status",
            "implementation_date",
            "last_major_upgrade",
            "planned_retirement_date",
            "technology_age_years",
        ],
        "governance": [
            "application_owner",
            "business_owner",
            "technical_owner",
            "product_manager",
            "development_team",
            "support_team",
        ],
        "risk": ["technical_risk", "business_risk", "vendor_risk", "obsolescence_risk"],
        "other": ["notes", "assessment_notes"],
    }

    # Build human-readable labels from field names
    def make_label(field_name):
        # Special cases
        labels = {
            "name": "Name *",
            "application_code": "APP ID / Application Code",
            "business_criticality": "Business Criticality",
            "total_cost_of_ownership": "Total Cost of Ownership",
            "user_base_size": "Number of Users",
            "roi_score": "ROI Score",
            "rpo_hours": "RPO (Hours)",
            "rto_hours": "RTO (Hours)",
            "pii_data_processed": "PII Data Processed",
            "gdpr_compliant": "GDPR Compliant",
        }
        if field_name in labels:
            return labels[field_name]
        # Convert snake_case to Title Case
        return field_name.replace("_", " ").title()

    # Build field list with metadata
    fields = [{"value": "", "label": "-- Skip Column --", "category": "skip"}]

    # Special fields that are used for import analysis but may not be model columns
    # These are virtual fields for ArchiMate, APQC PCF, and Vendor mapping
    special_import_fields = {
        "archimate_capability": {
            "label": "ArchiMate Capability (for linking)",
            "category": "capabilities_processes",
        },
        "functionality_capabilities": {
            "label": "APQC PCF Processes (Capabilities)",
            "category": "capabilities_processes",
        },
        "package_vendor": {
            "label": "Package Vendor",
            "category": "vendor",
        },
        "package_name": {
            "label": "Package Name (for vendor inference)",
            "category": "vendor",
        },
    }

    # Excluded fields (auto-generated, relationships, internal)
    excluded = {
        "id",
        "created_at",
        "updated_at",
        "discovered_by_ai",
        "discovery_confidence",
        "last_assessed",
        "archimate_element_id",
    }

    model_column_names = [c.name for c in columns]

    # Add fields by category for organized display
    for category, field_names in field_categories.items():
        for field_name in field_names:
            if field_name in excluded:
                continue
            # Check if it's a model column or a special import field
            if field_name in model_column_names:
                fields.append(
                    {
                        "value": field_name,
                        "label": make_label(field_name),
                        "category": category,
                        "required": field_name == "name",
                    }
                )
            elif field_name in special_import_fields:
                # Add special import fields (virtual fields for ArchiMate, APQC, Vendor)
                fields.append(
                    {
                        "value": field_name,
                        "label": special_import_fields[field_name]["label"],
                        "category": special_import_fields[field_name]["category"],
                        "required": False,
                    }
                )

    # Add any remaining model fields not in categories
    categorized = set()
    for cat_fields in field_categories.values():
        categorized.update(cat_fields)

    for col in columns:
        if col.name not in excluded and col.name not in categorized:
            fields.append(
                {
                    "value": col.name,
                    "label": make_label(col.name),
                    "category": "other",
                    "required": False,
                }
            )

    # Build auto-mapping aliases based on field names
    aliases = {}
    alias_patterns = {
        "name": ["name", "application name", "app name", "appname", "app_name"],
        "application_code": [
            "app id",
            "application id",
            "app code",
            "application_code",
            "appid",
        ],
        "description": ["description", "app description", "details"],
        "deployment_status": [
            "application status",
            "deployment status",
            "status",
            "deploy status",
        ],
        "application_category": ["category", "application category", "app category"],
        "business_criticality": ["business criticality", "criticality", "priority"],
        "lifecycle_status": ["lifecycle status", "lifecycle", "life cycle"],
        "vendor_name": ["vendor", "vendor name", "supplier"],
        "package_vendor": ["package vendor"],  # Explicit Package Vendor column
        "package_name": ["package name", "product name"],  # For vendor inference
        "business_owner": ["business owner", "app business owner", "owner"],
        "technology_stack": ["technology stack", "tech stack", "technologies"],
        "user_base_size": ["target users", "user count", "number of users", "users"],
        "total_cost_of_ownership": [
            "total run cost",
            "total cost",
            "tco",
            "annual cost",
        ],
        "notes": ["comments", "remarks", "notes"],
        "version": ["version", "app version"],
        "support_level": ["support level", "support hours"],
        "programming_languages": ["programming languages", "languages"],
        # ArchiMate and Capabilities/Processes mapping
        "archimate_capability": [
            "application capability (archimate)",
            "application capability(archimate)",  # Without space before parenthesis
            "application capability",
            "archimate",
            "archimate capability",
            "archimate element",
        ],
        "functionality_capabilities": [
            "functionality (capabilities)",
            "functionality(capabilities)",  # Without space before parenthesis
            "functionality",
            "capabilities",
            "pcf",
            "apqc",
            "apqc pcf",
            "pcf process",
            "processes",
            "business capabilities",
            "functional capabilities",
        ],
    }

    return jsonify({"fields": fields, "aliases": alias_patterns}), 200


@application_mgmt.route("/applications/preview-excel", methods=["POST"])
@login_required
def preview_excel_applications():
    """Preview Excel, CSV, or JSON file without importing"""
    import csv
    import io
    import json

    import openpyxl

    if "file" not in request.files:
        current_app.logger.warning("Preview: No file in request")
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        current_app.logger.warning("Preview: Empty filename")
        return jsonify({"error": "No file selected"}), 400

    filename = (file.filename or "").lower()
    current_app.logger.info(f"Preview: Processing file {file.filename}")

    try:
        headers = []
        rows = []
        total_rows = 0

        if filename.endswith((".xlsx", ".xls")):
            # Parse Excel file with data_only mode to handle corrupted stylesheets
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
            except ValueError as ve:
                # Handle corrupted stylesheet errors by reading without styling
                if "stylesheet" in str(ve).lower() or "color" in str(ve).lower():
                    file.seek(0)  # Reset file pointer
                    from warnings import filterwarnings

                    from openpyxl import load_workbook

                    filterwarnings("ignore", category=UserWarning, module="openpyxl")
                    # Try with read_only mode which is more permissive
                    wb = load_workbook(file, data_only=True, read_only=True)
                else:
                    raise
            ws = wb.active
            headers = [
                str(cell.value).strip() if cell.value else ""
                for cell in ws[1]
                if cell.value
            ]
            for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
                if any(cell for cell in row if cell):
                    rows.append(
                        [
                            str(cell).strip() if cell else ""
                            for cell in row[: len(headers)]
                        ]
                    )
            total_rows = ws.max_row - 1

        elif filename.endswith(".csv"):
            # Parse CSV file
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.reader(stream)
            all_rows = list(reader)
            if all_rows:
                headers = [str(h).strip() for h in all_rows[0] if h]
                for row in all_rows[1:11]:
                    if any(cell for cell in row if cell):
                        rows.append(
                            [
                                str(cell).strip() if cell else ""
                                for cell in row[: len(headers)]
                            ]
                        )
                total_rows = len(all_rows) - 1

        elif filename.endswith(".json"):
            # Parse JSON file
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list) and len(data) > 0:
                # Extract headers from first object
                first_item = data[0] if isinstance(data[0], dict) else {}
                headers = list(first_item.keys())
                for item in data[:10]:
                    if isinstance(item, dict):
                        rows.append([str(item.get(h, "")).strip() for h in headers])
                total_rows = len(data)
            else:
                return jsonify(
                    {"error": "Invalid JSON structure. Expected array of objects."}
                ), 400

        else:
            return (
                jsonify(
                    {
                        "error": "Unsupported file format. Use .xlsx, .xls, .csv, or .json"
                    }
                ),
                400,
            )

        current_app.logger.info(
            f"Preview: Returning {len(rows)} preview rows, {total_rows} total"
        )

        return jsonify(
            {"headers": headers, "rows": rows, "total_rows": total_rows}
        ), 200

    except Exception as e:
        current_app.logger.error(f"Error previewing file: {e}", exc_info=True)
        return jsonify({"error": "An internal error occurred"}), 500


@application_mgmt.route("/applications/analyze-import", methods=["POST"])
@login_required
def analyze_import():
    """
    Analyze an import file without actually importing.
    Returns counts of what will be created, updated, or skipped.
    Includes ArchiMate, APQC PCF, and Vendor analysis.
    """
    import csv
    import io
    import json
    import unicodedata

    import openpyxl

    def clean_unicode_text(text):
        """Clean Unicode text to prevent encoding errors"""
        if not text:
            return ""

        # Normalize Unicode characters
        normalized = unicodedata.normalize("NFKC", str(text))

        # Replace problematic characters with safe alternatives
        replacements = {
            "\u2011": "-",  # Non-breaking hyphen
            "\u2013": "-",  # En dash
            "\u2014": "-",  # Em dash
            "\u2018": "'",  # Left single quote
            "\u2019": "'",  # Right single quote
            "\u201c": '"',  # Left double quote
            "\u201d": '"',  # Right double quote
            "\u2026": "...",  # Ellipsis
            "\xa0": " ",  # Non-breaking space
        }

        for unicode_char, replacement in replacements.items():
            normalized = normalized.replace(unicode_char, replacement)

        # Remove any remaining control characters except newlines and tabs
        cleaned = "".join(
            char for char in normalized if ord(char) >= 32 or char in "\n\t"
        )

        return cleaned

    try:
        # Import AI Import Service for comprehensive AI analysis
        from ..services.ai_import_service import get_ai_import_service

        ai_service = get_ai_import_service()
        current_app.logger.info("AI Import Service loaded for comprehensive analysis")
    except ImportError as e:
        current_app.logger.error(f"Failed to import AI Import Service: {e}")
        ai_service = None

    # Import APQC Hierarchy Service for enhanced process matching
    try:
        from ..services.apqc_hierarchy_service import APQCHierarchyService

        apqc_hierarchy = APQCHierarchyService()
        current_app.logger.info(
            "APQC Hierarchy Service loaded for enhanced process matching"
        )
    except ImportError as e:
        current_app.logger.error(f"Failed to import APQC Hierarchy Service: {e}")
        apqc_hierarchy = None

    # Import models for analysis
    from ..models.apqc_process import APQCProcess
    from ..models.archimate_core import ArchiMateElement

    try:
        from ..models.vendor.vendor_organization import VendorOrganization
    except ImportError:
        VendorOrganization = None

    # Ensure we have a clean session state - proper PostgreSQL transaction cleanup
    try:
        # Force rollback to clear any failed transaction state
        db.session.rollback()
    except Exception as e:  # fabricated-values-ok
        logger.debug(f"Ignored: {e}")  # Ignore rollback errors

    # Start a fresh transaction
    try:
        db.session.begin()
    except Exception as e:  # fabricated-values-ok
        logger.debug(f"Ignored: {e}")  # If begin fails, continue with existing session

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = (file.filename or "").lower()
    duplicate_mode = request.form.get("duplicate_mode", "merge")

    # Get custom field mappings from frontend
    custom_mappings = {}
    field_mappings_json = request.form.get("field_mappings")
    if field_mappings_json:
        try:
            custom_mappings = json.loads(field_mappings_json)
        except json.JSONDecodeError:
            logger.exception("Failed to parse JSON response")

    try:
        headers = []
        data_rows = []

        # Parse file based on type
        if filename.endswith(".csv"):
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)
            headers = reader.fieldnames or []
            data_rows = list(reader)

            # Clean headers and data - IMPORTANT: strip whitespace to ensure keys match
            headers = [clean_unicode_text(h).strip() for h in headers]
            data_rows = [
                {
                    clean_unicode_text(k).strip(): clean_unicode_text(v)
                    for k, v in row.items()
                }
                for row in data_rows
            ]

        elif filename.endswith(".json"):
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list) and len(data) > 0:
                first_item = data[0] if isinstance(data[0], dict) else {}
                headers = list(first_item.keys())
                data_rows = data

                # Clean headers and data for JSON - strip whitespace for consistency
                headers = [clean_unicode_text(h).strip() for h in headers]
                data_rows = [
                    {
                        clean_unicode_text(k).strip(): clean_unicode_text(v)
                        for k, v in row.items()
                    }
                    for row in data_rows
                    if isinstance(row, dict)
                ]

        else:
            # Excel file - handle corrupted stylesheets gracefully
            try:
                wb = openpyxl.load_workbook(file)
            except ValueError as ve:
                # Handle corrupted stylesheet/color issues
                if "stylesheet" in str(ve).lower() or "color" in str(ve).lower():
                    file.seek(0)
                    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                else:
                    raise
            ws = wb.active
            headers = [
                clean_unicode_text(str(cell.value).strip()) if cell.value else ""
                for cell in ws[1]
                if cell.value
            ]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row if cell):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = (
                                clean_unicode_text(row[i]) if row[i] else ""
                            )
                    data_rows.append(row_dict)

        # Find column mappings - comprehensive detection for all important columns
        name_column = None
        app_id_column = None
        vendor_column = None
        package_name_column = None  # For vendor inference from package name
        archimate_column = None  # For "Application Capability (Archimate)"
        pcf_column = None  # For APQC PCF processes
        capabilities_column = None  # For "Functionality (Capabilities)"
        description_column = None

        # Check custom mappings first
        if custom_mappings:
            current_app.logger.info(f"Custom mappings received: {custom_mappings}")
            for header, target_field in custom_mappings.items():
                if target_field == "name":
                    name_column = header
                elif target_field == "application_code":
                    app_id_column = header
                elif target_field == "vendor_name":
                    vendor_column = header
                elif target_field == "package_name":
                    package_name_column = header
                elif target_field == "package_vendor":
                    vendor_column = header
                elif target_field in [
                    "_functional_capabilities",
                    "pcf_process",
                    "functionality_capabilities",
                ]:
                    capabilities_column = (
                        header  # This contains both capabilities AND PCF processes
                    )
                elif target_field in ["_capabilities", "capabilities"]:
                    capabilities_column = header
                elif target_field in [
                    "archimate_capability",
                    "application_capability_archimate",
                ]:
                    archimate_column = header
                elif target_field == "description":
                    description_column = header

        # Auto-detect columns based on actual Excel header patterns
        for h in headers:
            h_upper = str(h).upper().strip() if h else ""
            h_clean = h_upper.replace("(", "").replace(")", "").strip()

            # Name column - check for various name patterns
            # Note: h_upper is already stripped, so "Name " becomes "NAME"
            if not name_column:
                if (
                    h_upper == "NAME"
                    or h_upper == "APPLICATION NAME"
                    or h_upper == "APP NAME"
                ):
                    name_column = h
                # Also check if the header starts with "NAME" (for headers like "Name (column)")
                elif h_upper.startswith("NAME") and "PACKAGE" not in h_upper:
                    name_column = h

            # App ID column
            if not app_id_column:
                if (
                    h_upper == "APP ID"
                    or h_upper == "APPLICATION ID"
                    or h_upper == "ID"
                    or h_upper == "APP CODE"
                ):
                    app_id_column = h

            # ArchiMate column - "Application Capability (Archimate)"
            if not archimate_column:
                if "ARCHIMATE" in h_upper or "APPLICATION CAPABILITY" in h_upper:
                    archimate_column = h

            # Package Vendor column - explicit vendor info
            if not vendor_column:
                if "PACKAGE VENDOR" in h_upper:
                    vendor_column = h
                elif h_upper == "VENDOR" or h_upper == "SUPPLIER":
                    vendor_column = h

            # Package Name column - for vendor inference
            if not package_name_column:
                if "PACKAGE NAME" in h_upper or h_upper == "PACKAGE NAME":
                    package_name_column = h
                elif "PRODUCT NAME" in h_upper:
                    package_name_column = h

            # Functionality/Capabilities column - contains APQC PCF process codes like "2.2 Advertising..."
            # This is the PRIMARY source for APQC PCF mapping
            if not capabilities_column:
                if "FUNCTIONALITY" in h_upper or "CAPABILITIES" in h_clean:
                    capabilities_column = h

            # PCF-specific column (if separate from capabilities)
            if not pcf_column:
                if "PCF" in h_upper or "APQC" in h_upper:
                    pcf_column = h

            # Description column
            if not description_column:
                if h_upper == "DESCRIPTION":
                    description_column = h

        # If no explicit PCF column, use capabilities column for APQC extraction
        # The "Functionality (Capabilities)" column contains PCF codes like "2.2 Advertising and campaigns"
        if not pcf_column and capabilities_column:
            pcf_column = capabilities_column

        # Debug logging for column detection
        current_app.logger.info(f"Column detection results:")
        current_app.logger.info(f"  - Headers found: {headers}")
        current_app.logger.info(f"  - Name: '{name_column}'")
        current_app.logger.info(f"  - App ID: '{app_id_column}'")
        current_app.logger.info(f"  - ArchiMate: '{archimate_column}'")
        current_app.logger.info(f"  - Vendor: '{vendor_column}'")
        current_app.logger.info(f"  - Package Name: '{package_name_column}'")
        current_app.logger.info(f"  - Capabilities/PCF: '{capabilities_column}'")
        current_app.logger.info(f"  - PCF: '{pcf_column}'")
        current_app.logger.info(f"  - Total data rows: {len(data_rows)}")
        if data_rows and len(data_rows) > 0:
            current_app.logger.info(
                f"  - First row keys: {list(data_rows[0].keys()) if data_rows[0] else 'empty'}"
            )
            current_app.logger.info(
                f"  - First row sample: {dict(list(data_rows[0].items())[:5]) if data_rows[0] else 'empty'}"
            )

        # Pre-load existing data for efficient lookups
        try:
            # Ensure clean transaction state before queries
            try:
                db.session.rollback()
            except Exception as e:  # fabricated-values-ok
                logger.debug(f"Ignored: {e}")

            try:
                db.session.begin()
            except Exception as e:  # fabricated-values-ok
                logger.debug(f"Ignored: {e}")

            # Execute queries with proper error handling
            existing_archimate = {
                e.name.lower(): e
                for e in ArchiMateElement.query.filter_by(
                    type="ApplicationComponent"
                ).all()
            }
            # Prefetch all existing applications to avoid N+1 queries later
            existing_apps_list = ApplicationComponent.query.all()
            existing_apps_by_name = {
                app.name.lower(): app for app in existing_apps_list if app.name
            }
            _all_apqc = APQCProcess.query.all()
            existing_apqc = {p.process_code: p for p in _all_apqc}
            existing_apqc_by_name = {
                p.process_name.lower(): p for p in _all_apqc
            }
            existing_vendors = {}
            if VendorOrganization:
                existing_vendors = {
                    v.name.lower(): v for v in VendorOrganization.query.all()
                }

        except Exception as e:
            current_app.logger.error(f"Database query error: {e}")
            # Force rollback to clear any failed transaction state
            try:
                db.session.rollback()
            except Exception as e:  # fabricated-values-ok
                logger.debug(f"Ignored: {e}")
            return jsonify({"error": "An internal error occurred"}), 500

        # Initialize vendor mapping service for intelligent vendor detection
        try:
            from ..services.application_vendor_mapping_service import (
                ApplicationVendorMappingService,
            )

            vendor_mapping_service = ApplicationVendorMappingService()
        except ImportError:
            vendor_mapping_service = None

        # Analysis counters
        will_create = 0
        will_update = 0
        duplicates_in_file = 0
        no_name = 0
        validation_errors = 0

        # ArchiMate analysis
        archimate_will_create = 0
        archimate_will_link = 0
        archimate_details = {"create": [], "link": []}

        # APQC analysis
        apqc_will_link = 0
        apqc_not_found = 0
        apqc_details = {"link": [], "not_found": []}

        # Vendor analysis - now includes inferred vendors
        vendor_will_create = 0
        vendor_will_link = 0
        vendor_inferred = 0
        vendor_details = {"create": [], "link": [], "inferred": []}

        details = {
            "create": [],
            "update": [],
            "duplicates": [],
            "no_name": [],
            "validation_errors": [],
        }

        seen_identifiers = set()
        seen_archimate_names = set()
        seen_vendors = set()
        seen_apqc = set()

        # === COMPREHENSIVE AI ANALYSIS USING AI IMPORT SERVICE ===
        ai_analysis_results = None
        if ai_service and len(data_rows) > 0:
            try:
                current_app.logger.info(
                    f"Starting comprehensive AI analysis for {len(data_rows)} applications..."
                )

                # Performance optimization: Set timeout for large imports
                import threading
                import time
                from concurrent.futures import ThreadPoolExecutor
                from concurrent.futures import TimeoutError as FutureTimeoutError

                # Define timeout handler for large imports (Windows compatible)
                class TimeoutError(Exception):
                    pass

                def run_with_timeout(func, *args, timeout_seconds=300, **kwargs):
                    """Run function with timeout using ThreadPool (Windows compatible)"""
                    with ThreadPoolExecutor(max_workers=1) as executor:
                        future = executor.submit(func, *args, **kwargs)
                        try:
                            return future.result(timeout=timeout_seconds)
                        except FutureTimeoutError:
                            raise TimeoutError(
                                "AI analysis timeout - too many applications"
                            )

                # Set timeout for imports over 100 applications (5 minutes)
                timeout_seconds = 300 if len(data_rows) > 100 else 120

                # Prepare application data for AI analysis
                applications_data = []
                processing_start_time = time.time()

                for row_idx, row_data in enumerate(data_rows):
                    if not isinstance(row_data, dict):
                        continue

                    # Get application name using same logic as below
                    name = None
                    if custom_mappings:
                        for header, target in custom_mappings.items():
                            if target == "name":
                                stripped_header = header.strip()
                                key_mapping = {k.strip(): k for k in row_data.keys()}
                                if stripped_header in key_mapping:
                                    actual_key = key_mapping[stripped_header]
                                    name = (
                                        str(row_data[actual_key]).strip()
                                        if row_data[actual_key]
                                        else None
                                    )
                                    break

                    if not name and name_column:
                        key_mapping = {k.strip(): k for k in row_data.keys()}
                        stripped_name_column = name_column.strip()
                        if stripped_name_column in key_mapping:
                            actual_key = key_mapping[stripped_name_column]
                            name = (
                                str(row_data[actual_key]).strip()
                                if row_data[actual_key]
                                else None
                            )

                    if name:  # Only include applications with names
                        app_data = {
                            "name": name,
                            "description": row_data.get(description_column, "")
                            if description_column
                            else "",
                            "vendor_name": row_data.get(vendor_column, "")
                            if vendor_column
                            else "",
                            "technology_stack": row_data.get("technology_stack", ""),
                            "business_domain": row_data.get("business_domain", ""),
                            "business_functions": row_data.get(
                                "application_functions_text", ""
                            ),
                            "capabilities": row_data.get(capabilities_column, "")
                            if capabilities_column
                            else "",
                            "apqc_codes": row_data.get(pcf_column, "")
                            if pcf_column
                            else "",
                        }
                        applications_data.append(app_data)

                    # Progress tracking for large imports
                    if len(data_rows) > 50 and row_idx % 25 == 0:
                        progress = (row_idx / len(data_rows)) * 100
                        current_app.logger.info(
                            f"AI data preparation progress: {progress:.1f}% ({row_idx}/{len(data_rows)})"
                        )

                data_preparation_time = time.time() - processing_start_time
                current_app.logger.info(
                    f"Data preparation completed in {data_preparation_time:.2f}s for {len(applications_data)} applications"
                )

                if applications_data:
                    # Use AI Import Service for file data analysis with cross-platform timeout protection
                    current_app.logger.info(
                        f"Calling AI Import Service analyze_file_data_for_preview for {len(applications_data)} applications (timeout: {timeout_seconds}s)..."
                    )

                    try:
                        # Use Windows-compatible timeout for large imports
                        if len(applications_data) > 100:
                            ai_analysis_results = run_with_timeout(
                                ai_service.analyze_file_data_for_preview,
                                applications_data=applications_data,
                                confidence_threshold=0.6,
                                timeout_seconds=timeout_seconds,
                            )
                        else:
                            ai_analysis_results = (
                                ai_service.analyze_file_data_for_preview(
                                    applications_data=applications_data,
                                    confidence_threshold=0.6,
                                )
                            )

                        total_analysis_time = time.time() - processing_start_time
                        current_app.logger.info(
                            f"File data AI analysis complete: {ai_analysis_results['total_analyzed']} applications analyzed in {total_analysis_time:.2f}s"
                        )

                        # Performance metrics
                        if ai_analysis_results["total_analyzed"] > 0:
                            avg_time_per_app = (
                                total_analysis_time
                                / ai_analysis_results["total_analyzed"]
                            )
                            current_app.logger.info(
                                f"Performance: {avg_time_per_app:.3f}s per application"
                            )

                        # Add performance metrics to results
                        ai_analysis_results["performance_metrics"] = {
                            "data_preparation_time_ms": int(
                                data_preparation_time * 1000
                            ),
                            "total_analysis_time_ms": int(total_analysis_time * 1000),
                            "avg_time_per_app_ms": int(avg_time_per_app * 1000)
                            if ai_analysis_results["total_analyzed"] > 0
                            else 0,
                            "timeout_protection": len(applications_data) > 100,
                            "file_preview_mode": True,
                        }

                    except TimeoutError as te:
                        current_app.logger.error(f"AI analysis timeout: {te}")
                        ai_analysis_results = {
                            "error": "timeout",
                            "message": f"AI analysis timed out after {timeout_seconds} seconds for {len(applications_data)} applications",
                            "partial_results": {
                                "applications_processed": len(applications_data),
                                "recommendation": "Consider processing in smaller batches",
                            },
                        }
                    except Exception as ai_error:
                        current_app.logger.error(f"AI analysis failed: {ai_error}")
                        ai_analysis_results = {
                            "error": "analysis_failed",
                            "message": str(ai_error),
                            "applications_attempted": len(applications_data),
                        }
                else:
                    current_app.logger.warning(
                        "No valid application data for AI analysis"
                    )
                    ai_analysis_results = {
                        "error": "no_data",
                        "message": "No valid application data found for AI analysis",
                    }

            except ImportError as import_error:
                current_app.logger.error(
                    f"Missing required modules for AI analysis: {import_error}"
                )
                ai_analysis_results = {
                    "error": "missing_modules",
                    "message": "Required modules for AI analysis not available",
                    "details": str(import_error),
                }
            except Exception as general_error:
                current_app.logger.error(
                    f"Unexpected error in AI analysis setup: {general_error}"
                )
                ai_analysis_results = {
                    "error": "setup_failed",
                    "message": "Failed to set up AI analysis",
                    "details": str(general_error),
                }
        else:
            current_app.logger.info(
                "AI Import Service not available or no data rows - using basic analysis"
            )
            ai_analysis_results = {
                "error": "service_unavailable",
                "message": "AI Import Service not available or no data provided",
            }

        for row_idx, row_data in enumerate(data_rows, start=2):
            if not isinstance(row_data, dict):
                continue
            if not any(v for v in row_data.values() if v):
                continue

            # Get name
            name = None
            current_app.logger.info(
                f"Row {row_idx}: Processing row data. name_column='{name_column}', custom_mappings={custom_mappings}"
            )
            current_app.logger.info(
                f"Row {row_idx}: Available row_data keys: {list(row_data.keys())}"
            )

            # Create a mapping of stripped keys to actual keys for robust matching
            key_mapping = {k.strip(): k for k in row_data.keys()}

            if custom_mappings:
                for header, target in custom_mappings.items():
                    if target == "name":
                        stripped_header = header.strip()
                        if stripped_header in key_mapping:
                            actual_key = key_mapping[stripped_header]
                            name = (
                                str(row_data[actual_key]).strip()
                                if row_data[actual_key]
                                else None
                            )
                            current_app.logger.info(
                                f"Row {row_idx}: Using custom mapping for name: {header} -> '{name}' (actual key: '{actual_key}')"
                            )
                            break
            if not name and name_column:
                stripped_name_column = name_column.strip()
                if stripped_name_column in key_mapping:
                    actual_key = key_mapping[stripped_name_column]
                    name = (
                        str(row_data[actual_key]).strip()
                        if row_data[actual_key]
                        else None
                    )
                    current_app.logger.info(
                        f"Row {row_idx}: Using auto-detected name column: {name_column} -> '{name}' (actual key: '{actual_key}')"
                    )

            if not name:
                current_app.logger.info(
                    f"Row {row_idx}: No name found. custom_mappings={custom_mappings}, name_column={name_column}, row_data_keys={row_data.keys()}"
                )
                no_name += 1
                details["no_name"].append({"row": row_idx})
                continue

            # Get app_id
            app_id = None
            if app_id_column:
                stripped_app_id_column = app_id_column.strip()
                if stripped_app_id_column in key_mapping:
                    actual_key = key_mapping[stripped_app_id_column]
                    app_id = (
                        str(row_data[actual_key]).strip()
                        if row_data[actual_key]
                        else None
                    )

            # Check for duplicates in file
            identifier = (name.lower(), app_id.lower() if app_id else None)
            if identifier in seen_identifiers:
                duplicates_in_file += 1
                details["duplicates"].append({"name": name, "row": row_idx})
                continue
            seen_identifiers.add(identifier)

            # Check for existing application in database
            try:
                # Use prefetched lookup instead of query to avoid N+1
                existing_app = existing_apps_by_name.get(name.lower())

            except Exception as db_error:
                current_app.logger.error(
                    f"Row {row_idx}: Database error during duplicate check: {db_error}"
                )
                # Force rollback to clear any failed transaction state
                try:
                    db.session.rollback()
                except Exception as e:  # fabricated-values-ok
                    logger.debug(f"Ignored: {e}")
                # Assume it doesn't exist if database query fails
                existing_app = None

            if existing_app:
                will_update += 1
                details["update"].append(
                    {"name": name, "id": existing_app.id, "row": row_idx}
                )
            else:
                will_create += 1
                details["create"].append({"name": name, "row": row_idx})

            # === ArchiMate Analysis ===
            # Use dedicated ArchiMate column if available ("Application Capability (Archimate)")
            # Otherwise fall back to application name
            archimate_value = None
            if archimate_column:
                stripped_archimate_column = archimate_column.strip()
                if stripped_archimate_column in key_mapping:
                    actual_key = key_mapping[stripped_archimate_column]
                    archimate_value = (
                        str(row_data[actual_key]).strip()
                        if row_data[actual_key]
                        else None
                    )

            # Fall back to application name if no ArchiMate column value
            if not archimate_value:
                archimate_value = name

            archimate_name = archimate_value.lower()
            if archimate_name not in seen_archimate_names:
                seen_archimate_names.add(archimate_name)
                if archimate_name in existing_archimate:
                    archimate_will_link += 1
                    archimate_details["link"].append(
                        {
                            "name": archimate_value,
                            "source": "archimate_column"
                            if archimate_column
                            else "app_name",
                            "existing_id": existing_archimate[archimate_name].id,
                            "row": row_idx,
                        }
                    )
                else:
                    archimate_will_create += 1
                    archimate_details["create"].append(
                        {
                            "name": archimate_value,
                            "source": "archimate_column"
                            if archimate_column
                            else "app_name",
                            "row": row_idx,
                        }
                    )

            # === Vendor Analysis ===
            # 1. First check explicit vendor column (Package Vendor)
            vendor_name = None
            vendor_source = None
            if vendor_column:
                stripped_vendor_column = vendor_column.strip()
                if stripped_vendor_column in key_mapping:
                    actual_key = key_mapping[stripped_vendor_column]
                    vendor_name = (
                        str(row_data[actual_key]).strip()
                        if row_data[actual_key]
                        else None
                    )
                    vendor_source = "column"

            # 2. If no vendor column value, try to infer from Package Name column
            if not vendor_name and package_name_column:
                stripped_package_name_column = package_name_column.strip()
                if stripped_package_name_column in key_mapping:
                    actual_key = key_mapping[stripped_package_name_column]
                    package_name = (
                        str(row_data[actual_key]).strip()
                        if row_data[actual_key]
                        else None
                    )
                    if package_name and vendor_mapping_service:
                        # Try to extract vendor from package name (e.g., "SAP S/4HANA", "Oracle Database")
                        inferred_vendor = vendor_mapping_service.find_vendor_by_name(
                            package_name
                        )
                        if inferred_vendor:
                            vendor_name = inferred_vendor.name
                            vendor_source = "inferred_from_package"
                        else:
                            # Try extracting vendor patterns from package name
                            package_lower = package_name.lower()
                            for (
                                base_vendor,
                                aliases,
                            ) in ApplicationVendorMappingService.VENDOR_NAME_ALIASES.items():
                                for alias in aliases:
                                    if (
                                        alias in package_lower
                                        or package_lower.startswith(alias)
                                    ):
                                        matched = (
                                            vendor_mapping_service.find_vendor_by_name(
                                                base_vendor
                                            )
                                        )
                                        if matched:
                                            vendor_name = matched.name
                                            vendor_source = "inferred_from_package"
                                            break
                                if vendor_name:
                                    break

            # 3. If still no vendor, try to infer from application name
            if not vendor_name and vendor_mapping_service:
                # Try to extract vendor from application name using common patterns
                # E.g., "SAP S/4HANA", "Oracle EBS", "Microsoft Dynamics 365"
                inferred_vendor = vendor_mapping_service.find_vendor_by_name(name)
                if inferred_vendor:
                    vendor_name = inferred_vendor.name
                    vendor_source = "inferred_from_name"
                else:
                    # Try extracting first word/known vendor patterns from app name
                    name_lower = name.lower()
                    for (
                        base_vendor,
                        aliases,
                    ) in ApplicationVendorMappingService.VENDOR_NAME_ALIASES.items():
                        for alias in aliases:
                            if alias in name_lower or name_lower.startswith(alias):
                                # Found a vendor pattern in the name
                                matched = vendor_mapping_service.find_vendor_by_name(
                                    base_vendor
                                )
                                if matched:
                                    vendor_name = matched.name
                                    vendor_source = "inferred_pattern"
                                    break
                        if vendor_name:
                            break

            if vendor_name:
                vendor_key = vendor_name.lower()
                if vendor_key not in seen_vendors:
                    seen_vendors.add(vendor_key)

                    # Try fuzzy matching first using the service
                    matched_vendor = None
                    if vendor_mapping_service:
                        matched_vendor = vendor_mapping_service.find_vendor_by_name(
                            vendor_name
                        )

                    # Fall back to exact match
                    if not matched_vendor and vendor_key in existing_vendors:
                        matched_vendor = existing_vendors[vendor_key]

                    if matched_vendor:
                        if vendor_source in (
                            "inferred_from_name",
                            "inferred_pattern",
                            "inferred_from_package",
                        ):
                            vendor_inferred += 1
                            vendor_details["inferred"].append(
                                {
                                    "name": matched_vendor.name,
                                    "existing_id": matched_vendor.id,
                                    "app_name": name,
                                    "source": vendor_source,
                                    "row": row_idx,
                                }
                            )
                        else:
                            vendor_will_link += 1
                            vendor_details["link"].append(
                                {
                                    "name": matched_vendor.name,
                                    "existing_id": matched_vendor.id,
                                    "app_name": name,
                                    "row": row_idx,
                                }
                            )
                    else:
                        vendor_will_create += 1
                        vendor_details["create"].append(
                            {
                                "name": vendor_name,
                                "app_name": name,
                                "source": vendor_source or "column",
                                "row": row_idx,
                            }
                        )

            # === APQC PCF Analysis ===
            # Use AI-powered semantic analysis instead of just regex matching
            pcf_raw_value = None
            if pcf_column:
                stripped_pcf_column = pcf_column.strip()
                if stripped_pcf_column in key_mapping:
                    actual_key = key_mapping[stripped_pcf_column]
                    pcf_raw_value = (
                        str(row_data[actual_key]).strip()
                        if row_data[actual_key]
                        else None
                    )

            # Also check capabilities column if different from pcf_column
            if (
                not pcf_raw_value
                and capabilities_column
                and capabilities_column != pcf_column
            ):
                stripped_capabilities_column = capabilities_column.strip()
                if stripped_capabilities_column in key_mapping:
                    actual_key = key_mapping[stripped_capabilities_column]
                    pcf_raw_value = (
                        str(row_data[actual_key]).strip()
                        if row_data[actual_key]
                        else None
                    )

            # Use SemanticAPQC Service with REAL embeddings (the method that actually works)
            if pcf_raw_value and pcf_raw_value.strip():
                current_app.logger.info(
                    f"Row {row_idx}: Starting Semantic APQC analysis for text: {pcf_raw_value[:100]}..."
                )
                try:
                    # Use the ACTUAL working SemanticAPQCService method
                    from ..services.semantic_apqc_service import SemanticAPQCService

                    # Ensure clean transaction state
                    try:
                        db.session.commit()
                    except Exception:
                        db.session.rollback()

                    current_app.logger.info(
                        f"Row {row_idx}: Calling Semantic APQC classification"
                    )
                    # Get REAL semantic classification (the method that actually works)
                    semantic_service = SemanticAPQCService()
                    classification_result = semantic_service.classify_text(
                        pcf_raw_value, max_results=10
                    )

                    # Enhance with APQC Hierarchy Service for better matching
                    enhanced_matches = []
                    matches_list = []
                    if classification_result:
                        if hasattr(classification_result, "matches"):
                            # APQCClassificationResult object (from semantic service)
                            matches_list = classification_result.matches
                        elif isinstance(classification_result, list):
                            # Raw list of matches (from other services)
                            matches_list = classification_result

                    # Use APQC Hierarchy Service to enhance matches with hierarchy paths and auto-linking
                    if matches_list and apqc_hierarchy:
                        try:
                            for match in matches_list:
                                # Handle both dataclass objects and dictionaries
                                if hasattr(match, "process_code"):
                                    # APQCMatch dataclass object
                                    process_code = match.process_code
                                else:
                                    # Dictionary object
                                    process_code = match.get("process_code")
                                if process_code:
                                    # Get enhanced match with hierarchy information
                                    enhanced_match = apqc_hierarchy.search_processes(
                                        query=process_code, limit=1
                                    )
                                    if enhanced_match and len(enhanced_match) > 0:
                                        enhanced_data = enhanced_match[0]
                                        # Create enhanced match dictionary
                                        if hasattr(match, "process_code"):
                                            # APQCMatch dataclass object - convert to dict
                                            enhanced_match_dict = {
                                                "process_id": match.process_id,
                                                "process_code": match.process_code,
                                                "process_name": match.process_name,
                                                "similarity_score": getattr(
                                                    match, "similarity_score", 0
                                                ),
                                                "confidence": getattr(
                                                    match, "confidence", "medium"
                                                ),
                                                "hierarchy_path": getattr(
                                                    enhanced_data,
                                                    "hierarchy_path",
                                                    None,
                                                )
                                                if hasattr(
                                                    enhanced_data, "hierarchy_path"
                                                )
                                                else enhanced_data.get(
                                                    "hierarchy_path"
                                                ),
                                                "auto_link_parents": getattr(
                                                    enhanced_data,
                                                    "auto_link_parents",
                                                    None,
                                                )
                                                if hasattr(
                                                    enhanced_data, "auto_link_parents"
                                                )
                                                else enhanced_data.get(
                                                    "auto_link_parents"
                                                ),
                                                "rationale": getattr(
                                                    enhanced_data, "rationale", None
                                                )
                                                if hasattr(enhanced_data, "rationale")
                                                else enhanced_data.get("rationale"),
                                                "level": getattr(
                                                    enhanced_data, "level", None
                                                )
                                                if hasattr(enhanced_data, "level")
                                                else enhanced_data.get("level"),
                                            }
                                            enhanced_matches.append(enhanced_match_dict)
                                        else:
                                            # Dictionary object - update in place
                                            match.update(
                                                {
                                                    "hierarchy_path": getattr(
                                                        enhanced_data,
                                                        "hierarchy_path",
                                                        None,
                                                    )
                                                    if hasattr(
                                                        enhanced_data, "hierarchy_path"
                                                    )
                                                    else enhanced_data.get(
                                                        "hierarchy_path"
                                                    ),
                                                    "auto_link_parents": getattr(
                                                        enhanced_data,
                                                        "auto_link_parents",
                                                        None,
                                                    )
                                                    if hasattr(
                                                        enhanced_data,
                                                        "auto_link_parents",
                                                    )
                                                    else enhanced_data.get(
                                                        "auto_link_parents"
                                                    ),
                                                    "rationale": getattr(
                                                        enhanced_data, "rationale", None
                                                    )
                                                    if hasattr(
                                                        enhanced_data, "rationale"
                                                    )
                                                    else enhanced_data.get("rationale"),
                                                    "level": getattr(
                                                        enhanced_data, "level", None
                                                    )
                                                    if hasattr(enhanced_data, "level")
                                                    else enhanced_data.get("level"),
                                                }
                                            )
                                            enhanced_matches.append(match)
                            enhanced_matches = matches_list
                        except Exception as e:
                            current_app.logger.error(
                                f"Error enhancing matches with APQC hierarchy: {e}"
                            )
                            enhanced_matches = matches_list
                    else:
                        enhanced_matches = matches_list

                    current_app.logger.info(
                        f"Row {row_idx}: Semantic APQC classification returned {len(enhanced_matches)} matches"
                    )

                    # Process AI results
                    if enhanced_matches and len(enhanced_matches) > 0:
                        ai_matches_found = 0
                        for match in enhanced_matches:
                            # Handle both dataclass objects and dictionaries
                            if hasattr(match, "process_code"):
                                # APQCMatch dataclass object
                                process_code = match.process_code
                                similarity_score = getattr(match, "similarity_score", 0)
                            else:
                                # Dictionary object
                                process_code = match.get("process_code")
                                similarity_score = match.get("similarity_score", 0)

                            if (
                                process_code and float(similarity_score) > 0.3
                            ):  # Semantic similarity threshold
                                pcf_key = process_code.lower()
                                if pcf_key not in seen_apqc:
                                    seen_apqc.add(pcf_key)

                                    try:
                                        # Use prefetched lookup instead of query to avoid N+1
                                        existing_apqc = existing_apqc.get(process_code)
                                        if existing_apqc:
                                            apqc_will_link += 1
                                            ai_matches_found += 1

                                            # Get attributes based on object type
                                            if hasattr(match, "process_name"):
                                                # Use clean database name instead of semantic match name
                                                process_name = (
                                                    existing_apqc.process_name
                                                    or match.process_name
                                                )
                                            else:
                                                # Use clean database name instead of semantic match name
                                                process_name = (
                                                    existing_apqc.process_name
                                                    or match.get("process_name", "")
                                                )

                                            apqc_details["link"].append(
                                                {
                                                    "process_code": process_code,
                                                    "process_name": process_name,
                                                    "confidence": "high"
                                                    if float(similarity_score) > 0.7
                                                    else "medium",
                                                    "similarity_score": float(
                                                        similarity_score
                                                    ),
                                                    "source": "semantic_similarity",
                                                    "existing_id": existing_apqc.id,
                                                    "row": row_idx,
                                                    "input_text": pcf_raw_value[:100]
                                                    + "..."
                                                    if len(pcf_raw_value) > 100
                                                    else pcf_raw_value,
                                                    "hierarchy_path": getattr(
                                                        match, "hierarchy_path", None
                                                    ),
                                                    "auto_link_parents": getattr(
                                                        match, "auto_link_parents", None
                                                    ),
                                                    "rationale": getattr(
                                                        match, "rationale", None
                                                    ),
                                                    "level": getattr(
                                                        match, "level", None
                                                    ),
                                                }
                                            )
                                            current_app.logger.info(
                                                f"Row {row_idx}: Semantic APQC matched APQC {process_code} with similarity_score {float(similarity_score):.3f}"
                                            )
                                        else:
                                            # Process not found in database
                                            apqc_not_found += 1
                                            apqc_details["not_found"].append(
                                                {
                                                    "process_code": process_code,
                                                    "process_name": process_name,
                                                    "confidence": "high"
                                                    if float(similarity_score) > 0.7
                                                    else "medium",
                                                    "similarity_score": float(
                                                        similarity_score
                                                    ),
                                                    "source": "semantic_similarity",
                                                    "row": row_idx,
                                                    "input_text": pcf_raw_value[:100]
                                                    + "..."
                                                    if len(pcf_raw_value) > 100
                                                    else pcf_raw_value,
                                                }
                                            )
                                    except Exception as db_error:
                                        current_app.logger.error(
                                            f"Row {row_idx}: Database error during APQC lookup: {db_error}"
                                        )
                                        # Continue without this match
                                        continue

                        current_app.logger.info(
                            f"Row {row_idx}: Semantic APQC processing complete, found {ai_matches_found} matches"
                        )
                    else:
                        # Fallback to regex parsing if Semantic APQC returns no results
                        current_app.logger.warning(
                            f"Row {row_idx}: Semantic APQC classification returned no results, falling back to regex"
                        )
                        # Simple regex fallback
                        try:
                            pcf_pattern = r"(?=\b\d+\.\d+(?:\.\d+)*\s)"
                            parts = re.split(pcf_pattern, pcf_raw_value)

                            for part in parts:
                                part = part.strip()
                                if part:
                                    code_match = re.match(
                                        r"^(\d+\.\d+(?:\.\d+)*)\s*(.*)$", part
                                    )
                                    if code_match:
                                        pcf_code = code_match.group(1)
                                        pcf_key = pcf_code.lower()
                                        if pcf_key not in seen_apqc:
                                            seen_apqc.add(pcf_key)

                                            # Use prefetched lookup instead of query to avoid N+1
                                            existing_apqc_proc = existing_apqc.get(
                                                pcf_code
                                            )
                                            if existing_apqc_proc:
                                                apqc_will_link += 1
                                                apqc_details["link"].append(
                                                    {
                                                        "process_code": pcf_code,
                                                        "process_name": code_match.group(
                                                            2
                                                        ).strip(),
                                                        "source": "regex_fallback",
                                                        "existing_id": existing_apqc_proc.id,
                                                        "row": row_idx,
                                                    }
                                                )
                                            if existing_apqc_proc:
                                                apqc_will_link += 1
                                                apqc_details["link"].append(
                                                    {
                                                        "process_code": pcf_code,
                                                        "process_name": code_match.group(
                                                            2
                                                        ).strip(),
                                                        "source": "regex_fallback",
                                                        "existing_id": existing_apqc.id,
                                                        "row": row_idx,
                                                    }
                                                )
                                            else:
                                                apqc_not_found += 1
                                                apqc_details["not_found"].append(
                                                    {
                                                        "process_code": pcf_code,
                                                        "process_name": code_match.group(
                                                            2
                                                        ).strip(),
                                                        "source": "regex_fallback",
                                                        "row": row_idx,
                                                    }
                                                )
                        except Exception as regex_error:
                            current_app.logger.error(
                                f"Error in fallback regex analysis: {regex_error}"
                            )

                except Exception as apqc_error:
                    current_app.logger.error(
                        f"Error in APQC analysis for row {row_idx}: {apqc_error}"
                    )
                    # Fallback to regex parsing
                    try:
                        pcf_pattern = r"(?=\b\d+\.\d+(?:\.\d+)*\s)"
                        parts = re.split(pcf_pattern, pcf_raw_value)

                        for part in parts:
                            part = part.strip()
                            if part:
                                code_match = re.match(
                                    r"^(\d+\.\d+(?:\.\d+)*)\s*(.*)$", part
                                )
                                if code_match:
                                    pcf_code = code_match.group(1)
                                    pcf_key = pcf_code.lower()
                                    if pcf_key not in seen_apqc:
                                        seen_apqc.add(pcf_key)

                                        # Use prefetched lookup instead of query to avoid N+1
                                        existing_apqc_proc = existing_apqc.get(pcf_code)
                                        if existing_apqc_proc:
                                            apqc_will_link += 1
                                            apqc_details["link"].append(
                                                {
                                                    "process_code": pcf_code,
                                                    "process_name": code_match.group(
                                                        2
                                                    ).strip(),
                                                    "source": "regex_fallback_error",
                                                    "existing_id": existing_apqc_proc.id,
                                                    "row": row_idx,
                                                }
                                            )
                                        else:
                                            apqc_not_found += 1
                                            apqc_details["not_found"].append(
                                                {
                                                    "process_code": pcf_code,
                                                    "process_name": code_match.group(
                                                        2
                                                    ).strip(),
                                                    "source": "regex_fallback_error",
                                                    "row": row_idx,
                                                }
                                            )
                    except Exception as fallback_error:
                        current_app.logger.error(
                            f"Error in fallback analysis: {fallback_error}"
                        )
            else:
                # Try description-based AI analysis if no PCF column
                if name and description_column:
                    description = (
                        row_data.get(description_column, "")
                        if description_column in key_mapping
                        else ""
                    )
                    description = str(description).strip() if description else ""
                    if description:
                        try:
                            from ..services.semantic_apqc_service import (
                                SemanticAPQCService,
                            )

                            # Ensure clean transaction state for AI service
                            try:
                                db.session.commit()
                            except Exception:
                                db.session.rollback()

                            semantic_service = SemanticAPQCService()

                            # Use application description for APQC inference
                            classification_result = semantic_service.classify_text_sync(
                                f"Application: {name}. Description: {description}",
                                max_results=5,
                            )

                            # Handle different result types from different APQC services
                            matches_list = []
                            if classification_result:
                                if hasattr(classification_result, "matches"):
                                    # APQCClassificationResult object (from semantic service)
                                    matches_list = classification_result.matches
                                elif isinstance(classification_result, list):
                                    # Raw list of matches (from other services)
                                    matches_list = classification_result

                            if matches_list and len(matches_list) > 0:
                                for match in matches_list:
                                    # Handle both dataclass objects and dictionaries
                                    if hasattr(match, "process_code"):
                                        # APQCMatch dataclass object
                                        process_code = match.process_code
                                        process_name = match.process_name
                                        confidence = getattr(
                                            match, "confidence", "medium"
                                        )
                                        similarity_score = getattr(
                                            match, "similarity_score", 0
                                        )
                                    else:
                                        # Dictionary object
                                        process_code = match.get("process_code")
                                        process_name = match.get("process_name", "")
                                        confidence = match.get("confidence", "medium")
                                        similarity_score = match.get(
                                            "similarity_score", 0
                                        )

                                    if (
                                        confidence and float(similarity_score) > 0.6
                                    ):  # Higher threshold for inferred matches
                                        pcf_key = process_code.lower()
                                        if pcf_key not in seen_apqc:
                                            seen_apqc.add(pcf_key)

                                            try:
                                                # Use prefetched lookup instead of query to avoid N+1
                                                existing_apqc_proc = existing_apqc.get(
                                                    process_code
                                                )
                                                if existing_apqc_proc:
                                                    apqc_will_link += 1
                                                    apqc_details["link"].append(
                                                        {
                                                            "process_code": process_code,
                                                            "process_name": process_name,
                                                            "confidence": confidence,
                                                            "source": "ai_inferred_from_description",
                                                            "existing_id": existing_apqc_proc.id,
                                                            "row": row_idx,
                                                            "app_name": name,
                                                        }
                                                    )
                                            except Exception as db_error:
                                                current_app.logger.error(
                                                    f"Row {row_idx}: Database error during description-based APQC lookup: {db_error}"
                                                )
                                                continue
                        except Exception as desc_ai_error:
                            current_app.logger.error(
                                f"Error in description-based APQC analysis: {desc_ai_error}"
                            )
                            # Continue without APQC analysis if both methods fail

            # === Enhanced Vendor Analysis using AI ===
            if vendor_name:
                vendor_key = vendor_name.lower()
                if vendor_key not in seen_vendors:
                    seen_vendors.add(vendor_key)

                    # Try AI-powered vendor analysis if we have description
                    description = (
                        getattr(row_data.get(description_column, ""), "")
                        if description_column
                        else ""
                    )
                    vendor_ai_insights = {}

                    if description and vendor_source in [
                        "inferred_from_name",
                        "column",
                    ]:
                        current_app.logger.info(
                            f"Row {row_idx}: Starting AI vendor analysis for {vendor_name}"
                        )
                        try:
                            from ..services.llm_service import LLMService

                            # Ensure clean transaction state for AI service
                            try:
                                db.session.commit()
                            except Exception:
                                db.session.rollback()

                            llm_service = LLMService()

                            # Use AI to enhance vendor analysis
                            app_data = {
                                "name": name,
                                "description": description,
                                "vendor_name": vendor_name,
                                "application_category": getattr(
                                    row_data.get("application_category", ""), ""
                                ),
                                "business_criticality": getattr(
                                    row_data.get("business_criticality", ""), ""
                                ),
                            }

                            current_app.logger.info(
                                f"Row {row_idx}: Calling LLMService.analyze_application_semantically()"
                            )
                            ai_insight = llm_service.analyze_application_semantically(
                                app_data
                            )

                            current_app.logger.info(
                                f"Row {row_idx}: AI vendor analysis completed, status: {ai_insight.get('status', 'unknown')}"
                            )

                            if ai_insight and not ai_insight.get("error"):
                                # Add AI insights to vendor details
                                vendor_ai_insights = {
                                    "vendor_confidence": ai_insight.get(
                                        "risk_factors", {}
                                    ).get("overall_risk_level", "medium"),
                                    "vendor_risk_assessment": ai_insight.get(
                                        "risk_factors", {}
                                    ).get("overall_risk_level", "medium"),
                                    "vendor_recommendations": ai_insight.get(
                                        "recommendations", {}
                                    ).get("architecture_improvements", [])[
                                        :3
                                    ],  # Limit to 3 recommendations
                                    "business_function": ai_insight.get(
                                        "business_function", {}
                                    ).get("primary_function", ""),
                                    "technology_analysis": ai_insight.get(
                                        "technology_analysis", {}
                                    ),
                                    "ai_analysis_status": "success",
                                }
                                current_app.logger.info(
                                    f"Row {row_idx}: AI insights added for vendor {vendor_name}"
                                )
                            else:
                                vendor_ai_insights = {
                                    "ai_analysis_status": "failed",
                                    "ai_error": ai_insight.get(
                                        "error", "Unknown error"
                                    ),
                                }
                                current_app.logger.warning(
                                    f"Row {row_idx}: AI vendor analysis failed: {ai_insight.get('error', 'Unknown error')}"
                                )

                        except Exception as vendor_ai_error:
                            current_app.logger.error(
                                f"Row {row_idx}: Error in vendor AI analysis: {vendor_ai_error}"
                            )
                            vendor_ai_insights = {
                                "ai_analysis_status": "exception",
                                "ai_error": str(vendor_ai_error),
                            }

                    # Try fuzzy matching first using the service
                    matched_vendor = None
                    if vendor_mapping_service:
                        matched_vendor = vendor_mapping_service.find_vendor_by_name(
                            vendor_name
                        )

                    # Fall back to exact match
                    if not matched_vendor and vendor_key in existing_vendors:
                        matched_vendor = existing_vendors[vendor_key]

                    if matched_vendor:
                        if vendor_source in (
                            "inferred_from_name",
                            "inferred_pattern",
                            "inferred_from_package",
                        ):
                            vendor_inferred += 1
                            vendor_details["inferred"].append(
                                {
                                    "name": matched_vendor.name,
                                    "existing_id": matched_vendor.id,
                                    "app_name": name,
                                    "source": vendor_source,
                                    "row": row_idx,
                                    **vendor_ai_insights,  # Add AI insights
                                }
                            )
                        else:
                            vendor_will_link += 1
                            vendor_details["link"].append(
                                {
                                    "name": matched_vendor.name,
                                    "existing_id": matched_vendor.id,
                                    "app_name": name,
                                    "row": row_idx,
                                    **vendor_ai_insights,  # Add AI insights
                                }
                            )
                    else:
                        vendor_will_create += 1
                        vendor_details["create"].append(
                            {
                                "name": vendor_name,
                                "app_name": name,
                                "source": vendor_source or "column",
                                "row": row_idx,
                                **vendor_ai_insights,  # Add AI insights
                            }
                        )

        # Initialize AI analysis summary dictionary
        ai_analysis_summary = {
            "apqc_analyzed": 0,
            "vendors_analyzed": 0,
            "ai_matches": 0,
            "confidence_avg": 0,
        }

        # === ENHANCED AI ANALYSIS SUMMARY ===
        # Include comprehensive AI results if available
        comprehensive_ai_results = None
        if ai_analysis_results and "error" not in ai_analysis_results:
            # Extract vendor analysis from applications
            vendor_analysis_summary = {
                "vendors_found": 0,
                "vendor_types": {},
                "high_reliability_vendors": 0,
                "low_complexity_vendors": 0,
                "vendor_recommendations": [],
            }

            if ai_analysis_results.get("applications"):
                for app_result in ai_analysis_results["applications"]:
                    vendor_analysis = app_result.get("vendor_analysis", {})
                    if vendor_analysis and vendor_analysis.get("vendor_name"):
                        vendor_analysis_summary["vendors_found"] += 1

                        vendor_type = vendor_analysis.get("vendor_type", "unknown")
                        if vendor_type not in vendor_analysis_summary["vendor_types"]:
                            vendor_analysis_summary["vendor_types"][vendor_type] = 0
                        vendor_analysis_summary["vendor_types"][vendor_type] += 1

                        reliability = vendor_analysis.get("reliability_score", 0)
                        if reliability >= 0.8:
                            vendor_analysis_summary["high_reliability_vendors"] += 1

                        complexity = vendor_analysis.get(
                            "integration_complexity", "medium"
                        )
                        if complexity == "low":
                            vendor_analysis_summary["low_complexity_vendors"] += 1

                        recommendations = vendor_analysis.get("recommendations", [])
                        if recommendations:
                            vendor_analysis_summary["vendor_recommendations"].extend(
                                recommendations[:2]
                            )  # Limit recommendations

            comprehensive_ai_results = {
                "total_analyzed": ai_analysis_results["total_analyzed"],
                "capability_mappings_found": ai_analysis_results[
                    "capability_mappings_found"
                ],
                "process_mappings_found": ai_analysis_results["process_mappings_found"],
                "archimate_elements_generated": ai_analysis_results[
                    "archimate_elements_generated"
                ],
                "high_confidence_mappings": ai_analysis_results[
                    "high_confidence_mappings"
                ],
                "vendor_analysis_found": ai_analysis_results.get(
                    "vendor_analysis_found", 0
                ),
                "processing_stats": ai_analysis_results["processing_stats"],
                "applications": ai_analysis_results["applications"],
                "performance_metrics": ai_analysis_results.get(
                    "performance_metrics", {}
                ),
                "vendor_analysis_summary": vendor_analysis_summary,
            }

            # Update AI analysis summary with comprehensive results
            ai_analysis_summary.update(
                {
                    "comprehensive_ai_enabled": True,
                    "ai_service_used": "AIImportService",
                    "bulk_analysis_complete": True,
                    "total_applications_analyzed": ai_analysis_results[
                        "total_analyzed"
                    ],
                    "ai_models_used": ai_analysis_results["processing_stats"][
                        "ai_models_used"
                    ],
                    "avg_processing_time_ms": ai_analysis_results["processing_stats"][
                        "avg_processing_time_ms"
                    ],
                    "performance_optimized": ai_analysis_results.get(
                        "performance_metrics", {}
                    ).get("timeout_protection", False),
                    "data_preparation_time_ms": ai_analysis_results.get(
                        "performance_metrics", {}
                    ).get("data_preparation_time_ms", 0),
                    "vendor_analysis_enabled": True,
                    "vendors_analyzed": vendor_analysis_summary["vendors_found"],
                    # NEW: AI Generation Counts for Enhanced Visibility
                    "capability_mappings_found": ai_analysis_results[
                        "capability_mappings_found"
                    ],
                    "archimate_elements_generated": ai_analysis_results[
                        "archimate_elements_generated"
                    ],
                    "process_mappings_found": ai_analysis_results[
                        "process_mappings_found"
                    ],
                    "high_confidence_mappings": ai_analysis_results[
                        "high_confidence_mappings"
                    ],
                }
            )
        elif ai_analysis_results and "error" in ai_analysis_results:
            # Handle AI analysis errors
            comprehensive_ai_results = (
                ai_analysis_results  # Pass through error information
            )

            # Update AI analysis summary with error information
            ai_analysis_summary.update(
                {
                    "comprehensive_ai_enabled": False,
                    "ai_service_used": "AIImportService (Error)",
                    "bulk_analysis_complete": False,
                    "ai_error": {
                        "type": ai_analysis_results.get("error", "unknown"),
                        "message": ai_analysis_results.get(
                            "message", "Unknown error occurred"
                        ),
                        "details": ai_analysis_results.get("details", ""),
                    },
                    # NEW: AI Generation Counts (set to 0 on error)
                    "capability_mappings_found": 0,
                    "archimate_elements_generated": 0,
                    "process_mappings_found": 0,
                    "high_confidence_mappings": 0,
                }
            )
        else:
            ai_analysis_summary.update(
                {
                    "comprehensive_ai_enabled": False,
                    "ai_service_used": "Basic Semantic Analysis",
                    "bulk_analysis_complete": False,
                    # NEW: AI Generation Counts (set to 0 for fallback)
                    "capability_mappings_found": 0,
                    "archimate_elements_generated": 0,
                    "process_mappings_found": 0,
                    "high_confidence_mappings": 0,
                }
            )

        # Calculate AI analysis summary
        # Note: "semantic_similarity" is the source for AI-powered APQC matches
        ai_apqc_sources = [
            "semantic_similarity",
            "ai_semantic",
            "ai_inferred_from_description",
        ]
        ai_analysis_summary.update(
            {
                "apqc_analyzed": len(
                    [
                        d
                        for d in apqc_details["link"]
                        if d.get("source") in ai_apqc_sources
                    ]
                ),
                "vendors_analyzed": len(
                    [d for d in vendor_details["link"] if "ai_analysis_status" in d]
                ),
                "ai_matches": len(
                    [
                        d
                        for d in apqc_details["link"]
                        if d.get("source") in ai_apqc_sources
                    ]
                )
                + len([d for d in vendor_details["link"] if "ai_analysis_status" in d]),
                "confidence_avg": 0,
            }
        )

        # Calculate average confidence for AI matches
        ai_confidences = []
        for d in apqc_details["link"]:
            if d.get("source") in ai_apqc_sources and "similarity_score" in d:
                ai_confidences.append(float(d.get("similarity_score", 0)))
        for d in vendor_details["link"]:
            if "ai_analysis_status" in d and d.get("ai_analysis_status") == "success":
                # Convert vendor confidence to numeric (high=85, medium=60, low=35)
                vendor_conf = d.get("vendor_confidence", "medium")
                if vendor_conf == "high":
                    ai_confidences.append(85)
                elif vendor_conf == "medium":
                    ai_confidences.append(60)
                elif vendor_conf == "low":
                    ai_confidences.append(35)

        if ai_confidences:
            ai_analysis_summary["confidence_avg"] = round(
                sum(ai_confidences) / len(ai_confidences), 1
            )

        return (
            jsonify(
                {
                    "success": True,
                    "will_create": will_create,
                    "will_update": will_update,
                    "duplicates_in_file": duplicates_in_file,
                    "no_name": no_name,
                    "validation_errors": validation_errors,
                    "total_rows": len(data_rows),
                    "details": details,
                    # Column detection info - helps users see what was detected
                    "detected_columns": {
                        "name": name_column,
                        "app_id": app_id_column,
                        "archimate": archimate_column,
                        "vendor": vendor_column,
                        "package_name": package_name_column,
                        "capabilities_pcf": capabilities_column,
                        "description": description_column,
                    },
                    # ArchiMate analysis
                    "archimate": {
                        "will_create": archimate_will_create,
                        "will_link": archimate_will_link,
                        "detected_column": archimate_column,
                        "details": archimate_details,
                    },
                    # APQC PCF analysis
                    "apqc": {
                        "will_link": apqc_will_link,
                        "not_found": apqc_not_found,
                        "detected_column": pcf_column or capabilities_column,
                        "details": apqc_details,
                    },
                    # Vendor analysis
                    "vendors": {
                        "will_create": vendor_will_create,
                        "will_link": vendor_will_link,
                        "inferred": vendor_inferred,
                        "detected_vendor_column": vendor_column,
                        "detected_package_name_column": package_name_column,
                        "details": vendor_details,
                    },
                    # AI Analysis Summary
                    "ai_analysis": ai_analysis_summary,
                    # Comprehensive AI Results (NEW)
                    "comprehensive_ai_results": comprehensive_ai_results,
                    # Detailed breakdown for frontend
                    "apqc_details": apqc_details,
                    "vendor_details": vendor_details,
                    # Pre-organized APQC links by row index for execution endpoint
                    # Format: {"row_index": [{"existing_id": 123, "process_code": "3.1", ...}, ...]}
                    # Frontend should pass this as 'apqc_links' form data when executing import
                    "apqc_links_by_row": _organize_apqc_links_by_row(
                        apqc_details.get("link", [])
                    ),
                }
            ),
            200,
        )

    except Exception as e:
        current_app.logger.error(f"Error analyzing import: {e}", exc_info=True)
        return jsonify({"error": "An internal error occurred"}), 500


@application_mgmt.route("/applications/auto-create-vendors", methods=["POST"])
@login_required
def auto_create_vendors():
    """
    Auto-create missing vendors before import.
    Creates VendorOrganization records for vendors that don't exist.
    """
    try:
        from ..models.vendor.vendor_organization import VendorOrganization
    except ImportError:
        return jsonify({"error": "VendorOrganization model not available"}), 500

    data = request.get_json()
    if not data or "vendors" not in data:
        return jsonify({"error": "No vendors provided"}), 400

    vendor_names = data.get("vendors", [])
    if not vendor_names:
        return jsonify({"error": "Empty vendor list"}), 400

    created_count = 0
    skipped_count = 0
    created_vendors = []
    errors = []

    # Prefetch existing vendors to avoid N+1 query in loop
    existing_vendors_list = VendorOrganization.query.all()
    existing_vendors_by_name = {
        v.name.lower(): v for v in existing_vendors_list if v.name
    }

    for vendor_name in vendor_names:
        if not vendor_name or not vendor_name.strip():
            continue

        vendor_name = vendor_name.strip()

        # Check if vendor already exists (case-insensitive)
        # Use prefetched lookup instead of query to avoid N+1
        existing = existing_vendors_by_name.get(vendor_name.lower())

        if existing:
            skipped_count += 1
            continue

        try:
            # Create new vendor organization
            new_vendor = VendorOrganization(
                name=vendor_name,
                display_name=vendor_name,
                vendor_type="software_vendor",
                status="active",
                evaluation_status="approved",
                contract_status="catalog",
            )
            db.session.add(new_vendor)
            db.session.flush()  # Get the ID
            created_count += 1
            created_vendors.append({"id": new_vendor.id, "name": new_vendor.name})
        except Exception as e:
            errors.append(f"{vendor_name}: {str(e)}")

    if created_count > 0:
        db.session.commit()

    return (
        jsonify(
            {
                "success": True,
                "created": created_count,
                "skipped": skipped_count,
                "vendors": created_vendors,
                "errors": errors[:10] if errors else [],
            }
        ),
        200,
    )


@application_mgmt.route("/applications/upload-excel", methods=["POST"])
@login_required
def upload_excel_applications():
    """
    Upload and process Excel, CSV, or JSON file for application import.
    Supports duplicate detection (Name + App ID) and merging.
    Accepts optional field_mappings JSON from frontend for custom column mapping.
    """
    import csv
    import io
    import json
    from datetime import datetime

    import openpyxl

    from app.models.application_import_history import ApplicationImportHistory

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = (file.filename or "").lower()
    duplicate_mode = request.form.get("duplicate_mode", "merge")

    # Get custom field mappings from frontend (if provided)
    custom_mappings = {}
    field_mappings_json = request.form.get("field_mappings")
    if field_mappings_json:
        try:
            custom_mappings = json.loads(field_mappings_json)
        except json.JSONDecodeError:
            logger.exception("Failed to parse JSON response")

    # Get pre-computed APQC links from semantic classification (passed from preview)
    # Format: {"row_index": [{"existing_id": 123, "process_code": "3.1", ...}, ...], ...}
    apqc_links_by_row = {}
    apqc_links_json = request.form.get("apqc_links")
    if apqc_links_json:
        try:
            apqc_links_by_row = json.loads(apqc_links_json)
            current_app.logger.info(
                f"Import: Received {len(apqc_links_by_row)} rows with APQC links"
            )
        except json.JSONDecodeError:
            current_app.logger.warning("Import: Failed to parse apqc_links JSON")

    # Get ArchiMate generation option from frontend
    generate_archimate = (
        request.form.get("generate_archimate", "false").lower() == "true"
    )
    archimate_layers = (
        request.form.get("archimate_layers", "").split(",")
        if request.form.get("archimate_layers")
        else None
    )
    if archimate_layers:
        archimate_layers = [l.strip() for l in archimate_layers if l.strip()]
    current_app.logger.info(
        f"Import: ArchiMate generation: {generate_archimate}, layers: {archimate_layers}"
    )

    # Track created/updated application IDs for post-import processing
    created_app_ids = []
    updated_app_ids = []
    apqc_links_by_app = {}

    # NEW: Store original file data for AI analysis
    file_applications_data = []  # Store original file data with app IDs

    # Check if batch processing should be used for large imports
    use_batch_processing = False
    batch_job_id = None
    if len(data_rows) > 100:  # Use batch processing for imports >100 applications
        use_batch_processing = True
        current_app.logger.info(
            f"Large import detected ({len(data_rows)} rows), enabling batch processing"
        )

    # Determine import source based on file type
    if filename.endswith(".csv"):
        import_source = "csv"
    elif filename.endswith(".json"):
        import_source = "json"
    else:
        import_source = "excel"

    # Create import history record
    import_history = ApplicationImportHistory(
        imported_by_id=current_user.id if current_user.is_authenticated else None,
        imported_by_name=current_user.email
        if current_user.is_authenticated
        else "System",
        import_source=import_source,
        file_name=file.filename,
        file_size=len(file.read()),
        duplicate_mode=duplicate_mode,
        status="pending",
    )
    file.seek(0)  # Reset file pointer
    db.session.add(import_history)
    db.session.flush()

    # Create batch job if needed
    if use_batch_processing:
        try:
            from ..services.batch_processing_service import (
                BatchJobConfig,
                BatchJobType,
                BatchProcessingService,
            )

            batch_service = BatchProcessingService()

            # Prepare items for batch processing
            items = []
            for row_idx, row in enumerate(data_rows):
                items.append(
                    {
                        "row_index": row_idx,
                        "data": row,
                        "import_history_id": import_history.id,
                        "duplicate_mode": duplicate_mode,
                        "generate_archimate": generate_archimate,
                    }
                )

            batch_config = BatchJobConfig(
                job_name=f"AI Import - {file.filename}",
                job_type="ai_import",
                items=items,
                confidence_threshold=0.6,
                auto_retry=True,
                max_retries=3,
                parallel_processing=False,
                user_id=current_user.id if current_user.is_authenticated else None,
                config_data={
                    "import_history_id": import_history.id,
                    "file_name": file.filename,
                    "duplicate_mode": duplicate_mode,
                    "generate_archimate": generate_archimate,
                    "archimate_layers": archimate_layers,
                    "field_mappings": custom_mappings,
                    "apqc_links_by_row": apqc_links_by_row,
                },
            )

            batch_job_result = batch_service.create_batch_job(batch_config)
            if batch_job_result["success"]:
                batch_job_id = batch_job_result["batch_job_id"]
                import_history.batch_job_id = batch_job_id
                current_app.logger.info(
                    f"Created batch job {batch_job_id} for large import"
                )
            else:
                current_app.logger.error(
                    f"Failed to create batch job: {batch_job_result.get('error')}"
                )
                use_batch_processing = False  # Fallback to regular processing

        except ImportError as e:
            current_app.logger.error(f"Batch processing service not available: {e}")
            use_batch_processing = False  # Fallback to regular processing
        except Exception as e:
            current_app.logger.error(f"Error creating batch job: {e}")
            use_batch_processing = False  # Fallback to regular processing

    try:
        headers = []
        data_rows = []

        # Parse file based on type
        if filename.endswith(".csv"):
            # Parse CSV file
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)
            headers = reader.fieldnames or []
            data_rows = list(reader)

        elif filename.endswith(".json"):
            # Parse JSON file
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list) and len(data) > 0:
                first_item = data[0] if isinstance(data[0], dict) else {}
                headers = list(first_item.keys())
                data_rows = data
            else:
                return jsonify({"error": "Invalid JSON structure"}), 400

        else:
            # Parse Excel file (default) - handle corrupted stylesheets gracefully
            try:
                wb = openpyxl.load_workbook(file)
            except ValueError as ve:
                # Handle corrupted stylesheet/color issues
                if "stylesheet" in str(ve).lower() or "color" in str(ve).lower():
                    file.seek(0)
                    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                else:
                    raise
            ws = wb.active
            # Strip whitespace from headers to match preview endpoint behavior
            headers = [
                str(cell.value).strip() if cell.value else ""
                for cell in ws[1]
                if cell.value
            ]

        # Find name column using custom mappings or auto-detect
        name_column = None
        app_id_column = None

        # Use custom mappings if provided (from frontend field mapping UI)
        if custom_mappings:
            for header, target_field in custom_mappings.items():
                if target_field == "name":
                    name_column = header
                elif target_field == "application_code":
                    app_id_column = header

        # Fall back to auto-detection if no custom mapping for name
        if not name_column:
            for h in headers:
                h_upper = str(h).upper() if h else ""
                if "NAME" in h_upper and "APP" not in h_upper:
                    name_column = h
                    break
                elif h_upper in ["NAME", "APPLICATION NAME", "APP NAME"]:
                    name_column = h
                    break

        if not app_id_column:
            for h in headers:
                h_upper = str(h).upper() if h else ""
                if "APP ID" in h_upper or h_upper in [
                    "ID",
                    "APPLICATION ID",
                    "APP CODE",
                    "CODE",
                ]:
                    app_id_column = h
                    break

        name_idx = (
            headers.index(name_column)
            if name_column and name_column in headers
            else None
        )

        if name_idx is None and not custom_mappings:
            return (
                jsonify(
                    {
                        "error": 'Name column not found. Please map at least one column to "Name".'
                    }
                ),
                400,
            )

        # Process rows
        records_created = 0
        records_updated = 0
        records_skipped = 0
        records_failed = 0
        errors = []

        processed_identifiers = set()  # Track (name_lower, app_id) combinations

        # Prefetch all existing applications to avoid N+1 query in import loop
        existing_apps_list = ApplicationComponent.query.all()
        existing_apps_by_name = {
            app.name.lower(): app for app in existing_apps_list if app.name
        }

        # Prefetch ArchiMate elements for archimate_element_id lookups
        from ..models.archimate_core import ArchiMateElement

        existing_archimate_list = ArchiMateElement.query.all()
        existing_archimate_by_name = {
            elem.name.lower(): elem for elem in existing_archimate_list if elem.name
        }

        # Handle CSV/JSON imports using data_rows
        if data_rows:
            for row_idx, row_data in enumerate(data_rows, start=2):
                if not isinstance(row_data, dict):
                    continue
                if not any(v for v in row_data.values() if v):
                    continue

                try:
                    # Get name using custom mapping or detected column
                    name = None
                    # Create a mapping of stripped keys to actual keys for robust matching
                    key_mapping = {k.strip(): k for k in row_data.keys()}

                    if custom_mappings:
                        for header, target in custom_mappings.items():
                            if target == "name":
                                stripped_header = header.strip()
                                if stripped_header in key_mapping:
                                    actual_key = key_mapping[stripped_header]
                                    name = (
                                        str(row_data[actual_key]).strip()
                                        if row_data[actual_key]
                                        else None
                                    )
                                    break
                    if not name and name_column:
                        stripped_name_column = name_column.strip()
                        if stripped_name_column in key_mapping:
                            actual_key = key_mapping[stripped_name_column]
                            name = (
                                str(row_data[actual_key]).strip()
                                if row_data[actual_key]
                                else None
                            )

                    if not name:
                        records_failed += 1
                        errors.append(f"Row {row_idx}: Name is required")
                        continue

                    # Get app_id if available
                    app_id = None
                    if app_id_column:
                        stripped_app_id_column = app_id_column.strip()
                        if stripped_app_id_column in key_mapping:
                            actual_key = key_mapping[stripped_app_id_column]
                            app_id = (
                                str(row_data[actual_key]).strip()
                                if row_data[actual_key]
                                else None
                            )

                    # Duplicate detection
                    identifier = (name.lower(), app_id.lower() if app_id else None)
                    if identifier in processed_identifiers:
                        if duplicate_mode == "skip":
                            records_skipped += 1
                            continue

                    # Check existing application in database
                    try:
                        # Use prefetched lookup instead of query to avoid N+1
                        existing_app = existing_apps_by_name.get(name.lower())
                    except Exception as db_error:
                        current_app.logger.error(
                            f"Row {row_idx}: Database error during duplicate check: {db_error}"
                        )
                        # Rollback to clear failed transaction state (critical for PostgreSQL)
                        try:
                            db.session.rollback()
                        except Exception as e:  # fabricated-values-ok
                            logger.debug(f"Ignored: {e}")
                        # Assume it doesn't exist if database query fails
                        existing_app = None

                    # Build app_data using custom mappings
                    app_data = {"name": name}
                    if app_id:
                        app_data["application_code"] = app_id

                    for header, value in row_data.items():
                        if not value:
                            continue
                        # Use stripped header to match custom mappings
                        stripped_header = header.strip()
                        target_field = custom_mappings.get(stripped_header)
                        if target_field and target_field not in [
                            "",
                            "name",
                            "application_code",
                        ]:
                            # Convert integer fields using parse_integer_from_range
                            if target_field in INTEGER_RANGE_FIELDS:
                                app_data[target_field] = parse_integer_from_range(value)
                            else:
                                app_data[target_field] = str(value).strip()

                    # Create or update
                    if existing_app:
                        if duplicate_mode == "skip":
                            records_skipped += 1
                        elif duplicate_mode in ["merge", "update"]:
                            changed_fields = {}
                            for key, val in app_data.items():
                                if key == "name":
                                    continue  # Don't overwrite primary key field
                                if not hasattr(existing_app, key):  # model-safety-ok
                                    continue
                                # Skip empty/falsy values — preserve existing data
                                # Catches None, empty strings, 0, 0.0, False
                                if not val and val is not False:
                                    continue
                                if isinstance(val, str) and not val.strip():
                                    continue
                                old_val = getattr(existing_app, key)  # model-safety-ok
                                if old_val != val:
                                    changed_fields[key] = {
                                        "old": str(old_val) if old_val else None,
                                        "new": str(val),
                                    }
                                    setattr(existing_app, key, val)
                            if changed_fields:
                                current_app.logger.debug(
                                    "Import merge updated %s: %s",
                                    existing_app.name,
                                    changed_fields,
                                )
                            records_updated += 1
                            updated_app_ids.append(existing_app.id)
                            # Track APQC links for this app
                            row_key = str(row_idx)
                            if row_key in apqc_links_by_row:
                                apqc_links_by_app[existing_app.id] = apqc_links_by_row[
                                    row_key
                                ]
                        else:  # duplicate mode
                            new_app = ApplicationComponent(**app_data)
                            db.session.add(new_app)
                            db.session.flush()  # Get the ID
                            records_created += 1
                            created_app_ids.append(new_app.id)

                            # NEW: Store original file data for AI analysis
                            file_app_data = {
                                "application_id": new_app.id,
                                "name": app_data.get("name", ""),
                                "description": app_data.get("description", ""),
                                "vendor_name": app_data.get("vendor_name", ""),
                                "technology_stack": app_data.get(
                                    "technology_stack", ""
                                ),
                                "business_domain": app_data.get("business_domain", ""),
                                "business_functions": app_data.get(
                                    "application_functions_text", ""
                                ),
                                "capabilities": app_data.get(
                                    "imported_capabilities", ""
                                ),
                                "apqc_codes": app_data.get("imported_apqc_codes", ""),
                            }
                            file_applications_data.append(file_app_data)

                            # Track APQC links for this app
                            row_key = str(row_idx)
                            if row_key in apqc_links_by_row:
                                apqc_links_by_app[new_app.id] = apqc_links_by_row[
                                    row_key
                                ]
                    else:
                        new_app = ApplicationComponent(**app_data)
                        db.session.add(new_app)
                        db.session.flush()  # Get the ID
                        records_created += 1
                        created_app_ids.append(new_app.id)

                        # NEW: Store original file data for AI analysis
                        file_app_data = {
                            "application_id": new_app.id,
                            "name": app_data.get("name", ""),
                            "description": app_data.get("description", ""),
                            "vendor_name": app_data.get("vendor_name", ""),
                            "technology_stack": app_data.get("technology_stack", ""),
                            "business_domain": app_data.get("business_domain", ""),
                            "business_functions": app_data.get(
                                "application_functions_text", ""
                            ),
                            "capabilities": app_data.get("imported_capabilities", ""),
                            "apqc_codes": app_data.get("imported_apqc_codes", ""),
                        }
                        file_applications_data.append(file_app_data)

                        # Track APQC links for this app
                        row_key = str(row_idx)
                        if row_key in apqc_links_by_row:
                            apqc_links_by_app[new_app.id] = apqc_links_by_row[row_key]

                    processed_identifiers.add(identifier)

                    # Commit each successful row to prevent loss on subsequent failures
                    try:
                        db.session.commit()
                    except Exception as commit_error:
                        current_app.logger.error(
                            f"Row {row_idx}: Commit failed: {commit_error}"
                        )
                        db.session.rollback()
                        records_failed += 1
                        errors.append(
                            f"Row {row_idx}: Commit failed - {str(commit_error)}"
                        )
                        # Decrement the appropriate counter since this row failed to save
                        if existing_app and duplicate_mode in ["merge", "update"]:
                            records_updated -= 1
                            if existing_app.id in updated_app_ids:
                                updated_app_ids.remove(existing_app.id)
                        elif "new_app" in locals():
                            records_created -= 1
                            if created_app_ids and created_app_ids[-1]:
                                created_app_ids.pop()
                        continue

                except Exception as e:
                    # Row processing failed - rollback only this row's changes
                    try:
                        db.session.rollback()
                    except Exception as e:  # fabricated-values-ok
                        logger.debug(f"Ignored: {e}")
                    records_failed += 1
                    errors.append(f"Row {row_idx}: {str(e)}")
                    continue

            # Final commit for any remaining session changes (import_history, etc.)
            db.session.commit()
            import_history.records_created = records_created
            import_history.records_updated = records_updated
            import_history.records_skipped = records_skipped
            import_history.records_failed = records_failed
            import_history.status = "completed" if records_failed == 0 else "partial"
            import_history.error_summary = "; ".join(errors[:5]) if errors else None
            db.session.commit()

            # AI-powered import analysis if requested (replaces basic ArchiMate generation)
            ai_analysis_stats = None
            if generate_archimate and (created_app_ids or updated_app_ids):
                try:
                    from ..services.ai_import_service import get_ai_import_service

                    ai_service = get_ai_import_service()

                    all_app_ids = created_app_ids + updated_app_ids
                    current_app.logger.info(
                        f"Starting AI Import analysis for {len(all_app_ids)} applications"
                    )

                    ai_analysis_stats = {
                        "processed": 0,
                        "capability_mappings_created": 0,
                        "process_mappings_created": 0,
                        "archimate_elements_created": 0,
                        "high_confidence_mappings": 0,
                        "avg_confidence": 0.0,
                        "ai_models_used": [],
                        "failed": 0,
                        "errors": [],
                        "bulk_analysis_used": True,
                    }

                    try:
                        # Use FILE DATA AI analysis for better context preservation
                        current_app.logger.info(
                            f"Using analyze_file_data_for_preview for {len(file_applications_data)} applications..."
                        )

                        if file_applications_data:
                            # Perform file data AI analysis (preserves original context)
                            bulk_results = ai_service.analyze_file_data_for_preview(
                                applications_data=file_applications_data,
                                confidence_threshold=0.6,
                            )

                            current_app.logger.info(
                                f"File data AI analysis complete: {bulk_results['total_analyzed']} applications analyzed"
                            )

                            # Process bulk results and create mappings
                            for app_result in bulk_results["applications"]:
                                if "error" not in app_result:
                                    app_id = app_result["application_id"]

                                    # Create AI mappings in database
                                    mapping_result = ai_service.create_ai_mappings(
                                        application_id=app_id,
                                        capability_mappings=app_result[
                                            "capability_mappings"
                                        ],
                                        process_mappings=app_result["process_mappings"],
                                        archimate_elements=app_result[
                                            "archimate_elements"
                                        ],
                                        created_by=current_user.email
                                        if current_user.is_authenticated
                                        else "ai_import",
                                    )

                                    # Update statistics
                                    ai_analysis_stats["processed"] += 1
                                    ai_analysis_stats[
                                        "capability_mappings_created"
                                    ] += mapping_result["capability_mappings_created"]
                                    ai_analysis_stats["process_mappings_created"] += (
                                        mapping_result["process_mappings_created"]
                                    )
                                    ai_analysis_stats["archimate_elements_created"] += (
                                        mapping_result["archimate_elements_created"]
                                    )

                                    # Track high confidence mappings
                                    high_conf = app_result.get(
                                        "high_confidence_mappings", 0
                                    )
                                    ai_analysis_stats["high_confidence_mappings"] += (
                                        high_conf
                                    )

                                    # Track average confidence
                                    if (
                                        app_result.get("avg_capability_confidence", 0)
                                        > 0
                                    ):
                                        ai_analysis_stats["avg_confidence"] += (
                                            app_result["avg_capability_confidence"]
                                        )

                                    if mapping_result["errors"]:
                                        ai_analysis_stats["errors"].extend(
                                            mapping_result["errors"][:2]
                                        )
                                else:
                                    ai_analysis_stats["failed"] += 1
                                    ai_analysis_stats["errors"].append(
                                        f"App {app_result.get('application_id', 'unknown')}: {app_result.get('error', 'Unknown error')[:50]}"
                                    )

                            # Update AI models used
                            ai_analysis_stats["ai_models_used"] = bulk_results[
                                "processing_stats"
                            ]["ai_models_used"]

                            # Calculate average confidence
                            if ai_analysis_stats["processed"] > 0:
                                ai_analysis_stats["avg_confidence"] = round(
                                    ai_analysis_stats["avg_confidence"]
                                    / ai_analysis_stats["processed"],
                                    2,
                                )

                            # Add processing stats
                            ai_analysis_stats["bulk_processing_stats"] = {
                                "total_analyzed": bulk_results["total_analyzed"],
                                "capability_mappings_found": bulk_results[
                                    "capability_mappings_found"
                                ],
                                "process_mappings_found": bulk_results[
                                    "process_mappings_found"
                                ],
                                "archimate_elements_generated": bulk_results[
                                    "archimate_elements_generated"
                                ],
                                "avg_processing_time_ms": bulk_results[
                                    "processing_stats"
                                ]["avg_processing_time_ms"],
                            }

                            current_app.logger.info(
                                f"Bulk AI processing complete: {ai_analysis_stats}"
                            )
                        else:
                            current_app.logger.warning(
                                "No applications found for AI analysis"
                            )
                            ai_analysis_stats["errors"].append(
                                "No applications found for AI analysis"
                            )

                    except Exception as bulk_error:
                        current_app.logger.error(
                            f"Bulk AI analysis failed, falling back to individual processing: {bulk_error}"
                        )

                        # Fallback to individual processing
                        total_confidence = 0.0
                        confidence_count = 0

                        for app_id in all_app_ids:
                            try:
                                # Run comprehensive AI analysis
                                ai_result = (
                                    ai_service.analyze_application_for_ai_mapping(
                                        app_id
                                    )
                                )

                                if ai_result.warnings:
                                    ai_analysis_stats["errors"].extend(
                                        ai_result.warnings[:2]
                                    )

                                # Create AI mappings in database
                                mapping_result = ai_service.create_ai_mappings(
                                    application_id=app_id,
                                    capability_mappings=ai_result.capability_mappings,
                                    process_mappings=ai_result.process_mappings,
                                    archimate_elements=ai_result.archimate_elements,
                                    created_by=current_user.email
                                    if current_user.is_authenticated
                                    else "ai_import",
                                )

                                # Update statistics
                                ai_analysis_stats["processed"] += 1
                                ai_analysis_stats["capability_mappings_created"] += (
                                    mapping_result["capability_mappings_created"]
                                )
                                ai_analysis_stats["process_mappings_created"] += (
                                    mapping_result["process_mappings_created"]
                                )
                                ai_analysis_stats["archimate_elements_created"] += (
                                    mapping_result["archimate_elements_created"]
                                )

                                # Track high confidence mappings
                                high_conf = sum(
                                    1
                                    for m in ai_result.capability_mappings
                                    if m.get("confidence_score", 0) >= 0.7
                                )
                                ai_analysis_stats["high_confidence_mappings"] += (
                                    high_conf
                                )

                                # Track average confidence
                                if ai_result.avg_capability_confidence > 0:
                                    total_confidence += (
                                        ai_result.avg_capability_confidence
                                    )
                                    confidence_count += 1

                                # Track AI models used
                                for model in ai_result.ai_models_used:
                                    if model not in ai_analysis_stats["ai_models_used"]:
                                        ai_analysis_stats["ai_models_used"].append(
                                            model
                                        )

                                if mapping_result["errors"]:
                                    ai_analysis_stats["errors"].extend(
                                        mapping_result["errors"][:2]
                                    )

                            except Exception as app_error:
                                ai_analysis_stats["failed"] += 1
                                ai_analysis_stats["errors"].append(
                                    f"App {app_id}: {str(app_error)[:50]}"
                                )
                                current_app.logger.error(
                                    f"AI analysis failed for app {app_id}: {app_error}"
                                )

                        # Calculate average confidence
                        if confidence_count > 0:
                            ai_analysis_stats["avg_confidence"] = round(
                                total_confidence / confidence_count, 2
                            )

                        ai_analysis_stats["bulk_analysis_used"] = False

                    # Trim errors list
                    ai_analysis_stats["errors"] = ai_analysis_stats["errors"][:5]

                    current_app.logger.info(
                        f"AI Import analysis complete: {ai_analysis_stats}"
                    )

                except Exception as ai_error:
                    current_app.logger.error(f"Error in AI import analysis: {ai_error}")
                    ai_analysis_stats = {"error": str(ai_error)}

            return (
                jsonify(
                    {
                        "success": True,
                        "created": records_created,
                        "updated": records_updated,
                        "skipped": records_skipped,
                        "failed": records_failed,
                        "errors": errors[:10],
                        "partial_import": records_failed > 0
                        and (records_created + records_updated) > 0,
                        "ai_analysis": ai_analysis_stats,
                    }
                ),
                200,
            )

        # Excel processing (original code continues below)
        # Create header to index mapping for robust lookup
        header_to_index = {h.strip(): i for i, h in enumerate(headers)}

        app_id_idx = header_to_index.get("APP ID")
        name_idx = None
        if name_column:
            name_idx = header_to_index.get(name_column.strip())

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=False), start=2
        ):
            # Skip empty rows
            if not any(cell.value for cell in row if cell):
                continue

            try:
                # Extract values using key mapping
                name = None
                if name_idx is not None and name_idx < len(row) and row[name_idx].value:
                    name = str(row[name_idx].value).strip()

                app_id = None
                if (
                    app_id_idx is not None
                    and app_id_idx < len(row)
                    and row[app_id_idx].value
                ):
                    app_id = str(row[app_id_idx].value).strip()

                if not name:
                    records_failed += 1
                    errors.append(f"Row {row_idx}: Name is required")
                    continue

                # Create identifier for duplicate detection
                identifier = (name.lower(), app_id.lower() if app_id else None)
                if identifier in processed_identifiers:
                    if duplicate_mode == "skip":
                        records_skipped += 1
                        continue
                    # For merge/duplicate, continue processing

                # Check for existing application
                try:
                    # Use prefetched lookup instead of query to avoid N+1
                    existing_app = existing_apps_by_name.get(name.lower())
                except Exception as db_error:
                    current_app.logger.error(
                        f"Row {row_idx}: Database error during duplicate check: {db_error}"
                    )
                    # Rollback to clear failed transaction state (critical for PostgreSQL)
                    try:
                        db.session.rollback()
                    except Exception as e:  # fabricated-values-ok
                        logger.debug(f"Ignored: {e}")
                    # Assume it doesn't exist if database query fails
                    existing_app = None

                # Map Excel columns to model fields (complete mapping from template)
                app_data = {}
                for i, header in enumerate(headers):
                    if i < len(row) and row[i].value:
                        header_clean = str(header).strip()
                        value = str(row[i].value).strip()

                        # Complete field mapping from Excel template (only valid ApplicationComponent fields)
                        field_mapping = {
                            # 'APP ID': skipped - not a valid model field
                            "Name": "name",
                            "Description": "description",
                            "Application Status": "deployment_status",
                            # 'Managed type': skipped - not a valid model field
                            "Category": "application_category",
                            "Business Criticality": "business_criticality",
                            "Application Capability (Archimate)": "archimate_element_id",  # Will need lookup
                            # 'Access Mode': skipped - not a valid model field
                            "Functionality (Capabilities)": "notes",  # Store in notes for now
                            "Target Users": "user_type",
                            "First Go-Live (year)": "go_live_date",  # Will need date parsing
                            "Lifecycle Status": "lifecycle_status",
                            "Application Weight": "application_weight",
                            "Technology status": "technology_status",
                            "Vendor & Maintenance Risk": "vendor_maintenance_risk",
                            "Move to cloud Suitability": "cloud_suitability",
                            "Major Program Impact": "major_program_impact",
                            "Comments": "notes",
                            "Priority for Action": "priority_for_action",
                            "TIME Destiny (Manual)": "time_destiny",
                            "Retirement date": "retirement_date",
                            "Successor": "replacement_application",
                            "Deployment Scope": "deployment_scope",
                            "Frequency of Usage": "usage_frequency",
                            "Functional Complexity": "functional_complexity",
                            "Main URL": "main_url",
                            "User Satisfaction": "user_satisfaction",
                            "Hosting Environment": "hosting_environment",
                            "Identity Provider": "identity_provider",
                            "Based on": "based_on",
                            "SGEC Landing Zone": "sgec_landing_zone",
                            "SGEC Cloud Pattern": "sgec_cloud_pattern",
                            "SG Data Center": "sg_data_center",
                            "SG Managed Platform": "sg_managed_platform",
                            "Public Cloud Provider": "public_cloud_provider",
                            "Application Platform": "application_platform",
                            "Other Hosting Details": "other_hosting_details",
                            "Interfaces - Number": "interfaces_count",
                            "Interfaces - Description": "interfaces_description",
                            "Development Type": "development_type",
                            "Level of Customization": "level_of_customization",
                            "Development Provider": "development_provider",
                            "Package Name": "package_name",
                            "Package Vendor": "package_vendor",
                            "Programming Languages": "programming_languages",
                            "Source Code Availability": "source_code_availability",
                            "Technical Complexity": "technical_complexity",
                            "Risk Level": "risk_level",
                            "Availability of Documentation": "availability_of_documentation",
                            "Availability of Knowledge": "availability_of_knowledge",
                            "DRP Status": "drp_status",
                            "RPO": "rpo_hours",
                            "RTO": "rto_hours",
                            "Support Hours": "support_hours",
                            "Support Level": "support_level",
                            "Support Region": "support_region",
                            "Maintenance Provider": "maintenance_provider",
                            "Cost Curency": "cost_currency",
                            "Total Run Cost (Auto)": "total_run_cost",
                            "Hardware Cost": "hardware_cost",
                            "Software Cost": "software_cost",
                            "Facilities and Utilities Cost": "facilities_utilities_cost",
                            "Internal Labor Cost": "internal_labor_cost",
                            "External Labor Cost": "external_labor_cost",
                            "External Services Cost": "external_services_cost",
                            "Internal Services Cost": "internal_services_cost",
                            "Telecom Services Cost": "telecom_services_cost",
                            "Other Costs": "other_costs",
                            "IT Unit Managing the App": "it_unit_managing_app",
                            "Application Manager": "application_manager",
                            "App Business Owner": "business_owner",
                            "IT Security Officer": "it_security_officer",
                            "Business Security Officer": "business_security_officer",
                            "Business Unit Owner of the App": "business_unit_owner",
                            "Countries where the App is used": "countries_where_used",
                            "Apps Portal URL": "apps_portal_url",
                            "PSAT Status": "psat_status",
                            "Certified": "certified",
                            "Risk Assessment Status": "risk_assessment_status",
                            "Core Data OK": "core_data_ok",
                            "Operational Data OK": "operational_data_ok",
                            "Data Quality Analysis": "data_quality_analysis",
                            "Certified At": "certified_at",
                            "Certified By": "certified_by",
                            # PCF Process / Functional Capability fields - multiple aliases
                            "Functional Capabilities": "_functional_capabilities",
                            "PCF Processes": "_functional_capabilities",
                            "Business Processes": "_functional_capabilities",
                            "Process Support": "_functional_capabilities",
                            "Capabilities": "_capabilities",  # Separate for business capabilities
                            "Capability": "_capabilities",
                            "Business Capabilities": "_capabilities",
                            "Supported Capabilities": "_capabilities",
                            "Business Process": "_functional_capabilities",
                            "Process": "_functional_capabilities",
                            "PCF": "_functional_capabilities",
                            "PCF Process": "_functional_capabilities",
                            "Processes": "_functional_capabilities",
                        }

                        if header_clean in field_mapping:
                            field_name = field_mapping[header_clean]
                            # Type conversion for numeric and boolean fields
                            if field_name in [
                                "application_weight",
                                "total_run_cost",
                                "hardware_cost",
                                "software_cost",
                                "facilities_utilities_cost",
                                "internal_labor_cost",
                                "external_labor_cost",
                                "external_services_cost",
                                "internal_services_cost",
                                "telecom_services_cost",
                                "other_costs",
                            ]:
                                try:
                                    app_data[field_name] = (
                                        float(value.replace(",", "")) if value else None
                                    )
                                except (ValueError, TypeError):
                                    app_data[field_name] = None
                            elif field_name in INTEGER_RANGE_FIELDS:
                                # Use parse_integer_from_range to handle ranges like "101 - 500"
                                app_data[field_name] = parse_integer_from_range(value)
                            elif field_name in [
                                "certified",
                                "core_data_ok",
                                "operational_data_ok",
                            ]:
                                app_data[field_name] = str(value).lower() in [
                                    "true",
                                    "yes",
                                    "1",
                                    "y",
                                ]
                            elif field_name in [
                                "go_live_date",
                                "retirement_date",
                                "certified_at",
                            ]:
                                # Date parsing - try multiple formats
                                try:
                                    # Try dateutil first
                                    try:
                                        from dateutil import parser

                                        app_data[field_name] = parser.parse(
                                            value
                                        ).date()
                                    except ImportError:
                                        # Fallback to datetime.strptime
                                        for fmt in [
                                            "%Y-%m-%d",
                                            "%d/%m/%Y",
                                            "%m/%d/%Y",
                                            "%Y",
                                        ]:
                                            try:
                                                if (
                                                    fmt == "%Y"
                                                    and value.isdigit()
                                                    and len(value) == 4
                                                ):
                                                    app_data[field_name] = datetime(
                                                        int(value), 1, 1
                                                    ).date()
                                                    break
                                                else:
                                                    app_data[field_name] = (
                                                        datetime.strptime(
                                                            value, fmt
                                                        ).date()
                                                    )
                                                    break
                                            except (ValueError, TypeError):
                                                continue
                                        else:
                                            app_data[field_name] = None
                                except (ValueError, TypeError):
                                    app_data[field_name] = None
                            elif field_name == "archimate_element_id":
                                # Special handling: lookup ArchiMate element by name
                                if value:
                                    # Use prefetched lookup instead of query to avoid N+1
                                    archimate_element = existing_archimate_by_name.get(
                                        value.lower()
                                    )
                                    if archimate_element:
                                        app_data[field_name] = archimate_element.id
                                    else:
                                        # Try partial match in prefetched data
                                        value_lower = value.lower()
                                        archimate_element = next(
                                            (
                                                elem
                                                for name, elem in existing_archimate_by_name.items()
                                                if value_lower in name
                                            ),
                                            None,
                                        )
                                        if archimate_element:
                                            app_data[field_name] = archimate_element.id
                                        else:
                                            # Leave as None if not found
                                            app_data[field_name] = None
                                else:
                                    app_data[field_name] = None
                            elif field_name in [
                                "access_mode",  # Added due to original VARCHAR(50) error
                                "interface_type",
                                "protocol",
                                "data_format",
                                "message_pattern",
                                "authentication_method",
                                "authorization_model",
                                "sla_uptime_hours",
                                "data_classification",
                                "api_version_strategy",
                                "pricing_tier",
                                "integration_pattern",
                                "event_type",
                                "trigger_type",
                                "message_broker",
                                "schema_format",
                                "delivery_guarantee",
                                "ordering_guarantee",
                                "collaboration_type",
                                "architecture_pattern",
                                "orchestration_pattern",
                                "service_mesh",
                                "api_gateway",
                                "service_discovery",
                                "container_platform",
                                "primary_protocol",
                                "sync_async_mix",
                                "component_type",
                                "architecture_style",
                                "version",
                                "repository_type",
                                "deployment_model",
                                "user_type",
                            ]:
                                # VARCHAR(50) constraint - truncate long values
                                if value and len(value) > 50:
                                    app_data[field_name] = (
                                        value[:47] + "..."
                                    )  # Truncate with ellipsis
                                else:
                                    app_data[field_name] = value
                            elif field_name in FIELD_MAX_LENGTHS:
                                # Use defined max lengths for fields with known constraints
                                max_len = FIELD_MAX_LENGTHS[field_name]
                                if value and len(value) > max_len:
                                    app_data[field_name] = value[: max_len - 3] + "..."
                                else:
                                    app_data[field_name] = value
                            else:
                                # Default safety truncation for unknown string fields (255 chars)
                                if isinstance(value, str) and len(value) > 255:
                                    app_data[field_name] = value[:252] + "..."
                                else:
                                    app_data[field_name] = value

                # Extract functional capabilities and business capabilities before creating/updating app
                functional_capabilities = app_data.pop("_functional_capabilities", None)
                business_capabilities = app_data.pop("_capabilities", None)

                if existing_app and duplicate_mode == "merge":
                    # Update existing
                    changed_fields = {}
                    for key, value in app_data.items():
                        if not hasattr(existing_app, key):  # model-safety-ok
                            continue
                        # Skip empty/falsy values — preserve existing data
                        # Catches None, empty strings, 0, 0.0, False
                        if not value and value is not False:
                            continue
                        if isinstance(value, str) and not value.strip():
                            continue
                        old_val = getattr(existing_app, key)  # model-safety-ok
                        if old_val != value:
                            changed_fields[key] = {
                                "old": str(old_val) if old_val else None,
                                "new": str(value),
                            }
                            setattr(existing_app, key, value)
                    if changed_fields:
                        current_app.logger.debug(
                            "Import merge updated %s: %s",
                            existing_app.name,
                            changed_fields,
                        )
                    records_updated += 1
                    target_app = existing_app
                    updated_app_ids.append(existing_app.id)
                    # Track APQC links for this app
                    row_key = str(row_idx)
                    if row_key in apqc_links_by_row:
                        apqc_links_by_app[existing_app.id] = apqc_links_by_row[row_key]
                elif existing_app and duplicate_mode == "skip":
                    records_skipped += 1
                    continue
                else:
                    # Create new
                    app = ApplicationComponent(**app_data)
                    db.session.add(app)
                    db.session.flush()  # Get the ID for relationship creation
                    records_created += 1
                    target_app = app
                    created_app_ids.append(app.id)
                    # Track APQC links for this app
                    row_key = str(row_idx)
                    if row_key in apqc_links_by_row:
                        apqc_links_by_app[app.id] = apqc_links_by_row[row_key]

                # Link to PCF processes / functional capabilities (fuzzy string matching fallback)
                if functional_capabilities and target_app:
                    _link_application_to_processes(target_app, functional_capabilities)

                # Link to business capabilities
                if business_capabilities and target_app:
                    _link_application_to_capabilities(target_app, business_capabilities)

                # Link APQC processes using pre-computed semantic matches (AI-powered)
                # These come from the preview/analyze endpoint's semantic classification
                row_key = str(row_idx)
                if row_key in apqc_links_by_row and target_app:
                    apqc_matches = apqc_links_by_row[row_key]
                    if apqc_matches:
                        link_result = _link_application_to_apqc_by_ids(
                            target_app, apqc_matches
                        )
                        if link_result["linked"] > 0:
                            current_app.logger.info(
                                f"Row {row_idx}: AI-linked {link_result['linked']} APQC processes "
                                f"(skipped: {link_result['skipped']})"
                            )

                processed_identifiers.add(identifier)

                # Commit each successful row to prevent loss on subsequent failures
                # This ensures previously processed rows are saved even if later rows fail
                try:
                    db.session.commit()
                except Exception as commit_error:
                    current_app.logger.error(
                        f"Row {row_idx}: Commit failed: {commit_error}"
                    )
                    db.session.rollback()
                    records_failed += 1
                    errors.append(f"Row {row_idx}: Commit failed - {str(commit_error)}")
                    # Decrement the appropriate counter since this row failed to save
                    if existing_app and duplicate_mode in ["merge", "update"]:
                        records_updated -= 1
                        if existing_app.id in updated_app_ids:
                            updated_app_ids.remove(existing_app.id)
                    else:
                        records_created -= 1
                        if created_app_ids and created_app_ids[-1]:
                            created_app_ids.pop()
                    continue

            except Exception as e:
                # Row processing failed - rollback only this row's changes, not all previous rows
                # Since we commit each successful row above, rollback here only affects the current failed row
                try:
                    db.session.rollback()
                except Exception as e:  # fabricated-values-ok
                    logger.debug(
                        f"Ignored: {e}"
                    )  # Rollback may fail if session is already clean
                records_failed += 1
                errors.append(f"Row {row_idx}: {str(e)}")
                current_app.logger.error(f"Error processing row {row_idx}: {e}")

        # Update import history with comprehensive audit data
        import_history.total_records = (
            records_created + records_updated + records_skipped + records_failed
        )
        import_history.records_created = records_created
        import_history.records_updated = records_updated
        import_history.records_skipped = records_skipped
        import_history.records_failed = records_failed
        import_history.status = "completed" if records_failed == 0 else "partial"
        import_history.error_summary = (
            f"{records_failed} records failed" if records_failed > 0 else None
        )
        import_history.error_details = json.dumps(errors) if errors else None

        # Store import settings and statistics for audit trail
        import_settings = {
            "duplicate_mode": duplicate_mode,
            "generate_archimate": generate_archimate,
            "ai_analysis_enabled": bool(ai_analysis_stats),
            "file_type": filename.split(".")[-1] if filename else "unknown",
            "processing_time_seconds": (
                datetime.utcnow() - import_history.imported_at
            ).total_seconds()
            if import_history.imported_at
            else None,
        }

        # Include AI analysis statistics in import settings
        if ai_analysis_stats:
            import_settings["ai_analysis"] = ai_analysis_stats

        import_history.import_settings = json.dumps(import_settings)

        # Link created/updated application IDs to this import batch for traceability
        if created_app_ids or updated_app_ids:
            linked_applications = {
                "created_ids": created_app_ids,
                "updated_ids": updated_app_ids,
                "total_processed": len(created_app_ids) + len(updated_app_ids),
            }
            # Store in error_details field as JSON for now (could be separate table in future)
            existing_details = json.loads(import_history.error_details or "{}")
            if isinstance(existing_details, list):
                # Convert to dict if it's a list of errors
                existing_details = {"errors": existing_details}
            existing_details["linked_applications"] = linked_applications
            import_history.error_details = json.dumps(existing_details)

        db.session.commit()

        # AI-powered import analysis if requested (Excel path - replaces basic ArchiMate generation)
        ai_analysis_stats = None
        if generate_archimate and (created_app_ids or updated_app_ids):
            try:
                from ..services.ai_import_service import get_ai_import_service

                ai_service = get_ai_import_service()

                all_app_ids = created_app_ids + updated_app_ids
                current_app.logger.info(
                    f"Starting AI Import analysis (Excel) for {len(all_app_ids)} applications"
                )

                ai_analysis_stats = {
                    "processed": 0,
                    "capability_mappings_created": 0,
                    "process_mappings_created": 0,
                    "archimate_elements_created": 0,
                    "high_confidence_mappings": 0,
                    "avg_confidence": 0.0,
                    "ai_models_used": [],
                    "failed": 0,
                    "errors": [],
                }

                total_confidence = 0.0
                confidence_count = 0

                for app_id in all_app_ids:
                    try:
                        # Run comprehensive AI analysis
                        ai_result = ai_service.analyze_application_for_ai_mapping(
                            app_id
                        )

                        if ai_result.warnings:
                            ai_analysis_stats["errors"].extend(ai_result.warnings[:2])

                        # Create AI mappings in database
                        mapping_result = ai_service.create_ai_mappings(
                            application_id=app_id,
                            capability_mappings=ai_result.capability_mappings,
                            process_mappings=ai_result.process_mappings,
                            archimate_elements=ai_result.archimate_elements,
                            created_by=current_user.email
                            if current_user.is_authenticated
                            else "ai_import",
                        )

                        # Update statistics
                        ai_analysis_stats["processed"] += 1
                        ai_analysis_stats["capability_mappings_created"] += (
                            mapping_result["capability_mappings_created"]
                        )
                        ai_analysis_stats["process_mappings_created"] += mapping_result[
                            "process_mappings_created"
                        ]
                        ai_analysis_stats["archimate_elements_created"] += (
                            mapping_result["archimate_elements_created"]
                        )

                        # Track high confidence mappings
                        high_conf = sum(
                            1
                            for m in ai_result.capability_mappings
                            if m.get("confidence_score", 0) >= 0.7
                        )
                        ai_analysis_stats["high_confidence_mappings"] += high_conf

                        # Track average confidence
                        if ai_result.avg_capability_confidence > 0:
                            total_confidence += ai_result.avg_capability_confidence
                            confidence_count += 1

                        # Track AI models used
                        for model in ai_result.ai_models_used:
                            if model not in ai_analysis_stats["ai_models_used"]:
                                ai_analysis_stats["ai_models_used"].append(model)

                        if mapping_result["errors"]:
                            ai_analysis_stats["errors"].extend(
                                mapping_result["errors"][:2]
                            )

                    except Exception as app_error:
                        ai_analysis_stats["failed"] += 1
                        ai_analysis_stats["errors"].append(
                            f"App {app_id}: {str(app_error)[:50]}"
                        )
                        current_app.logger.error(
                            f"AI analysis failed for app {app_id} (Excel): {app_error}"
                        )

                # Calculate average confidence
                if confidence_count > 0:
                    ai_analysis_stats["avg_confidence"] = round(
                        total_confidence / confidence_count, 2
                    )

                # Trim errors list
                ai_analysis_stats["errors"] = ai_analysis_stats["errors"][:5]

                current_app.logger.info(
                    f"AI Import analysis complete (Excel): {ai_analysis_stats}"
                )

            except Exception as ai_error:
                current_app.logger.error(
                    f"Error in AI import analysis (Excel): {ai_error}"
                )
                ai_analysis_stats = {"error": str(ai_error)}

        return (
            jsonify(
                {
                    "success": True,
                    "created": records_created,
                    "updated": records_updated,
                    "skipped": records_skipped,
                    "failed": records_failed,
                    "errors": errors[:10],  # Return first 10 errors
                    "ai_analysis": ai_analysis_stats,
                }
            ),
            200,
        )

    except Exception as e:
        db.session.rollback()
        import_history.status = "failed"
        import_history.error_summary = str(e)
        db.session.commit()
        current_app.logger.error(f"Error importing Excel: {e}")
        return jsonify({"error": "An internal error occurred"}), 500


@application_mgmt.route("/applications/import-manual", methods=["POST"])
@login_required
def import_manual_applications():
    """Process manual entry applications"""
    import json

    from app.models.application_import_history import ApplicationImportHistory

    data = request.get_json()
    applications = data.get("applications", [])
    duplicate_mode = data.get("duplicate_mode", "merge")

    # Create import history
    import_history = ApplicationImportHistory(
        imported_by_id=current_user.id if current_user.is_authenticated else None,
        imported_by_name=current_user.email
        if current_user.is_authenticated
        else "System",
        import_source="manual",
        duplicate_mode=duplicate_mode,
        status="pending",
    )
    db.session.add(import_history)
    db.session.flush()

    records_created = 0
    records_updated = 0
    records_skipped = 0
    records_failed = 0
    errors = []

    # Prefetch all existing applications to avoid N+1 query in loop
    existing_apps_list = ApplicationComponent.query.all()
    existing_apps_by_name = {
        app.name.lower(): app for app in existing_apps_list if app.name
    }

    for idx, app_data in enumerate(applications, start=1):
        try:
            name = app_data.get("name", "").strip()
            app_id = app_data.get("app_id", "").strip() or None

            if not name:
                records_failed += 1
                errors.append(f"Row {idx}: Name is required")
                continue

            # Check for duplicates
            try:
                # Ensure clean transaction state before query
                db.session.commit()
            except Exception:
                db.session.rollback()

            try:
                # Use prefetched lookup instead of query to avoid N+1
                existing_app = existing_apps_by_name.get(name.lower())
            except Exception as db_error:
                current_app.logger.error(
                    f"Row {idx}: Database error during duplicate check: {db_error}"
                )
                # Assume it doesn't exist if database query fails
                existing_app = None

            if existing_app and duplicate_mode == "merge":
                # Update existing
                changed_fields = {}
                for key, value in app_data.items():
                    if not hasattr(existing_app, key):  # model-safety-ok
                        continue
                    # Skip empty/falsy values — preserve existing data
                    # Catches None, empty strings, 0, 0.0, False
                    if not value and value is not False:
                        continue
                    if isinstance(value, str) and not value.strip():
                        continue
                    old_val = getattr(existing_app, key)  # model-safety-ok
                    if old_val != value:
                        changed_fields[key] = {
                            "old": str(old_val) if old_val else None,
                            "new": str(value),
                        }
                        setattr(existing_app, key, value)
                if changed_fields:
                    current_app.logger.debug(
                        "Import merge updated %s: %s", existing_app.name, changed_fields
                    )
                records_updated += 1
            elif existing_app and duplicate_mode == "skip":
                records_skipped += 1
                continue
            else:
                # Create new - filter to only valid ApplicationComponent fields
                valid_fields = {
                    col.name for col in ApplicationComponent.__table__.columns
                }
                filtered_data = {
                    k: v for k, v in app_data.items() if k in valid_fields and k != "id"
                }
                app = ApplicationComponent(**filtered_data)
                db.session.add(app)
                records_created += 1

        except Exception as e:
            records_failed += 1
            errors.append(f"Row {idx}: {str(e)}")

    # Update import history with comprehensive audit data
    import_history.total_records = len(applications)
    import_history.records_created = records_created
    import_history.records_updated = records_updated
    import_history.records_skipped = records_skipped
    import_history.records_failed = records_failed
    import_history.status = "completed" if records_failed == 0 else "partial"
    import_history.error_summary = (
        f"{records_failed} records failed" if records_failed > 0 else None
    )
    import_history.error_details = json.dumps(errors) if errors else None

    # Store import settings and statistics for audit trail
    import_settings = {
        "duplicate_mode": duplicate_mode,
        "import_source": "manual",
        "processing_time_seconds": (
            datetime.utcnow() - import_history.imported_at
        ).total_seconds()
        if import_history.imported_at
        else None,
        "total_applications": len(applications),
    }

    # Link created/updated application IDs to this import batch for traceability
    if created_app_ids or updated_app_ids:
        linked_applications = {
            "created_ids": created_app_ids,
            "updated_ids": updated_app_ids,
            "total_processed": len(created_app_ids) + len(updated_app_ids),
        }
        import_settings["linked_applications"] = linked_applications

    import_history.import_settings = json.dumps(import_settings)

    db.session.commit()

    # If using batch processing, start the batch job and return job info
    if use_batch_processing and batch_job_id:
        try:
            from ..services.batch_processing_service import BatchProcessingService

            batch_service = BatchProcessingService()

            # Add all applications to batch job for processing
            for idx, app_data in enumerate(file_applications_data):
                batch_service.add_batch_item(
                    batch_job_id=batch_job_id,
                    item_data=app_data,
                    item_type="application_import",
                    priority=1,  # Normal priority
                )

            # Start the batch job
            start_result = batch_service.start_batch_job(batch_job_id)
            if start_result["success"]:
                current_app.logger.info(
                    f"Started batch job {batch_job_id} for import processing"
                )

                return (
                    jsonify(
                        {
                            "success": True,
                            "batch_processing": True,
                            "batch_job_id": batch_job_id,
                            "message": f"Large import started as batch job {batch_job_id}",
                            "total_records": len(applications),
                            "status_url": f"/api/batch/jobs/{batch_job_id}/progress",
                        }
                    ),
                    200,
                )
            else:
                current_app.logger.error(
                    f"Failed to start batch job: {start_result.get('error')}"
                )
                # Fallback to regular response
                use_batch_processing = False

        except Exception as e:
            current_app.logger.error(f"Error starting batch job: {e}")
            # Fallback to regular response
            use_batch_processing = False

    return (
        jsonify(
            {
                "success": True,
                "batch_processing": use_batch_processing,
                "batch_job_id": batch_job_id,
                "created": records_created,
                "updated": records_updated,
                "skipped": records_skipped,
                "failed": records_failed,
                "errors": errors,
            }
        ),
        200,
    )


@application_mgmt.route("/applications/import-history", methods=["GET"])
@login_required
def get_import_history():
    """Get comprehensive import history with audit trail"""
    import json

    from app.models.application_import_history import ApplicationImportHistory

    history = (
        ApplicationImportHistory.query.order_by(
            ApplicationImportHistory.imported_at.desc()
        )
        .limit(50)
        .all()
    )

    # Enhanced history with audit data
    enhanced_history = []
    for h in history:
        history_dict = h.to_dict()

        # Parse import settings for additional context
        if h.import_settings:
            try:
                settings = json.loads(h.import_settings)
                history_dict.update(
                    {
                        "processing_time_seconds": settings.get(
                            "processing_time_seconds"
                        ),
                        "file_type": settings.get("file_type"),
                        "ai_analysis_enabled": settings.get(
                            "ai_analysis_enabled", False
                        ),
                        "duplicate_mode": settings.get("duplicate_mode"),
                        "linked_applications": settings.get("linked_applications", {}),
                        "ai_analysis_stats": settings.get("ai_analysis"),
                    }
                )
            except json.JSONDecodeError:
                history_dict["import_settings_parse_error"] = True

        # Parse error details to separate errors from linked applications
        if h.error_details:
            try:
                details = json.loads(h.error_details)
                if isinstance(details, dict):
                    history_dict["errors"] = details.get("errors", [])
                    history_dict["linked_applications"] = details.get(
                        "linked_applications",
                        history_dict.get("linked_applications", {}),
                    )
                else:
                    history_dict["errors"] = details
            except json.JSONDecodeError:
                history_dict["errors"] = [h.error_details] if h.error_details else []

        # Add rollback eligibility (within 7 days for completed imports)
        if h.status == "completed" and h.imported_at:
            days_since_import = (datetime.utcnow() - h.imported_at).days
            history_dict["can_rollback"] = days_since_import <= 7
        else:
            history_dict["can_rollback"] = False

        enhanced_history.append(history_dict)

    return jsonify({"history": enhanced_history}), 200


@application_mgmt.route("/applications/download-template", methods=["GET"])
@login_required
def download_application_template():
    """Download Excel template with current application data"""
    from io import BytesIO

    import openpyxl

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Get template headers from the actual template file
    try:
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "Uploads",
            "Appliations Template.xlsx",
        )
        template_wb = openpyxl.load_workbook(template_path)
        template_ws = template_wb.active
        headers = [cell.value for cell in template_ws[1] if cell.value]
    except Exception:
        current_app.logger.debug("Failed to load template", exc_info=True)
        # Fallback headers if template not found
        headers = [
            "APP ID",
            "Name",
            "Description",
            "Application Status",
            "Managed type",
            "Category",
            "Business Criticality",
        ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Get current applications and write data
    applications = ApplicationComponent.query.limit(1000).all()

    # Map model fields to Excel columns (simplified - expand as needed)
    field_mapping = {
        "app_id": "APP ID",
        "name": "Name",
        "description": "Description",
        "deployment_status": "Application Status",
        "managed_type": "Managed type",
        "application_category": "Category",
        "business_criticality": "Business Criticality",
    }

    for row_idx, app in enumerate(applications, start=2):
        for col_idx, header in enumerate(headers, start=1):
            # Find matching field
            field_name = None
            for model_field, excel_header in field_mapping.items():
                if excel_header == header:
                    field_name = model_field
                    break

            if field_name and hasattr(app, field_name):
                value = getattr(app, field_name)
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"applications_template_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )

