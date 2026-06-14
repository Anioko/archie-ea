"""RATA-018: Rationalization business case export service.

Generates PDF (HTML), Excel (openpyxl), and CSV exports of portfolio rationalization data.
"""

import csv
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class RationalizationExportService:
    """Generate rationalization business case exports."""

    @staticmethod
    def get_scored_apps(scope=None):
        """Fetch scored apps with optional disposition filter."""
        from app import db
        from app.models.application_portfolio import ApplicationComponent
        from app.models.application_rationalization import ApplicationRationalizationScore

        query = (
            db.session.query(ApplicationRationalizationScore, ApplicationComponent)
            .join(
                ApplicationComponent,
                ApplicationRationalizationScore.application_component_id == ApplicationComponent.id,
            )
        )

        if scope and scope.get("disposition"):
            dispositions = scope["disposition"]
            if isinstance(dispositions, str):
                dispositions = [d.strip() for d in dispositions.split(",")]
            query = query.filter(
                ApplicationRationalizationScore.disposition_action.in_(dispositions)
            )

        return query.all()

    @staticmethod
    def generate_csv(scope=None):
        """Generate CSV export of scored applications."""
        rows = RationalizationExportService.get_scored_apps(scope)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "App ID", "App Name", "Overall Score", "Technical Health",
            "Business Value", "Cost Efficiency", "Vendor Risk",
            "TIME Action", "Disposition", "Confidence",
            "Estimated Annual Savings", "Review Status",
        ])

        for score, app in rows:
            writer.writerow([
                app.id,
                app.name,
                score.overall_health_score,
                score.technical_health_score,
                score.business_value_score,
                score.cost_efficiency_score,
                score.vendor_risk_score,
                score.rationalization_action,
                score.disposition_action,
                score.disposition_confidence,
                float(score.estimated_annual_savings or 0),
                score.review_status,
            ])

        output.seek(0)
        return output.getvalue().encode("utf-8")

    @staticmethod
    def generate_excel(scope=None):
        """Generate Excel workbook with multiple sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Sheet 1: All Scored Apps
        ws = wb.active
        ws.title = "Scored Applications"
        headers = [
            "App ID", "App Name", "Overall Score", "Technical Health",
            "Business Value", "Cost Efficiency", "Vendor Risk",
            "TIME Action", "Disposition", "Confidence",
            "Est. Annual Savings", "Review Status",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        rows = RationalizationExportService.get_scored_apps(scope)
        for score, app in rows:
            ws.append([
                app.id,
                app.name,
                score.overall_health_score,
                score.technical_health_score,
                score.business_value_score,
                score.cost_efficiency_score,
                score.vendor_risk_score,
                score.rationalization_action,
                score.disposition_action,
                score.disposition_confidence,
                float(score.estimated_annual_savings or 0),
                score.review_status,
            ])

        # RAG conditional formatting
        from openpyxl.styles import Font as XlFont
        for row_idx in range(2, ws.max_row + 1):
            score_cell = ws.cell(row=row_idx, column=3)
            if score_cell.value is not None:
                if score_cell.value >= 70:
                    score_cell.font = XlFont(color="16A34A")
                elif score_cell.value >= 40:
                    score_cell.font = XlFont(color="D97706")
                else:
                    score_cell.font = XlFont(color="DC2626")

        # Auto-size columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        ws.freeze_panes = "A2"

        # Sheet 2: Summary
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Metric", "Value"])
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].font = Font(bold=True)

        total = len(rows)
        time_dist = {}
        total_savings = 0
        for score, app in rows:
            action = score.rationalization_action or "UNKNOWN"
            time_dist[action] = time_dist.get(action, 0) + 1
            total_savings += float(score.estimated_annual_savings or 0)

        ws2.append(["Total Scored Applications", total])
        ws2.append(["Total Projected Savings", total_savings])
        ws2.append([])
        ws2.append(["TIME Distribution", ""])
        for action in ["TOLERATE", "INVEST", "MIGRATE", "ELIMINATE"]:
            ws2.append([action, time_dist.get(action, 0)])

        ws2.append([])
        ws2.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_pdf(scope=None):
        """Generate PDF as styled HTML (ReportLab not available)."""
        rows = RationalizationExportService.get_scored_apps(scope)

        total = len(rows)
        time_dist = {}
        total_savings = 0
        for score, app in rows:
            action = score.rationalization_action or "UNKNOWN"
            time_dist[action] = time_dist.get(action, 0) + 1
            total_savings += float(score.estimated_annual_savings or 0)

        # Top 10 candidates by savings
        sorted_rows = sorted(rows, key=lambda r: float(r[0].estimated_annual_savings or 0), reverse=True)[:10]

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Rationalization Business Case</title>
<style>
body {{ font-family: system-ui, sans-serif; max-width: 800px; margin: 0 auto; padding: 40px; color: #1a1a1a; }}
h1 {{ font-size: 24px; border-bottom: 2px solid #3b82f6; padding-bottom: 8px; }}
h2 {{ font-size: 18px; margin-top: 30px; color: #374151; }}
table {{ width: 100%; border-collapse: collapse; margin: 16px 0; font-size: 13px; }}
th {{ background: #f1f5f9; text-align: left; padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; }}
td {{ padding: 8px 12px; border: 1px solid #e2e8f0; }}
.metric {{ display: inline-block; background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px 24px; margin: 8px; text-align: center; }}
.metric-value {{ font-size: 28px; font-weight: 700; color: #1e40af; }}
.metric-label {{ font-size: 12px; color: #6b7280; margin-top: 4px; }}
.footer {{ margin-top: 40px; padding-top: 16px; border-top: 1px solid #e2e8f0; font-size: 11px; color: #9ca3af; }}
</style></head><body>
<h1>Application Rationalization — Business Case</h1>

<div style="display: flex; flex-wrap: wrap;">
<div class="metric"><div class="metric-value">{total}</div><div class="metric-label">Apps Scored</div></div>
<div class="metric"><div class="metric-value">{total_savings:,.0f}</div><div class="metric-label">Projected Savings</div></div>
<div class="metric"><div class="metric-value">{time_dist.get('ELIMINATE', 0)}</div><div class="metric-label">Eliminate</div></div>
<div class="metric"><div class="metric-value">{time_dist.get('MIGRATE', 0)}</div><div class="metric-label">Migrate</div></div>
</div>

<h2>TIME Distribution</h2>
<table>
<tr><th>Action</th><th>Count</th><th>%</th></tr>
"""
        for action in ["TOLERATE", "INVEST", "MIGRATE", "ELIMINATE"]:
            count = time_dist.get(action, 0)
            pct = round(count / total * 100, 1) if total > 0 else 0
            html += f"<tr><td>{action}</td><td>{count}</td><td>{pct}%</td></tr>\n"

        html += """</table>
<h2>Top 10 Rationalization Candidates</h2>
<table>
<tr><th>Application</th><th>Score</th><th>Disposition</th><th>Est. Savings</th></tr>
"""
        for score, app in sorted_rows:
            html += f"<tr><td>{app.name}</td><td>{score.overall_health_score}</td><td>{score.disposition_action or '—'}</td><td>{float(score.estimated_annual_savings or 0):,.0f}</td></tr>\n"

        html += f"""</table>
<div class="footer">Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} by A.R.C.H.I.E. Platform</div>
</body></html>"""

        return html.encode("utf-8")
