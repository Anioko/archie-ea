"""
Export Service for Vendor Analysis

Generates Excel and PDF exports of vendor analysis results for executive presentations
and stakeholder review.
"""

import io
import json
import logging
from datetime import datetime
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import OptionsAnalysis, VendorOption

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting vendor analysis results to various formats."""

    def __init__(self):
        """Initialize export service."""
        pass

    def export_to_excel(self, analysis: OptionsAnalysis) -> io.BytesIO:
        """
        Export vendor analysis to Excel workbook.

        Creates a comprehensive Excel workbook with:
        - Executive Summary
        - Vendor Comparison Matrix
        - Detailed Scoring
        - TCO Breakdown
        - Charts and Visualizations

        Args:
            analysis: OptionsAnalysis to export

        Returns:
            BytesIO: Excel file in memory
        """
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_executive_summary(wb, analysis)
        self._create_comparison_matrix(wb, analysis)
        self._create_detailed_scoring(wb, analysis)
        self._create_tco_breakdown(wb, analysis)
        self._create_charts(wb, analysis)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _create_executive_summary(self, wb: openpyxl.Workbook, analysis: OptionsAnalysis):
        """Create Executive Summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)

        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)

        # Title
        ws["A1"] = "Vendor Analysis Report"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:E1")

        # Analysis Info
        row = 3
        info_items = [
            ("Analysis Name:", analysis.name),
            ("Capability:", analysis.capability.name if analysis.capability else "N/A"),
            ("Created By:", analysis.created_by.name if analysis.created_by else "Unknown"),
            (
                "Created Date:",
                analysis.created_at.strftime("%Y-%m-%d") if analysis.created_at else "N/A",
            ),
            ("Status:", analysis.status),
            ("Analysis Type:", analysis.analysis_type),
            ("Vendors Analyzed:", str(analysis.total_vendors_analyzed)),
        ]

        # Add context info if available
        if analysis.organization_size:
            info_items.append(("Organization Size:", analysis.organization_size))
        if analysis.industry_vertical:
            info_items.append(("Industry:", analysis.industry_vertical))
        if analysis.deployment_scale:
            info_items.append(("Deployment Scale:", analysis.deployment_scale))

        for label, value in info_items:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Winner section
        row += 2
        ws[f"A{row}"] = "RECOMMENDED VENDOR"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws.merge_cells(f"A{row}:E{row}")

        row += 1
        winner = analysis.get_winner()
        if winner:
            ws[f"A{row}"] = winner.vendor_name
            ws[f"A{row}"].font = Font(bold=True, size=14, color="006100")
            ws[f"B{row}"] = f"Score: {winner.total_score:.1f}/100"
            ws[f"C{row}"] = f"TCO: ${winner.tco_total:,.0f}" if winner.tco_total else "N/A"
        else:
            ws[f"A{row}"] = "No recommendation available"

        # Criteria Weights
        row += 3
        ws[f"A{row}"] = "SCORING CRITERIA WEIGHTS"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        weights = analysis.get_criteria_weights()
        for criterion, weight in weights.items():
            ws[f"A{row}"] = criterion.replace("_", " ").title()
            ws[f"B{row}"] = f"{weight * 100:.0f}%"
            row += 1

        # Auto-size columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20

    def _create_comparison_matrix(self, wb: openpyxl.Workbook, analysis: OptionsAnalysis):
        """Create Vendor Comparison Matrix sheet."""
        ws = wb.create_sheet("Comparison Matrix")

        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "Rank",
            "Vendor",
            "Total Score",
            "Cost Score",
            "Capability Score",
            "Risk Score",
            "Strategic Fit",
            "Implementation",
            "TCO",
            "Recommendation",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Data
        vendors = analysis.vendor_options

        for row_idx, vendor in enumerate(vendors, start=2):
            ws.cell(row=row_idx, column=1).value = vendor.ranking or row_idx - 1
            ws.cell(row=row_idx, column=2).value = vendor.vendor_name
            ws.cell(row=row_idx, column=3).value = vendor.total_score
            ws.cell(row=row_idx, column=4).value = vendor.cost_score
            ws.cell(row=row_idx, column=5).value = vendor.capability_coverage_score
            ws.cell(row=row_idx, column=6).value = vendor.risk_score
            ws.cell(row=row_idx, column=7).value = vendor.strategic_fit_score
            ws.cell(row=row_idx, column=8).value = vendor.implementation_score
            ws.cell(row=row_idx, column=9).value = (
                float(vendor.tco_total) if vendor.tco_total else 0
            )
            ws.cell(row=row_idx, column=10).value = "✓ Recommended" if vendor.ranking == 1 else ""

            # Apply borders
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border

            # Highlight winner
            if vendor.ranking == 1:
                winner_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = winner_fill

        # Format numbers
        for row in range(2, len(vendors) + 2):
            for col in [3, 4, 5, 6, 7, 8]:  # Score columns
                ws.cell(row=row, column=col).number_format = "0.0"
            ws.cell(row=row, column=9).number_format = "$#,##0"  # TCO column

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_detailed_scoring(self, wb: openpyxl.Workbook, analysis: OptionsAnalysis):
        """Create Detailed Scoring sheet with all scoring breakdowns."""
        ws = wb.create_sheet("Detailed Scoring")

        vendors = analysis.vendor_options

        row = 1
        for vendor in vendors:
            # Vendor header
            ws[f"A{row}"] = vendor.vendor_name
            ws[f"A{row}"].font = Font(bold=True, size=14)
            ws.merge_cells(f"A{row}:D{row}")
            row += 1

            # Cost breakdown
            ws[f"A{row}"] = "Cost Analysis"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            if vendor.tco_breakdown:
                try:
                    tco_data = json.loads(vendor.tco_breakdown)
                    ws[f"A{row}"] = "Year"
                    ws[f"B{row}"] = "License"
                    ws[f"C{row}"] = "Support"
                    ws[f"D{row}"] = "Infrastructure"
                    ws[f"E{row}"] = "Total"
                    row += 1

                    for year_key, year_data in tco_data.items():
                        ws[f"A{row}"] = year_key.replace("year", "Year ")
                        ws[f"B{row}"] = f"${year_data.get('license', 0):,.0f}"
                        ws[f"C{row}"] = f"${year_data.get('support', 0):,.0f}"
                        ws[f"D{row}"] = f"${year_data.get('infrastructure', 0):,.0f}"
                        ws[f"E{row}"] = f"${year_data.get('total', 0):,.0f}"
                        row += 1
                except Exception as e:
                    logger.debug("Failed to parse TCO breakdown JSON: %s", e)

            # Risk breakdown
            row += 1
            ws[f"A{row}"] = "Risk Assessment"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            risk_items = [
                ("Vendor Lock-in Risk:", vendor.vendor_lock_in_risk),
                ("Market Position Risk:", vendor.market_position_risk),
                ("Support Continuity Risk:", vendor.support_continuity_risk),
                ("Technology Maturity Risk:", vendor.technology_maturity_risk),
                ("Compliance Risk:", vendor.compliance_risk),
            ]

            for label, value in risk_items:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = f"{value}/10" if value else "N/A"
                row += 1

            row += 2  # Space before next vendor

    def _create_tco_breakdown(self, wb: openpyxl.Workbook, analysis: OptionsAnalysis):
        """Create TCO Breakdown sheet."""
        ws = wb.create_sheet("TCO Breakdown")

        vendors = analysis.vendor_options

        # Create summary table
        ws["A1"] = "Vendor TCO Summary"
        ws["A1"].font = Font(bold=True, size=14)

        row = 3
        ws[f"A{row}"] = "Vendor"
        ws[f"B{row}"] = "Total TCO (5yr)"
        ws[f"C{row}"] = "Annual Average"
        ws[f"D{row}"] = "Implementation Cost"

        for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"], ws[f"D{row}"]]:
            cell.font = Font(bold=True)

        row += 1
        for vendor in vendors:
            ws[f"A{row}"] = vendor.vendor_name
            ws[f"B{row}"] = float(vendor.tco_total) if vendor.tco_total else 0
            ws[f"C{row}"] = float(vendor.tco_total / 5) if vendor.tco_total else 0

            # Extract implementation cost from TCO breakdown
            impl_cost = 0
            if vendor.tco_breakdown:
                try:
                    tco_data = json.loads(vendor.tco_breakdown)
                    impl_cost = tco_data.get("year1", {}).get("implementation", 0)
                except Exception as e:
                    logger.debug("Failed to parse vendor TCO breakdown: %s", e)
            ws[f"D{row}"] = impl_cost
            row += 1

        # Format as currency
        for r in range(4, row):
            for col in ["B", "C", "D"]:
                ws[f"{col}{r}"].number_format = "$#,##0"

    def _create_charts(self, wb: openpyxl.Workbook, analysis: OptionsAnalysis):
        """Create Charts sheet with visualizations."""
        ws = wb.create_sheet("Charts")

        vendors = analysis.vendor_options

        # Prepare data for chart
        ws["A1"] = "Vendor"
        ws["B1"] = "Total Score"

        for idx, vendor in enumerate(vendors, start=2):
            ws[f"A{idx}"] = vendor.vendor_name
            ws[f"B{idx}"] = vendor.total_score or 0

        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Vendor Comparison - Total Scores"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Vendor"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(vendors) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(vendors) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        ws.add_chart(chart, "D2")

        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
