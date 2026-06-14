"""
Advanced TCO Calculation Engine - LLM-PRD - 03 Implementation

Comprehensive Total Cost of Ownership calculation engine with 12 cost categories,
industry benchmarks, sensitivity analysis, and advanced reporting capabilities.

Key Features:
- 12 - category comprehensive cost model
- Industry-specific benchmarks for 5 industries × 3 org sizes
- ±20% sensitivity analysis with tornado charts
- Year-by-year cost breakdown with visualization
- Excel export with formulas and pivot tables
- Confidence level assessment
"""

import base64
import io
import json
import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

import numpy as np

# Excel export dependencies
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from sqlalchemy import func

from app import db
from app.models.business_capabilities import BusinessCapability
from app.models.vendor.vendor_organization import (
    TCOCalculation,
    VendorOrganization,
    VendorProduct,
    VendorProductPricing,
)

logger = logging.getLogger(__name__)


@dataclass
class TCOCostCategory:
    """Represents a TCO cost category with its properties."""

    name: str
    cost_type: str  # "one_time", "recurring", "end_of_period"
    weight: float
    description: str
    typical_percentage: float  # Typical percentage of total TCO


@dataclass
class SensitivityAnalysis:
    """Results of sensitivity analysis."""

    base_tco: Decimal
    best_case_tco: Decimal
    worst_case_tco: Decimal
    sensitivity_factors: Dict[str, float]
    tornado_data: List[Dict[str, Any]]
    confidence_level: str


@dataclass
class IndustryBenchmark:
    """Industry-specific TCO benchmark."""

    industry: str
    organization_size: str
    median_tco_per_user: Decimal
    implementation_months: int
    cost_distribution: Dict[str, float]  # Percentage by category


class AdvancedTCOEngine:
    """
    Advanced TCO calculation engine with comprehensive cost modeling and analysis.
    """

    # Configurable default multipliers — override via constructor or config.
    # These are industry-average estimates used ONLY when real data is unavailable.
    DEFAULT_COST_MULTIPLIERS = {
        "support_maintenance_pct": 0.20,       # % of annual license
        "cloud_infrastructure_pct": {"cloud": 0.15, "hybrid": 0.08, "on-premise": 0.03},
        "internal_labor_pct": 0.30,            # % of annual license
        "ongoing_enhancements_pct": 0.10,      # % of annual license per year
        "implementation_services_pct": 2.0,    # % of annual license (one-time)
        "data_migration_per_user": 500,        # $ per user (one-time)
        "integration_development_pct": 1.5,    # % of annual license (one-time)
        "customization_pct": 0.8,              # % of annual license (one-time)
        "training_per_user": 1000,             # $ per user (one-time)
        "change_management_pct": 0.3,          # % of annual license (one-time)
        "exit_costs_pct": 0.5,                 # % of annual license (end of period)
        "default_license_per_user": 1000,      # $ per user when no pricing data
    }

    # 12 TCO cost categories as specified in PRD-V03
    TCO_CATEGORIES = {
        "software_licensing": TCOCostCategory(
            "Software Licensing",
            "recurring",
            0.25,
            "Annual software license fees and subscriptions",
            30.0,
        ),
        "support_maintenance": TCOCostCategory(
            "Support & Maintenance",
            "recurring",
            0.15,
            "Annual support, maintenance, and update fees",
            15.0,
        ),
        "cloud_infrastructure": TCOCostCategory(
            "Cloud Infrastructure",
            "recurring",
            0.10,
            "Cloud hosting, storage, and infrastructure costs",
            10.0,
        ),
        "implementation_services": TCOCostCategory(
            "Implementation Services",
            "one_time",
            0.20,
            "Professional services for system implementation",
            20.0,
        ),
        "data_migration": TCOCostCategory(
            "Data Migration",
            "one_time",
            0.08,
            "Data extraction, transformation, and loading costs",
            8.0,
        ),
        "integration_development": TCOCostCategory(
            "Integration Development",
            "one_time",
            0.12,
            "Custom integration development and configuration",
            12.0,
        ),
        "customization": TCOCostCategory(
            "Customization", "one_time", 0.10, "System customization and configuration costs", 10.0
        ),
        "training": TCOCostCategory(
            "Training", "one_time", 0.05, "User training and change management programs", 5.0
        ),
        "change_management": TCOCostCategory(
            "Change Management",
            "one_time",
            0.03,
            "Organizational change management and communication",
            3.0,
        ),
        "internal_labor": TCOCostCategory(
            "Internal Labor", "recurring", 0.15, "Internal staff costs for system management", 15.0
        ),
        "ongoing_enhancements": TCOCostCategory(
            "Ongoing Enhancements",
            "recurring",
            0.10,
            "Continuous improvements and enhancements",
            10.0,
        ),
        "exit_costs": TCOCostCategory(
            "Exit Costs",
            "end_of_period",
            0.05,
            "Data extraction and migration costs at contract end",
            5.0,
        ),
    }

    # Industry benchmarks (5 industries × 3 org sizes = 15 benchmarks)
    INDUSTRY_BENCHMARKS = {
        "manufacturing": {
            "small": IndustryBenchmark(
                "manufacturing",
                "small",
                Decimal("800"),
                12,
                {"software_licensing": 35, "implementation_services": 18, "internal_labor": 20},
            ),
            "medium": IndustryBenchmark(
                "manufacturing",
                "medium",
                Decimal("1500"),
                18,
                {"software_licensing": 30, "implementation_services": 20, "internal_labor": 18},
            ),
            "large": IndustryBenchmark(
                "manufacturing",
                "large",
                Decimal("3000"),
                24,
                {"software_licensing": 28, "implementation_services": 22, "internal_labor": 17},
            ),
        },
        "financial_services": {
            "small": IndustryBenchmark(
                "financial_services",
                "small",
                Decimal("1200"),
                15,
                {
                    "software_licensing": 32,
                    "implementation_services": 20,
                    "support_maintenance": 18,
                },
            ),
            "medium": IndustryBenchmark(
                "financial_services",
                "medium",
                Decimal("2200"),
                20,
                {
                    "software_licensing": 30,
                    "implementation_services": 18,
                    "support_maintenance": 17,
                },
            ),
            "large": IndustryBenchmark(
                "financial_services",
                "large",
                Decimal("4500"),
                30,
                {
                    "software_licensing": 28,
                    "implementation_services": 17,
                    "support_maintenance": 16,
                },
            ),
        },
        "healthcare": {
            "small": IndustryBenchmark(
                "healthcare",
                "small",
                Decimal("900"),
                14,
                {
                    "software_licensing": 33,
                    "implementation_services": 19,
                    "integration_development": 16,
                },
            ),
            "medium": IndustryBenchmark(
                "healthcare",
                "medium",
                Decimal("1800"),
                19,
                {
                    "software_licensing": 31,
                    "implementation_services": 18,
                    "integration_development": 15,
                },
            ),
            "large": IndustryBenchmark(
                "healthcare",
                "large",
                Decimal("3800"),
                28,
                {
                    "software_licensing": 29,
                    "implementation_services": 17,
                    "integration_development": 14,
                },
            ),
        },
        "retail": {
            "small": IndustryBenchmark(
                "retail",
                "small",
                Decimal("600"),
                10,
                {"software_licensing": 36, "cloud_infrastructure": 15, "support_maintenance": 16},
            ),
            "medium": IndustryBenchmark(
                "retail",
                "medium",
                Decimal("1200"),
                15,
                {"software_licensing": 32, "cloud_infrastructure": 14, "support_maintenance": 15},
            ),
            "large": IndustryBenchmark(
                "retail",
                "large",
                Decimal("2800"),
                22,
                {"software_licensing": 30, "cloud_infrastructure": 13, "support_maintenance": 14},
            ),
        },
        "technology": {
            "small": IndustryBenchmark(
                "technology",
                "small",
                Decimal("700"),
                8,
                {
                    "software_licensing": 34,
                    "integration_development": 18,
                    "ongoing_enhancements": 14,
                },
            ),
            "medium": IndustryBenchmark(
                "technology",
                "medium",
                Decimal("1400"),
                12,
                {
                    "software_licensing": 32,
                    "integration_development": 16,
                    "ongoing_enhancements": 13,
                },
            ),
            "large": IndustryBenchmark(
                "technology",
                "large",
                Decimal("3200"),
                20,
                {
                    "software_licensing": 30,
                    "integration_development": 15,
                    "ongoing_enhancements": 12,
                },
            ),
        },
    }

    def __init__(self):
        """Initialize the advanced TCO engine."""
        self.logger = logging.getLogger(__name__)

    def calculate_comprehensive_tco(
        self,
        vendor_product_id: int,
        user_count: int,
        tco_period_years: int = 5,
        deployment_model: str = "cloud",
        organization_size: str = "medium",
        industry: str = "manufacturing",
        include_sensitivity_analysis: bool = True,
        sensitivity_range: float = 0.20,
    ) -> Dict[str, Any]:
        """
        Calculate comprehensive TCO with all 12 cost categories and analysis.

        Args:
            vendor_product_id: ID of the vendor product
            user_count: Number of users
            tco_period_years: TCO calculation period in years
            deployment_model: "cloud", "on-premise", or "hybrid"
            organization_size: "small", "medium", "large", or "enterprise"
            industry: Industry sector
            include_sensitivity_analysis: Whether to include sensitivity analysis
            sensitivity_range: Range for sensitivity analysis (±20% default)

        Returns:
            Comprehensive TCO calculation results
        """

        # Input validation
        if user_count < 1:
            raise ValueError(f"user_count must be >= 1, got {user_count}")
        if tco_period_years < 2:
            raise ValueError(f"tco_period_years must be >= 2, got {tco_period_years}")
        if not (0.01 <= sensitivity_range <= 1.0):
            raise ValueError(f"sensitivity_range must be 0.01-1.0, got {sensitivity_range}")

        try:
            # Get vendor product information
            vendor_product = self._get_vendor_product(vendor_product_id)
            if not vendor_product:
                raise ValueError(f"Vendor product {vendor_product_id} not found")

            # Get pricing information
            pricing_tiers = self._get_pricing_tiers(vendor_product_id)

            # Calculate base TCO
            base_tco = self._calculate_base_tco(
                vendor_product,
                pricing_tiers,
                user_count,
                tco_period_years,
                deployment_model,
                organization_size,
                industry,
            )

            # Get industry benchmarks
            benchmark = self._get_industry_benchmark(industry, organization_size)

            # Calculate comparative metrics
            comparative_metrics = self._calculate_comparative_metrics(
                base_tco, benchmark, user_count, tco_period_years
            )

            # Generate year-by-year breakdown
            yearly_breakdown = self._generate_yearly_breakdown(base_tco, tco_period_years)

            # Perform sensitivity analysis
            sensitivity_analysis = None
            if include_sensitivity_analysis:
                sensitivity_analysis = self._perform_sensitivity_analysis(
                    base_tco, sensitivity_range
                )

            # Calculate confidence level
            confidence_level = self._calculate_confidence_level(
                vendor_product, pricing_tiers, base_tco
            )

            # Create comprehensive result
            result = {
                "vendor_product": {
                    "id": vendor_product.id,
                    "name": vendor_product.name,
                    "vendor_name": vendor_product.vendor_organization.name,
                    "category": vendor_product.category,
                },
                "calculation_parameters": {
                    "user_count": user_count,
                    "tco_period_years": tco_period_years,
                    "deployment_model": deployment_model,
                    "organization_size": organization_size,
                    "industry": industry,
                },
                "cost_breakdown": base_tco,
                "yearly_breakdown": yearly_breakdown,
                "comparative_metrics": comparative_metrics,
                "sensitivity_analysis": sensitivity_analysis,
                "confidence_level": confidence_level,
                "calculation_timestamp": datetime.utcnow().isoformat(),
            }

            # Save TCO calculation to database
            self._save_tco_calculation(result)

            return result

        except Exception as e:
            self.logger.error(f"TCO calculation failed: {e}")
            raise

    def _calculate_base_tco(
        self,
        vendor_product: VendorProduct,
        pricing_tiers: List[VendorProductPricing],
        user_count: int,
        tco_period_years: int,
        deployment_model: str,
        organization_size: str,
        industry: str,
    ) -> Dict[str, Any]:
        """Calculate base TCO with all 12 cost categories."""

        # Get appropriate pricing tier
        pricing_tier = self._select_pricing_tier(pricing_tiers, user_count)

        # Base annual license cost
        base_annual_cost = self._calculate_annual_license_cost(pricing_tier, user_count)

        # Calculate each cost category using configurable multipliers
        m = self.DEFAULT_COST_MULTIPLIERS
        costs = {}
        is_estimated = not bool(pricing_tier)

        # Recurring costs
        costs["software_licensing"] = base_annual_cost * tco_period_years

        costs["support_maintenance"] = base_annual_cost * m["support_maintenance_pct"] * tco_period_years

        # Cloud infrastructure varies by deployment model
        cloud_multiplier = m["cloud_infrastructure_pct"].get(deployment_model, 0.15)
        costs["cloud_infrastructure"] = base_annual_cost * cloud_multiplier * tco_period_years

        costs["internal_labor"] = base_annual_cost * m["internal_labor_pct"] * tco_period_years

        # Ongoing enhancements (starts year 2)
        enhancement_years = max(tco_period_years - 1, 0)
        costs["ongoing_enhancements"] = base_annual_cost * m["ongoing_enhancements_pct"] * enhancement_years

        # One-time costs (Year 1)
        costs["implementation_services"] = base_annual_cost * m["implementation_services_pct"]
        costs["data_migration"] = user_count * m["data_migration_per_user"]
        costs["integration_development"] = base_annual_cost * m["integration_development_pct"]
        costs["customization"] = base_annual_cost * m["customization_pct"]
        costs["training"] = user_count * m["training_per_user"]
        costs["change_management"] = base_annual_cost * m["change_management_pct"]

        # Exit costs (Final year)
        costs["exit_costs"] = base_annual_cost * m["exit_costs_pct"]

        if is_estimated:
            logger.warning(
                f"TCO calculation uses DEFAULT estimated multipliers (no real pricing data). "
                f"user_count={user_count}, base_annual_cost={base_annual_cost}"
            )

        # Calculate totals
        one_time_total = sum(
            costs[cat]
            for cat, category in self.TCO_CATEGORIES.items()
            if category.cost_type == "one_time"
        )

        recurring_total = sum(
            costs[cat]
            for cat, category in self.TCO_CATEGORIES.items()
            if category.cost_type == "recurring"
        )

        end_of_period_total = costs["exit_costs"]

        total_tco = one_time_total + recurring_total + end_of_period_total

        # Create cost breakdown structure
        cost_breakdown = {
            "costs": costs,
            "summary": {
                "one_time_total": one_time_total,
                "recurring_total": recurring_total,
                "end_of_period_total": end_of_period_total,
                "total_tco": total_tco,
                "annual_average": total_tco / tco_period_years,
                "per_user_annual": total_tco / (user_count * tco_period_years),
                "per_user_total": total_tco / user_count,
            },
            "cost_distribution": {
                cat: {
                    "amount": costs[cat],
                    "percentage": (costs[cat] / total_tco * 100) if total_tco > 0 else 0,
                    "type": category.cost_type,
                    "description": category.description,
                }
                for cat, category in self.TCO_CATEGORIES.items()
            },
        }

        return cost_breakdown

    def _perform_sensitivity_analysis(
        self, base_tco: Dict[str, Any], sensitivity_range: float = 0.20
    ) -> SensitivityAnalysis:
        """Perform sensitivity analysis on TCO calculation."""

        base_total = base_tco["summary"]["total_tco"]

        # Calculate sensitivity for each cost category
        sensitivity_factors = {}
        tornado_data = []

        for category_name, category in self.TCO_CATEGORIES.items():
            category_cost = base_tco["costs"][category_name]

            # Calculate impact of ±20% change
            positive_impact = category_cost * (1 + sensitivity_range)
            negative_impact = category_cost * (1 - sensitivity_range)

            # Net impact on total TCO
            positive_change = positive_impact - category_cost
            negative_change = negative_impact - category_cost

            sensitivity_factors[category_name] = {
                "positive_impact": positive_change,
                "negative_impact": negative_change,
                "percentage_impact": abs(positive_change / base_total) * 100
                if base_total > 0
                else 0,
            }

            # Add to tornado chart data
            tornado_data.append(
                {
                    "category": category_name,
                    "negative_change": float(negative_change),
                    "positive_change": float(positive_change),
                    "base_cost": float(category_cost),
                    "impact_percentage": sensitivity_factors[category_name]["percentage_impact"],
                }
            )

        # Sort by impact for tornado chart
        tornado_data.sort(key=lambda x: x["impact_percentage"], reverse=True)

        # Calculate best and worst case scenarios
        best_case_tco = base_total + sum(
            factor["negative_impact"] for factor in sensitivity_factors.values()
        )
        worst_case_tco = base_total + sum(
            factor["positive_impact"] for factor in sensitivity_factors.values()
        )

        # Determine confidence level based on sensitivity
        total_variation = (worst_case_tco - best_case_tco) / base_total
        if total_variation < 0.15:
            confidence_level = "high"
        elif total_variation < 0.30:
            confidence_level = "medium"
        else:
            confidence_level = "low"

        return SensitivityAnalysis(
            base_tco=base_total,
            best_case_tco=best_case_tco,
            worst_case_tco=worst_case_tco,
            sensitivity_factors=sensitivity_factors,
            tornado_data=tornado_data,
            confidence_level=confidence_level,
        )

    def _generate_yearly_breakdown(
        self, base_tco: Dict[str, Any], tco_period_years: int
    ) -> List[Dict[str, Any]]:
        """Generate year-by-year cost breakdown."""

        yearly_data = []
        costs = base_tco["costs"]

        for year in range(1, tco_period_years + 1):
            year_cost = {"year": year, "costs": {}}

            # One-time costs (Year 1 only)
            if year == 1:
                for category_name, category in self.TCO_CATEGORIES.items():
                    if category.cost_type == "one_time":
                        year_cost["costs"][category_name] = costs[category_name]
            else:
                for category_name, category in self.TCO_CATEGORIES.items():
                    if category.cost_type == "one_time":
                        year_cost["costs"][category_name] = 0

            # Recurring costs (every year)
            for category_name, category in self.TCO_CATEGORIES.items():
                if category.cost_type == "recurring":
                    if category_name == "ongoing_enhancements":
                        # Ongoing enhancements start year 2
                        year_cost["costs"][category_name] = (
                            costs[category_name] / max(tco_period_years - 1, 1) if year > 1 else 0
                        )
                    else:
                        year_cost["costs"][category_name] = costs[category_name] / tco_period_years

            # Exit costs (final year only)
            if year == tco_period_years:
                for category_name, category in self.TCO_CATEGORIES.items():
                    if category.cost_type == "end_of_period":
                        year_cost["costs"][category_name] = costs[category_name]
            else:
                for category_name, category in self.TCO_CATEGORIES.items():
                    if category.cost_type == "end_of_period":
                        year_cost["costs"][category_name] = 0

            # Calculate year total
            year_cost["total"] = sum(year_cost["costs"].values())

            # Calculate cumulative total
            if year == 1:
                year_cost["cumulative_total"] = year_cost["total"]
            else:
                year_cost["cumulative_total"] = (
                    yearly_data[-1]["cumulative_total"] + year_cost["total"]
                )

            yearly_data.append(year_cost)

        return yearly_data

    def _calculate_comparative_metrics(
        self,
        base_tco: Dict[str, Any],
        benchmark: IndustryBenchmark,
        user_count: int,
        tco_period_years: int,
    ) -> Dict[str, Any]:
        """Calculate comparative metrics against industry benchmarks."""

        total_tco = base_tco["summary"]["total_tco"]
        per_user_annual = base_tco["summary"]["per_user_annual"]

        # Benchmark comparisons
        benchmark_total_tco = benchmark.median_tco_per_user * user_count * tco_period_years
        benchmark_per_user_annual = benchmark.median_tco_per_user

        # Percentage differences
        vs_industry_percentage = (
            ((total_tco - benchmark_total_tco) / benchmark_total_tco * 100)
            if benchmark_total_tco > 0
            else 0
        )
        vs_per_user_percentage = (
            ((per_user_annual - benchmark_per_user_annual) / benchmark_per_user_annual * 100)
            if benchmark_per_user_annual > 0
            else 0
        )

        # Cost category comparison
        category_comparison = {}
        for category, benchmark_percentage in benchmark.cost_distribution.items():
            actual_cost = base_tco["costs"].get(category, 0)
            actual_percentage = (actual_cost / total_tco * 100) if total_tco > 0 else 0

            category_comparison[category] = {
                "actual_percentage": actual_percentage,
                "benchmark_percentage": benchmark_percentage,
                "difference": actual_percentage - benchmark_percentage,
                "assessment": self._assess_category_difference(
                    actual_percentage - benchmark_percentage
                ),
            }

        return {
            "industry_benchmark": {
                "industry": benchmark.industry,
                "organization_size": benchmark.organization_size,
                "median_tco_per_user": benchmark.median_tco_per_user,
                "implementation_months": benchmark.implementation_months,
            },
            "cost_comparison": {
                "total_tco_vs_benchmark": vs_industry_percentage,
                "per_user_annual_vs_benchmark": vs_per_user_annual,
                "assessment": self._assess_cost_performance(vs_industry_percentage),
            },
            "category_comparison": category_comparison,
        }

    def _assess_cost_performance(self, vs_benchmark_percentage: float) -> str:
        """Assess cost performance against benchmark."""
        if vs_benchmark_percentage <= -10:
            return "excellent_cost_advantage"
        elif vs_benchmark_percentage <= 0:
            return "good_cost_position"
        elif vs_benchmark_percentage <= 15:
            return "slightly_premium"
        elif vs_benchmark_percentage <= 30:
            return "premium_pricing"
        else:
            return "significant_premium"

    def _assess_category_difference(self, difference: float) -> str:
        """Assess cost category difference from benchmark."""
        if abs(difference) <= 5:
            return "aligned"
        elif difference > 5:
            return "higher_than_benchmark"
        else:
            return "lower_than_benchmark"

    def _calculate_confidence_level(
        self,
        vendor_product: VendorProduct,
        pricing_tiers: List[VendorProductPricing],
        base_tco: Dict[str, Any],
    ) -> str:
        """Calculate confidence level in TCO calculation."""

        confidence_factors = []

        # Data availability
        if pricing_tiers:
            confidence_factors.append(0.3)  # Pricing data available
        else:
            confidence_factors.append(0.1)  # No pricing data

        # Vendor maturity
        if vendor_product.product_lifecycle == "leader":
            confidence_factors.append(0.3)
        elif vendor_product.product_lifecycle == "mature":
            confidence_factors.append(0.2)
        else:
            confidence_factors.append(0.1)

        # Market intelligence
        if vendor_product.market_intelligence:
            confidence_factors.append(0.2)
        else:
            confidence_factors.append(0.1)

        # Cost complexity
        total_tco = base_tco["summary"]["total_tco"]
        if total_tco > 1000000:  # Large, complex TCO
            confidence_factors.append(0.1)
        else:
            confidence_factors.append(0.2)

        # Calculate overall confidence
        overall_confidence = sum(confidence_factors)

        if overall_confidence >= 0.8:
            return "high"
        elif overall_confidence >= 0.6:
            return "medium"
        else:
            return "low"

    def _get_vendor_product(self, vendor_product_id: int) -> Optional[VendorProduct]:
        """Get vendor product by ID."""
        return db.session.query(VendorProduct).filter_by(id=vendor_product_id).first()

    def _get_pricing_tiers(self, vendor_product_id: int) -> List[VendorProductPricing]:
        """Get pricing tiers for vendor product."""
        return db.session.query(VendorProductPricing).filter_by(product_id=vendor_product_id).all()

    def _select_pricing_tier(
        self, pricing_tiers: List[VendorProductPricing], user_count: int
    ) -> Optional[VendorProductPricing]:
        """Select appropriate pricing tier based on user count."""

        if not pricing_tiers:
            return None

        # Find tier that matches user count
        suitable_tiers = []
        for tier in pricing_tiers:
            if tier.minimum_users and user_count >= tier.minimum_users:
                if not tier.maximum_users or user_count <= tier.maximum_users:
                    suitable_tiers.append(tier)

        # Return the tier with the minimum users that fits
        if suitable_tiers:
            return min(suitable_tiers, key=lambda x: x.minimum_users)

        # If no suitable tier, return the one with lowest minimum users
        return min(pricing_tiers, key=lambda x: x.minimum_users or 0)

    def _calculate_annual_license_cost(
        self, pricing_tier: Optional[VendorProductPricing], user_count: int
    ) -> Decimal:
        """Calculate annual license cost."""

        if not pricing_tier:
            # Default estimation if no pricing data — flagged as estimate
            default_per_user = self.DEFAULT_COST_MULTIPLIERS["default_license_per_user"]
            logger.warning(
                f"No pricing tier found — using default ${default_per_user}/user estimate "
                f"for {user_count} users"
            )
            return Decimal(str(user_count * default_per_user))

        base_price = pricing_tier.base_price or Decimal(str(self.DEFAULT_COST_MULTIPLIERS["default_license_per_user"]))

        # Apply volume discounts if available
        volume_discounts = pricing_tier.volume_discounts or {}
        discount_percentage = 0

        for tier_name, discount_info in volume_discounts.items():
            if user_count >= discount_info.get("min_users", 0):
                discount_percentage = max(discount_percentage, discount_info.get("discount", 0))

        # Apply discount
        discounted_price = base_price * (1 - Decimal(str(discount_percentage)) / 100)

        return discounted_price * user_count

    def _get_industry_benchmark(self, industry: str, organization_size: str) -> IndustryBenchmark:
        """Get industry benchmark for given industry and organization size."""

        industry_benchmarks = self.INDUSTRY_BENCHMARKS.get(industry, {})
        benchmark = industry_benchmarks.get(organization_size)

        if not benchmark:
            # Return default benchmark if specific one not found
            return self.INDUSTRY_BENCHMARKS["manufacturing"]["medium"]

        return benchmark

    def _save_tco_calculation(self, tco_result: Dict[str, Any]) -> None:
        """Save TCO calculation to database."""

        try:
            vendor_product_id = tco_result["vendor_product"]["id"]
            params = tco_result["calculation_parameters"]
            costs = tco_result["cost_breakdown"]["costs"]
            summary = tco_result["cost_breakdown"]["summary"]
            comparative = tco_result["comparative_metrics"]
            sensitivity = tco_result.get("sensitivity_analysis")

            # Check if calculation already exists
            existing = TCOCalculation.query.filter_by(
                vendor_product_id=vendor_product_id,
                user_count=params["user_count"],
                tco_period_years=params["tco_period_years"],
            ).first()

            if existing:
                # Update existing calculation
                existing.total_tco = summary["total_tco"]
                existing.annual_average = summary["annual_average"]
                existing.per_user_annual = summary["per_user_annual"]
                existing.industry_median_tco = (
                    comparative["industry_benchmark"]["median_tco_per_user"]
                    * params["user_count"]
                    * params["tco_period_years"]
                )
                existing.vs_industry_percentage = comparative["cost_comparison"][
                    "total_tco_vs_benchmark"
                ]
                existing.confidence_level = tco_result["confidence_level"]

                if sensitivity:
                    existing.min_tco = sensitivity.best_case_tco
                    existing.max_tco = sensitivity.worst_case_tco
                    existing.sensitivity_factors = json.dumps(sensitivity.sensitivity_factors)

                existing.updated_at = datetime.utcnow()
            else:
                # Create new calculation
                tco_calc = TCOCalculation(
                    vendor_product_id=vendor_product_id,
                    user_count=params["user_count"],
                    tco_period_years=params["tco_period_years"],
                    deployment_model=params["deployment_model"],
                    # One-time costs
                    implementation_costs=costs["implementation_services"],
                    data_migration_costs=costs["data_migration"],
                    integration_development_costs=costs["integration_development"],
                    customization_costs=costs["customization"],
                    training_costs=costs["training"],
                    change_management_costs=costs["change_management"],
                    one_time_total=summary["one_time_total"],
                    # Recurring costs
                    software_licensing_total=costs["software_licensing"],
                    support_maintenance_total=costs["support_maintenance"],
                    cloud_infrastructure_total=costs["cloud_infrastructure"],
                    internal_labor_total=costs["internal_labor"],
                    ongoing_enhancements_total=costs["ongoing_enhancements"],
                    recurring_total=summary["recurring_total"],
                    # Exit costs
                    exit_costs_total=costs["exit_costs"],
                    # Summary
                    total_tco=summary["total_tco"],
                    annual_average=summary["annual_average"],
                    per_user_annual=summary["per_user_annual"],
                    # Benchmarks
                    industry_median_tco=comparative["industry_benchmark"]["median_tco_per_user"]
                    * params["user_count"]
                    * params["tco_period_years"],
                    vs_industry_percentage=comparative["cost_comparison"]["total_tco_vs_benchmark"],
                    # Sensitivity
                    min_tco=sensitivity.best_case_tco if sensitivity else None,
                    max_tco=sensitivity.worst_case_tco if sensitivity else None,
                    sensitivity_factors=json.dumps(sensitivity.sensitivity_factors)
                    if sensitivity
                    else None,
                    # Confidence
                    confidence_level=tco_result["confidence_level"],
                )

                db.session.add(tco_calc)

            db.session.commit()

        except Exception as e:
            self.logger.error(f"Failed to save TCO calculation: {e}")
            db.session.rollback()
            raise

    def export_tco_to_excel(
        self,
        tco_calculation_id: int,
        include_charts: bool = True,
        include_sensitivity: bool = True,
        include_pivot_tables: bool = True,
    ) -> Dict[str, Any]:
        """
        Export TCO calculation to Excel with formulas, charts, and pivot tables.

        Args:
            tco_calculation_id: ID of the TCO calculation
            include_charts: Whether to include charts
            include_sensitivity: Whether to include sensitivity analysis
            include_pivot_tables: Whether to include pivot tables

        Returns:
            Dictionary with Excel file data and metadata
        """

        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export functionality")

        try:
            # Get TCO calculation
            tco_calc = db.session.query(TCOCalculation).filter_by(id=tco_calculation_id).first()
            if not tco_calc:
                raise ValueError(f"TCO calculation {tco_calculation_id} not found")

            # Get vendor product
            vendor_product = (
                db.session.query(VendorProduct).filter_by(id=tco_calc.vendor_product_id).first()
            )
            if not vendor_product:
                raise ValueError(f"Vendor product {tco_calc.vendor_product_id} not found")

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets
            summary_sheet = wb.create_sheet("Executive Summary")
            cost_breakdown_sheet = wb.create_sheet("Cost Breakdown")
            yearly_sheet = wb.create_sheet("Yearly Breakdown")
            benchmark_sheet = wb.create_sheet("Benchmark Analysis")

            if include_sensitivity:
                sensitivity_sheet = wb.create_sheet("Sensitivity Analysis")

            if include_pivot_tables:
                pivot_sheet = wb.create_sheet("Pivot Analysis")

            # Populate Executive Summary
            self._create_executive_summary(summary_sheet, tco_calc, vendor_product)

            # Populate Cost Breakdown
            self._create_cost_breakdown(cost_breakdown_sheet, tco_calc)

            # Populate Yearly Breakdown
            self._create_yearly_breakdown(yearly_sheet, tco_calc)

            # Populate Benchmark Analysis
            self._create_benchmark_analysis(benchmark_sheet, tco_calc)

            # Populate Sensitivity Analysis
            if include_sensitivity:
                self._create_sensitivity_analysis(sensitivity_sheet, tco_calc)

            # Create Pivot Tables
            if include_pivot_tables:
                self._create_pivot_tables(pivot_sheet, tco_calc)

            # Add Charts
            if include_charts:
                self._add_charts_to_workbook(wb, tco_calc)

            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Encode for JSON response
            excel_data = base64.b64encode(excel_buffer.read()).decode("utf-8")

            return {
                "success": True,
                "excel_data": excel_data,
                "filename": f"tco_analysis_{vendor_product.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "file_size": len(excel_data),
                "sheets_created": len(wb.worksheets),
                "includes_charts": include_charts,
                "includes_sensitivity": include_sensitivity,
                "includes_pivot_tables": include_pivot_tables,
            }

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise

    def _create_executive_summary(self, sheet, tco_calc, vendor_product):
        """Create executive summary sheet."""

        # Title
        sheet["A1"] = "TCO Analysis Executive Summary"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("A1:D1")

        # Product Information
        row = 3
        sheet[f"A{row}"] = "Product Information"
        sheet[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        sheet[f"A{row}"] = "Product Name:"
        sheet[f"B{row}"] = vendor_product.name
        sheet[f"C{row}"] = "Vendor:"
        sheet[f"D{row}"] = (
            vendor_product.vendor_organization.name
            if vendor_product.vendor_organization
            else "Unknown"
        )
        row += 1

        sheet[f"A{row}"] = "User Count:"
        sheet[f"B{row}"] = tco_calc.user_count
        sheet[f"C{row}"] = "TCO Period:"
        sheet[f"D{row}"] = f"{tco_calc.tco_period_years} years"
        row += 1

        sheet[f"A{row}"] = "Deployment Model:"
        sheet[f"B{row}"] = tco_calc.deployment_model
        sheet[f"C{row}"] = "Confidence Level:"
        sheet[f"D{row}"] = tco_calc.confidence_level.title()
        row += 2

        # Financial Summary
        sheet[f"A{row}"] = "Financial Summary"
        sheet[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        # Key metrics with formulas
        sheet[f"A{row}"] = "Total TCO:"
        sheet[f"B{row}"] = tco_calc.total_tco
        sheet[f"B{row}"].number_format = "$#,##0"
        sheet[f"B{row}"].font = Font(bold=True)
        row += 1

        sheet[f"A{row}"] = "Annual Average:"
        sheet[f"B{row}"] = f"=B{row - 1}/{tco_calc.tco_period_years}"
        sheet[f"B{row}"].number_format = "$#,##0"
        row += 1

        sheet[f"A{row}"] = "Per User Annual:"
        sheet[f"B{row}"] = tco_calc.per_user_annual
        sheet[f"B{row}"].number_format = "$#,##0"
        row += 1

        sheet[f"A{row}"] = "Per User Total:"
        sheet[f"B{row}"] = f"=B{row - 3}/{tco_calc.user_count}"
        sheet[f"B{row}"].number_format = "$#,##0"
        row += 2

        # Cost Structure
        sheet[f"A{row}"] = "Cost Structure"
        sheet[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        sheet[f"A{row}"] = "One-time Costs:"
        sheet[f"B{row}"] = tco_calc.one_time_total
        sheet[f"B{row}"].number_format = "$#,##0"
        sheet[f"C{row}"] = f"=B{row}/B{row - 6}"
        sheet[f"C{row}"].number_format = "0.0%"
        row += 1

        sheet[f"A{row}"] = "Recurring Costs:"
        sheet[f"B{row}"] = tco_calc.recurring_total
        sheet[f"B{row}"].number_format = "$#,##0"
        sheet[f"C{row}"] = f"=B{row}/B{row - 7}"
        sheet[f"C{row}"].number_format = "0.0%"
        row += 1

        sheet[f"A{row}"] = "Exit Costs:"
        sheet[f"B{row}"] = tco_calc.exit_costs_total
        sheet[f"B{row}"].number_format = "$#,##0"
        sheet[f"C{row}"] = f"=B{row}/B{row - 8}"
        sheet[f"C{row}"].number_format = "0.0%"

        # Column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 20

    def _create_cost_breakdown(self, sheet, tco_calc):
        """Create cost breakdown sheet."""

        # Title
        sheet["A1"] = "Cost Breakdown Analysis"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("A1:E1")

        # Headers
        headers = ["Cost Category", "Amount", "Percentage", "Type", "Description"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Cost categories data
        cost_data = [
            (
                "Software Licensing",
                tco_calc.software_licensing_total,
                "recurring",
                "Annual software license fees",
            ),
            (
                "Support & Maintenance",
                tco_calc.support_maintenance_total,
                "recurring",
                "Annual support and maintenance",
            ),
            (
                "Cloud Infrastructure",
                tco_calc.cloud_infrastructure_total,
                "recurring",
                "Cloud hosting and infrastructure",
            ),
            ("Internal Labor", tco_calc.internal_labor_total, "recurring", "Internal staff costs"),
            (
                "Ongoing Enhancements",
                tco_calc.ongoing_enhancements_total,
                "recurring",
                "Continuous improvements",
            ),
            (
                "Implementation Services",
                tco_calc.implementation_costs,
                "one_time",
                "System implementation",
            ),
            ("Data Migration", tco_calc.data_migration_costs, "one_time", "Data migration costs"),
            (
                "Integration Development",
                tco_calc.integration_development_costs,
                "one_time",
                "Custom integrations",
            ),
            ("Customization", tco_calc.customization_costs, "one_time", "System customization"),
            ("Training", tco_calc.training_costs, "one_time", "User training programs"),
            (
                "Change Management",
                tco_calc.change_management_costs,
                "one_time",
                "Change management",
            ),
            ("Exit Costs", tco_calc.exit_costs_total, "end_of_period", "Data extraction at exit"),
        ]

        row = 4
        for category, amount, cost_type, description in cost_data:
            sheet.cell(row=row, column=1, value=category)
            sheet.cell(row=row, column=2, value=amount)
            sheet.cell(row=row, column=3, value=f"=B{row}/$B${len(cost_data) + 4}")
            sheet.cell(row=row, column=4, value=cost_type)
            sheet.cell(row=row, column=5, value=description)
            row += 1

        # Total row
        sheet.cell(row=row, column=1, value="TOTAL")
        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=f"=SUM(B4:B{row - 1})")
        sheet.cell(row=row, column=2).font = Font(bold=True)
        sheet.cell(row=row, column=2).number_format = "$#,##0"
        sheet.cell(row=row, column=3, value="100%")
        sheet.cell(row=row, column=3).font = Font(bold=True)

        # Format columns
        for r in range(4, row + 1):
            sheet.cell(row=r, column=2).number_format = "$#,##0"
            sheet.cell(row=r, column=3).number_format = "0.0%"

        # Column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 35

    def _create_yearly_breakdown(self, sheet, tco_calc):
        """Create yearly breakdown sheet."""

        # Title
        sheet["A1"] = "Year-by-Year Cost Breakdown"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("A1:M1")

        # Headers
        headers = (
            ["Year"] + [f"Year {i}" for i in range(1, tco_calc.tco_period_years + 1)] + ["Total"]
        )
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Cost categories
        cost_categories = [
            ("One-time Costs", "one_time"),
            ("Software Licensing", "recurring"),
            ("Support & Maintenance", "recurring"),
            ("Cloud Infrastructure", "recurring"),
            ("Internal Labor", "recurring"),
            ("Ongoing Enhancements", "recurring"),
            ("Exit Costs", "end_of_period"),
        ]

        row = 4
        for category_name, cost_type in cost_categories:
            sheet.cell(row=row, column=1, value=category_name)
            sheet.cell(row=row, column=1).font = Font(bold=True)

            # Calculate yearly amounts (simplified allocation)
            for year in range(1, tco_calc.tco_period_years + 1):
                if cost_type == "one_time" and year == 1:
                    amount = tco_calc.one_time_total
                elif cost_type == "recurring":
                    if category_name == "Ongoing Enhancements" and year == 1:
                        amount = 0
                    else:
                        amount = (
                            getattr(
                                tco_calc,
                                f"{category_name.lower().replace(' & ', '_').replace(' ', '_')}_total",
                                0,
                            )
                            / tco_calc.tco_period_years
                        )
                elif cost_type == "end_of_period" and year == tco_calc.tco_period_years:
                    amount = tco_calc.exit_costs_total
                else:
                    amount = 0

                sheet.cell(row=row, column=year + 1, value=amount)
                sheet.cell(row=row, column=year + 1).number_format = "$#,##0"

            # Total for category
            sheet.cell(
                row=row,
                column=tco_calc.tco_period_years + 2,
                value=f"=SUM(B{row}:{chr(64 + tco_calc.tco_period_years + 1)}{row})",
            )
            sheet.cell(row=row, column=tco_calc.tco_period_years + 2).number_format = "$#,##0"
            row += 1

        # Year totals
        sheet.cell(row=row, column=1, value="Year Total")
        sheet.cell(row=row, column=1).font = Font(bold=True)
        for year in range(1, tco_calc.tco_period_years + 1):
            sheet.cell(row=row, column=year + 1, value=f"=SUM(B{4}:{chr(64 + year)}{row - 1})")
            sheet.cell(row=row, column=year + 1).number_format = "$#,##0"

        sheet.cell(
            row=row,
            column=tco_calc.tco_period_years + 2,
            value=f"=SUM(B{row}:{chr(64 + tco_calc.tco_period_years + 1)}{row})",
        )
        sheet.cell(row=row, column=tco_calc.tco_period_years + 2).number_format = "$#,##0"
        sheet.cell(row=row, column=tco_calc.tco_period_years + 2).font = Font(bold=True)

        # Column widths
        sheet.column_dimensions["A"].width = 20
        for col in range(2, tco_calc.tco_period_years + 3):
            sheet.column_dimensions[chr(64 + col)].width = 12

    def _create_benchmark_analysis(self, sheet, tco_calc):
        """Create benchmark analysis sheet."""

        # Title
        sheet["A1"] = "Industry Benchmark Analysis"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("A1:E1")

        # Benchmark comparison
        row = 3
        sheet[f"A{row}"] = "Benchmark Comparison"
        sheet[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        sheet[f"A{row}"] = "Industry Median TCO:"
        sheet[f"B{row}"] = tco_calc.industry_median_tco
        sheet[f"B{row}"].number_format = "$#,##0"
        row += 1

        sheet[f"A{row}"] = "Calculated TCO:"
        sheet[f"B{row}"] = tco_calc.total_tco
        sheet[f"B{row}"].number_format = "$#,##0"
        row += 1

        sheet[f"A{row}"] = "Difference:"
        sheet[f"B{row}"] = f"=B{row}-B{row - 2}"
        sheet[f"B{row}"].number_format = "$#,##0"
        row += 1

        sheet[f"A{row}"] = "Percentage vs Industry:"
        sheet[f"B{row}"] = tco_calc.vs_industry_percentage
        sheet[f"B{row}"].number_format = "0.0%"
        row += 2

        # Performance assessment
        sheet[f"A{row}"] = "Performance Assessment"
        sheet[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        if tco_calc.vs_industry_percentage <= -10:
            assessment = "Excellent Cost Advantage"
        elif tco_calc.vs_industry_percentage <= 0:
            assessment = "Good Cost Position"
        elif tco_calc.vs_industry_percentage <= 15:
            assessment = "Slightly Premium"
        else:
            assessment = "Premium Pricing"

        sheet[f"A{row}"] = "Assessment:"
        sheet[f"B{row}"] = assessment
        sheet[f"B{row}"].font = Font(bold=True)

        # Column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 20

    def _create_sensitivity_analysis(self, sheet, tco_calc):
        """Create sensitivity analysis sheet."""

        # Title
        sheet["A1"] = "Sensitivity Analysis"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("A1:F1")

        # Sensitivity factors
        if tco_calc.sensitivity_factors:
            sensitivity_data = json.loads(tco_calc.sensitivity_factors)

            # Headers
            headers = [
                "Cost Category",
                "Base Cost",
                "Best Case",
                "Worst Case",
                "Impact %",
                "Assessment",
            ]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row = 4
            for category, factors in sensitivity_data.items():
                sheet.cell(row=row, column=1, value=category.replace("_", " ").title())
                sheet.cell(row=row, column=2, value=tco_calc.total_tco)
                sheet.cell(row=row, column=3, value=tco_calc.min_tco)
                sheet.cell(row=row, column=4, value=tco_calc.max_tco)
                sheet.cell(row=row, column=5, value=factors.get("percentage_impact", 0))
                sheet.cell(
                    row=row,
                    column=6,
                    value="High Impact"
                    if factors.get("percentage_impact", 0) > 5
                    else "Low Impact",
                )
                row += 1

            # Summary
            row += 2
            sheet[f"A{row}"] = "Sensitivity Summary"
            sheet[f"A{row}"].font = Font(size=14, bold=True)
            row += 1

            sheet[f"A{row}"] = "Base TCO:"
            sheet[f"B{row}"] = tco_calc.total_tco
            sheet[f"B{row}"].number_format = "$#,##0"
            row += 1

            sheet[f"A{row}"] = "Best Case TCO:"
            sheet[f"B{row}"] = tco_calc.min_tco
            sheet[f"B{row}"].number_format = "$#,##0"
            row += 1

            sheet[f"A{row}"] = "Worst Case TCO:"
            sheet[f"B{row}"] = tco_calc.max_tco
            sheet[f"B{row}"].number_format = "$#,##0"
            row += 1

            variation = (
                ((tco_calc.max_tco - tco_calc.min_tco) / tco_calc.total_tco) * 100
                if tco_calc.total_tco > 0
                else 0
            )
            sheet[f"A{row}"] = "Variation Range:"
            sheet[f"B{row}"] = f"{variation:.1f}%"

        # Column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 12
        sheet.column_dimensions["F"].width = 15

    def _create_pivot_tables(self, sheet, tco_calc):
        """Create pivot table analysis sheet."""

        # Title
        sheet["A1"] = "Pivot Table Analysis"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("A1:D1")

        # Cost Type Summary
        row = 3
        sheet[f"A{row}"] = "Cost Type Summary"
        sheet[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        # Headers
        headers = ["Cost Type", "Amount", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        row += 1

        # Data
        cost_types = [
            ("One-time Costs", tco_calc.one_time_total),
            ("Recurring Costs", tco_calc.recurring_total),
            ("Exit Costs", tco_calc.exit_costs_total),
        ]

        for cost_type, amount in cost_types:
            sheet.cell(row=row, column=1, value=cost_type)
            sheet.cell(row=row, column=2, value=amount)
            sheet.cell(row=row, column=2).number_format = "$#,##0"
            sheet.cell(row=row, column=3, value=f"=B{row}/B{row + 3}")
            sheet.cell(row=row, column=3).number_format = "0.0%"
            row += 1

        # Total
        sheet.cell(row=row, column=1, value="Total")
        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=f"=SUM(B{row - 3}:B{row - 1})")
        sheet.cell(row=row, column=2).number_format = "$#,##0"
        sheet.cell(row=row, column=2).font = Font(bold=True)
        sheet.cell(row=row, column=3, value="100%")
        sheet.cell(row=row, column=3).font = Font(bold=True)

        # Column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15

    def _add_charts_to_workbook(self, wb, tco_calc):
        """Add charts to the workbook."""

        try:
            # Cost breakdown chart
            cost_breakdown_sheet = wb["Cost Breakdown"]

            # Create bar chart for cost breakdown
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "TCO Cost Breakdown"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Cost Category"

            # Data range
            data = Reference(cost_breakdown_sheet, min_col=2, max_col=2, min_row=4, max_row=15)
            categories = Reference(cost_breakdown_sheet, min_col=1, min_row=4, max_row=15)

            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)

            # Position chart
            cost_breakdown_sheet.add_chart(chart, "G4")

        except Exception as e:
            logger.warning(f"Failed to create charts: {e}")


# Convenience function for direct usage
def calculate_advanced_tco(
    vendor_product_id: int,
    user_count: int,
    tco_period_years: int = 5,
    deployment_model: str = "cloud",
    organization_size: str = "medium",
    industry: str = "manufacturing",
) -> Dict[str, Any]:
    """
    Convenience function to calculate advanced TCO.

    Args:
        vendor_product_id: ID of the vendor product
        user_count: Number of users
        tco_period_years: TCO calculation period in years
        deployment_model: Deployment model
        organization_size: Organization size
        industry: Industry sector

    Returns:
        TCO calculation results
    """
    engine = AdvancedTCOEngine()
    return engine.calculate_comprehensive_tco(
        vendor_product_id=vendor_product_id,
        user_count=user_count,
        tco_period_years=tco_period_years,
        deployment_model=deployment_model,
        organization_size=organization_size,
        industry=industry,
    )


if __name__ == "__main__":
    # Test the advanced TCO engine
    logging.basicConfig(level=logging.INFO)

    print("Testing Advanced TCO Engine...")

    # Test cost categories
    engine = AdvancedTCOEngine()
    print(f"TCO Categories: {len(engine.TCO_CATEGORIES)}")
    print(f"Industry Benchmarks: {len(engine.INDUSTRY_BENCHMARKS)} industries")

    # Test benchmark retrieval
    benchmark = engine._get_industry_benchmark("manufacturing", "medium")
    print(f"Manufacturing medium benchmark: ${benchmark.median_tco_per_user} per user")
