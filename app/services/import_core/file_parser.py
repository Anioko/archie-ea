"""
Shared File Parser for Import Systems

Common file parsing functionality for CSV, Excel, and JSON files.
Used by both unified applications import and batch import systems.
"""

import csv
import io
import json
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from werkzeug.datastructures import FileStorage


class ImportFileParser:
    """
    Unified file parser for import operations.
    
    Supports CSV, Excel (.xlsx, .xls), and JSON files.
    Provides consistent parsing interface across all import systems.
    """
    
    @staticmethod
    def parse_file(file: FileStorage, filename: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Parse import file and return data rows with headers.
        
        Args:
            file: FileStorage object
            filename: Original filename
            
        Returns:
            Tuple of (data_rows, headers)
            
        Raises:
            ValueError: If file format is not supported or parsing fails
        """
        if filename.endswith('.csv'):
            return ImportFileParser._parse_csv(file)
        elif filename.endswith('.json'):
            return ImportFileParser._parse_json(file)
        elif filename.endswith(('.xlsx', '.xls')):
            return ImportFileParser._parse_excel(file)
        else:
            raise ValueError(f"Unsupported file format: {filename}")
    
    @staticmethod
    def _parse_csv(file: FileStorage) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse CSV file."""
        try:
            # Reset file pointer
            file.seek(0)
            
            # Read and decode content
            raw_content = file.stream.read().decode('utf-8-sig')
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)
            
            # Get headers and clean them
            raw_headers = reader.fieldnames or []
            headers = [h.strip() for h in raw_headers if h.strip()]
            
            # Parse rows
            data_rows = []
            for row in reader:
                # Clean row data
                cleaned_row = {}
                for key, value in row.items():
                    if key:  # Skip empty keys
                        cleaned_key = key.strip()
                        cleaned_value = value.strip() if isinstance(value, str) else value
                        cleaned_row[cleaned_key] = cleaned_value
                
                # Only include non-empty rows
                if any(v for v in cleaned_row.values() if v):
                    data_rows.append(cleaned_row)
            
            return data_rows, headers
            
        except Exception as e:
            raise ValueError(f"Failed to parse CSV file: {e}")
    
    @staticmethod
    def _parse_json(file: FileStorage) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse JSON file."""
        try:
            # Reset file pointer
            file.seek(0)
            
            # Load JSON data
            data = json.load(file)
            
            # Handle different JSON structures
            if isinstance(data, dict):
                # Look for common keys that might contain the array
                for key in ['applications', 'data', 'records', 'items']:
                    if key in data and isinstance(data[key], list):
                        data = data[key]
                        break
                else:
                    # If no array found, treat the dict as a single record
                    data = [data]
            
            if not isinstance(data, list):
                raise ValueError("JSON must contain an array of objects")
            
            if not data:
                return [], []
            
            # Get headers from first object
            first_item = data[0] if isinstance(data[0], dict) else {}
            headers = [str(k).strip() for k in first_item.keys() if k]
            
            # Parse rows
            data_rows = []
            for item in data:
                if isinstance(item, dict):
                    # Clean item data
                    cleaned_item = {}
                    for key, value in item.items():
                        if key:  # Skip empty keys
                            cleaned_key = str(key).strip()
                            cleaned_value = str(value).strip() if isinstance(value, str) else value
                            cleaned_item[cleaned_key] = cleaned_value
                    
                    # Only include non-empty items
                    if any(v for v in cleaned_item.values() if v):
                        data_rows.append(cleaned_item)
            
            return data_rows, headers
            
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON format: {e}")
        except Exception as e:
            raise ValueError(f"Failed to parse JSON file: {e}")
    
    @staticmethod
    def _parse_excel(file: FileStorage) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse Excel file (.xlsx, .xls)."""
        try:
            # Reset file pointer
            file.seek(0)
            
            # Load workbook
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    header = str(cell.value).strip()
                    if header:
                        headers.append(header)
            
            if not headers:
                raise ValueError("Excel file has no headers in first row")
            
            # Parse data rows
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not any(cell for cell in row if cell):
                    continue
                
                # Create row dictionary
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        value = row[i]
                        if value is not None:
                            # Convert to string and clean
                            cleaned_value = str(value).strip()
                            row_dict[header] = cleaned_value
                
                # Only include non-empty rows
                if any(v for v in row_dict.values() if v):
                    data_rows.append(row_dict)
            
            return data_rows, headers
            
        except Exception as e:
            raise ValueError(f"Failed to parse Excel file: {e}")
    
    @staticmethod
    def validate_file_structure(
        data_rows: List[Dict[str, Any]], 
        headers: List[str],
        required_columns: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Validate parsed file structure.
        
        Args:
            data_rows: Parsed data rows
            headers: File headers
            required_columns: List of required column names
            
        Returns:
            Validation result with errors and warnings
        """
        result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'stats': {
                'total_rows': len(data_rows),
                'total_columns': len(headers),
                'empty_rows': 0,
                'rows_with_missing_data': 0
            }
        }
        
        # Check for empty file
        if not data_rows:
            result['valid'] = False
            result['errors'].append("File contains no data rows")
            return result
        
        # Check required columns
        if required_columns:
            missing_columns = []
            for required_col in required_columns:
                if required_col not in headers:
                    missing_columns.append(required_col)
            
            if missing_columns:
                result['valid'] = False
                result['errors'].append(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Analyze data quality
        empty_rows = 0
        rows_with_missing_data = 0
        
        for i, row in enumerate(data_rows, start=2):  # Start from row 2 (after header)
            # Check for completely empty rows
            if not any(v for v in row.values() if v):
                empty_rows += 1
                result['warnings'].append(f"Row {i} is completely empty")
                continue
            
            # Check for missing required data
            missing_data = [col for col in required_columns or [] 
                          if col not in row or not row[col]]
            if missing_data:
                rows_with_missing_data += 1
                result['warnings'].append(f"Row {i} missing data for: {', '.join(missing_data)}")
        
        result['stats']['empty_rows'] = empty_rows
        result['stats']['rows_with_missing_data'] = rows_with_missing_data
        
        # Add summary warnings
        if empty_rows > len(data_rows) * 0.1:  # More than 10% empty rows
            result['warnings'].append(f"High number of empty rows: {empty_rows}/{len(data_rows)}")
        
        if rows_with_missing_data > len(data_rows) * 0.2:  # More than 20% with missing data
            result['warnings'].append(f"High number of rows with missing data: {rows_with_missing_data}/{len(data_rows)}")
        
        return result
    
    @staticmethod
    def get_file_info(file: FileStorage, filename: str) -> Dict[str, Any]:
        """
        Get file information for validation and logging.
        
        Args:
            file: FileStorage object
            filename: Original filename
            
        Returns:
            File information dictionary
        """
        # Get file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        # Determine file type
        file_type = 'unknown'
        if filename.endswith('.csv'):
            file_type = 'csv'
        elif filename.endswith('.json'):
            file_type = 'json'
        elif filename.endswith(('.xlsx', '.xls')):
            file_type = 'excel'
        
        return {
            'filename': filename,
            'file_size': file_size,
            'file_type': file_type,
            'mime_type': getattr(file, 'content_type', None),
            'size_mb': round(file_size / (1024 * 1024), 2)
        }


# Convenience functions for common operations
def parse_import_file(file: FileStorage, filename: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Convenience function to parse import file."""
    return ImportFileParser.parse_file(file, filename)


def validate_import_data(
    data_rows: List[Dict[str, Any]], 
    headers: List[str],
    required_columns: Optional[List[str]] = None
) -> Dict[str, Any]:
    """Convenience function to validate import data."""
    return ImportFileParser.validate_file_structure(data_rows, headers, required_columns)


def get_import_file_info(file: FileStorage, filename: str) -> Dict[str, Any]:
    """Convenience function to get file information."""
    return ImportFileParser.get_file_info(file, filename)
