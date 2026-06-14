"""Sophisticated import routes: field mapping, duplicate analysis, Excel upload, AI preview, manual import, history, rollback, template download."""

import csv
import hashlib
import io
import json
from datetime import datetime, timedelta

import openpyxl
from flask import current_app, jsonify, request, send_file
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename

from app import db
from app.decorators import audit_log
from app.models.application_portfolio import ApplicationComponent
from app.models.import_audit import ImportAuditService, ImportSessionLog
from app.services.unified_import.duplicate_detector import DuplicateDetector

from . import unified_applications_bp

# --- Import Security Configuration ---
ALLOWED_IMPORT_EXTENSIONS = {"csv", "xlsx", "xls", "json"}
ALLOWED_IMPORT_MIMETYPES = {
    "text/csv",
    "text/plain",
    "application/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/json",
    "application/octet-stream",
}
MAX_IMPORT_FILE_SIZE = 50 * 1024 * 1024  # 50MB

from app.utils.import_rate_limiter import import_rate_limit


def validate_import_file(file):
    """
    Validate uploaded import file for security.
    Returns (sanitized_filename_lower, error_message) tuple.
    error_message is None on success.
    """
    if not file or file.filename == "":
        return None, "No file selected"

    sanitized = secure_filename(file.filename)
    if not sanitized:
        return None, "Invalid filename"

    ext = sanitized.rsplit(".", 1)[-1].lower() if "." in sanitized else ""
    if ext not in ALLOWED_IMPORT_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_IMPORT_EXTENSIONS))
        return None, f"Unsupported file type '.{ext}'. Allowed: {allowed}"

    # MIME type check (warn only — browsers are inconsistent)
    content_type = getattr(file, "content_type", "") or ""
    if content_type and content_type not in ALLOWED_IMPORT_MIMETYPES:
        current_app.logger.warning(
            "Import file MIME mismatch: %s for %s", content_type, sanitized
        )

    # File size check
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > MAX_IMPORT_FILE_SIZE:
        mb = MAX_IMPORT_FILE_SIZE // (1024 * 1024)
        return None, f"File too large. Maximum size: {mb}MB"
    if size == 0:
        return None, "File is empty"

    return sanitized.lower(), None


# Date fields that need special parsing
DATE_FIELDS = {
    "implementation_date",
    "contract_expiry_date",
    "planned_retirement_date",
    "last_major_upgrade",
    "last_backup_date",
    "last_security_audit_date",
    "last_penetration_test_date",
    "last_assessed",
    "created_at",
    "updated_at",
    "go_live_date",
    "end_of_life_date",
}


import logging as _logging

logger = _logging.getLogger(__name__)
_date_logger = _logging.getLogger(__name__)


def parse_flexible_date(value, date_order="iso"):
    """
    Parse dates flexibly with configurable locale ordering.

    Args:
        value: Raw date string to parse
        date_order: One of "iso" (YYYY-MM-DD first), "dmy" (DD/MM/YYYY first, UK/EU),
                    or "mdy" (MM/DD/YYYY first, US). Default "iso".

    Returns a datetime object or None if parsing fails.
    """
    if not value:
        return None

    value_str = str(value).strip()
    if not value_str:
        return None

    # Year-only values are too ambiguous — could be version, ID, or count.
    # Skip silently instead of converting "2009" to 2009-01-01.
    if value_str.isdigit() and len(value_str) == 4:
        _date_logger.debug("Skipping year-only value '%s' (too ambiguous)", value_str)
        return None

    # Build format list with locale-preferred order
    # ISO 8601 is always tried first (unambiguous)
    iso_formats = [
        "%Y-%m-%d",  # 2024-01-15
        "%Y/%m/%d",  # 2024/01/15
        "%Y-%m-%dT%H:%M:%S",  # ISO datetime
        "%Y-%m-%d %H:%M:%S",  # Standard datetime
        "%Y%m%d",  # 20240115
    ]

    dmy_formats = [
        "%d/%m/%Y",  # 15/01/2024
        "%d-%m-%Y",  # 15-01-2024
    ]

    mdy_formats = [
        "%m/%d/%Y",  # 01/15/2024
        "%m-%d-%Y",  # 01-15-2024
    ]

    # Named month formats are unambiguous
    named_formats = [
        "%d %b %Y",  # 15 Jan 2024
        "%d %B %Y",  # 15 January 2024
        "%b %d, %Y",  # Jan 15, 2024
        "%B %d, %Y",  # January 15, 2024
    ]

    # Assemble format list based on locale preference
    if date_order == "dmy":
        date_formats = iso_formats + dmy_formats + mdy_formats + named_formats
    elif date_order == "mdy":
        date_formats = iso_formats + mdy_formats + dmy_formats + named_formats
    else:  # "iso" default
        date_formats = iso_formats + dmy_formats + mdy_formats + named_formats

    for fmt in date_formats:
        try:
            parsed = datetime.strptime(value_str, fmt)
            # Warn when date is ambiguous (day and month both <= 12)
            if fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
                if parsed.day <= 12 and parsed.month <= 12 and parsed.day != parsed.month:
                    _date_logger.warning(
                        "Ambiguous date '%s' parsed as %s using format '%s' "
                        "(date_order=%s). Consider ISO 8601 (YYYY-MM-DD).",
                        value_str, parsed.strftime("%Y-%m-%d"), fmt, date_order,
                    )
            return parsed
        except ValueError:
            continue

    # If all parsing fails, return None (will skip setting the field)
    return None


def clean_import_data(app_data: dict) -> dict:
    """
    Clean and enrich application data at import time.

    This function:
    1. Detects and removes garbage vendor_name values (e.g., "Stable", "PHP", "React")
    2. Attempts to extract vendor from product name patterns in name/description

    Returns:
        Cleaned app_data dictionary with improved vendor_name
    """
    from app.services.application_architecture_mapper import (
        INVALID_VENDOR_NAMES,
        PRODUCT_TO_VENDOR_MAP,
    )

    # 1. Clean vendor_name - detect garbage values
    vendor_name = app_data.get("vendor_name", "")
    if vendor_name and vendor_name.strip().lower() in INVALID_VENDOR_NAMES:
        app_data["vendor_name"] = None  # Clear garbage
        vendor_name = None

    # 2. Try to extract vendor from product name patterns if no valid vendor
    if not app_data.get("vendor_name"):
        searchable = " ".join(
            [
                app_data.get("name", "") or "",
                app_data.get("description", "") or "",
            ]
        ).lower()

        # Sort patterns by length (longer matches first for accuracy)
        for pattern, vendor in sorted(PRODUCT_TO_VENDOR_MAP.items(), key=lambda x: -len(x[0])):
            if pattern in searchable:
                app_data["vendor_name"] = vendor
                break

    return app_data


@unified_applications_bp.route("/import-fields", methods=["GET"])
@login_required
def get_import_fields():
    """
    Return all valid ApplicationComponent model fields for import mapping.
    This allows the frontend to dynamically show available fields based on the actual model.
    Comprehensive mapping for enterprise application portfolios.
    """
    from sqlalchemy import inspect

    # Get model columns via SQLAlchemy inspection
    mapper = inspect(ApplicationComponent)
    columns = mapper.columns

    # Define comprehensive field categories for enterprise import
    field_categories = {
        "identity": [
            "name",
            "description",
            "application_code",
            "version",
        ],
        "classification": [
            "component_type",
            "application_type",
            "application_category",
            "deployment_model",
            "deployment_status",
            "criticality",
            "business_criticality",
            "lifecycle_status",
            "strategic_importance",
            "business_value",
            "differentiation_level",
        ],
        "capabilities_processes": [
            "archimate_capability",
            "functionality_capabilities",
            "imported_capabilities",
            "application_functions_text",
            "imported_apqc_codes",
        ],
        "business": [
            "business_domain",
            "business_purpose",
            "business_functions",
            "user_base_size",
            "user_types",
            "user_count",
            "user_type",
            "concurrent_users_max",
            "average_daily_users",
        ],
        "vendor": [
            "vendor_name",
            "package_vendor",
            "package_name",
            "vendor_type",
            "contract_type",
            "support_level",
            "contract_expiry_date",
        ],
        "technology": [
            "technology_stack",
            "programming_languages",
            "frameworks",
            "database_platforms",
            "primary_database",
            "cache_technology",
            "message_queue",
            "integration_methods",
            "api_available",
            "api_documentation",
            "exposes_api",
            "integration_pattern",
            "architecture_style",
        ],
        "financial": [
            "total_cost_of_ownership",
            "license_cost",
            "license_type",
            "license_cost_annual",
            "maintenance_cost",
            "infrastructure_cost",
            "infrastructure_cost_monthly",
            "support_cost",
            "implementation_cost",
            "roi_score",
        ],
        "lifecycle": [
            "implementation_date",
            "go_live_date",
            "last_major_upgrade",
            "planned_retirement_date",
            "end_of_life_date",
            "technology_age_years",
        ],
        "performance": [
            "availability_target",
            "availability_actual",
            "performance_rating",
            "user_satisfaction_score",
            "sla_availability_percentage",
            "current_uptime_percentage",
            "response_time_target_ms",
            "throughput_target_tps",
        ],
        "integration": [
            "integration_complexity",
            "number_of_integrations",
            "interfaces_count",
            "dependencies_count",
            "data_architecture",
        ],
        "security": [
            "data_classification",
            "security_level",
            "compliance_requirements",
            "security_certifications",
            "authentication_method",
            "authorization_model",
            "encryption_at_rest",
            "encryption_in_transit",
            "pii_data_processed",
            "gdpr_compliant",
            "last_security_audit_date",
            "last_penetration_test_date",
        ],
        "governance": [
            "application_owner",
            "business_owner",
            "technical_owner",
            "technical_lead",
            "product_manager",
            "development_team",
            "support_team",
            "architecture_domain",
        ],
        "risk": [
            "technical_risk",
            "business_risk",
            "vendor_risk",
            "obsolescence_risk",
        ],
        "infrastructure": [
            "cloud_provider",
            "deployment_region",
            "container_image",
            "kubernetes_namespace",
            "scalability_model",
            "disaster_recovery_enabled",
            "rpo_hours",
            "rto_hours",
            "backup_frequency",
        ],
        "notes": [
            "notes",
            "assessment_notes",
        ],
    }

    # Build human-readable labels from field names
    def make_label(field_name):
        labels = {
            "name": "Name *",
            "application_code": "APP ID / Application Code",
            "archimate_capability": "ArchiMate Capability (for linking)",
            "functionality_capabilities": "APQC PCF Processes (Capabilities)",
            "package_vendor": "Package Vendor",
            "package_name": "Package Name (for vendor inference)",
            "total_cost_of_ownership": "Total Run Cost / TCO",
            "license_cost": "Software / License Cost",
            "infrastructure_cost": "Hardware / Infrastructure Cost",
            "implementation_cost": "Internal Labor Cost",
            "support_cost": "External Labor Cost",
            "maintenance_cost": "External Services Cost",
            "user_satisfaction_score": "User Satisfaction (0 - 100)",
            "interfaces_count": "Number of Interfaces",
            "rpo_hours": "RPO (Hours)",
            "rto_hours": "RTO (Hours)",
            "go_live_date": "First Go-Live Date",
            "planned_retirement_date": "Retirement Date",
            "cloud_provider": "Hosting Environment / Cloud Provider",
            "deployment_region": "Data Center / Region",
            "authentication_method": "Identity Provider / Auth Method",
            "programming_languages": "Programming Languages",
            "integration_complexity": "Technical / Functional Complexity",
            "disaster_recovery_enabled": "DRP Status",
        }
        if field_name in labels:
            return labels[field_name]
        return field_name.replace("_", " ").title()

    # Build field list with metadata
    fields = [{"value": "", "label": "-- Skip Column --", "category": "skip"}]

    # Excluded fields (auto-generated, relationships, internal)
    excluded = {
        "id",
        "created_at",
        "updated_at",
        "discovered_by_ai",
        "discovery_confidence",
        "last_assessed",
        "archimate_element_id",
        "vendor_product_id",
    }

    model_column_names = [c.name for c in columns]

    # Special fields that are not in model but handled specially
    special_fields = {
        "archimate_capability",
        "functionality_capabilities",
        "package_vendor",
        "package_name",
    }

    # Add fields by category
    for category, field_names in field_categories.items():
        for field_name in field_names:
            if field_name in excluded:
                continue
            if field_name in model_column_names or field_name in special_fields:
                fields.append(
                    {
                        "value": field_name,
                        "label": make_label(field_name),
                        "category": category,
                        "required": field_name == "name",
                    }
                )

    # Comprehensive auto-mapping aliases for enterprise CSV headers
    alias_patterns = {
        # Identity
        "name": ["name", "application name", "app name"],
        "application_code": ["app id", "application id", "app code", "id"],
        "description": ["description", "app description"],
        "version": ["version", "app version"],
        # Classification
        "deployment_status": ["application status", "status", "app status"],
        "application_category": ["category", "app category", "application category"],
        "business_criticality": ["business criticality", "criticality", "critical level"],
        "lifecycle_status": [
            "lifecycle status",
            "lifecycle",
            "time destiny (manual)",
            "time destiny",
        ],
        "strategic_importance": ["application weight", "weight", "strategic importance"],
        "vendor_type": ["managed type", "development type", "management type"],
        "deployment_model": ["deployment scope", "deployment model", "deployment"],
        # Capabilities & Processes
        "archimate_capability": [
            "application capability (archimate)",
            "archimate capability",
            "archimate",
            "capability (archimate)",
        ],
        "functionality_capabilities": [
            "functionality (capabilities)",
            "functionality capabilities",
            "capabilities",
            "pcf",
            "apqc",
            "apqc processes",
        ],
        # Business
        "business_domain": [
            "business unit owner of the app",
            "business unit",
            "business domain",
        ],
        "user_types": ["target users", "users", "user types"],
        "user_satisfaction_score": ["user satisfaction", "satisfaction", "user score"],
        # Vendor
        "vendor_name": ["vendor", "vendor name", "supplier", "maintenance provider"],
        "package_vendor": ["package vendor"],
        "package_name": ["package name", "product name"],
        "contract_type": ["contract type", "contract"],
        "support_level": ["support level", "support hours", "support"],
        # Technology
        "programming_languages": ["programming languages", "languages", "programming language"],
        "technology_stack": [
            "application platform",
            "platform",
            "technology stack",
            "based on",
        ],
        "api_documentation": ["main url", "url", "apps portal url", "portal url"],
        "authentication_method": ["identity provider", "idp", "auth method", "access mode"],
        # Financial
        "total_cost_of_ownership": [
            "total run cost (auto)",
            "total run cost",
            "tco",
            "total cost",
        ],
        "license_cost": ["software cost", "license cost", "software"],
        "infrastructure_cost": ["hardware cost", "infrastructure cost", "hardware"],
        "infrastructure_cost_monthly": [
            "facilities and utilities cost",
            "facilities cost",
            "utilities cost",
        ],
        "implementation_cost": ["internal labor cost", "internal labor"],
        "support_cost": ["external labor cost", "external labor"],
        "maintenance_cost": ["external services cost", "external services"],
        # Lifecycle
        "go_live_date": [
            "first go-live (year)",
            "first go-live",
            "go-live",
            "go live date",
            "go live",
        ],
        "planned_retirement_date": ["retirement date", "retirement", "planned retirement"],
        # Performance & Integration
        "interfaces_count": [
            "interfaces - number",
            "interfaces number",
            "number of interfaces",
        ],
        "integration_complexity": [
            "technical complexity",
            "functional complexity",
            "complexity",
        ],
        # Security & DR
        "disaster_recovery_enabled": ["drp status", "drp", "disaster recovery"],
        "rpo_hours": ["rpo", "recovery point objective"],
        "rto_hours": ["rto", "recovery time objective"],
        # Risk
        "technical_risk": ["technology status", "risk level", "tech risk"],
        "vendor_risk": ["vendor & maintenance risk", "vendor risk", "maintenance risk"],
        # Infrastructure
        "cloud_provider": [
            "hosting environment",
            "public cloud provider",
            "cloud provider",
            "hosting",
        ],
        "deployment_region": [
            "sg data center",
            "data center",
            "support region",
            "region",
            "countries where the app is used",
        ],
        # Governance
        "application_owner": ["application manager", "app manager", "application owner"],
        "business_owner": ["app business owner", "business owner"],
        "technical_owner": ["it security officer", "technical owner"],
        "development_team": ["development provider", "it unit managing the app", "it unit"],
        "support_team": ["support team", "maintenance provider"],
        # Notes (catch-all for unmapped fields)
        "notes": [
            "comments",
            "notes",
            "other hosting details",
            "interfaces - description",
            "other costs",
            "major program impact",
            "successor",
        ],
    }

    return jsonify({"fields": fields, "aliases": alias_patterns}), 200


@unified_applications_bp.route("/analyze-import-duplicates", methods=["POST"])
@login_required
@audit_log("analyze_import_duplicates")
@import_rate_limit(max_calls_per_minute=10, max_calls_per_hour=50, max_calls_per_day=200)
def analyze_import_duplicates():
    """
    Analyze an import file and return detailed duplicate analysis.
    Shows exactly which records will be created, updated, or are duplicates.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename, validation_error = validate_import_file(file)
    if validation_error:
        return jsonify({"error": validation_error}), 400

    date_order = request.form.get("date_format", "iso")  # iso, dmy, mdy
    if date_order not in ("iso", "dmy", "mdy"):
        date_order = "iso"

    # Get custom field mappings from frontend
    custom_mappings = {}
    field_mappings_json = request.form.get("field_mappings")
    if field_mappings_json:
        try:
            custom_mappings = json.loads(field_mappings_json)
        except json.JSONDecodeError:
            logger.exception("Failed to JSON parsing")
            pass

    try:
        headers = []
        data_rows = []

        # Parse file based on type
        if filename.endswith(".csv"):
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)
            raw_headers = reader.fieldnames or []
            headers = [h.strip() for h in raw_headers]
            for row in reader:
                cleaned_row = {k.strip(): v for k, v in row.items()}
                data_rows.append(cleaned_row)

        elif filename.endswith(".json"):
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list) and len(data) > 0:
                first_item = data[0] if isinstance(data[0], dict) else {}
                headers = [str(k).strip() for k in first_item.keys()]
                for item in data:
                    if isinstance(item, dict):
                        cleaned_item = {str(k).strip(): v for k, v in item.items()}
                        data_rows.append(cleaned_item)

        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [
                str(cell.value).strip() if cell.value else "" for cell in ws[1] if cell.value
            ]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row if cell):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = str(row[i]).strip() if row[i] else ""
                    data_rows.append(row_dict)
        else:
            return jsonify({"error": "Unsupported file format"}), 400

        # Find name column
        name_column = None
        if custom_mappings:
            for header, target in custom_mappings.items():
                if target == "name":
                    name_column = header
                    break

        if not name_column:
            for h in headers:
                h_upper = str(h).upper() if h else ""
                if h_upper in ["NAME", "APPLICATION NAME", "APP NAME"] or (
                    "NAME" in h_upper and "APP" not in h_upper
                ):
                    name_column = h
                    break

        if not name_column:
            return jsonify({"error": "Name column not found"}), 400

        # Analyze each row
        will_create = []
        will_update = []
        duplicates_in_file = []
        no_name = []
        validation_errors = []  # Track field validation errors

        # Pre-load existing application names for O(1) lookup (avoid N+1)
        lookup = DuplicateDetector.preload_existing_apps()  # model-safety-ok: prefetched
        existing_apps_by_name = lookup["by_name"]

        seen_names = {}  # Track names we've seen in this file

        for row_idx, row_data in enumerate(data_rows, start=2):
            name = None
            if name_column and name_column in row_data:
                name = str(row_data[name_column]).strip() if row_data[name_column] else None

            if not name:
                no_name.append({"row": row_idx, "data": row_data})
                continue

            name_lower = name.lower()

            # Validate mapped fields
            row_errors = []
            for header, value in row_data.items():
                if not value:
                    continue
                target_field = custom_mappings.get(header)
                if target_field and target_field in DATE_FIELDS:
                    # Validate date field
                    parsed = parse_flexible_date(value, date_order)
                    if parsed is None:
                        row_errors.append(
                            {
                                "field": target_field,
                                "header": header,
                                "value": str(value)[:50],
                                "error": "Invalid date format",
                            }
                        )

            if row_errors:
                validation_errors.append({"row": row_idx, "name": name, "errors": row_errors})

            # Check if duplicate in file
            if name_lower in seen_names:
                duplicates_in_file.append(
                    {"row": row_idx, "name": name, "duplicate_of_row": seen_names[name_lower]}
                )
                continue

            seen_names[name_lower] = row_idx

            # Check if exists in database (pre-loaded lookup)
            existing = existing_apps_by_name.get(name_lower)

            if existing:
                will_update.append(
                    {
                        "row": row_idx,
                        "name": name,
                        "existing_id": existing["id"],
                        "existing_name": existing["name"],
                    }
                )
            else:
                will_create.append({"row": row_idx, "name": name})

        # Log successful analysis
        log_import_analysis(
            import_type='unified_applications',
            analysis_results={
                'filename': filename,
                'total_rows': len(data_rows),
                'will_create': len(will_create),
                'will_update': len(will_update),
                'duplicates_in_file': len(duplicates_in_file),
                'no_name': len(no_name),
                'validation_errors': len(validation_errors),
                'operation': 'duplicate_analysis_complete'
            }
        )

        return (
            jsonify(
                {
                    "success": True,
                    "total_rows": len(data_rows),
                    "will_create": len(will_create),
                    "will_update": len(will_update),
                    "duplicates_in_file": len(duplicates_in_file),
                    "no_name": len(no_name),
                    "validation_errors": len(validation_errors),
                    "details": {
                        "create": will_create[:20],  # First 20 for preview
                        "update": will_update[:20],
                        "duplicates": duplicates_in_file[:20],
                        "missing_name": no_name[:10],
                        "validation": validation_errors[:20],  # Show first 20 validation errors
                    },
                }
            ),
            200,
        )

    except Exception as e:
        # Log failed analysis
        log_import_analysis(
            import_type='unified_applications',
            analysis_results={
                'filename': filename,
                'error': str(e),
                'operation': 'duplicate_analysis_failed'
            },
            success=False,
            error_message=str(e)
        )
        current_app.logger.error(f"Error analyzing import: {e}", exc_info=True)
        return jsonify({"error": "Failed to analyze import file. Please check the file format and try again."}), 500


@unified_applications_bp.route("/upload-excel", methods=["POST"])
@login_required
@audit_log("upload_excel_applications")
@import_rate_limit(max_calls_per_minute=5, max_calls_per_hour=30, max_calls_per_day=150)
def upload_excel_applications():
    """
    Simplified upload and process Excel, CSV, or JSON file for application import.
    Supports basic duplicate detection and field mapping.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename, validation_error = validate_import_file(file)
    if validation_error:
        return jsonify({"error": validation_error}), 400

    # Idempotency check: hash file content to detect duplicate uploads
    file_content = file.read()
    file.seek(0)  # Reset for downstream parsing
    file_hash = hashlib.sha256(file_content).hexdigest()[:16]
    idempotency_window = 300  # 5 minutes
    recent_dup = ImportSessionLog.query.filter(
        ImportSessionLog.user_id == current_user.id,
        ImportSessionLog.filename == filename,
        ImportSessionLog.started_at >= datetime.utcnow() - timedelta(seconds=idempotency_window),
    ).first()
    if recent_dup:
        return jsonify({
            "error": "This file was already imported within the last 5 minutes. "
                     "If this was intentional, please wait and try again.",
            "previous_import_id": recent_dup.session_id,
        }), 409

    duplicate_mode = request.form.get("duplicate_mode", "merge")
    date_order = request.form.get("date_format", "iso")  # iso, dmy, mdy
    if date_order not in ("iso", "dmy", "mdy"):
        date_order = "iso"

    # AI options from frontend
    enable_ai = request.form.get("enable_ai", "false").lower() == "true"
    map_capabilities = request.form.get("map_capabilities", "true").lower() == "true"
    map_processes = request.form.get("map_processes", "true").lower() == "true"
    generate_archimate = request.form.get("generate_archimate", "false").lower() == "true"
    confidence_threshold = float(request.form.get("confidence_threshold", "0.7"))

    # Get custom field mappings from frontend (if provided)
    custom_mappings = {}
    field_mappings_json = request.form.get("field_mappings")
    if field_mappings_json:
        try:
            custom_mappings = json.loads(field_mappings_json)
        except json.JSONDecodeError:
            logger.exception("Failed to JSON parsing")
            pass

    # Create audit session
    audit_log = ImportAuditService.create_audit_session(
        user_id=current_user.id,
        import_source='unified_applications',
        filename=filename,
        operation_type='import',
        duplicate_mode=duplicate_mode
    )

    try:

        headers = []
        data_rows = []

        # Parse file based on type
        if filename.endswith(".csv"):
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)
            raw_headers = reader.fieldnames or []
            headers = [h.strip() for h in raw_headers]
            for row in reader:
                cleaned_row = {k.strip(): v for k, v in row.items()}
                data_rows.append(cleaned_row)

        elif filename.endswith(".json"):
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list) and len(data) > 0:
                first_item = data[0] if isinstance(data[0], dict) else {}
                headers = [str(k).strip() for k in first_item.keys()]
                for item in data:
                    if isinstance(item, dict):
                        cleaned_item = {str(k).strip(): v for k, v in item.items()}
                        data_rows.append(cleaned_item)
            else:
                return jsonify({"error": "Invalid JSON structure"}), 400

        else:
            # Parse Excel file (default)
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [
                str(cell.value).strip() if cell.value else "" for cell in ws[1] if cell.value
            ]

        # Find name column using custom mappings or auto-detect
        name_column = None
        app_id_column = None

        # Use custom mappings if provided
        if custom_mappings:
            for header, target_field in custom_mappings.items():
                if target_field == "name":
                    name_column = header
                elif target_field == "application_code":
                    app_id_column = header

        # Fall back to auto-detection
        if not name_column:
            for h in headers:
                h_upper = str(h).upper() if h else ""
                if "NAME" in h_upper and "APP" not in h_upper:
                    name_column = h
                    break
                elif h_upper in ["NAME", "APPLICATION NAME", "APP NAME"]:
                    name_column = h
                    break

        if not name_column:
            return (
                jsonify(
                    {
                        "error": 'Name column not found. Please map at least one column to "Name".'
                    }
                ),
                400,
            )

        # Process rows
        records_created = 0
        records_updated = 0
        records_skipped = 0
        records_failed = 0
        errors = []
        skipped_fields = []  # Track silently dropped values for user visibility
        audit_changes = []  # Track before/after for audit trail

        # Pre-load existing apps for O(1) duplicate detection (avoid N+1)
        lookup = DuplicateDetector.preload_existing_apps()  # model-safety-ok: prefetched

        # Batch pre-load full ApplicationComponent objects for matched rows
        # to avoid N+1 query.get() calls in the main loop
        matched_ids = set()
        for row_data in data_rows:
            if not isinstance(row_data, dict):
                continue
            _name = None
            if custom_mappings:
                for header, target in custom_mappings.items():
                    if target == "name" and header in row_data:
                        _name = str(row_data[header]).strip() if row_data[header] else None
                        break
            if not _name and name_column and name_column in row_data:
                _name = str(row_data[name_column]).strip() if row_data[name_column] else None
            if not _name:
                continue
            _app_id = None
            if app_id_column and app_id_column in row_data:
                _app_id = str(row_data[app_id_column]).strip() if row_data[app_id_column] else None
            _match = DuplicateDetector.find_existing_app(_name, lookup, _app_id)
            if _match:
                matched_ids.add(_match["id"])

        existing_apps_by_id = {}
        if matched_ids:
            existing_apps_by_id = {  # model-safety-ok: single batch query
                app.id: app
                for app in ApplicationComponent.query.filter(
                    ApplicationComponent.id.in_(matched_ids)
                ).all()
            }

        processed_identifiers = set()

        for row_idx, row_data in enumerate(data_rows, start=2):
            if not isinstance(row_data, dict):
                continue
            if not any(v for v in row_data.values() if v):
                continue

            try:
                # Get name using custom mapping
                name = None
                if custom_mappings:
                    for header, target in custom_mappings.items():
                        if target == "name" and header in row_data:
                            name = str(row_data[header]).strip() if row_data[header] else None
                            break
                if not name and name_column and name_column in row_data:
                    name = str(row_data[name_column]).strip() if row_data[name_column] else None

                if not name:
                    records_failed += 1
                    errors.append(f"Row {row_idx}: Name is required")
                    continue

                # Get app_id if available
                app_id = None
                if app_id_column and app_id_column in row_data:
                    app_id = (
                        str(row_data[app_id_column]).strip()
                        if row_data[app_id_column]
                        else None
                    )

                # Duplicate detection
                identifier = (name.lower(), app_id.lower() if app_id else None)
                if identifier in processed_identifiers:
                    if duplicate_mode == "skip":
                        records_skipped += 1
                        continue

                # Check existing application (shared detector, pre-loaded)
                match = DuplicateDetector.find_existing_app(name, lookup, app_id)
                existing_app_id = match["id"] if match else None
                existing_app = (
                    existing_apps_by_id.get(existing_app_id)
                    if existing_app_id else None
                )

                # Build app_data using custom mappings
                app_data = {"name": name}
                if app_id:
                    app_data["application_code"] = app_id

                # Define numeric fields that need type conversion
                NUMERIC_FIELDS = {
                    "user_base_size",
                    "user_count",
                    "concurrent_users_max",
                    "average_daily_users",
                    "total_cost_of_ownership",
                    "license_cost",
                    "license_cost_annual",
                    "maintenance_cost",
                    "infrastructure_cost",
                    "infrastructure_cost_monthly",
                    "support_cost",
                    "implementation_cost",
                    "roi_score",
                    "interfaces_count",
                    "number_of_integrations",
                    "dependencies_count",
                    "availability_target",
                    "availability_actual",
                    "user_satisfaction_score",
                    "rpo_hours",
                    "rto_hours",
                    "technology_age_years",
                    "response_time_target_ms",
                    "throughput_target_tps",
                }
                BOOLEAN_FIELDS = {
                    "api_available",
                    "exposes_api",
                    "competitive_advantage",
                    "manufacturing_critical",
                    "shop_floor_system",
                    "real_time_requirements",
                    "disaster_recovery_enabled",
                    "encryption_at_rest",
                    "encryption_in_transit",
                    "pii_data_processed",
                    "gdpr_compliant",
                }

                for header, value in row_data.items():
                    if not value:
                        continue
                    target_field = custom_mappings.get(header)
                    if target_field and target_field not in ["", "name", "application_code"]:
                        value_str = str(value).strip()

                        # Handle special capability fields
                        if target_field == "archimate_capability":
                            # Store ArchiMate capability text for AI processing
                            app_data["imported_capabilities"] = value_str
                        elif target_field == "functionality_capabilities":
                            # Store APQC/PCF codes for AI processing
                            app_data["imported_apqc_codes"] = value_str
                            app_data["application_functions_text"] = value_str
                        # Handle date fields
                        elif target_field in DATE_FIELDS:
                            parsed_date = parse_flexible_date(value, date_order)
                            if parsed_date:
                                app_data[target_field] = parsed_date
                            else:
                                current_app.logger.warning(
                                    "Import row %d: could not parse '%s' as date for field '%s'",
                                    row_idx, value_str[:50], target_field,
                                )
                                skipped_fields.append({
                                    "row": row_idx, "field": target_field,
                                    "value": value_str[:50], "reason": "unparseable date",
                                })
                        # Handle numeric fields
                        elif target_field in NUMERIC_FIELDS:
                            try:
                                # Remove currency symbols, commas, and whitespace
                                clean_value = (
                                    value_str.replace(",", "")
                                    .replace("$", "")
                                    .replace("\u20ac", "")
                                    .replace("\u00a3", "")
                                    .strip()
                                )
                                if clean_value:
                                    if "." in clean_value:
                                        app_data[target_field] = float(clean_value)
                                    else:
                                        app_data[target_field] = int(clean_value)
                            except ValueError:  # fabricated-values-ok
                                current_app.logger.warning(
                                    "Import row %d: could not parse '%s' as number for field '%s'",
                                    row_idx, value_str[:50], target_field,
                                )
                                skipped_fields.append({
                                    "row": row_idx, "field": target_field,
                                    "value": value_str[:50], "reason": "unparseable number",
                                })
                        # Handle boolean fields
                        elif target_field in BOOLEAN_FIELDS:
                            app_data[target_field] = value_str.lower() in (
                                "yes",
                                "true",
                                "1",
                                "y",
                                "enabled",
                                "active",
                            )
                        else:
                            app_data[target_field] = value_str

                # Clean and enrich data at import time
                app_data = clean_import_data(app_data)

                # Create or update
                if existing_app:
                    if duplicate_mode == "skip":
                        records_skipped += 1
                    elif duplicate_mode in ["merge", "update"]:
                        changed_fields = {}
                        for key, val in app_data.items():
                            if key == "name":
                                continue  # Don't overwrite primary key field
                            if not hasattr(existing_app, key):  # model-safety-ok
                                continue
                            # Skip empty/falsy values — preserve existing data
                            # Catches None, empty strings, 0, 0.0, False
                            if not val and val is not False:
                                continue
                            if isinstance(val, str) and not val.strip():
                                continue
                            old_val = getattr(existing_app, key)  # model-safety-ok
                            if old_val != val:
                                changed_fields[key] = {"old": str(old_val) if old_val else None, "new": str(val)}
                                setattr(existing_app, key, val)
                        if changed_fields:
                            current_app.logger.debug(
                                "Import merge updated %s: %s",
                                existing_app.name, ", ".join(changed_fields.keys())
                            )
                            audit_changes.append({
                                "app_name": name, "app_id": existing_app.id,
                                "action": "updated", "changed_fields": changed_fields,
                            })
                        records_updated += 1
                    else:  # duplicate mode
                        new_app = ApplicationComponent(**app_data)
                        db.session.add(new_app)
                        audit_changes.append({"app_name": name, "action": "created"})
                        records_created += 1
                else:
                    new_app = ApplicationComponent(**app_data)
                    db.session.add(new_app)
                    audit_changes.append({"app_name": name, "action": "created"})
                    records_created += 1

                processed_identifiers.add(identifier)

            except Exception as e:
                records_failed += 1
                current_app.logger.warning(f"Import row {row_idx} failed: {e}")
                errors.append(f"Row {row_idx}: Could not process this record")
                continue

        # Commit changes
        db.session.commit()

        # Update audit log with results and detailed changes
        try:
            # Add detailed changes to audit log
            for change in audit_changes:
                if change.get("action") == "created":
                    audit_log.add_record_change(
                        record_id=change.get("app_id"),
                        record_type="ApplicationComponent",
                        operation="create",
                        before_data={},
                        after_data={"name": change.get("app_name")},
                        changed_fields=["name"]
                    )
                elif change.get("action") == "updated":
                    audit_log.add_record_change(
                        record_id=change.get("app_id"),
                        record_type="ApplicationComponent",
                        operation="update",
                        before_data={k: v.get("old") for k, v in change.get("changed_fields", {}).items()},
                        after_data={k: v.get("new") for k, v in change.get("changed_fields", {}).items()},
                        changed_fields=list(change.get("changed_fields", {}).keys())
                    )

            # Complete the audit session
            audit_log.records_processed = len(data_rows)
            audit_log.records_created = records_created
            audit_log.records_updated = records_updated
            audit_log.records_skipped = records_skipped
            audit_log.records_failed = records_failed
            audit_log.changes_summary = audit_changes[:500]  # Cap at 500 entries
            audit_log.complete_import(status='completed')

            db.session.add(audit_log)
            db.session.commit()

        except Exception as audit_err:
            db.session.rollback()
            current_app.logger.warning("Failed to write import audit log: %s", audit_err)
            # Don't fail the import if audit logging fails

        return (
            jsonify(
                {
                    "success": True,
                    "created": records_created,
                    "updated": records_updated,
                    "skipped": records_skipped,
                    "failed": records_failed,
                    "errors": errors[:10],
                    "skipped_fields": skipped_fields[:100],
                }
            ),
            200,
        )

    except Exception as e:
        db.session.rollback()

        # Mark audit log as failed
        try:
            if 'audit_log' in locals():
                audit_log.complete_import(status='failed', error_message=str(e))
                db.session.add(audit_log)
                db.session.commit()
        except Exception as audit_err:
            current_app.logger.warning("Failed to mark audit log as failed: %s", audit_err)

        current_app.logger.error(f"Error importing file: {e}", exc_info=True)
        return jsonify({"error": "Import failed. Please check your file format and try again."}), 500


@unified_applications_bp.route("/preview-ai-analysis", methods=["POST"])
@login_required
@audit_log("preview_ai_analysis")
@import_rate_limit(max_calls_per_minute=3, max_calls_per_hour=20, max_calls_per_day=100)
def preview_ai_analysis():
    """Preview AI analysis results BEFORE committing import."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    filename, validation_error = validate_import_file(file)
    if validation_error:
        return jsonify({"error": validation_error}), 400
    confidence_threshold = float(request.form.get("confidence_threshold", "0.7"))
    max_preview = int(request.form.get("max_preview", "10"))

    try:
        data_rows = []

        # Parse file based on type
        if filename.endswith(".csv"):
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)
            for row in reader:
                cleaned_row = {k.strip(): v for k, v in row.items()}
                data_rows.append(cleaned_row)

        elif filename.endswith(".json"):
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list) and len(data) > 0:
                for item in data:
                    if isinstance(item, dict):
                        cleaned_item = {str(k).strip(): v for k, v in item.items()}
                        data_rows.append(cleaned_item)

        else:
            # Parse Excel file (default)
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [
                str(cell.value).strip() if cell.value else "" for cell in ws[1] if cell.value
            ]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row if cell):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = str(row[i]).strip() if row[i] else ""
                    data_rows.append(row_dict)

        # Identify name column (simple detection)
        name_column = None
        if data_rows:
            first_row = data_rows[0]
            for key in first_row.keys():
                if key.upper() in ["NAME", "APPLICATION NAME", "APP NAME"]:
                    name_column = key
                    break
            if not name_column:
                for key in first_row.keys():
                    if "NAME" in key.upper():
                        name_column = key
                        break

        # Prepare applications data for AI
        applications_data = []
        for row_data in data_rows[:max_preview]:
            name = None
            if name_column and name_column in row_data:
                name = str(row_data[name_column]).strip() if row_data[name_column] else None

            if not name:
                continue

            app_data = {"name": name}

            # Add other fields as description/context
            description_parts = []
            for k, v in row_data.items():
                if k != name_column and v:
                    description_parts.append(f"{k}: {v}")

            if description_parts:
                app_data["description"] = " ".join(description_parts)

            applications_data.append(app_data)

        # Call AI Service
        from app.services.ai_import_service import get_ai_import_service

        ai_service = get_ai_import_service()

        preview = ai_service.analyze_file_data_for_preview(
            applications_data=applications_data, confidence_threshold=confidence_threshold
        )

        return (
            jsonify(
                {
                    "success": True,
                    "total_applications": len(data_rows),
                    "summary": {
                        "total_capabilities_found": preview.get("capability_mappings_found", 0),
                        "total_processes_found": preview.get("process_mappings_found", 0),
                        "total_archimate_elements": preview.get(
                            "archimate_elements_generated", 0
                        ),
                    },
                    "applications": preview.get("applications", []),
                }
            ),
            200,
        )

    except Exception as e:
        current_app.logger.error(f"Error previewing AI analysis: {e}", exc_info=True)
        return jsonify({"error": "AI analysis preview failed. Please try again."}), 500


@unified_applications_bp.route("/import-manual", methods=["POST"])
@login_required
@audit_log("import_manual_applications")
def import_manual_applications():
    """Process manual entry applications"""
    data = request.get_json()
    applications = data.get("applications", [])
    duplicate_mode = data.get("duplicate_mode", "merge")
    date_order = data.get("date_format", "iso")
    if date_order not in ("iso", "dmy", "mdy"):
        date_order = "iso"

    records_created = 0
    records_updated = 0
    records_skipped = 0
    records_failed = 0
    errors = []
    skipped_fields = []  # Track silently dropped values for user visibility
    audit_changes = []  # Track before/after for audit trail

    # Pre-load existing apps for consistent case-insensitive matching
    lookup = DuplicateDetector.preload_existing_apps()  # model-safety-ok: prefetched

    # Batch pre-load full ApplicationComponent objects for matched entries
    # to avoid N+1 query.get() calls in the loop
    matched_ids = set()
    for app_entry in applications:
        _name = app_entry.get("name", "").strip()
        _app_id = app_entry.get("app_id", "").strip() or None
        if _name:
            _match = DuplicateDetector.find_existing_app(_name, lookup, _app_id)
            if _match:
                matched_ids.add(_match["id"])

    existing_apps_by_id = {}
    if matched_ids:
        existing_apps_by_id = {  # model-safety-ok: single batch query
            app.id: app
            for app in ApplicationComponent.query.filter(
                ApplicationComponent.id.in_(matched_ids)
            ).all()
        }

    for idx, app_data in enumerate(applications, start=1):
        try:
            name = app_data.get("name", "").strip()
            app_id = app_data.get("app_id", "").strip() or None

            if not name:
                records_failed += 1
                errors.append(f"Row {idx}: Name is required")
                continue

            # Check for duplicates (shared detector, pre-loaded)
            match = DuplicateDetector.find_existing_app(name, lookup, app_id)
            existing_app = (
                existing_apps_by_id.get(match["id"])
                if match else None
            )

            # Process date fields
            processed_data = {}
            for key, value in app_data.items():
                if key in DATE_FIELDS and value:
                    parsed_date = parse_flexible_date(value, date_order)
                    if parsed_date:
                        processed_data[key] = parsed_date
                    else:
                        value_str = str(value).strip()
                        current_app.logger.warning(
                            "Manual import row %d: could not parse '%s' as date for field '%s'",
                            idx, value_str[:50], key,
                        )
                        skipped_fields.append({
                            "row": idx, "field": key,
                            "value": value_str[:50], "reason": "unparseable date",
                        })
                elif value:
                    processed_data[key] = value

            # Clean and enrich data at import time
            processed_data = clean_import_data(processed_data)

            if existing_app and duplicate_mode in ("merge", "update"):
                # Update existing — skip empty strings to preserve data
                changed_fields = {}
                for key, value in processed_data.items():
                    if key == "name":
                        continue
                    if not hasattr(existing_app, key):  # model-safety-ok
                        continue
                    # Skip empty/falsy values — preserve existing data
                    # Catches None, empty strings, 0, 0.0, False
                    if not value and value is not False:
                        continue
                    if isinstance(value, str) and not value.strip():
                        continue
                    old_val = getattr(existing_app, key)  # model-safety-ok
                    if old_val != value:
                        changed_fields[key] = {"old": str(old_val) if old_val else None, "new": str(value)}
                        setattr(existing_app, key, value)
                if changed_fields:
                    audit_changes.append({
                        "app_name": name, "app_id": existing_app.id,
                        "action": "updated", "changed_fields": changed_fields,
                    })
                records_updated += 1
            elif existing_app and duplicate_mode == "skip":
                records_skipped += 1
                continue
            else:
                # Create new - filter to only valid ApplicationComponent fields, exclude system fields
                SYSTEM_FIELDS = {"id", "created_at", "updated_at", "created_by", "updated_by"}
                valid_fields = {col.name for col in ApplicationComponent.__table__.columns} - SYSTEM_FIELDS
                filtered_data = {
                    k: v for k, v in processed_data.items() if k in valid_fields
                }
                app = ApplicationComponent(**filtered_data)
                db.session.add(app)
                audit_changes.append({"app_name": name, "action": "created"})
                records_created += 1

        except Exception as e:
            records_failed += 1
            current_app.logger.warning(f"Manual import row {idx} failed: {e}")
            errors.append(f"Row {idx}: Could not process this record")

    db.session.commit()

    # Write audit trail
    try:
        audit = ImportSessionLog(
            user_id=current_user.id,
            user_email=current_user.email if hasattr(current_user, "email") else None,  # model-safety-ok
            import_type="manual",
            records_created=records_created,
            records_updated=records_updated,
            records_skipped=records_skipped,
            records_failed=records_failed,
            duplicate_mode=duplicate_mode,
            changes=audit_changes[:500],
            errors=errors[:50],
        )
        db.session.add(audit)
        db.session.commit()
    except Exception as audit_err:
        current_app.logger.warning("Failed to write import audit log: %s", audit_err)

    return (
        jsonify(
            {
                "success": True,
                "created": records_created,
                "updated": records_updated,
                "skipped": records_skipped,
                "failed": records_failed,
                "errors": errors,
                "skipped_fields": skipped_fields[:100],
            }
        ),
        200,
    )


@unified_applications_bp.route("/import-history", methods=["GET"])
@login_required
def import_history():
    """Get import history from audit trail."""
    page = request.args.get("page", 1, type=int)
    per_page = min(request.args.get("per_page", 20, type=int), 100)
    import_source = request.args.get("import_source", "unified_applications")

    try:
        # Get import history using the audit service
        audit_logs = ImportAuditService.get_import_history(
            user_id=current_user.id,
            import_source=import_source,
            limit=per_page,
            offset=(page - 1) * per_page
        )

        # Get total count for pagination
        total_query = ImportSessionLog.query.filter(
            ImportSessionLog.user_id == current_user.id,
            ImportSessionLog.import_source == import_source
        )
        total = total_query.count()

        return jsonify({
            "history": [log.to_dict() for log in audit_logs],
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": (total + per_page - 1) // per_page,
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.debug("Import history query failed: %s", e)
        return jsonify({
            "history": [],
            "total": 0,
            "page": page,
            "per_page": per_page,
            "pages": 0,
        }), 200


@unified_applications_bp.route("/import-history/<string:session_id>/rollback", methods=["POST"])
@login_required
@audit_log("rollback_import_by_session")
def rollback_import_by_session(session_id):
    """Rollback a specific import by restoring previous values (by audit session ID)."""
    from flask_login import current_user

    # Get audit record
    audit_log = ImportAuditService.get_audit_session(session_id)
    if not audit_log:
        return jsonify({"error": "Import session not found"}), 404

    # Check if user has permission (admin or original importer)
    if not (hasattr(current_user, 'is_admin') and current_user.is_admin) and current_user.id != audit_log.user_id:
        return jsonify({"error": "Permission denied. Only admins or the original importer can rollback."}), 403

    # Check if audit has rollback data
    if not audit_log.rollback_data:
        return jsonify({"error": "No rollback data available for this import."}), 400

    # Get rollback reason from request
    request_data = request.get_json() or {}
    rollback_reason = request_data.get("reason", "Manual rollback by user")

    try:
        # Perform rollback using audit service
        rollback_results = ImportAuditService.rollback_import(
            session_id=session_id,
            user_id=current_user.id,
            reason=rollback_reason
        )

        return jsonify({
            "success": True,
            "message": f"Rollback completed. {rollback_results['records_restored']} records restored.",
            "results": rollback_results
        }), 200

    except Exception as e:
        current_app.logger.error(f"Rollback failed for session {session_id}: {e}", exc_info=True)
        return jsonify({"error": "Rollback failed. Please contact an administrator."}), 500


@unified_applications_bp.route("/download-template", methods=["GET"])
@login_required
def download_application_template():
    """Download comprehensive Excel template with current application data"""
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Comprehensive headers matching enterprise import format
    headers = [
        # Identity
        "APP ID",
        "Name",
        "Description",
        # Classification
        "Application Capability (Archimate)",
        "Application Status",
        "Managed type",
        "Category",
        "Business Criticality",
        "Lifecycle Status",
        # Business
        "Target Users",
        "Business Unit Owner of the App",
        # Vendor
        "Package Name",
        "Package Vendor",
        "Vendor",
        # Lifecycle
        "First Go-Live (year)",
        "Retirement date",
        # Technology
        "Application Platform",
        "Programming Languages",
        "Main URL",
        "Identity Provider",
        "Hosting Environment",
        "Public Cloud Provider",
        "SG Data Center",
        # Integration
        "Interfaces - Number",
        "Interfaces - Description",
        "Functional Complexity",
        # Risk
        "Technology status",
        "Vendor & Maintenance Risk",
        "Risk Level",
        # Disaster Recovery
        "DRP Status",
        "RPO",
        "RTO",
        # Performance
        "User Satisfaction",
        # Support
        "Support Level",
        "Support Hours",
        "Support Region",
        "Maintenance Provider",
        # Financial
        "Total Run Cost (Auto)",
        "Hardware Cost",
        "Software Cost",
        "Facilities and Utilities Cost",
        "Internal Labor Cost",
        "External Labor Cost",
        "External Services Cost",
        # Governance
        "IT Unit Managing the App",
        "Application Manager",
        "App Business Owner",
        "IT Security Officer",
        # Capabilities
        "Functionality (Capabilities)",
        # Notes
        "Comments",
        "Other Hosting Details",
    ]

    # Map model fields to Excel columns
    field_mapping = {
        "application_code": "APP ID",
        "name": "Name",
        "description": "Description",
        "deployment_status": "Application Status",
        "vendor_type": "Managed type",
        "application_category": "Category",
        "business_criticality": "Business Criticality",
        "lifecycle_status": "Lifecycle Status",
        "user_types": "Target Users",
        "business_domain": "Business Unit Owner of the App",
        "vendor_name": "Vendor",
        "go_live_date": "First Go-Live (year)",
        "planned_retirement_date": "Retirement date",
        "technology_stack": "Application Platform",
        "programming_languages": "Programming Languages",
        "api_documentation": "Main URL",
        "authentication_method": "Identity Provider",
        "cloud_provider": "Hosting Environment",
        "deployment_region": "SG Data Center",
        "interfaces_count": "Interfaces - Number",
        "integration_complexity": "Functional Complexity",
        "technical_risk": "Technology status",
        "vendor_risk": "Vendor & Maintenance Risk",
        "disaster_recovery_enabled": "DRP Status",
        "rpo_hours": "RPO",
        "rto_hours": "RTO",
        "user_satisfaction_score": "User Satisfaction",
        "support_level": "Support Level",
        "support_team": "Maintenance Provider",
        "total_cost_of_ownership": "Total Run Cost (Auto)",
        "infrastructure_cost": "Hardware Cost",
        "license_cost": "Software Cost",
        "implementation_cost": "Internal Labor Cost",
        "support_cost": "External Labor Cost",
        "maintenance_cost": "External Services Cost",
        "development_team": "IT Unit Managing the App",
        "application_owner": "Application Manager",
        "business_owner": "App Business Owner",
        "technical_owner": "IT Security Officer",
        "notes": "Comments",
    }

    # Style for header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Write headers with styling
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        # Set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    # Get current applications and write sample data
    applications = ApplicationComponent.query.limit(20).all()

    for row_idx, app in enumerate(applications, start=2):
        for col_idx, header in enumerate(headers, start=1):
            # Find matching field
            field_name = None
            for model_field, excel_header in field_mapping.items():
                if excel_header == header:
                    field_name = model_field
                    break

            if field_name and hasattr(app, field_name):
                value = getattr(app, field_name)
                # Handle date fields
                if value is not None:
                    if hasattr(value, "isoformat"):
                        value = (
                            value.strftime("%Y-%m-%d")
                            if hasattr(value, "strftime")
                            else str(value)
                        )
                    elif isinstance(value, bool):
                        value = "Yes" if value else "No"
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f'applications_template_{datetime.now().strftime("%Y%m%d")}.xlsx',
    )
