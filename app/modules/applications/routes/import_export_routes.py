"""
Application import/export routes.

Extracted from app/routes/unified_applications_routes.py
(lines 218-371, 3818-3999, 4002-4213, 4216-4302, 4305-4462,
 4465-4475, 4477-5820, 5823-7160).

Routes:
    - analyze_import_stream()          POST "/analyze-import-stream"
    - export_csv()                     GET  "/export/csv"
    - analyze_import()                 POST "/analyze-import"
    - preview_excel()                  POST "/preview-excel"
    - rollback_import(history_id)      POST "/rollback-import/<int:history_id>"
    - application_import_page()        GET  "/import"
    - application_import()             POST "/import"
    - import_with_ai_review()          POST "/import-with-ai-review"
"""

import csv
import io
import json
import logging
import re
from datetime import datetime

from flask import (
    Response,
    current_app,
    flash,
    jsonify,
    make_response,
    redirect,
    request,
    stream_with_context,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl import load_workbook
from app import db
from app.models.application_portfolio import ApplicationComponent
from app.models.audit_log import AuditLog
from app.services.ai_import_service import get_ai_import_service
from app.utils.file_validation import (
    InvalidFileTypeError,
    get_allowed_extensions_display,
    validate_mime_type,
)
from app.utils.import_rate_limiter import add_rate_limit_headers, import_rate_limit
from app.decorators import audit_log

from . import unified_applications_bp
from ._helpers import _sanitize_csv_value

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 1. analyze_import_stream  (POST /analyze-import-stream)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/analyze-import-stream", methods=["POST"])
@login_required
@audit_log("analyze_import_stream")
@import_rate_limit(max_calls_per_minute=5, max_calls_per_hour=30, max_calls_per_day=150)
def analyze_import_stream():
    """Stream AI analysis progress in real-time using Server-Sent Events."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    confidence_threshold = float(request.form.get("confidence_threshold", "0.7"))
    archimate_mode = request.form.get("archimate_mode", "standard")

    # Read file into memory before starting the generator (request context closes file)
    filename = (file.filename or "").lower()
    file_content = file.read()

    def generate_progress():
        try:
            applications_data = []

            yield f"data: {json.dumps({'stage': 'parsing', 'message': 'Reading file...', 'progress': 0})}\n\n"

            if filename.endswith((".xlsx", ".xls")):
                file_stream = io.BytesIO(file_content)
                wb = load_workbook(file_stream, data_only=True)
                ws = wb.active
                headers = [
                    str(cell.value).strip() if cell.value else ""
                    for cell in ws[1]
                    if cell.value
                ]

                row_count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(cell for cell in row if cell):
                        row_data = {}
                        for i, header in enumerate(headers):
                            if i < len(row) and row[i]:
                                row_data[header.lower().replace(" ", "_")] = str(
                                    row[i]
                                ).strip()

                        app_name = (
                            row_data.get("name")
                            or row_data.get("application_name")
                            or row_data.get("app_name")
                            or row_data.get("application")
                            or row_data.get("system_name")
                        )

                        if app_name:
                            row_data["name"] = app_name
                            applications_data.append(row_data)
                        row_count += 1

                        if row_count % 10 == 0:
                            yield f"data: {json.dumps({'stage': 'parsing', 'message': f'Parsed {row_count} rows...', 'progress': 5})}\n\n"

            elif filename.endswith(".csv"):
                raw_content = file_content.decode("utf-8-sig")
                stream = io.StringIO(raw_content)
                reader = csv.DictReader(stream)

                row_count = 0
                for row in reader:
                    row_data = {
                        k.strip().lower().replace(" ", "_"): v.strip()
                        for k, v in row.items()
                        if v and v.strip()
                    }

                    app_name = (
                        row_data.get("name")
                        or row_data.get("application_name")
                        or row_data.get("app_name")
                        or row_data.get("application")
                        or row_data.get("system_name")
                    )

                    if app_name:
                        row_data["name"] = app_name
                        applications_data.append(row_data)
                    row_count += 1

                    if row_count % 10 == 0:
                        yield f"data: {json.dumps({'stage': 'parsing', 'message': f'Parsed {row_count} rows...', 'progress': 5})}\n\n"

            elif filename.endswith(".json"):
                data = json.loads(file_content.decode("utf-8"))
                if isinstance(data, list):
                    applications_data = data

            total_apps = len(applications_data)
            yield f"data: {json.dumps({'stage': 'parsing', 'message': f'Parsed {total_apps} applications', 'progress': 10, 'total_apps': total_apps})}\n\n"

            if total_apps == 0:
                yield f"data: {json.dumps({'stage': 'error', 'message': 'No applications found in file', 'progress': 0})}\n\n"
                return

            yield f"data: {json.dumps({'stage': 'analyzing', 'message': 'Starting AI analysis...', 'progress': 15})}\n\n"

            ai_service = get_ai_import_service()
            preview_results = ai_service.analyze_file_data_for_preview(
                applications_data=applications_data,
                confidence_threshold=confidence_threshold,
                archimate_mode=archimate_mode,
            )

            yield f"data: {json.dumps({'stage': 'analyzing', 'message': 'AI analysis in progress...', 'progress': 50})}\n\n"

            results = {
                "total_analyzed": preview_results.get("total_analyzed", 0),
                "capability_mappings_found": preview_results.get(
                    "capability_mappings_found", 0
                ),
                "process_mappings_found": preview_results.get(
                    "process_mappings_found", 0
                ),
                "archimate_elements_generated": preview_results.get(
                    "archimate_elements_generated", 0
                ),
                "vendor_matches": preview_results.get("vendor_matches", 0),
                "high_confidence_count": preview_results.get(
                    "high_confidence_count", 0
                ),
                "applications": preview_results.get("applications", []),
            }

            yield f"data: {json.dumps({'stage': 'analyzing', 'message': 'Finalizing results...', 'progress': 90, 'capabilities': results['capability_mappings_found'], 'processes': results['process_mappings_found'], 'archimate': results['archimate_elements_generated']})}\n\n"

            yield f"data: {json.dumps({'stage': 'complete', 'message': 'Analysis complete', 'progress': 100, 'results': results})}\n\n"

        except Exception as e:
            current_app.logger.error(f"Error in streaming analysis: {e}")
            yield f"data: {json.dumps({'stage': 'error', 'message': 'An error occurred during analysis.', 'progress': 0})}\n\n"

    return Response(
        stream_with_context(generate_progress()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
        # Set a longer timeout for streaming responses (10 minutes)
        direct_passthrough=True,
    )


# ---------------------------------------------------------------------------
# 2. export_csv  (GET /export/csv)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/export/csv")
@login_required
def export_csv():
    """Export applications as comprehensive CSV matching import template format"""
    try:
        import csv
        from io import StringIO

        # Bounded query to prevent OOM on large datasets
        MAX_EXPORT = 10000
        query = ApplicationComponent.query.order_by(ApplicationComponent.name)

        ids_param = request.args.get("ids", "").strip()
        if ids_param:
            try:
                id_list = [int(x) for x in ids_param.split(",") if x.strip()]
                if id_list:
                    query = query.filter(ApplicationComponent.id.in_(id_list))
            except (ValueError, TypeError):
                logger.exception("Failed to compute id_list")
                pass

        apps = query.limit(MAX_EXPORT).all()

        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)

        # Comprehensive header matching import template
        headers = [
            "APP ID",
            "Name",
            "Description",
            "Application Capability (Archimate)",
            "Application Status",
            "Managed type",
            "Category",
            "Business Criticality",
            "Lifecycle Status",
            "Target Users",
            "Business Unit Owner of the App",
            "Package Name",
            "Package Vendor",
            "Vendor",
            "First Go-Live (year)",
            "Retirement date",
            "Application Platform",
            "Programming Languages",
            "Main URL",
            "Identity Provider",
            "Hosting Environment",
            "Public Cloud Provider",
            "SG Data Center",
            "Interfaces - Number",
            "Functional Complexity",
            "Technology status",
            "Vendor & Maintenance Risk",
            "Risk Level",
            "DRP Status",
            "RPO",
            "RTO",
            "User Satisfaction",
            "Support Level",
            "Maintenance Provider",
            "Total Run Cost (Auto)",
            "Hardware Cost",
            "Software Cost",
            "Internal Labor Cost",
            "External Labor Cost",
            "External Services Cost",
            "IT Unit Managing the App",
            "Application Manager",
            "App Business Owner",
            "IT Security Officer",
            "Functionality (Capabilities)",
            "Comments",
            "Created At",
            "Updated At",
        ]
        writer.writerow(headers)

        # Batch-load ArchiMate elements to avoid N+1 queries
        archimate_map = {}
        archimate_ids = [a.archimate_element_id for a in apps if a.archimate_element_id]
        if archimate_ids:
            from app.models.archimate_core import ArchiMateElement

            elements = ArchiMateElement.query.filter(
                ArchiMateElement.id.in_(archimate_ids)
            ).all()
            archimate_map = {e.id: e.name or "" for e in elements}

        # Write data for each application
        for app in apps:
            archimate_cap = archimate_map.get(app.archimate_element_id, "")

            # Format dates
            go_live = app.go_live_date.strftime("%Y") if app.go_live_date else ""
            retirement = (
                app.planned_retirement_date.strftime("%Y-%m-%d")
                if app.planned_retirement_date
                else ""
            )
            created_at = (
                app.created_at.strftime("%Y-%m-%d %H:%M:%S") if app.created_at else ""
            )
            updated_at = (
                app.updated_at.strftime("%Y-%m-%d %H:%M:%S") if app.updated_at else ""
            )

            # Get APQC codes if available
            functionality = (
                app.imported_apqc_codes or app.application_functions_text or ""
            )

            # Format boolean fields
            drp_status = (
                "Yes"
                if app.disaster_recovery_enabled
                else "No"
                if app.disaster_recovery_enabled is not None
                else ""
            )

            row = [
                app.application_code or "",
                app.name or "",
                app.description or "",
                archimate_cap,
                app.deployment_status or "",
                app.vendor_type or "",
                app.application_category or "",
                app.business_criticality or "",
                app.lifecycle_status or "",
                app.user_types or "",
                app.business_domain or "",
                "",  # Package Name - would need vendor product lookup
                "",  # Package Vendor - would need vendor product lookup
                app.vendor_name or "",
                go_live,
                retirement,
                app.technology_stack or "",
                app.programming_languages or "",
                app.api_documentation or "",
                app.authentication_method or "",
                app.cloud_provider or "",
                app.cloud_provider or "",  # Public Cloud Provider
                app.deployment_region or "",
                app.interfaces_count or "",
                app.integration_complexity or "",
                app.technical_risk or "",
                app.vendor_risk or "",
                app.technical_risk or "",  # Risk Level
                drp_status,
                app.rpo_hours or "",
                app.rto_hours or "",
                app.user_satisfaction_score or "",
                app.support_level or "",
                app.support_team or "",
                app.total_cost_of_ownership or "",
                app.infrastructure_cost or "",
                app.license_cost or "",
                app.implementation_cost or "",
                app.support_cost or "",
                app.maintenance_cost or "",
                app.development_team or "",
                app.application_owner or "",
                app.business_owner or "",
                app.technical_owner or "",
                functionality,
                app.notes or "",
                created_at,
                updated_at,
            ]
            writer.writerow([_sanitize_csv_value(str(v)) for v in row])

        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = (
            f"attachment; filename=applications_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )

        return response

    except Exception as e:
        current_app.logger.error(f"Error exporting CSV: {str(e)}")
        flash("Error exporting applications. Please try again.", "error")
        return redirect(url_for("unified_applications.application_list"))


# ---------------------------------------------------------------------------
# 3. analyze_import  (POST /analyze-import)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/analyze-import", methods=["POST"])
@login_required
@audit_log("analyze_import")
@import_rate_limit(max_calls_per_minute=5, max_calls_per_hour=30, max_calls_per_day=150)
def analyze_import():
    """Preview AI analysis results BEFORE committing import."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    confidence_threshold = float(request.form.get("confidence_threshold", "0.7"))
    # Analyze ALL applications - no artificial limit
    max_preview = int(
        request.form.get("max_preview", "999999")
    )  # Effectively unlimited
    archimate_mode = request.form.get(
        "archimate_mode", "standard"
    )  # quick, standard, or comprehensive

    try:
        filename = (file.filename or "").lower()
        applications_data = []

        current_app.logger.info(f"[ANALYZE-IMPORT] Processing file: {filename}")

        # Parse file based on type
        if filename.endswith((".xlsx", ".xls")):
            from openpyxl import load_workbook

            wb = load_workbook(file, data_only=True)
            ws = wb.active
            headers = [
                str(cell.value).strip() if cell.value else ""
                for cell in ws[1]
                if cell.value
            ]
            current_app.logger.info(f"[ANALYZE-IMPORT] Excel headers: {headers}")

            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row if cell):
                    row_data = {}
                    for i, header in enumerate(headers):
                        if i < len(row) and row[i]:
                            row_data[header.lower().replace(" ", "_")] = str(
                                row[i]
                            ).strip()

                    # Flexible name matching - check multiple possible name columns
                    app_name = (
                        row_data.get("name")
                        or row_data.get("application_name")
                        or row_data.get("app_name")
                        or row_data.get("application")
                        or row_data.get("system_name")
                    )

                    if app_name:
                        # Ensure 'name' key exists for downstream processing
                        row_data["name"] = app_name
                        applications_data.append(row_data)
                        current_app.logger.info(
                            f"[ANALYZE-IMPORT] Added application: {app_name}"
                        )
                    else:
                        current_app.logger.warning(
                            f"[ANALYZE-IMPORT] Row {row_count} has no name field. Keys: {list(row_data.keys())}"
                        )
                    row_count += 1

        elif filename.endswith(".csv"):
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)

            current_app.logger.info(
                f"[ANALYZE-IMPORT] CSV headers: {reader.fieldnames}"
            )

            row_count = 0
            for row in reader:
                # Strip spaces from keys AND values, then normalize keys
                row_data = {
                    k.strip().lower().replace(" ", "_"): v.strip()
                    for k, v in row.items()
                    if v and v.strip()
                }

                # Flexible name matching - check multiple possible name columns
                app_name = (
                    row_data.get("name")
                    or row_data.get("application_name")
                    or row_data.get("app_name")
                    or row_data.get("application")
                    or row_data.get("system_name")
                )

                if app_name:
                    # Ensure 'name' key exists for downstream processing
                    row_data["name"] = app_name
                    applications_data.append(row_data)
                    current_app.logger.info(
                        f"[ANALYZE-IMPORT] Added application: {app_name}"
                    )
                else:
                    current_app.logger.warning(
                        f"[ANALYZE-IMPORT] Row {row_count} has no name field. Keys: {list(row_data.keys())}"
                    )
                row_count += 1

        elif filename.endswith(".json"):
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list):
                applications_data = data

        current_app.logger.info(
            f"[ANALYZE-IMPORT] Parsed {len(applications_data)} applications from file"
        )

        if len(applications_data) == 0:
            current_app.logger.warning(
                f"[ANALYZE-IMPORT] No applications found in file. Filename: {filename}"
            )
            return (
                jsonify(
                    {
                        "success": True,
                        "total_rows": 0,
                        "will_create": 0,
                        "will_update": 0,
                        "duplicates_in_file": 0,
                        "comprehensive_ai_results": {
                            "total_analyzed": 0,
                            "capability_mappings_found": 0,
                            "process_mappings_found": 0,
                            "archimate_elements_generated": 0,
                            "high_confidence_mappings": 0,
                            "processing_stats": {"avg_processing_time_ms": 0},
                        },
                        "applications": [],
                        "warning": "No applications found in file. Please check that your file has a 'Name' column with data.",
                    }
                ),
                200,
            )

        # Call AI preview service
        from app.services.ai_import_service import get_ai_import_service

        ai_service = get_ai_import_service()

        preview = ai_service.analyze_file_data_for_preview(
            applications_data=applications_data[:max_preview],
            confidence_threshold=confidence_threshold,
            archimate_mode=archimate_mode,
        )

        # Check if we're only analyzing a subset
        is_preview_limited = len(applications_data) > max_preview
        analyzed_count = min(len(applications_data), max_preview)

        # Format response to match frontend expectations
        response_data = {
            "success": True,
            "total_rows": len(applications_data),
            "analyzed_count": analyzed_count,
            "is_preview_limited": is_preview_limited,
            "will_create": len(
                applications_data
            ),  # Simplified - actual logic would check duplicates
            "will_update": 0,
            "duplicates_in_file": 0,
            "comprehensive_ai_results": {
                "total_analyzed": preview.get("total_analyzed", analyzed_count),
                "capability_mappings_found": preview.get(
                    "capability_mappings_found", 0
                ),
                "process_mappings_found": preview.get("process_mappings_found", 0),
                "archimate_elements_generated": preview.get(
                    "archimate_elements_generated", 0
                ),
                "vendor_matches": preview.get("vendor_matches", 0),
                "high_confidence_mappings": preview.get("high_confidence_count", 0),
                "processing_stats": {
                    "avg_processing_time_ms": preview.get("avg_processing_time_ms", 0)
                },
            },
            "applications": preview.get("applications", []),
        }

        # Add warning if only analyzing subset
        if is_preview_limited:
            response_data["warning"] = (
                f"Preview mode: Analyzed first {analyzed_count} of {len(applications_data)} applications. Full import will process all {len(applications_data)} apps."
            )

        # Add rate limit headers
        response = jsonify(response_data)
        add_rate_limit_headers(response)
        return response, 200

    except Exception as e:
        current_app.logger.error(f"Error analyzing import: {e}")
        return jsonify({"error": "An internal error occurred"}), 500


# ---------------------------------------------------------------------------
# 4. preview_excel  (POST /preview-excel)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/preview-excel", methods=["POST"])
@login_required
@audit_log("preview_excel_import")
def preview_excel():
    """Preview Excel, CSV, or JSON file without importing"""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = (file.filename or "").lower()

    try:
        headers = []
        rows = []
        total_rows = 0

        if filename.endswith((".xlsx", ".xls")):
            from openpyxl import load_workbook

            wb = load_workbook(file, data_only=True)
            ws = wb.active
            headers = [
                str(cell.value).strip() if cell.value else ""
                for cell in ws[1]
                if cell.value
            ]
            for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
                if any(cell for cell in row if cell):
                    rows.append(
                        [
                            str(cell).strip() if cell else ""
                            for cell in row[: len(headers)]
                        ]
                    )
            total_rows = ws.max_row - 1

        elif filename.endswith(".csv"):
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.reader(stream)
            all_rows = list(reader)
            if all_rows:
                headers = [str(h).strip() for h in all_rows[0] if h]
                for row in all_rows[1:11]:
                    if any(cell for cell in row if cell):
                        rows.append(
                            [
                                str(cell).strip() if cell else ""
                                for cell in row[: len(headers)]
                            ]
                        )
                total_rows = len(all_rows) - 1

        elif filename.endswith(".json"):
            data = json.load(file)
            if isinstance(data, dict):
                data = data.get("applications", data.get("data", [data]))
            if isinstance(data, list) and len(data) > 0:
                first_item = data[0] if isinstance(data[0], dict) else {}
                headers = list(first_item.keys())
                for item in data[:10]:
                    if isinstance(item, dict):
                        rows.append([str(item.get(h, "")).strip() for h in headers])
                total_rows = len(data)
            else:
                return jsonify({"error": "Invalid JSON structure"}), 400
        else:
            return (
                jsonify(
                    {
                        "error": "Unsupported file format. Use .xlsx, .xls, .csv, or .json"
                    }
                ),
                400,
            )

        return jsonify(
            {"headers": headers, "rows": rows, "total_rows": total_rows}
        ), 200

    except Exception as e:
        current_app.logger.error(f"Error previewing file: {e}")
        return jsonify({"error": "An internal error occurred"}), 500


# ---------------------------------------------------------------------------
# 5. rollback_import  (POST /rollback-import/<int:history_id>)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/rollback-import/<int:history_id>", methods=["POST"])
@login_required
@audit_log("rollback_import")
@import_rate_limit(max_calls_per_minute=10, max_calls_per_hour=50, max_calls_per_day=200)
def rollback_import(history_id):
    """Rollback an import batch - delete all applications and mappings created in that batch.

    Optimized: Uses bulk queries and deletes to avoid N + 1 query patterns.
    """
    try:
        from app.models.application_import_history import ApplicationImportHistory
        from app.models.application_portfolio import ApplicationComponent
        from app.models.apqc_process import ProcessApplicationMapping
        from app.models.archimate_core import ArchiMateElement
        from app.models.unified_application_capability_mapping import (
            UnifiedApplicationCapabilityMapping,
        )

        history = ApplicationImportHistory.query.get(history_id)
        if not history:
            return jsonify({"success": False, "error": "Import history not found"}), 404

        # Ownership check: only the importer or admin can rollback
        is_admin = getattr(current_user, "is_admin", False)
        if history.imported_by_id is None:
            # Legacy imports have no owner - only admins can rollback
            if not is_admin:
                return jsonify(
                    {
                        "success": False,
                        "error": "Only administrators can rollback legacy imports",
                    }
                ), 403
        elif history.imported_by_id != current_user.id and not is_admin:
            return jsonify(
                {"success": False, "error": "Not authorized to rollback this import"}
            ), 403

        # Parse import settings to get application IDs
        import json

        settings = (
            json.loads(history.import_settings) if history.import_settings else {}
        )
        app_ids = settings.get("application_ids", [])

        if not app_ids:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No application IDs found in import history",
                    }
                ),
                400,
            )

        # Delete in reverse order: mappings first, then applications
        deleted_counts = {
            "capability_mappings": 0,
            "process_mappings": 0,
            "archimate_elements": 0,
            "applications": 0,
        }

        # OPTIMIZATION: Bulk delete capability mappings (single query instead of N queries)
        cap_count = UnifiedApplicationCapabilityMapping.query.filter(
            UnifiedApplicationCapabilityMapping.application_id.in_(app_ids)
        ).delete(synchronize_session="fetch")
        deleted_counts["capability_mappings"] = cap_count

        # OPTIMIZATION: Bulk delete process mappings (single query instead of N queries)
        proc_count = ProcessApplicationMapping.query.filter(
            ProcessApplicationMapping.application_id.in_(app_ids)
        ).delete(synchronize_session="fetch")
        deleted_counts["process_mappings"] = proc_count

        # Fetch all applications in batch to get their ArchiMate element IDs
        apps = ApplicationComponent.query.filter(
            ApplicationComponent.id.in_(app_ids)
        ).all()

        # AUDIT-IMP-006: Capture application names before deletion for audit trail
        rolled_back_app_names = [app.name for app in apps]

        # Collect ArchiMate element IDs to delete
        archimate_ids = [
            app.archimate_element_id for app in apps if app.archimate_element_id
        ]

        # OPTIMIZATION: Bulk delete ArchiMate elements
        if archimate_ids:
            elem_count = ArchiMateElement.query.filter(
                ArchiMateElement.id.in_(archimate_ids)
            ).delete(synchronize_session="fetch")
            deleted_counts["archimate_elements"] = elem_count

        # OPTIMIZATION: Bulk delete applications
        app_count = ApplicationComponent.query.filter(
            ApplicationComponent.id.in_(app_ids)
        ).delete(synchronize_session="fetch")
        deleted_counts["applications"] = app_count

        # Mark history as rolled back
        history.status = "rolled_back"

        db.session.commit()

        # AUDIT-IMP-006: Create audit log entry for the rollback action.
        # This provides accountability for destructive rollback operations.
        try:
            from datetime import datetime
            from app.models.batch_import import ImportAuditLog

            rollback_audit = ImportAuditLog(
                user_id=current_user.id,
                user_email=getattr(current_user, "email", None),  # model-safety-ok
                import_type="rollback",
                filename=history.file_name or f"import_batch_{history_id}",
                records_created=0,
                records_updated=0,
                records_skipped=0,
                records_failed=0,
                duplicate_mode="rollback",
                changes=[
                    {
                        "action": "rollback",
                        "import_history_id": history_id,
                        "applications_deleted": rolled_back_app_names[:500],
                        "deleted_counts": deleted_counts,
                        "original_importer": history.imported_by_name,
                        "original_import_date": history.imported_at.isoformat()
                        if history.imported_at
                        else None,
                    }
                ],
                errors=None,
            )
            db.session.add(rollback_audit)
            db.session.commit()
        except Exception as audit_err:
            current_app.logger.warning(
                f"Failed to write rollback audit log: {audit_err}"
            )

        return (
            jsonify(
                {
                    "success": True,
                    "message": f"Successfully rolled back import batch {history_id}",
                    "deleted": deleted_counts,
                }
            ),
            200,
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Rollback error: {e}")
        return jsonify({"success": False, "error": "An internal error occurred"}), 500


# ---------------------------------------------------------------------------
# 6. application_import_page  (GET /import)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/import", methods=["GET"])
@login_required
def application_import_page():
    """Redirect to application list where import modal is available"""
    flash(
        "Use the Import button on the Applications page to import applications.", "info"
    )
    return redirect(url_for("unified_applications.application_list"))


# ---------------------------------------------------------------------------
# 7. application_import  (POST /import)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/import", methods=["POST"])
@login_required
@audit_log("application_import")
@import_rate_limit(max_calls_per_minute=5, max_calls_per_hour=30, max_calls_per_day=150)
def application_import():
    """Import Application Components from CSV or JSON with AI-powered analysis"""
    if "file" not in request.files:
        flash("No file uploaded", "error")
        return redirect(url_for("unified_applications.application_list"))

    file = request.files["file"]
    original_filename = file.filename  # Store for audit logging

    if file.filename == "":
        flash("No file selected", "error")
        return redirect(url_for("unified_applications.application_list"))

    # AUDIT-IMP-004: Validate file extension to prevent uploading of unsafe file types
    import os

    allowed_extensions = {".csv", ".xlsx", ".xls"}
    file_ext = os.path.splitext(file.filename or "")[1].lower()
    if file_ext not in allowed_extensions:
        # IMP-002: Audit failed upload
        AuditLog.log_file_upload(
            user_id=current_user.id,
            filename=original_filename,
            ip_address=request.remote_addr,
            route=request.endpoint,
            status="failed",
            error_message=f"Invalid file extension: {file_ext}"
        )
        flash(
            f"Invalid file type '{file_ext}'. Only CSV and Excel files (.csv, .xlsx, .xls) are allowed.",
            "error",
        )
        return redirect(url_for("unified_applications.application_list"))

    # IMP-001: Validate MIME type by magic numbers (file header), not just extension
    try:
        mime_type = validate_mime_type(file, file.filename)
    except InvalidFileTypeError as e:
        # IMP-002: Audit failed upload
        AuditLog.log_file_upload(
            user_id=current_user.id,
            filename=original_filename,
            ip_address=request.remote_addr,
            route=request.endpoint,
            status="failed",
            error_message=f"Invalid MIME type: {str(e)}"
        )
        flash(
            f"Invalid file type. Allowed: {get_allowed_extensions_display()}",
            "error",
        )
        return redirect(url_for("unified_applications.application_list"))

    # AUDIT-IMP-005: Content-hash idempotency check to prevent double-click race condition.
    # Read file content once into memory so we can hash it for dedup and still parse it later.
    from app.services.unified_import.import_orchestrator import (
        check_import_idempotency,
        store_import_idempotency,
    )

    _file_content_bytes = file.read()
    file.seek(0)  # Reset stream so downstream parsers can re-read

    _cached_result = check_import_idempotency(_file_content_bytes, current_user.id)
    if _cached_result is not None:
        flash(
            "This file was already imported within the last 5 minutes. "
            "Returning previous result.",
            "warning",
        )
        return redirect(url_for("unified_applications.application_list"))

    # Legacy DB-based idempotency check (kept for backward compatibility with import_token)
    import_token = request.form.get("import_token")
    if import_token:
        from datetime import datetime, timedelta
        from app.models.batch_import import ImportAuditLog

        idempotency_window = 300  # 5 minutes
        recent_dup = ImportAuditLog.query.filter(
            ImportAuditLog.user_id == current_user.id,
            ImportAuditLog.filename == (file.filename or ""),
            ImportAuditLog.timestamp
            >= datetime.utcnow() - timedelta(seconds=idempotency_window),
        ).first()
        if recent_dup:
            flash(
                "This file was already imported within the last 5 minutes. Please wait and try again.",
                "warning",
            )
            return redirect(url_for("unified_applications.application_list"))

    # Get import mode: 'skip', 'update', or 'duplicate' (default to skip for safety)
    import_mode = request.form.get("import_mode", "skip")

    # Get AI import options
    enable_ai = request.form.get("enable_ai", "false").lower() == "true"
    map_capabilities = request.form.get("map_capabilities", "true").lower() == "true"
    map_processes = request.form.get("map_processes", "true").lower() == "true"
    generate_archimate = (
        request.form.get("generate_archimate", "false").lower() == "true"
    )
    clone_vendor = request.form.get("clone_vendor_archimate", "false").lower() == "true"
    confidence_threshold = float(request.form.get("confidence_threshold", "0.7"))

    # Quick import mode: skip AI to prevent timeout (for E2E tests and fast imports)
    quick_mode = request.form.get("quick_mode", "false").lower() == "true"
    if quick_mode:
        current_app.logger.info(
            "Quick mode enabled - skipping AI analysis to prevent timeout"
        )
        enable_ai = False

    try:
        filename = (file.filename or "").lower()

        def _clean(value):
            if isinstance(value, str):
                value = value.strip()
                return value or None
            return value

        def _parse_user_count(raw_value):
            if raw_value in (None, ""):
                return None

            if isinstance(raw_value, bool):
                return int(raw_value)

            if isinstance(raw_value, (int, float)):
                try:
                    return int(raw_value)
                except (TypeError, ValueError):
                    return None

            cleaned = _clean(raw_value)
            if cleaned is None:
                return None

            sanitized = cleaned.replace(",", "")
            if sanitized.isdigit():
                return int(sanitized)

            matches = re.findall(r"\d+", sanitized)
            if len(matches) == 1:
                return int(matches[0])

            return None

        if filename.endswith(".csv"):
            # Parse CSV
            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)

            # Skip leading blank lines that break DictReader header detection
            while True:
                pos = stream.tell()
                line = stream.readline()
                if not line:
                    stream.seek(0)
                    break
                if line.strip():
                    stream.seek(pos)
                    break

            reader = csv.DictReader(stream)

            if not reader.fieldnames:
                raise ValueError("CSV file is missing headers.")

            # Find the name column (flexible matching)
            name_column = None
            name_variants = [
                "name",
                "application name",
                "app name",
                "application_name",
                "app_name",
                "appname",
            ]
            for field in reader.fieldnames:
                if field and field.lower().strip() in name_variants:
                    name_column = field
                    break

            if not name_column:
                raise ValueError(
                    "CSV file must include a 'Name' column (or 'Application Name', 'App Name')."
                )

            # Helper to find column value with flexible matching
            def get_csv_field(row, *variants):
                """Get field value trying multiple column name variants"""
                for variant in variants:
                    for key in row.keys():
                        if key and key.lower().strip() == variant.lower():
                            return row[key]
                        if key and key.lower().replace(
                            " ", "_"
                        ) == variant.lower().replace(" ", "_"):
                            return row[key]
                        if key and key.lower().replace(
                            "_", " "
                        ) == variant.lower().replace("_", " "):
                            return row[key]
                return None

            count = 0
            skipped = 0
            updated = 0
            non_numeric_user_counts = []
            # Track names processed in THIS import to detect duplicates within the same CSV file
            processed_names = set()

            # Batch prefetch all existing application names for case-insensitive lookup
            _all_existing_apps = ApplicationComponent.query.limit(
                10000
            ).all()  # Limit to prevent OOM on large datasets
            _existing_app_by_name = {
                a.name.lower(): a for a in _all_existing_apps if a.name
            }

            for index, row in enumerate(
                reader, start=2
            ):  # start=2 accounts for header row
                if not any(
                    (value or "").strip() for value in row.values() if value is not None
                ):
                    continue  # skip empty rows

                name = _clean(row.get(name_column))
                if not name:
                    raise ValueError(f"Row {index}: 'Name' is required.")

                # Normalize name for duplicate checking (case-insensitive)
                name_lower = name.lower()

                # Check if this name was already processed in THIS import (within same CSV)
                if name_lower in processed_names:
                    if import_mode == "skip":
                        skipped += 1
                        continue
                    elif import_mode != "duplicate":
                        # For 'update' mode, we already updated it earlier in this import
                        skipped += 1
                        continue
                    # For 'duplicate' mode, allow it to continue and create duplicate

                # Check if application exists (case-insensitive match, using prefetched data)
                existing_app = _existing_app_by_name.get(name_lower)

                if existing_app and import_mode == "skip":
                    skipped += 1
                    processed_names.add(name_lower)  # Track that we've seen this name
                    continue
                elif existing_app and import_mode == "update":
                    # Update existing record with flexible column matching
                    existing_app.component_type = _clean(
                        get_csv_field(
                            row,
                            "Type",
                            "component_type",
                            "Component Type",
                            "App Type",
                            "Application Type",
                        )
                    )
                    existing_app.application_category = _clean(
                        get_csv_field(
                            row, "Category", "application_category", "App Category"
                        )
                    )
                    existing_app.technology_stack = _clean(
                        get_csv_field(
                            row,
                            "Technology Stack",
                            "technology_stack",
                            "Technology",
                            "Tech Stack",
                            "Stack",
                        )
                    )
                    existing_app.version = _clean(get_csv_field(row, "Version"))
                    existing_app.deployment_status = (
                        _clean(
                            get_csv_field(
                                row,
                                "Status",
                                "deployment_status",
                                "Deployment Status",
                                "App Status",
                                "Application Status",
                            )
                        )
                        or "planned"
                    )
                    existing_app.business_domain = _clean(
                        get_csv_field(
                            row, "Business Domain", "business_domain", "Domain"
                        )
                    )
                    existing_app.business_owner = _clean(
                        get_csv_field(
                            row,
                            "Owner",
                            "business_owner",
                            "Business Owner",
                            "App Owner",
                        )
                    )
                    existing_app.development_team = _clean(
                        get_csv_field(
                            row,
                            "Team",
                            "development_team",
                            "Development Team",
                            "Dev Team",
                        )
                    )
                    existing_app.user_base_size = _parse_user_count(
                        get_csv_field(
                            row, "Users", "user_count", "User Count", "Number of Users"
                        )
                    )
                    existing_app.business_criticality = _clean(
                        get_csv_field(
                            row,
                            "Criticality",
                            "business_criticality",
                            "Business Criticality",
                        )
                    )
                    existing_app.description = _clean(
                        get_csv_field(row, "Description", "App Description")
                    )
                    existing_app.lifecycle_status = _clean(
                        get_csv_field(
                            row, "Lifecycle Status", "lifecycle_status", "Lifecycle"
                        )
                    )

                    # Add import fields for auto-mapping functionality
                    existing_app.imported_capabilities = _clean(
                        get_csv_field(
                            row,
                            "Capabilities",
                            "imported_capabilities",
                            "Business Capabilities",
                            "Capability List",
                        )
                    )
                    existing_app.application_services = _clean(
                        get_csv_field(
                            row,
                            "Services",
                            "application_services",
                            "Application Services",
                            "Service List",
                        )
                    )
                    existing_app.application_functions_text = _clean(
                        get_csv_field(
                            row,
                            "Functions",
                            "application_functions_text",
                            "Application Functions",
                            "Function List",
                        )
                    )
                    existing_app.imported_apqc_codes = _clean(
                        get_csv_field(
                            row,
                            "APQC Codes",
                            "imported_apqc_codes",
                            "APQC",
                            "Process Codes",
                        )
                    )

                    updated += 1
                    processed_names.add(name_lower)  # Track that we've seen this name
                else:
                    # Create new record (duplicate mode or doesn't exist) with flexible column matching
                    app = ApplicationComponent(
                        name=name,
                        component_type=_clean(
                            get_csv_field(
                                row,
                                "Type",
                                "component_type",
                                "Component Type",
                                "App Type",
                                "Application Type",
                            )
                        ),
                        application_category=_clean(
                            get_csv_field(
                                row, "Category", "application_category", "App Category"
                            )
                        ),
                        technology_stack=_clean(
                            get_csv_field(
                                row,
                                "Technology Stack",
                                "technology_stack",
                                "Technology",
                                "Tech Stack",
                                "Stack",
                            )
                        ),
                        version=_clean(get_csv_field(row, "Version")),
                        deployment_status=_clean(
                            get_csv_field(
                                row,
                                "Status",
                                "deployment_status",
                                "Deployment Status",
                                "App Status",
                                "Application Status",
                            )
                        )
                        or "planned",
                        business_domain=_clean(
                            get_csv_field(
                                row, "Business Domain", "business_domain", "Domain"
                            )
                        ),
                        business_owner=_clean(
                            get_csv_field(
                                row,
                                "Owner",
                                "business_owner",
                                "Business Owner",
                                "App Owner",
                            )
                        ),
                        development_team=_clean(
                            get_csv_field(
                                row,
                                "Team",
                                "development_team",
                                "Development Team",
                                "Dev Team",
                            )
                        ),
                        user_base_size=_parse_user_count(
                            get_csv_field(
                                row,
                                "Users",
                                "user_count",
                                "User Count",
                                "Number of Users",
                            )
                        ),
                        business_criticality=_clean(
                            get_csv_field(
                                row,
                                "Criticality",
                                "business_criticality",
                                "Business Criticality",
                            )
                        ),
                        description=_clean(
                            get_csv_field(row, "Description", "App Description")
                        ),
                        lifecycle_status=_clean(
                            get_csv_field(
                                row, "Lifecycle Status", "lifecycle_status", "Lifecycle"
                            )
                        ),
                        # Add import fields for auto-mapping functionality
                        imported_capabilities=_clean(
                            get_csv_field(
                                row,
                                "Capabilities",
                                "imported_capabilities",
                                "Business Capabilities",
                                "Capability List",
                            )
                        ),
                        application_services=_clean(
                            get_csv_field(
                                row,
                                "Services",
                                "application_services",
                                "Application Services",
                                "Service List",
                            )
                        ),
                        application_functions_text=_clean(
                            get_csv_field(
                                row,
                                "Functions",
                                "application_functions_text",
                                "Application Functions",
                                "Function List",
                            )
                        ),
                        imported_apqc_codes=_clean(
                            get_csv_field(
                                row,
                                "APQC Codes",
                                "imported_apqc_codes",
                                "APQC",
                                "Process Codes",
                            )
                        ),
                    )
                    db.session.add(app)
                    count += 1
                    processed_names.add(name_lower)  # Track that we've seen this name

            db.session.commit()

            # AI-POWERED IMPORT: Run intelligent analysis and mapping if enabled
            ai_results = {
                "total_capabilities_mapped": 0,
                "total_processes_mapped": 0,
                "total_archimate_created": 0,
                "vendor_matches": 0,
            }

            if enable_ai and (count > 0 or updated > 0):
                try:
                    from app.services.ai_import_service import get_ai_import_service

                    ai_service = get_ai_import_service()

                    # Get recently imported/updated applications
                    recent_apps = (
                        ApplicationComponent.query.order_by(
                            ApplicationComponent.id.desc()
                        )
                        .limit(count + updated)
                        .all()
                    )

                    current_app.logger.info(
                        f"Running AI analysis on {len(recent_apps)} applications"
                    )

                    # Batch prefetch existing capability and process mappings for all recent apps
                    _recent_app_ids = [a.id for a in recent_apps]
                    _existing_cap_mappings = set()
                    _existing_proc_mappings = set()
                    if _recent_app_ids:
                        if map_capabilities:
                            from app.models.unified_application_capability_mapping import (
                                UnifiedApplicationCapabilityMapping,
                            )

                            _cap_rows = UnifiedApplicationCapabilityMapping.query.filter(
                                UnifiedApplicationCapabilityMapping.application_id.in_(
                                    _recent_app_ids
                                )
                            ).all()
                            _existing_cap_mappings = {
                                (r.application_id, r.capability_id) for r in _cap_rows
                            }
                        if map_processes:
                            from app.models.apqc_process import (
                                ProcessApplicationMapping,
                            )

                            _proc_rows = ProcessApplicationMapping.query.filter(
                                ProcessApplicationMapping.application_id.in_(
                                    _recent_app_ids
                                )
                            ).all()
                            _existing_proc_mappings = {
                                (r.application_id, r.process_id) for r in _proc_rows
                            }

                    for app in recent_apps:
                        try:
                            # Run comprehensive AI analysis
                            ai_result = ai_service.analyze_application_for_ai_mapping(
                                app.id
                            )

                            # Create capability mappings for high-confidence matches
                            if map_capabilities and ai_result.capability_mappings:
                                from app.models.unified_application_capability_mapping import (
                                    UnifiedApplicationCapabilityMapping,
                                )

                                for cap_mapping in ai_result.capability_mappings:
                                    if (
                                        cap_mapping.get("confidence_score", 0)
                                        >= confidence_threshold
                                    ):
                                        _cap_key = (
                                            app.id,
                                            cap_mapping["capability_id"],
                                        )
                                        if _cap_key not in _existing_cap_mappings:
                                            db.session.add(
                                                UnifiedApplicationCapabilityMapping(
                                                    application_id=app.id,
                                                    capability_id=cap_mapping[
                                                        "capability_id"
                                                    ],
                                                    confidence_score=cap_mapping.get(
                                                        "confidence_score", 0.5
                                                    ),
                                                    mapping_method="ai_import_integrated",
                                                    rationale=cap_mapping.get(
                                                        "rationale", ""
                                                    ),
                                                    created_by=current_user.email
                                                    if current_user.is_authenticated
                                                    else "ai_import",
                                                )
                                            )
                                            _existing_cap_mappings.add(_cap_key)
                                            ai_results["total_capabilities_mapped"] += 1

                            # Create APQC process mappings for high-confidence matches
                            if map_processes and ai_result.process_mappings:
                                from app.models.apqc_process import (
                                    ProcessApplicationMapping,
                                )

                                for proc_mapping in ai_result.process_mappings:
                                    if (
                                        proc_mapping.get("similarity_score", 0)
                                        >= confidence_threshold
                                    ):
                                        _proc_key = (app.id, proc_mapping["process_id"])
                                        if _proc_key not in _existing_proc_mappings:
                                            db.session.add(
                                                ProcessApplicationMapping(
                                                    application_id=app.id,
                                                    process_id=proc_mapping[
                                                        "process_id"
                                                    ],
                                                    confidence_score=proc_mapping.get(
                                                        "similarity_score", 0.5
                                                    ),
                                                    mapping_method="ai_import_integrated",
                                                    created_by=current_user.email
                                                    if current_user.is_authenticated
                                                    else "ai_import",
                                                )
                                            )
                                            _existing_proc_mappings.add(_proc_key)
                                            ai_results["total_processes_mapped"] += 1

                            # Generate ArchiMate elements if requested
                            if generate_archimate and ai_result.archimate_elements:
                                try:
                                    archimate_service = (
                                        ai_service._get_archimate_service()
                                    )
                                    for elem_data in ai_result.archimate_elements:
                                        elem = (
                                            archimate_service.create_element_from_dict(
                                                elem_data,
                                                created_by=current_user.email
                                                if current_user.is_authenticated
                                                else "ai_import",
                                            )
                                        )
                                        if elem:
                                            ai_results["total_archimate_created"] += 1
                                            # Link first ApplicationComponent element to app
                                            if (
                                                elem_data.get("type")
                                                == "ApplicationComponent"
                                                and not app.archimate_element_id
                                            ):
                                                app.archimate_element_id = elem.id
                                except Exception as e:
                                    current_app.logger.warning(
                                        f"ArchiMate generation error for {app.name}: {e}"
                                    )

                            # Track vendor matches
                            if ai_result.vendor_product_identified:
                                ai_results["vendor_matches"] += 1

                        except Exception as e:
                            current_app.logger.error(
                                f"AI analysis error for app {app.name}: {e}"
                            )
                            continue

                    db.session.commit()
                    current_app.logger.info(f"AI import complete: {ai_results}")

                except Exception as e:
                    current_app.logger.error(f"AI import service error: {e}")
                    db.session.rollback()
                    flash(
                        "Import succeeded but AI analysis encountered an error.",
                        "warning",
                    )

            # Build success message
            msg_parts = []
            if count > 0:
                msg_parts.append(f"{count} created")
            if updated > 0:
                msg_parts.append(f"{updated} updated")
            if skipped > 0:
                msg_parts.append(f"{skipped} skipped")

            flash(f"Import complete: {', '.join(msg_parts)}", "success")

            # Add AI results to flash message if AI was enabled
            if enable_ai and (
                ai_results["total_capabilities_mapped"] > 0
                or ai_results["total_processes_mapped"] > 0
            ):
                ai_msg = f"AI Analysis: {ai_results['total_capabilities_mapped']} capabilities, {ai_results['total_processes_mapped']} processes"
                if ai_results["total_archimate_created"] > 0:
                    ai_msg += (
                        f", {ai_results['total_archimate_created']} ArchiMate elements"
                    )
                if ai_results["vendor_matches"] > 0:
                    ai_msg += f", {ai_results['vendor_matches']} vendor matches"
                flash(ai_msg, "success")

            if len(non_numeric_user_counts) > 0:
                preview = ", ".join(non_numeric_user_counts[:5])
                if len(non_numeric_user_counts) > 5:
                    preview += ", ..."
                flash(
                    "Some user counts were non-numeric and were left blank: " + preview,
                    "warning",
                )

        elif filename.endswith(".json"):
            # Parse JSON
            data = json.load(file.stream)

            if not isinstance(data, list):
                raise ValueError("JSON payload must be a list of application objects.")

            count = 0
            skipped = 0
            updated = 0
            # Track names processed in THIS import to detect duplicates within the same JSON file
            processed_names = set()

            # Batch prefetch all existing application names for case-insensitive lookup
            _all_existing_apps_json = ApplicationComponent.query.limit(
                10000
            ).all()  # Limit to prevent OOM on large datasets
            _existing_app_by_name_json = {
                a.name.lower(): a for a in _all_existing_apps_json if a.name
            }

            for index, item in enumerate(data, start=1):
                name = _clean(item.get("name"))
                if not name:
                    raise ValueError(f"Record {index}: 'name' is required.")

                # Normalize name for duplicate checking (case-insensitive)
                name_lower = name.lower()

                # Check if this name was already processed in THIS import (within same JSON)
                if name_lower in processed_names:
                    if import_mode == "skip":
                        skipped += 1
                        continue
                    elif import_mode != "duplicate":
                        # For 'update' mode, we already updated it earlier in this import
                        skipped += 1
                        continue
                    # For 'duplicate' mode, allow it to continue and create duplicate

                # Check if application exists (case-insensitive match, using prefetched data)
                existing_app = _existing_app_by_name_json.get(name_lower)

                if existing_app and import_mode == "skip":
                    skipped += 1
                    processed_names.add(name_lower)  # Track that we've seen this name
                    continue
                elif existing_app and import_mode == "update":
                    # Update existing record
                    existing_app.component_type = _clean(item.get("component_type"))
                    existing_app.application_category = _clean(
                        item.get("application_category")
                    )
                    existing_app.technology_stack = _clean(item.get("technology_stack"))
                    existing_app.version = _clean(item.get("version"))
                    existing_app.deployment_status = (
                        _clean(item.get("deployment_status")) or "planned"
                    )
                    existing_app.business_domain = _clean(item.get("business_domain"))
                    existing_app.business_owner = _clean(item.get("business_owner"))
                    existing_app.development_team = _clean(item.get("development_team"))
                    existing_app.user_base_size = _parse_user_count(
                        item.get("user_count")
                    )
                    existing_app.business_criticality = _clean(
                        item.get("business_criticality")
                    )

                    # Add import fields for auto-mapping functionality
                    existing_app.imported_capabilities = _clean(
                        item.get("imported_capabilities")
                    )
                    existing_app.application_services = _clean(
                        item.get("application_services")
                    )
                    existing_app.application_functions_text = _clean(
                        item.get("application_functions_text")
                    )
                    existing_app.imported_apqc_codes = _clean(
                        item.get("imported_apqc_codes")
                    )

                    updated += 1
                    processed_names.add(name_lower)  # Track that we've seen this name
                else:
                    # Create new record (duplicate mode or doesn't exist)
                    app = ApplicationComponent(
                        name=name,
                        component_type=_clean(item.get("component_type")),
                        application_category=_clean(item.get("application_category")),
                        technology_stack=_clean(item.get("technology_stack")),
                        version=_clean(item.get("version")),
                        deployment_status=_clean(item.get("deployment_status"))
                        or "planned",
                        business_domain=_clean(item.get("business_domain")),
                        business_owner=_clean(item.get("business_owner")),
                        development_team=_clean(item.get("development_team")),
                        user_base_size=_parse_user_count(item.get("user_count")),
                        business_criticality=_clean(item.get("business_criticality")),
                        # Add import fields for auto-mapping functionality
                        imported_capabilities=_clean(item.get("imported_capabilities")),
                        application_services=_clean(item.get("application_services")),
                        application_functions_text=_clean(
                            item.get("application_functions_text")
                        ),
                        imported_apqc_codes=_clean(item.get("imported_apqc_codes")),
                    )
                    db.session.add(app)
                    count += 1
                    processed_names.add(name_lower)  # Track that we've seen this name

            db.session.commit()

            # AI-POWERED IMPORT: Run intelligent analysis and mapping if enabled (JSON path)
            ai_results = {
                "total_capabilities_mapped": 0,
                "total_processes_mapped": 0,
                "total_archimate_created": 0,
                "vendor_matches": 0,
            }

            if enable_ai and (count > 0 or updated > 0):
                try:
                    from app.services.ai_import_service import get_ai_import_service

                    ai_service = get_ai_import_service()

                    recent_apps = (
                        ApplicationComponent.query.order_by(
                            ApplicationComponent.id.desc()
                        )
                        .limit(count + updated)
                        .all()
                    )

                    current_app.logger.info(
                        f"Running AI analysis on {len(recent_apps)} applications (JSON)"
                    )

                    # Batch prefetch existing capability and process mappings for JSON AI analysis
                    _json_app_ids = [a.id for a in recent_apps]
                    _json_existing_cap_mappings = set()
                    _json_existing_proc_mappings = set()
                    if _json_app_ids:
                        if map_capabilities:
                            from app.models.unified_application_capability_mapping import (
                                UnifiedApplicationCapabilityMapping,
                            )

                            _json_cap_rows = UnifiedApplicationCapabilityMapping.query.filter(
                                UnifiedApplicationCapabilityMapping.application_id.in_(
                                    _json_app_ids
                                )
                            ).all()
                            _json_existing_cap_mappings = {
                                (r.application_id, r.capability_id)
                                for r in _json_cap_rows
                            }
                        if map_processes:
                            from app.models.apqc_process import (
                                ProcessApplicationMapping,
                            )

                            _json_proc_rows = ProcessApplicationMapping.query.filter(
                                ProcessApplicationMapping.application_id.in_(
                                    _json_app_ids
                                )
                            ).all()
                            _json_existing_proc_mappings = {
                                (r.application_id, r.process_id)
                                for r in _json_proc_rows
                            }

                    for app in recent_apps:
                        try:
                            ai_result = ai_service.analyze_application_for_ai_mapping(
                                app.id
                            )

                            if map_capabilities and ai_result.capability_mappings:
                                from app.models.unified_application_capability_mapping import (
                                    UnifiedApplicationCapabilityMapping,
                                )

                                for cap_mapping in ai_result.capability_mappings:
                                    if (
                                        cap_mapping.get("confidence_score", 0)
                                        >= confidence_threshold
                                    ):
                                        _json_cap_key = (
                                            app.id,
                                            cap_mapping["capability_id"],
                                        )
                                        if (
                                            _json_cap_key
                                            not in _json_existing_cap_mappings
                                        ):
                                            db.session.add(
                                                UnifiedApplicationCapabilityMapping(
                                                    application_id=app.id,
                                                    capability_id=cap_mapping[
                                                        "capability_id"
                                                    ],
                                                    confidence_score=cap_mapping.get(
                                                        "confidence_score", 0.5
                                                    ),
                                                    mapping_method="ai_import_integrated",
                                                    rationale=cap_mapping.get(
                                                        "rationale", ""
                                                    ),
                                                    created_by=current_user.email
                                                    if current_user.is_authenticated
                                                    else "ai_import",
                                                )
                                            )
                                            _json_existing_cap_mappings.add(
                                                _json_cap_key
                                            )
                                            ai_results["total_capabilities_mapped"] += 1

                            if map_processes and ai_result.process_mappings:
                                from app.models.apqc_process import (
                                    ProcessApplicationMapping,
                                )

                                for proc_mapping in ai_result.process_mappings:
                                    if (
                                        proc_mapping.get("similarity_score", 0)
                                        >= confidence_threshold
                                    ):
                                        _json_proc_key = (
                                            app.id,
                                            proc_mapping["process_id"],
                                        )
                                        if (
                                            _json_proc_key
                                            not in _json_existing_proc_mappings
                                        ):
                                            db.session.add(
                                                ProcessApplicationMapping(
                                                    application_id=app.id,
                                                    process_id=proc_mapping[
                                                        "process_id"
                                                    ],
                                                    confidence_score=proc_mapping.get(
                                                        "similarity_score", 0.5
                                                    ),
                                                    mapping_method="ai_import_integrated",
                                                    created_by=current_user.email
                                                    if current_user.is_authenticated
                                                    else "ai_import",
                                                )
                                            )
                                            _json_existing_proc_mappings.add(
                                                _json_proc_key
                                            )
                                            ai_results["total_processes_mapped"] += 1

                            if generate_archimate and ai_result.archimate_elements:
                                try:
                                    archimate_service = (
                                        ai_service._get_archimate_service()
                                    )
                                    for elem_data in ai_result.archimate_elements:
                                        elem = (
                                            archimate_service.create_element_from_dict(
                                                elem_data,
                                                created_by=current_user.email
                                                if current_user.is_authenticated
                                                else "ai_import",
                                            )
                                        )
                                        if elem:
                                            ai_results["total_archimate_created"] += 1
                                            if (
                                                elem_data.get("type")
                                                == "ApplicationComponent"
                                                and not app.archimate_element_id
                                            ):
                                                app.archimate_element_id = elem.id
                                except Exception as e:
                                    current_app.logger.warning(
                                        f"ArchiMate generation error for {app.name}: {e}"
                                    )

                            if ai_result.vendor_product_identified:
                                ai_results["vendor_matches"] += 1

                        except Exception as e:
                            current_app.logger.error(
                                f"AI analysis error for app {app.name}: {e}"
                            )
                            continue

                    db.session.commit()
                    current_app.logger.info(f"AI import complete (JSON): {ai_results}")

                except Exception as e:
                    current_app.logger.error(f"AI import service error (JSON): {e}")
                    db.session.rollback()
                    flash(
                        "Import succeeded but AI analysis encountered an error.",
                        "warning",
                    )

            # Build success message
            msg_parts = []
            if count > 0:
                msg_parts.append(f"{count} created")
            if updated > 0:
                msg_parts.append(f"{updated} updated")
            if skipped > 0:
                msg_parts.append(f"{skipped} skipped")

            flash(f"Import complete: {', '.join(msg_parts)}", "success")

            if enable_ai and (
                ai_results["total_capabilities_mapped"] > 0
                or ai_results["total_processes_mapped"] > 0
            ):
                ai_msg = f"AI Analysis: {ai_results['total_capabilities_mapped']} capabilities, {ai_results['total_processes_mapped']} processes"
                if ai_results["total_archimate_created"] > 0:
                    ai_msg += (
                        f", {ai_results['total_archimate_created']} ArchiMate elements"
                    )
                if ai_results["vendor_matches"] > 0:
                    ai_msg += f", {ai_results['vendor_matches']} vendor matches"
                flash(ai_msg, "success")

        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            # Parse Excel file
            try:
                from openpyxl import load_workbook
            except ImportError:
                flash(
                    "Excel support requires openpyxl library. Please install it with: pip install openpyxl",
                    "error",
                )
                return redirect(url_for("unified_applications.application_list"))

            # Load workbook from file stream
            workbook = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
            sheet = workbook.active

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else "")

            # Find the name column (flexible matching)
            name_column = None
            name_variants = [
                "name",
                "application name",
                "app name",
                "application_name",
                "app_name",
                "appname",
            ]
            for header in headers:
                if header:
                    # Strip whitespace and normalize
                    normalized = str(header).strip().lower()
                    if normalized in name_variants:
                        name_column = header
                        break
                    # Also check if it contains 'name' as a word
                    if (
                        normalized == "name"
                        or normalized.endswith(" name")
                        or normalized.startswith("name ")
                    ):
                        name_column = header
                        break

            if not name_column:
                # Log available headers for debugging
                current_app.logger.error(f"Excel headers found: {headers}")
                raise ValueError(
                    f"Excel file must include a 'Name' column. Found headers: {', '.join([h for h in headers if h][:10])}..."
                )

            count = 0
            skipped = 0
            updated = 0
            non_numeric_user_counts = []
            processed_names = set()

            # Batch prefetch all existing application names for case-insensitive lookup
            _all_existing_apps_xlsx = ApplicationComponent.query.limit(
                10000
            ).all()  # Limit to prevent OOM on large datasets
            _existing_app_by_name_xlsx = {
                a.name.lower(): a for a in _all_existing_apps_xlsx if a.name
            }

            # Process rows starting from row 2 (after header)
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                # Create dict from row values
                row_dict = dict(zip(headers, row))

                # Skip empty rows
                if not any(v for v in row if v is not None and str(v).strip()):
                    continue

                name = _clean(row_dict.get(name_column))
                if not name:
                    raise ValueError(f"Row {row_idx}: 'Name' is required.")

                name_lower = name.lower()

                # Check for duplicates within this import
                if name_lower in processed_names:
                    if import_mode == "skip":
                        skipped += 1
                        continue
                    elif import_mode != "duplicate":
                        # For 'update' mode, we already updated it earlier in this import
                        skipped += 1
                        continue

                # Check if application exists in database (using prefetched data)
                existing_app = _existing_app_by_name_xlsx.get(name_lower)

                # Map Excel column names to model fields with flexible matching
                def get_field(*field_variants):
                    """Get field value trying multiple column name variants"""
                    for variant in field_variants:
                        for header, value in row_dict.items():
                            if not header:
                                continue
                            h = str(header).strip().lower()
                            v = str(variant).strip().lower()
                            # Direct match
                            if h == v:
                                return value
                            # Match with spaces/underscores normalized
                            if h.replace(" ", "_") == v.replace(" ", "_"):
                                return value
                            if h.replace(" ", "") == v.replace("_", "").replace(
                                " ", ""
                            ):
                                return value
                    return None

                if existing_app and import_mode == "skip":
                    skipped += 1
                    processed_names.add(name_lower)
                    continue
                elif existing_app and import_mode == "update":
                    # Update existing record with your template column names
                    existing_app.description = _clean(get_field("Description"))
                    existing_app.application_category = _clean(
                        get_field("Category", "application_category")
                    )
                    existing_app.business_criticality = _clean(
                        get_field("Business Criticality", "business_criticality")
                    )
                    existing_app.deployment_status = (
                        _clean(
                            get_field(
                                "Application Status", "Status", "deployment_status"
                            )
                        )
                        or "planned"
                    )
                    existing_app.lifecycle_status = _clean(
                        get_field("Lifecycle Status", "lifecycle_status")
                    )
                    existing_app.business_owner = _clean(
                        get_field(
                            "App Business Owner", "Business Owner", "business_owner"
                        )
                    )
                    existing_app.application_owner = _clean(
                        get_field("Application Manager", "application_owner")
                    )
                    existing_app.user_types = _clean(
                        get_field("Target Users", "user_types")
                    )
                    existing_app.programming_languages = _clean(
                        get_field("Programming Languages", "programming_languages")
                    )
                    existing_app.support_level = _clean(
                        get_field("Support Level", "support_level")
                    )
                    existing_app.vendor_name = _clean(
                        get_field("Package Vendor", "Vendor", "vendor_name")
                    )
                    existing_app.technical_risk = _clean(
                        get_field("Risk Level", "technical_risk")
                    )
                    existing_app.number_of_integrations = _parse_user_count(
                        get_field("Interfaces - Number", "number_of_integrations")
                    )
                    existing_app.cloud_provider = _clean(
                        get_field("Public Cloud Provider", "cloud_provider")
                    )
                    existing_app.rpo_hours = _parse_user_count(
                        get_field("RPO", "rpo_hours")
                    )
                    existing_app.rto_hours = _parse_user_count(
                        get_field("RTO", "rto_hours")
                    )
                    updated += 1
                    processed_names.add(name_lower)
                else:
                    # Create new record with your template column names
                    app = ApplicationComponent(
                        name=name,
                        description=_clean(get_field("Description")),
                        application_category=_clean(
                            get_field("Category", "application_category")
                        ),
                        business_criticality=_clean(
                            get_field("Business Criticality", "business_criticality")
                        ),
                        deployment_status=_clean(
                            get_field(
                                "Application Status", "Status", "deployment_status"
                            )
                        )
                        or "planned",
                        lifecycle_status=_clean(
                            get_field("Lifecycle Status", "lifecycle_status")
                        ),
                        business_owner=_clean(
                            get_field(
                                "App Business Owner", "Business Owner", "business_owner"
                            )
                        ),
                        application_owner=_clean(
                            get_field("Application Manager", "application_owner")
                        ),
                        user_types=_clean(get_field("Target Users", "user_types")),
                        programming_languages=_clean(
                            get_field("Programming Languages", "programming_languages")
                        ),
                        support_level=_clean(
                            get_field("Support Level", "support_level")
                        ),
                        vendor_name=_clean(
                            get_field("Package Vendor", "Vendor", "vendor_name")
                        ),
                        technical_risk=_clean(
                            get_field("Risk Level", "technical_risk")
                        ),
                        number_of_integrations=_parse_user_count(
                            get_field("Interfaces - Number", "number_of_integrations")
                        ),
                        cloud_provider=_clean(
                            get_field("Public Cloud Provider", "cloud_provider")
                        ),
                        rpo_hours=_parse_user_count(get_field("RPO", "rpo_hours")),
                        rto_hours=_parse_user_count(get_field("RTO", "rto_hours")),
                    )
                    db.session.add(app)
                    count += 1
                    processed_names.add(name_lower)

            db.session.commit()

            # Build success message
            msg_parts = []
            if count > 0:
                msg_parts.append(f"{count} created")
            if updated > 0:
                msg_parts.append(f"{updated} updated")
            if skipped > 0:
                msg_parts.append(f"{skipped} skipped")

            flash(f"Import complete: {', '.join(msg_parts)}", "success")

            # Store successful import in idempotency cache to prevent duplicate processing
            store_import_idempotency(
                _file_content_bytes,
                current_user.id,
                {"created": count, "updated": updated, "skipped": skipped},
            )

            # IMP-002: Audit successful upload
            AuditLog.log_file_upload(
                user_id=current_user.id,
                filename=original_filename,
                sanitized_filename=file.filename,
                file_size_bytes=len(_file_content_bytes),
                mime_type=mime_type,
                ip_address=request.remote_addr,
                route=request.endpoint,
                status="success"
            )

        else:
            flash(
                "Invalid file format. Please upload CSV, JSON, or Excel (.xlsx)",
                "error",
            )

    except ValueError as exc:
        db.session.rollback()
        flash(f"Invalid import data: {exc}", "error")
    except Exception as e:
        db.session.rollback()
        flash("Error importing file. Please try again.", "error")

    return redirect(url_for("unified_applications.application_list"))


# ---------------------------------------------------------------------------
# 8. import_with_ai_review  (POST /import-with-ai-review)
# ---------------------------------------------------------------------------
@unified_applications_bp.route("/import-with-ai-review", methods=["POST"])
@login_required
@audit_log("import_with_ai_review")
@import_rate_limit(max_calls_per_minute=3, max_calls_per_hour=20, max_calls_per_day=100)
def import_with_ai_review():
    """
    Sophisticated AI-powered import with incremental saves, detailed feedback,
    and review workflow. Returns detailed per-app analysis for user approval.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    # AUDIT-IMP-004: Validate file extension to prevent uploading of unsafe file types
    import os

    allowed_extensions = {".csv", ".xlsx", ".xls"}
    file_ext = os.path.splitext(file.filename or "")[1].lower()
    if file_ext not in allowed_extensions:
        return jsonify(
            {
                "error": f"Invalid file type '{file_ext}'. Only CSV and Excel files (.csv, .xlsx, .xls) are allowed."
            }
        ), 400

    # AUDIT-IMP-005: Content-hash idempotency check to prevent double-click race condition.
    # Read file content once into memory so we can hash it for dedup and still parse it later.
    from app.services.unified_import.import_orchestrator import (
        check_import_idempotency as _check_ai_idempotency,
        store_import_idempotency as _store_ai_idempotency,
    )

    _ai_file_content_bytes = file.read()
    file.seek(0)  # Reset stream so downstream parsers can re-read

    _ai_cached_result = _check_ai_idempotency(_ai_file_content_bytes, current_user.id)
    if _ai_cached_result is not None:
        return jsonify(
            {
                "success": True,
                "duplicate_detected": True,
                "message": "This file was already imported within the last 5 minutes. "
                "Returning previous result.",
                "previous_result": _ai_cached_result,
            }
        ), 200

    # Legacy DB-based idempotency check (kept for backward compatibility with import_token)
    import_token = request.form.get("import_token")
    if import_token:
        from datetime import datetime, timedelta
        from app.models.batch_import import ImportAuditLog

        idempotency_window = 300  # 5 minutes
        recent_dup = ImportAuditLog.query.filter(
            ImportAuditLog.user_id == current_user.id,
            ImportAuditLog.filename == (file.filename or ""),
            ImportAuditLog.timestamp
            >= datetime.utcnow() - timedelta(seconds=idempotency_window),
        ).first()
        if recent_dup:
            return jsonify(
                {
                    "error": "This file was already imported within the last 5 minutes. "
                    "If this was intentional, please wait and try again.",
                    "previous_import_id": recent_dup.id,
                }
            ), 409

    # Get AI options
    enable_ai = request.form.get("enable_ai", "false").lower() == "true"
    map_capabilities = request.form.get("map_capabilities", "true").lower() == "true"
    map_processes = request.form.get("map_processes", "true").lower() == "true"
    generate_archimate = (
        request.form.get("generate_archimate", "false").lower() == "true"
    )
    confidence_threshold = float(request.form.get("confidence_threshold", "0.7"))
    import_mode = request.form.get("import_mode", "skip")

    # Get custom field mappings if provided
    field_mappings = {}
    field_mappings_json = request.form.get("field_mappings")
    if field_mappings_json:
        try:
            field_mappings = json.loads(field_mappings_json)
        except json.JSONDecodeError:
            logger.exception("Failed to JSON parsing")
            pass

    try:
        filename = (file.filename or "").lower()

        # Helper functions
        def _clean(value):
            if isinstance(value, str):
                value = value.strip()
                return value or None
            return value

        # Parse file and import applications
        applications_imported = []
        count = 0
        updated = 0
        skipped = 0
        errors = []

        if filename.endswith(".csv"):
            import csv

            raw_content = file.stream.read().decode("utf-8-sig")
            stream = io.StringIO(raw_content)
            reader = csv.DictReader(stream)

            processed_names = set()

            # Batch prefetch all existing application names for case-insensitive lookup
            _all_existing_apps_review_csv = ApplicationComponent.query.limit(
                10000
            ).all()  # Limit to prevent OOM on large datasets
            _existing_app_by_name_review_csv = {
                a.name.lower(): a for a in _all_existing_apps_review_csv if a.name
            }

            for row in reader:
                name = _clean(
                    row.get("Name") or row.get("Application Name") or row.get("name")
                )
                if not name:
                    continue

                name_lower = name.lower()
                if name_lower in processed_names:
                    skipped += 1
                    continue

                # Check for existing application (using prefetched data)
                existing = _existing_app_by_name_review_csv.get(name_lower)

                if existing and import_mode == "skip":
                    skipped += 1
                    processed_names.add(name_lower)
                    continue
                elif existing and import_mode == "update":
                    existing.description = (
                        _clean(row.get("Description")) or existing.description
                    )
                    db.session.flush()  # AUDIT-IMP-003: flush validates without committing
                    updated += 1
                    applications_imported.append(existing)
                    processed_names.add(name_lower)
                else:
                    # Create new
                    app = ApplicationComponent(
                        name=name,
                        description=_clean(row.get("Description")),
                        application_category=_clean(row.get("Category")),
                        business_criticality=_clean(row.get("Business Criticality")),
                        lifecycle_status=_clean(row.get("Lifecycle Status")),
                        imported_capabilities=_clean(row.get("Capabilities")),
                        imported_apqc_codes=_clean(row.get("APQC Codes")),
                    )
                    db.session.add(app)
                    db.session.flush()  # AUDIT-IMP-003: flush assigns IDs without committing
                    count += 1
                    applications_imported.append(app)
                    processed_names.add(name_lower)

        elif filename.endswith((".xlsx", ".xls")):
            from openpyxl import load_workbook

            wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
            ws = wb.active

            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            name_column = None
            for h in headers:
                if h and "name" in h.lower():
                    name_column = h
                    break

            if not name_column:
                return jsonify({"error": "Excel file must have a 'Name' column"}), 400

            processed_names = set()

            # Batch prefetch all existing application names for case-insensitive lookup
            _all_existing_apps_review_xlsx = ApplicationComponent.query.limit(
                10000
            ).all()  # Limit to prevent OOM on large datasets
            _existing_app_by_name_review_xlsx = {
                a.name.lower(): a for a in _all_existing_apps_review_xlsx if a.name
            }

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {
                    headers[i]: row[i] for i in range(len(headers)) if i < len(row)
                }
                name = _clean(row_dict.get(name_column))

                if not name:
                    continue

                name_lower = name.lower()
                if name_lower in processed_names:
                    skipped += 1
                    continue

                # Check for existing application (using prefetched data)
                existing = _existing_app_by_name_review_xlsx.get(name_lower)

                if existing and import_mode == "skip":
                    skipped += 1
                    processed_names.add(name_lower)
                    continue
                elif existing and import_mode == "update":
                    existing.description = (
                        _clean(row_dict.get("Description")) or existing.description
                    )
                    db.session.flush()  # AUDIT-IMP-003: flush validates without committing
                    updated += 1
                    applications_imported.append(existing)
                    processed_names.add(name_lower)
                else:
                    app = ApplicationComponent(
                        name=name,
                        description=_clean(row_dict.get("Description")),
                        application_category=_clean(row_dict.get("Category")),
                        business_criticality=_clean(
                            row_dict.get("Business Criticality")
                        ),
                        lifecycle_status=_clean(row_dict.get("Lifecycle Status")),
                        imported_capabilities=_clean(row_dict.get("Capabilities")),
                        imported_apqc_codes=_clean(row_dict.get("APQC Codes")),
                    )
                    db.session.add(app)
                    db.session.flush()  # AUDIT-IMP-003: flush assigns IDs without committing
                    count += 1
                    applications_imported.append(app)
                    processed_names.add(name_lower)

        else:
            return jsonify({"error": "Unsupported file format"}), 400

        # AUDIT-IMP-003: Single atomic commit for all imported/updated applications.
        # Flush was used per-item above to validate and assign IDs, but nothing is
        # persisted until this commit.  If any row failed, the outer except will
        # rollback the entire batch so no orphaned records remain.
        db.session.commit()

        # NOW RUN AI ANALYSIS WITH INCREMENTAL SAVES AND DETAILED FEEDBACK
        ai_analysis_results = []

        if enable_ai and len(applications_imported) > 0:
            try:
                from app.services.ai_import_service import get_ai_import_service

                ai_service = get_ai_import_service()

                current_app.logger.info(
                    f"Running AI analysis on {len(applications_imported)} applications"
                )

                # Batch prefetch existing capability and process mappings for review AI analysis
                _review_app_ids = [a.id for a in applications_imported]
                _review_existing_cap_mappings = set()
                _review_existing_proc_mappings = set()
                if _review_app_ids:
                    if map_capabilities:
                        from app.models.unified_application_capability_mapping import (
                            UnifiedApplicationCapabilityMapping,
                        )

                        _review_cap_rows = (
                            UnifiedApplicationCapabilityMapping.query.filter(
                                UnifiedApplicationCapabilityMapping.application_id.in_(
                                    _review_app_ids
                                )
                            ).all()
                        )
                        _review_existing_cap_mappings = {
                            (r.application_id, r.capability_id)
                            for r in _review_cap_rows
                        }
                    if map_processes:
                        from app.models.apqc_process import ProcessApplicationMapping

                        _review_proc_rows = ProcessApplicationMapping.query.filter(
                            ProcessApplicationMapping.application_id.in_(
                                _review_app_ids
                            )
                        ).all()
                        _review_existing_proc_mappings = {
                            (r.application_id, r.process_id) for r in _review_proc_rows
                        }

                for app in applications_imported:
                    app_result = {
                        "application_id": app.id,
                        "application_name": app.name,
                        "status": "analyzing",
                        "capabilities_found": [],
                        "processes_found": [],
                        "archimate_elements": [],
                        "capabilities_saved": 0,
                        "processes_saved": 0,
                        "archimate_saved": 0,
                        "saved_to_db": False,
                        "error": None,
                    }

                    try:
                        # Run comprehensive AI analysis
                        ai_result = ai_service.analyze_application_for_ai_mapping(
                            app.id
                        )

                        # Build detailed review data for user approval
                        if ai_result.capability_mappings:
                            for cap in ai_result.capability_mappings:
                                app_result["capabilities_found"].append(
                                    {
                                        "capability_id": cap.get("capability_id"),
                                        "capability_name": cap.get("capability_name"),
                                        "confidence": cap.get("confidence_score", 0.0),
                                        "rationale": cap.get("rationale", ""),
                                        "meets_threshold": cap.get(
                                            "confidence_score", 0.0
                                        )
                                        >= confidence_threshold,
                                    }
                                )

                        if ai_result.process_mappings:
                            for proc in ai_result.process_mappings:
                                app_result["processes_found"].append(
                                    {
                                        "process_id": proc.get("process_id"),
                                        "process_code": proc.get("process_code"),
                                        "process_name": proc.get("process_name"),
                                        "confidence": proc.get("similarity_score", 0.0),
                                        "meets_threshold": proc.get(
                                            "similarity_score", 0.0
                                        )
                                        >= confidence_threshold,
                                    }
                                )

                        if ai_result.archimate_elements:
                            for elem in ai_result.archimate_elements:
                                app_result["archimate_elements"].append(
                                    {
                                        "type": elem.get("type"),
                                        "name": elem.get("name"),
                                        "description": elem.get("description", ""),
                                    }
                                )

                        # SAVE HIGH-CONFIDENCE MAPPINGS INCREMENTALLY (auto-approve)
                        mappings_saved = {
                            "capabilities": 0,
                            "processes": 0,
                            "archimate": 0,
                        }

                        try:
                            if map_capabilities and ai_result.capability_mappings:
                                from app.models.unified_application_capability_mapping import (
                                    UnifiedApplicationCapabilityMapping,
                                )

                                for cap_mapping in ai_result.capability_mappings:
                                    if (
                                        cap_mapping.get("confidence_score", 0)
                                        >= confidence_threshold
                                    ):
                                        cap_id = cap_mapping.get("capability_id")
                                        if cap_id:
                                            _review_cap_key = (app.id, cap_id)
                                            if (
                                                _review_cap_key
                                                not in _review_existing_cap_mappings
                                            ):
                                                db.session.add(
                                                    UnifiedApplicationCapabilityMapping(
                                                        application_id=app.id,
                                                        capability_id=cap_id,
                                                        confidence_score=cap_mapping.get(
                                                            "confidence_score", 0.5
                                                        ),
                                                        mapping_method="ai_import_review",
                                                        rationale=cap_mapping.get(
                                                            "rationale", ""
                                                        ),
                                                        created_by=current_user.email
                                                        if current_user.is_authenticated
                                                        else "ai_import",
                                                    )
                                                )
                                                _review_existing_cap_mappings.add(
                                                    _review_cap_key
                                                )
                                                mappings_saved["capabilities"] += 1

                                db.session.commit()  # Commit capabilities immediately
                                current_app.logger.info(
                                    f"Saved {mappings_saved['capabilities']} capabilities for {app.name}"
                                )

                            if map_processes and ai_result.process_mappings:
                                from app.models.apqc_process import (
                                    ProcessApplicationMapping,
                                )

                                for proc_mapping in ai_result.process_mappings:
                                    if (
                                        proc_mapping.get("similarity_score", 0)
                                        >= confidence_threshold
                                    ):
                                        proc_id = proc_mapping.get("process_id")
                                        if proc_id:
                                            _review_proc_key = (app.id, proc_id)
                                            if (
                                                _review_proc_key
                                                not in _review_existing_proc_mappings
                                            ):
                                                db.session.add(
                                                    ProcessApplicationMapping(
                                                        application_id=app.id,
                                                        process_id=proc_id,
                                                        confidence_score=proc_mapping.get(
                                                            "similarity_score", 0.5
                                                        ),
                                                        mapping_method="ai_import_review",
                                                        created_by=current_user.email
                                                        if current_user.is_authenticated
                                                        else "ai_import",
                                                    )
                                                )
                                                _review_existing_proc_mappings.add(
                                                    _review_proc_key
                                                )
                                                mappings_saved["processes"] += 1

                                db.session.commit()  # Commit processes immediately
                                current_app.logger.info(
                                    f"Saved {mappings_saved['processes']} processes for {app.name}"
                                )

                            if generate_archimate and ai_result.archimate_elements:
                                try:
                                    archimate_service = (
                                        ai_service._get_archimate_service()
                                    )
                                    for elem_data in ai_result.archimate_elements:
                                        elem = (
                                            archimate_service.create_element_from_dict(
                                                elem_data,
                                                created_by=current_user.email
                                                if current_user.is_authenticated
                                                else "ai_import",
                                            )
                                        )
                                        if elem:
                                            mappings_saved["archimate"] += 1

                                    db.session.commit()  # Commit ArchiMate immediately
                                    current_app.logger.info(
                                        f"Saved {mappings_saved['archimate']} ArchiMate elements for {app.name}"
                                    )
                                except Exception as e:
                                    current_app.logger.warning(
                                        f"ArchiMate generation error for {app.name}: {e}"
                                    )

                            app_result["capabilities_saved"] = mappings_saved[
                                "capabilities"
                            ]
                            app_result["processes_saved"] = mappings_saved["processes"]
                            app_result["archimate_saved"] = mappings_saved["archimate"]
                            app_result["saved_to_db"] = (
                                mappings_saved["capabilities"] > 0
                                or mappings_saved["processes"] > 0
                                or mappings_saved["archimate"] > 0
                            )
                            app_result["status"] = "completed"

                        except Exception as save_error:
                            current_app.logger.error(
                                f"Error saving mappings for {app.name}: {save_error}"
                            )
                            db.session.rollback()
                            app_result["status"] = "save_failed"
                            app_result["error"] = str(save_error)

                    except Exception as analysis_error:
                        current_app.logger.error(
                            f"AI analysis error for {app.name}: {analysis_error}"
                        )
                        app_result["status"] = "analysis_failed"
                        app_result["error"] = str(analysis_error)

                    ai_analysis_results.append(app_result)

            except Exception as e:
                current_app.logger.error(f"AI import service error: {e}")
                return (
                    jsonify(
                        {
                            "error": "AI analysis failed. Please try again.",
                            "import_stats": {
                                "created": count,
                                "updated": updated,
                                "skipped": skipped,
                            },
                        }
                    ),
                    500,
                )

        # Store successful import in idempotency cache to prevent duplicate processing
        _ai_import_result_summary = {
            "created": count,
            "updated": updated,
            "skipped": skipped,
            "total_analyzed": len(applications_imported),
        }
        _store_ai_idempotency(
            _ai_file_content_bytes, current_user.id, _ai_import_result_summary
        )

        # Return detailed response for review UI
        return jsonify(
            {
                "success": True,
                "import_stats": _ai_import_result_summary,
                "ai_enabled": enable_ai,
                "applications": ai_analysis_results,
                "summary": {
                    "total_capabilities_saved": sum(
                        app["capabilities_saved"] for app in ai_analysis_results
                    ),
                    "total_processes_saved": sum(
                        app["processes_saved"] for app in ai_analysis_results
                    ),
                    "total_archimate_saved": sum(
                        app["archimate_saved"] for app in ai_analysis_results
                    ),
                    "apps_with_mappings": sum(
                        1 for app in ai_analysis_results if app["saved_to_db"]
                    ),
                    "apps_with_errors": sum(
                        1 for app in ai_analysis_results if app.get("error")
                    ),
                },
            }
        )

    except Exception as e:
        current_app.logger.error(f"Import error: {e}")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred"}), 500
