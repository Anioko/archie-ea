"""Parse Excel/CSV files for data import into deployed solutions.

Uses openpyxl (already in requirements.txt) for Excel and csv stdlib for CSV.
Extends existing FileParser patterns from app/modules/import_batch/.
"""
import csv
import io
import logging
import re
from typing import Any

from werkzeug.datastructures import FileStorage

logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_ROWS_PER_SHEET = 10_000

# Patterns for type inference
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}T")
_DECIMAL_RE = re.compile(r"^-?\d+\.\d+$")
_INTEGER_RE = re.compile(r"^-?\d+$")
_CURRENCY_RE = re.compile(r"^[$£€¥]\s?\d")


class DataIngestionService:
    """Parse uploaded files into structured sheet data with type inference."""

    def parse_file(self, file: FileStorage) -> list[dict]:
        """Parse an uploaded file into sheet descriptors.

        Returns list of dicts, each with:
            name, headers, row_count, sample_rows, column_types, all_rows.
        """
        raw = file.read()
        if len(raw) > MAX_FILE_SIZE:
            raise ValueError(
                f"File size ({len(raw) / 1024 / 1024:.1f} MB) exceeds "
                f"maximum ({MAX_FILE_SIZE / 1024 / 1024:.0f} MB)."
            )
        file.seek(0)

        filename = file.filename or ""
        if filename.lower().endswith((".xlsx", ".xls")):
            return self._parse_excel(raw, filename)
        elif filename.lower().endswith(".csv"):
            return self._parse_csv(raw, filename)
        else:
            raise ValueError(
                f"Unsupported file type: {filename}. Use .xlsx, .xls, or .csv"
            )

    def _parse_csv(self, raw: bytes, filename: str) -> list[dict]:
        """Parse CSV bytes into a single-sheet descriptor."""
        # Handle UTF-8 BOM
        if raw[:3] == b"\xef\xbb\xbf":
            text = raw.decode("utf-8-sig")
        else:
            text = raw.decode("utf-8", errors="replace")

        reader = csv.DictReader(io.StringIO(text))
        headers = [h.strip() for h in (reader.fieldnames or [])]
        rows: list[dict[str, str]] = []
        for i, row in enumerate(reader):
            if i >= MAX_ROWS_PER_SHEET:
                break
            rows.append(
                {
                    h.strip(): (row.get(h) or "").strip()
                    for h in (reader.fieldnames or [])
                }
            )

        name = filename.rsplit(".", 1)[0] if "." in filename else filename
        column_types = {
            h: self.infer_column_type(h, [r.get(h, "") for r in rows[:100]])
            for h in headers
        }

        return [
            {
                "name": name,
                "headers": headers,
                "row_count": len(rows),
                "sample_rows": rows[:5],
                "column_types": column_types,
                "all_rows": rows,
            }
        ]

    def _parse_excel(self, raw: bytes, filename: str) -> list[dict]:
        """Parse Excel bytes into multiple sheet descriptors."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        sheets: list[dict] = []
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                continue
            headers = [
                str(h).strip() if h else f"column_{i}"
                for i, h in enumerate(header_row)
            ]
            rows: list[dict[str, str]] = []
            for i, row in enumerate(rows_iter):
                if i >= MAX_ROWS_PER_SHEET:
                    break
                rows.append(
                    {
                        headers[j]: str(v).strip() if v is not None else ""
                        for j, v in enumerate(row)
                        if j < len(headers)
                    }
                )

            column_types = {
                h: self.infer_column_type(h, [r.get(h, "") for r in rows[:100]])
                for h in headers
            }
            sheets.append(
                {
                    "name": ws.title,
                    "headers": headers,
                    "row_count": len(rows),
                    "sample_rows": rows[:5],
                    "column_types": column_types,
                    "all_rows": rows,
                }
            )
        wb.close()
        return sheets

    def infer_types(self, sheet: dict) -> list[dict]:
        """Infer column types from a sheet's sample data.

        Args:
            sheet: A sheet descriptor from parse_file() with 'headers' and 'all_rows'.

        Returns:
            List of dicts, each with:
                column (str), inferred_type (str), confidence (float), sample_values (list).
            Valid types: string, integer, decimal, date, boolean, email, enum.
        """
        headers = sheet.get("headers", [])
        all_rows = sheet.get("all_rows", [])
        results = []
        for header in headers:
            values = [r.get(header, "") for r in all_rows[:100]]
            inferred_type, confidence = self._infer_type_with_confidence(header, values)
            non_empty = [v for v in values if v and v.strip()]
            results.append({
                "column": header,
                "inferred_type": inferred_type,
                "confidence": round(confidence, 2),
                "sample_values": non_empty[:5],
            })
        return results

    def infer_column_type(self, header: str, values: list[str]) -> str:
        """Infer the data type of a column from header name and sample values.

        Returns one of: string, integer, decimal, date, boolean, email, enum.
        """
        inferred_type, _confidence = self._infer_type_with_confidence(header, values)
        return inferred_type

    def _infer_type_with_confidence(
        self, header: str, values: list[str]
    ) -> tuple[str, float]:
        """Core type inference returning (type, confidence).

        Valid types: string, integer, decimal, date, boolean, email, enum.
        """
        non_empty = [v for v in values if v and v.strip()]
        if not non_empty:
            return ("string", 1.0)

        # Check patterns on non-empty values
        type_votes: dict[str, int] = {
            "string": 0,
            "integer": 0,
            "decimal": 0,
            "date": 0,
            "boolean": 0,
            "email": 0,
            "enum": 0,
        }
        for v in non_empty[:50]:
            v = v.strip()
            if v.lower() in ("true", "false", "yes", "no"):
                type_votes["boolean"] += 1
            elif _EMAIL_RE.match(v):
                type_votes["email"] += 1
            elif _ISO_DATE_RE.match(v) or _DATE_RE.match(v):
                type_votes["date"] += 1
            elif _INTEGER_RE.match(v):
                type_votes["integer"] += 1
            elif _DECIMAL_RE.match(v) or _CURRENCY_RE.match(v):
                type_votes["decimal"] += 1
            else:
                type_votes["string"] += 1

        # Enum detection: if string values repeat with few distinct values
        # relative to total count, treat as enum (max 20 distinct values)
        sample = non_empty[:50]
        distinct = set(sample)
        if (
            type_votes["string"] > 0
            and len(distinct) <= 20
            and len(sample) >= 3
            and len(distinct) < len(sample) * 0.8
        ):
            type_votes["enum"] = type_votes["string"]
            type_votes["string"] = 0

        # Majority vote — need >60% confidence to override string default
        total = sum(type_votes.values())
        if total == 0:
            return ("string", 1.0)
        best_type = max(type_votes, key=lambda k: type_votes[k])
        confidence = type_votes[best_type] / total
        if confidence > 0.6:
            return (best_type, confidence)
        return ("string", confidence)
