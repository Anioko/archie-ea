"""
ArchiMate XML export route — SA-005.

Provides:
    GET /archimate/export/xml  — returns full model as OEF XML (application/xml)

SA-002:
    GET /archimate/traceability        — renders traceability_chain.html
    GET /api/archimate/traceability    — returns JSON chain

SA-007:
    GET /archimate/viewpoints                       — renders viewpoints.html
    GET /api/archimate/viewpoints                   — list of viewpoint definitions
    GET /api/archimate/viewpoints/<id>/data         — filtered elements + layout hints
"""

import concurrent.futures
import json
import re as _re

from flask import Blueprint, Response, abort, current_app, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from sqlalchemy.orm import joinedload

from app import db
from app.modules.architecture.routes.lucidchart_import_routes import (
    register_lucidchart_import_routes,
)
from app.modules.architecture.services.archimate_xml_export_service import export_to_xml
from app.utils.response_helpers import api_error, api_success

archimate_bp = Blueprint("archimate", __name__, url_prefix="/archimate")
register_lucidchart_import_routes(archimate_bp)

_LLM_TIMEOUT_SECONDS = 45


# ── CMP-025: RBAC ownership check ─────────────────────────────────────────
def _check_solution_access(solution_id):
    """Abort 403 if the current user may not edit the given solution."""
    if not solution_id:
        return  # Scratch diagrams (no solution) are unrestricted
    try:
        from app.models.solution_models import Solution
        sol = db.session.get(Solution, int(solution_id))
    except Exception:  # noqa: BLE001
        return  # Model not available — skip check gracefully
    if not sol:
        return  # Solution doesn't exist — let downstream handle
    if hasattr(current_user, "is_admin") and current_user.is_admin():
        return
    if getattr(sol, "owner_id", None) and sol.owner_id == current_user.id:
        return
    if getattr(sol, "created_by", None) and sol.created_by == current_user.id:
        return
    abort(403, description="You do not have permission to modify this solution")


# ── CMP-034: Input validation helper ──────────────────────────────────────
_HTML_TAG_RE = _re.compile(r"<[^>]+>")


def _validate_string(value, field_name, max_length=255, required=False):
    """Validate and sanitize a string input. Returns (cleaned, error_msg)."""
    if value is None:
        value = ""
    if not isinstance(value, str):
        return "", f"{field_name} must be a string"
    value = value.strip()
    if required and not value:
        return "", f"{field_name} is required"
    if len(value) > max_length:
        return "", f"{field_name} must be {max_length} characters or fewer"
    # Strip HTML tags for XSS prevention
    value = _HTML_TAG_RE.sub("", value)
    return value, None


def _run_archimate_llm_generation(requirements, context, target_layer="complete"):
    """Run ArchiMate generation with an application-context timeout guard."""
    from app.modules.architecture.services.archimate_llm_service import ArchiMateLLMService

    svc = ArchiMateLLMService()
    app_obj = current_app._get_current_object()

    def _call_llm():
        with app_obj.app_context():
            try:
                model_data, _ = svc.generate_archimate_from_requirements(
                    requirements=requirements,
                    context=context,
                    target_layer=target_layer,
                )
                return model_data
            finally:
                try:
                    db.session.remove()
                except Exception as exc:  # noqa: BLE001
                    current_app.logger.debug("CMP-036: session remove failed after LLM call: %s", exc)

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=1, thread_name_prefix="archimate-llm")
    future = executor.submit(_call_llm)
    try:
        return future.result(timeout=_LLM_TIMEOUT_SECONDS)
    except concurrent.futures.TimeoutError:
        future.cancel()
        current_app.logger.warning("SEC-003: LLM generation timed out after %ss", _LLM_TIMEOUT_SECONDS)
    finally:
        executor.shutdown(wait=False, cancel_futures=True)
    return None


@archimate_bp.route("/export/xml", methods=["GET"])
@login_required
def export_xml():
    """Export all ArchiMate elements/relationships as OEF XML.

    Query Parameters:
        model_id (int): Optional — scope export to a specific ArchitectureModel.

    Returns:
        application/xml with OEF XML content.
    """
    model_id = request.args.get("model_id", type=int)
    try:
        xml_content = export_to_xml(model_id=model_id)
        return Response(
            f'<?xml version="1.0" encoding="UTF-8"?>\n{xml_content}',
            status=200,
            mimetype="application/xml",
            headers={
                "Content-Disposition": "attachment; filename=archimate_export.xml",
            },
        )
    except Exception as exc:  # noqa: BLE001
        current_app.logger.error(f"[SA-005] export_to_xml failed: {exc}")
        return Response(str(exc), status=500, mimetype="text/plain")


# ── SA-002 — Cross-layer traceability chain ──────────────────────────────────

@archimate_bp.route("/traceability", methods=["GET"])
@login_required
def traceability_chain():
    """Render the 8-layer traceability chain page with cross-layer highlighting."""
    from app.services.archimate_traceability_service import (
        get_traceability_chain, get_gap_analysis, get_element_solution_map_by_layer,
    )
    from app.services.traceability_graph_service import build_traceability_graph

    solution_id = request.args.get("solution_id", type=int)
    chain = get_traceability_chain(solution_id=solution_id) or {}
    relationship_maps = chain.pop("relationship_maps", {})
    graph = build_traceability_graph()
    gap_analysis = get_gap_analysis()

    # GLB-056 / TRC-026: Element-to-solution mapping by layer (correct for all layers)
    element_ids_by_layer = {}
    for layer_key, items in chain.items():
        if isinstance(items, list):
            ids = [item.get("id") for item in items if isinstance(item, dict) and "id" in item]
            if ids:
                element_ids_by_layer[layer_key] = ids
    element_solutions_by_layer = get_element_solution_map_by_layer(element_ids_by_layer)
    # JSON: { "layer": { "element_id": [solutions] } } for template lookup by selected.layer + selected.id
    element_solutions_json = json.dumps({
        layer: {str(eid): sols for eid, sols in by_id.items()}
        for layer, by_id in element_solutions_by_layer.items()
    })

    # GLB-057: Solutions list for filter dropdown
    from app.models.solution_models import Solution
    solutions_list = Solution.query.order_by(Solution.name).all()

    return render_template(
        "archimate/traceability_chain.html",
        chain=chain,
        graph=graph,
        relationship_maps=relationship_maps,
        gap_analysis=gap_analysis,
        element_solutions_by_layer=element_solutions_by_layer,
        element_solutions_json=element_solutions_json,
        solutions_list=solutions_list,
        current_solution_id=solution_id,
    )


# ── GLB-059 — Element inline edit from traceability ──────────────────────────

@archimate_bp.route("/api/elements/<int:element_id>", methods=["PATCH"])
@login_required
def patch_element(element_id):
    """Update name/description of an ArchiMate element with audit trail."""
    from app.models.archimate_core import ArchiMateElement as AE
    from app.models.architecture_review_board import ARBAuditLog

    element = db.session.get(AE, element_id)
    if not element:
        return jsonify({"error": "Element not found"}), 404

    data = request.get_json(silent=True) or {}

    # GAP-CMP-009: Handle custom_properties merge (data classification, PII)
    # GAP-CMP-004: Handle lifecycle_history append
    incoming_cp = data.get("custom_properties")
    if incoming_cp and isinstance(incoming_cp, dict):
        existing_cp = getattr(element, "custom_properties", None) or {}
        if isinstance(existing_cp, str):
            try:
                existing_cp = json.loads(existing_cp)
            except (json.JSONDecodeError, TypeError):
                existing_cp = {}
        # GAP-CMP-004: lifecycle history append — accumulate transitions
        history_entry = incoming_cp.pop("_lifecycle_history_append", None)
        if history_entry and isinstance(history_entry, dict):
            lifecycle_history = existing_cp.get("lifecycle_history", [])
            if not isinstance(lifecycle_history, list):
                lifecycle_history = []
            lifecycle_history.append(history_entry)
            existing_cp["lifecycle_history"] = lifecycle_history
        existing_cp.update(incoming_cp)
        element.custom_properties = existing_cp

    name = data.get("name")
    description = data.get("description")

    # If name/description are provided, validate and update them
    if name is not None:
        name = name.strip()
        if not name:
            return jsonify({"error": "Name is required"}), 400
        if len(name) > 100:
            return jsonify({"error": "Name must be 100 characters or less"}), 400
    else:
        name = element.name

    old_name = element.name
    old_desc = element.description
    element.name = name
    element.description = (description or "").strip() if description is not None else element.description

    try:
        audit = ARBAuditLog(
            action="element_updated",
            entity_type="ArchiMateElement",
            entity_id=element_id,
            details=json.dumps({
                "old_name": old_name, "new_name": name,
                "old_description": old_desc, "new_description": description,
            }),
            user_id=current_user.id if hasattr(current_user, "id") else None,
        )
        db.session.add(audit)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    return jsonify({
        "id": element.id, "name": element.name,
        "description": element.description, "type": element.type, "layer": element.layer,
    })


# ── TRC-028 — Gap link actions (create missing traceability links) ─────────────

@archimate_bp.route("/api/link/driver-to-goal", methods=["POST"])
@login_required
def link_driver_to_goal():
    """Link an orphan driver to a goal. Body: {driver_id, goal_id}. Updates Goal.driver_id."""
    data = request.get_json(silent=True) or {}
    driver_id = data.get("driver_id")
    goal_id = data.get("goal_id")
    if driver_id is None or goal_id is None:
        return jsonify({"error": "driver_id and goal_id required"}), 400
    from app.models.motivation import Driver, Goal
    driver = db.session.get(Driver, int(driver_id))
    goal = db.session.get(Goal, int(goal_id))
    if not driver or not goal:
        return jsonify({"error": "Driver or Goal not found"}), 404
    goal.driver_id = driver.id
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "driver_id": driver.id, "goal_id": goal.id}), 200


@archimate_bp.route("/api/link/capability-to-application", methods=["POST"])
@login_required
def link_capability_to_application():
    """Link an orphan capability to an application. Body: {capability_id, application_id}. Creates ApplicationCapabilityMapping."""
    data = request.get_json(silent=True) or {}
    capability_id = data.get("capability_id")
    application_id = data.get("application_id")
    if capability_id is None or application_id is None:
        return jsonify({"error": "capability_id and application_id required"}), 400
    from app.models.business_capabilities import BusinessCapability
    from app.models.application_portfolio import ApplicationComponent
    from app.models.application_capability import ApplicationCapabilityMapping
    cap = db.session.get(BusinessCapability, int(capability_id))
    app = db.session.get(ApplicationComponent, int(application_id))
    if not cap or not app:
        return jsonify({"error": "Capability or Application not found"}), 404
    existing = ApplicationCapabilityMapping.query.filter_by(
        business_capability_id=cap.id,
        application_component_id=app.id,
    ).first()
    if existing:
        return jsonify({"ok": True, "message": "Already linked"}), 200
    mapping = ApplicationCapabilityMapping(
        business_capability_id=cap.id,
        application_component_id=app.id,
    )
    db.session.add(mapping)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "capability_id": cap.id, "application_id": app.id}), 200


# ── SA-007 — ArchiMate viewpoint diagram renderer ────────────────────────────

@archimate_bp.route("/viewpoints", methods=["GET"])
@login_required
def viewpoints_page():
    """Redirect to unified composer — viewpoints are now a mode within the composer."""
    active = request.args.get("active", "")
    solution_id = request.args.get("solution_id", type=int)
    params = {}
    if active:
        params["viewpoint"] = active
    if solution_id:
        params["solution_id"] = solution_id
    return redirect(url_for("archimate.composer_page", **params), code=302)


@archimate_bp.route("/composer", methods=["GET"])
@login_required
def composer_page():
    """Render the unified ArchiMate Composer — diagram editor + viewpoint viewer.

    Query Parameters:
        solution_id (int): Scope to a specific solution (required for save).
        viewpoint (str): Pre-select a viewpoint (opens in View mode).
    """
    from app.services.archimate_viewpoint_service import get_available_viewpoints, get_viewpoint_counts

    solution_id = request.args.get("solution_id", type=int)
    viewpoint = request.args.get("viewpoint", "")
    solution_name = None
    if solution_id:
        # tenant-filtered: scoped via parent FK (solution_id from request)
        row = db.session.execute(  # tenant-filtered: scoped via parent FK (solution_id from request)
            db.text("SELECT name FROM solutions WHERE id = :sid"),  # tenant-filtered
            {"sid": solution_id},
        ).fetchone()
        solution_name = row.name if row else f"Solution #{solution_id}"

    viewpoints = get_available_viewpoints()
    categories = {}
    for vp in viewpoints:
        cat = vp.get("category", "other")
        categories.setdefault(cat, []).append(vp)

    vp_counts = get_viewpoint_counts(solution_id) if solution_id else {}

    return render_template(
        "archimate/composer.html",
        solution_id=solution_id or 0,
        solution_name=solution_name or "Enterprise",
        viewpoints=viewpoints,
        viewpoint_categories=categories,
        viewpoint_counts=vp_counts,
        initial_viewpoint=viewpoint,
    )


@archimate_bp.route("/api/valid-relationship-types", methods=["GET"])
@login_required
def api_valid_relationship_types():
    """Return valid ArchiMate relationship types for a source+target element pair.

    Query Parameters:
        source_id (int): Source ArchiMate element ID.
        target_id (int): Target ArchiMate element ID.

    Returns JSON with valid_types (list of type strings for backward compat)
    and valid_types_detailed (list of {type, tier, description} dicts).
    """
    from app.models.archimate_core import ArchiMateElement
    from app.services.archimate_validity_service import ArchimateValidityService

    source_id = request.args.get("source_id", type=int)
    target_id = request.args.get("target_id", type=int)
    if not source_id or not target_id:
        return jsonify({"error": "source_id and target_id required"}), 400

    src = db.session.get(ArchiMateElement, source_id)
    tgt = db.session.get(ArchiMateElement, target_id)
    if not src or not tgt:
        return jsonify({"error": "Element not found"}), 404

    svc = ArchimateValidityService()
    detailed = svc.get_valid_relationships(src.type or "", tgt.type or "")

    return jsonify({
        "source_type": src.type,
        "target_type": tgt.type,
        "valid_types": [r["type"] for r in detailed],
        "valid_types_detailed": detailed,
    })


@archimate_bp.route("/viewpoints-api", methods=["GET"])
@login_required
def api_list_viewpoints():
    """Return JSON list of all viewpoint definitions."""
    from app.services.archimate_viewpoint_service import get_available_viewpoints
    return jsonify(get_available_viewpoints())


@archimate_bp.route("/viewpoints-api/<viewpoint_id>/data", methods=["GET"])
@login_required
def api_viewpoint_data(viewpoint_id: str):
    """Return filtered elements and layout hints for a viewpoint."""
    from app.services.archimate_viewpoint_service import get_viewpoint_data
    solution_id = request.args.get("solution_id", type=int)
    data = get_viewpoint_data(viewpoint_id=viewpoint_id, solution_id=solution_id)
    return jsonify(data)


# ── Relationship CRUD API ────────────────────────────────────────────────────

ARCHIMATE_RELATIONSHIP_TYPES = [
    "composition", "aggregation", "assignment", "realization",
    "serving", "access", "influence",
    "triggering", "flow",
    "specialization", "association",
]

# Canonical aliases: non-standard names that appear in legacy/seeded data
_REL_TYPE_ALIASES = {
    "realizes": "realization",
    "serves": "serving",
    "uses": "serving",
    "triggers": "triggering",
    "flows": "flow",
    "composes": "composition",
    "aggregates": "aggregation",
    "assigns": "assignment",
    "specializes": "specialization",
    "associates": "association",
}


def _normalize_rel_type(raw: str) -> str:
    """Normalise any legacy or non-canonical relationship type string to the
    lowercase canonical form accepted by ARCHIMATE_RELATIONSHIP_TYPES.

    Handles:
      - ``CompositionRelationship``  → ``composition``
      - ``Realization``              → ``realization``
      - ``Realizes``                 → ``realization``
      - ``composition``              → ``composition``  (already canonical)
    """
    if not raw:
        return ""
    # Strip trailing "Relationship" suffix (case-insensitive)
    normalised = _re.sub(r"(?i)relationship$", "", raw).strip()
    normalised = normalised.lower()
    # Map known aliases
    return _REL_TYPE_ALIASES.get(normalised, normalised)


@archimate_bp.route("/api/relationships", methods=["GET"])
@login_required
def api_list_relationships():
    """List relationships with optional filters.

    Query Parameters:
        solution_id (int): Filter by solution.
        element_id (int): Filter where element is source OR target.
        source_element_id (int): Filter by source element.
        target_element_id (int): Filter by target element.
        relationship_type (str): Filter by ArchiMate relationship type.
        include_enterprise (bool): Include enterprise-level (solution_id=NULL) relationships. Default true.
        page (int): Page number (default 1).
        per_page (int): Items per page (default 50, max 200).
    """
    from app.models.archimate_core import ArchiMateRelationship

    solution_id = request.args.get("solution_id", type=int)
    element_id = request.args.get("element_id", type=int)
    source_element_id = request.args.get("source_element_id", type=int)
    target_element_id = request.args.get("target_element_id", type=int)
    relationship_type = request.args.get("relationship_type")
    include_enterprise = request.args.get("include_enterprise", "true").lower() != "false"
    page = request.args.get("page", 1, type=int)
    per_page = min(request.args.get("per_page", 50, type=int), 200)

    query = ArchiMateRelationship.query

    if solution_id is not None:
        if include_enterprise:
            query = query.filter(
                db.or_(
                    ArchiMateRelationship.architecture_id == solution_id,
                    ArchiMateRelationship.architecture_id.is_(None),
                )
            )
        else:
            query = query.filter(ArchiMateRelationship.architecture_id == solution_id)

    if element_id is not None:
        query = query.filter(
            db.or_(
                ArchiMateRelationship.source_id == element_id,
                ArchiMateRelationship.target_id == element_id,
            )
        )

    # Bulk filter: all relationships where BOTH source and target are in the given set
    element_ids_param = request.args.get("element_ids")
    if element_ids_param:
        try:
            el_ids = [int(x) for x in element_ids_param.split(",") if x.strip().isdigit()]
        except ValueError:
            el_ids = []
        if el_ids:
            query = query.filter(
                ArchiMateRelationship.source_id.in_(el_ids),
                ArchiMateRelationship.target_id.in_(el_ids),
            )

    if source_element_id is not None:
        query = query.filter(ArchiMateRelationship.source_id == source_element_id)

    if target_element_id is not None:
        query = query.filter(ArchiMateRelationship.target_id == target_element_id)

    if relationship_type:
        query = query.filter(ArchiMateRelationship.type == relationship_type)

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    relationships = [
        {
            "id": r.id,
            "source_id": r.source_id,
            "target_id": r.target_id,
            "type": _normalize_rel_type(r.type or ""),
            "name": getattr(r, "name", None),
            "description": getattr(r, "description", None),
            "solution_id": r.architecture_id,
            "created_at": r.created_at.isoformat() if getattr(r, "created_at", None) else None,
        }
        for r in pagination.items
    ]

    return api_success({
        "relationships": relationships,
        "total": pagination.total,
        "page": pagination.page,
        "pages": pagination.pages,
    })


@archimate_bp.route("/api/relationships", methods=["POST"])
@login_required
# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
def api_create_relationship():
    """Create a new ArchiMate relationship.

    JSON Body:
        source_element_id (int): Required.
        target_element_id (int): Required.
        relationship_type (str): Required — must be a valid ArchiMate relationship type.
        solution_id (int): Optional — mapped to architecture_id for scoping.
    """
    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship

    data = request.get_json(silent=True) or {}

    source_id = data.get("source_element_id")
    target_id = data.get("target_element_id")
    rel_type = _normalize_rel_type(data.get("relationship_type") or "")

    # --- Validation ---
    if not source_id or not target_id:
        return api_error("source_element_id and target_element_id are required", 400)

    if not rel_type:
        return api_error("relationship_type is required", 400)

    if rel_type not in ARCHIMATE_RELATIONSHIP_TYPES:
        return api_error(f"Invalid relationship_type. Must be one of: {ARCHIMATE_RELATIONSHIP_TYPES}", 400)

    if source_id == target_id:
        return api_error("source_element_id and target_element_id must be different", 400)

    # Verify both elements exist
    source_el = db.session.get(ArchiMateElement, source_id)
    target_el = db.session.get(ArchiMateElement, target_id)
    if not source_el:
        return api_error(f"Source element {source_id} not found", 404)
    if not target_el:
        return api_error(f"Target element {target_id} not found", 404)

    # solution_id from the client maps to architecture_id on the model
    arch_id = data.get("solution_id")

    # Duplicate check using actual model columns
    existing = ArchiMateRelationship.query.filter_by(
        source_id=source_id,
        target_id=target_id,
        type=rel_type,
        architecture_id=arch_id,
    ).first()
    if existing:
        return api_error("Duplicate relationship already exists", 409)

    # BUG-CMP-002: Accept relationship metadata on creation
    rel = ArchiMateRelationship(
        source_id=source_id,
        target_id=target_id,
        type=rel_type,
        architecture_id=arch_id,
        description=data.get("description") or None,
        access_mode=data.get("access_mode") or None,
        flow_label=data.get("flow_label") or None,
        custom_label=data.get("custom_label") or None,
        created_by_id=getattr(current_user, "id", None) if hasattr(current_user, "id") else None,
        connection_spec=data.get("connection_spec") or None,
    )

    db.session.add(rel)
    db.session.commit()

    # Return flat JSON (no api_success wrapper) — all JS callers check data.id directly.
    return jsonify({
        "id": rel.id,
        "source_id": rel.source_id,
        "target_id": rel.target_id,
        "type": rel.type,
        "solution_id": rel.architecture_id,
        "description": rel.description,
        "access_mode": rel.access_mode,
        "flow_label": rel.flow_label,
        "custom_label": rel.custom_label,
        "connection_spec": rel.connection_spec,
    }), 201


@archimate_bp.route("/api/relationships/<int:rel_id>", methods=["PUT"])
@login_required
# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
def api_update_relationship(rel_id):
    """Update an existing ArchiMate relationship.

    JSON Body (all optional):
        name (str): Display name.
        description (str): Description.
        relationship_type (str): Must be a valid ArchiMate relationship type.
    """
    from app.models.archimate_core import ArchiMateRelationship

    rel = db.session.get(ArchiMateRelationship, rel_id)
    if not rel:
        return api_error("Relationship not found", 404)

    data = request.get_json(silent=True) or {}

    if "relationship_type" in data:
        normalised = _normalize_rel_type(data["relationship_type"] or "")
        if normalised not in ARCHIMATE_RELATIONSHIP_TYPES:
            return api_error(f"Invalid relationship_type. Must be one of: {ARCHIMATE_RELATIONSHIP_TYPES}", 400)
        rel.type = normalised

    if "name" in data:
        if hasattr(rel, "name"):
            rel.name = data["name"]

    # BUG-CMP-002: Persist relationship metadata
    if "description" in data:
        rel.description = data["description"] or None
    if "access_mode" in data:
        rel.access_mode = data["access_mode"] or None
    if "flow_label" in data:
        rel.flow_label = data["flow_label"] or None
    if "custom_label" in data:
        rel.custom_label = data["custom_label"] or None
    # GAP-INT-001: Connection specification (structured integration metadata)
    if "connection_spec" in data:
        rel.connection_spec = data["connection_spec"] or None

    db.session.commit()

    return api_success({
        "id": rel.id,
        "source_id": rel.source_id,
        "target_id": rel.target_id,
        "type": rel.type,
        "solution_id": rel.architecture_id,
        "description": rel.description,
        "access_mode": rel.access_mode,
        "flow_label": rel.flow_label,
        "custom_label": rel.custom_label,
        "connection_spec": rel.connection_spec,
    })


@archimate_bp.route("/api/relationships/<int:rel_id>", methods=["DELETE"])
@login_required
# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
def api_delete_relationship(rel_id):
    """Delete an ArchiMate relationship by ID."""
    from app.models.archimate_core import ArchiMateRelationship

    rel = db.session.get(ArchiMateRelationship, rel_id)
    if not rel:
        return api_error("Relationship not found", 404)

    db.session.delete(rel)
    db.session.commit()

    return api_success({"deleted": True, "id": rel_id})


# ── Saved Viewpoint CRUD API ──────────────────────────────────────────────────

@archimate_bp.route("/api/saved-viewpoints", methods=["GET"])
@login_required
def api_list_saved_viewpoints():
    """List saved diagrams, optionally filtered by solution_id."""
    from app.models.archimate_core import SavedDiagram

    solution_id = request.args.get("solution_id", type=int)
    query = SavedDiagram.query
    if solution_id:
        query = query.filter(
            db.or_(
                SavedDiagram.solution_id == solution_id,
                SavedDiagram.solution_id.is_(None),
            )
        )
    query = query.order_by(SavedDiagram.updated_at.desc())
    viewpoints = query.all()

    return jsonify({
        "viewpoints": [vp.to_dict() for vp in viewpoints],
    })


def _materialize_canvas_items(data):
    """Turn imported (not-yet-persisted) canvas items into real model rows.

    Lucid/CSV/OEF imports place elements on the canvas with their SOURCE ids
    (strings like "eu2WgWPngIX_"), but SavedDiagramElement.element_id is an
    Integer FK to archimate_elements — so saving an imported diagram crashed on
    the FK. Fix: any element whose id is not an existing integer is materialized
    as a real ArchiMateElement (and likewise relationships as
    ArchiMateRelationship), then the payload ids are rewritten in place.

    Returns (element_id_map, relationship_id_map) of source-id → new DB id so the
    client can update its canvas cells (otherwise a re-save would duplicate).
    """
    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship

    def _as_int(value):
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, str) and value.isdigit():
            return int(value)
        return None

    element_id_map = {}
    for el_data in (data.get("elements") or []):
        el_id = el_data.get("element_id")
        if not el_id or _as_int(el_id) is not None:
            continue
        element = ArchiMateElement(
            name=(el_data.get("name") or str(el_id))[:100],
            type=el_data.get("el_type") or "ApplicationComponent",
            layer=(el_data.get("layer") or "application")[:30],
        )
        db.session.add(element)
        db.session.flush()
        element_id_map[str(el_id)] = element.id
        el_data["element_id"] = element.id

    relationship_id_map = {}
    for rel_data in (data.get("relationships") or []):
        rel_id = rel_data.get("relationship_id")
        if not rel_id or _as_int(rel_id) is not None:
            continue
        src = rel_data.get("source_element_id")
        tgt = rel_data.get("target_element_id")
        src_id = element_id_map.get(str(src)) or _as_int(src)
        tgt_id = element_id_map.get(str(tgt)) or _as_int(tgt)
        if not src_id or not tgt_id:
            # endpoints unknown — drop rather than fabricate a dangling FK
            rel_data["relationship_id"] = None
            continue
        relationship = ArchiMateRelationship(
            type=(rel_data.get("rel_type") or "association")[:30],
            source_id=src_id,
            target_id=tgt_id,
            custom_label=(rel_data.get("label") or None),
        )
        db.session.add(relationship)
        db.session.flush()
        relationship_id_map[str(rel_id)] = relationship.id
        rel_data["relationship_id"] = relationship.id

    return element_id_map, relationship_id_map


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/saved-viewpoints", methods=["POST"])
@login_required
def api_create_saved_viewpoint():
    """Create a new saved viewpoint with element positions and relationships.

    JSON Body:
        name (str): Required.
        viewpoint_type (str): Optional viewpoint category.
        solution_id (int): Optional solution scope.
        description (str): Optional description.
        elements (list): Array of {element_id, x, y, width?, height?, rendering_mode?}.
        relationships (list): Array of {relationship_id, waypoints?, routing_style?}.
    """
    from app.models.archimate_core import (
        SavedDiagram, SavedDiagramElement, SavedDiagramRelationship,
    )

    data = request.get_json(silent=True) or {}

    # CMP-025: RBAC check
    _check_solution_access(data.get("solution_id"))

    # CMP-034: Input validation
    name, err = _validate_string(data.get("name"), "name", max_length=255, required=True)
    if err:
        return jsonify({"error": err}), 400
    desc, _ = _validate_string(data.get("description"), "description", max_length=5000)

    vp = SavedDiagram(
        name=name,
        viewpoint_type=data.get("viewpoint_type"),
        solution_id=data.get("solution_id"),
        description=desc,
    )
    db.session.add(vp)
    db.session.flush()

    # Materialize imported canvas items (string source ids) into real model rows
    # so the FK inserts below cannot fail and the data joins the model.
    element_id_map, relationship_id_map = _materialize_canvas_items(data)

    for el_data in (data.get("elements") or []):
        el_id = el_data.get("element_id")
        if not el_id:
            continue
        ve = SavedDiagramElement(
            diagram_id=vp.id,
            element_id=el_id,
            position_x=el_data.get("x", 0),
            position_y=el_data.get("y", 0),
            width=el_data.get("width", 180),
            height=el_data.get("height", 64),
            rendering_mode=el_data.get("rendering_mode", "black_box"),
        )
        db.session.add(ve)

    for rel_data in (data.get("relationships") or []):
        rel_id = rel_data.get("relationship_id")
        if not rel_id:
            continue
        import json as _json
        waypoints = rel_data.get("waypoints")
        vr = SavedDiagramRelationship(
            diagram_id=vp.id,
            relationship_id=rel_id,
            waypoints_json=_json.dumps(waypoints) if waypoints else None,
            routing_style=rel_data.get("routing_style", "manhattan"),
        )
        db.session.add(vr)

    db.session.commit()

    body = vp.to_dict()
    if element_id_map or relationship_id_map:
        body["element_id_map"] = element_id_map
        body["relationship_id_map"] = relationship_id_map
    return jsonify(body), 201


@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>", methods=["GET"])
@login_required
def api_get_saved_viewpoint(vp_id):
    """Load a saved viewpoint with element positions and relationships.

    Returns full viewpoint data: metadata, elements with positions/rendering,
    and relationships with waypoints/routing from the diagram junction tables.
    """
    import json as _json

    from app.models.archimate_core import (
        ArchiMateElement, ArchiMateRelationship, SavedDiagram,
    )

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return jsonify({"error": "Diagram not found"}), 404

    positions = vp.positions.all()
    element_ids = [p.element_id for p in positions]

    elements = []
    pos_map = {p.element_id: p for p in positions}

    if element_ids:
        els = ArchiMateElement.query.filter(ArchiMateElement.id.in_(element_ids)).all()
        for el in els:
            p = pos_map.get(el.id)
            elements.append({
                "id": el.id,
                "name": el.name,
                "type": el.type,
                "layer": el.layer,
                "description": el.description,
                "x": p.position_x if p else 0,
                "y": p.position_y if p else 0,
                "width": p.width if p else 180,
                "height": p.height if p else 64,
                "rendering_mode": p.rendering_mode if p else "black_box",
            })

        # Get relationships between these elements
        rels = ArchiMateRelationship.query.filter(
            ArchiMateRelationship.source_id.in_(element_ids),
            ArchiMateRelationship.target_id.in_(element_ids),
        ).all()
    else:
        rels = []

    # Build waypoint/routing map from diagram junction table
    rel_positions = vp.rel_positions.all()
    rel_pos_map = {rp.relationship_id: rp for rp in rel_positions}

    relationships = []
    for r in rels:
        rp = rel_pos_map.get(r.id)
        waypoints_raw = rp.waypoints_json if rp else None
        relationships.append({
            "id": r.id,
            "source_id": r.source_id,
            "target_id": r.target_id,
            "type": r.type,
            "name": getattr(r, "name", None),
            "waypoints": _json.loads(waypoints_raw) if waypoints_raw else None,
            "routing_style": rp.routing_style if rp else "manhattan",
        })

    return jsonify({
        "id": vp.id,
        "name": vp.name,
        "viewpoint_type": vp.viewpoint_type,
        "solution_id": vp.solution_id,
        "description": vp.description,
        "created_at": vp.created_at.isoformat() if vp.created_at else None,
        "updated_at": vp.updated_at.isoformat() if vp.updated_at else None,
        "elements": elements,
        "relationships": relationships,
    })


@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>/relationship-health", methods=["GET"])
@login_required
def api_viewpoint_relationship_health(vp_id):
    """Return stale relationships for a viewpoint after Abacus sync.

    A relationship is stale when one or both endpoint elements link to an
    ApplicationComponent that was NOT updated in the latest Abacus sync
    (ApplicationComponent.abacus_source=True AND
     ApplicationComponent.last_sync_from_abacus < ExternalSystem.last_sync_at).
    Returns an empty list when Abacus is not configured.
    """
    from datetime import datetime, timezone

    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship, SavedDiagram, SavedDiagramRelationship
    from app.models.application_portfolio import ApplicationComponent
    from app.models.models import ExternalSystem

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return jsonify({"error": "Diagram not found"}), 404

    # Resolve last Abacus sync time; return empty list if Abacus not configured
    abacus_sys = ExternalSystem.query.filter_by(system_type="abacus").first()
    if not abacus_sys or not abacus_sys.last_sync_at:
        return jsonify({
            "stale_relationships": [],
            "last_sync_at": None,
            "checked_at": datetime.now(timezone.utc).isoformat(),
        })

    last_sync_at = abacus_sys.last_sync_at

    # Gather element IDs present in this viewpoint
    positions = vp.positions.all()
    element_ids = [p.element_id for p in positions]
    if not element_ids:
        return jsonify({
            "stale_relationships": [],
            "last_sync_at": last_sync_at.isoformat(),
            "checked_at": datetime.now(timezone.utc).isoformat(),
        })

    # Load elements and their linked ApplicationComponents
    elements = ArchiMateElement.query.filter(
        ArchiMateElement.id.in_(element_ids),
        ArchiMateElement.application_component_id.isnot(None),
    ).all()

    app_component_ids = [el.application_component_id for el in elements]
    app_components = {}
    if app_component_ids:
        comps = ApplicationComponent.query.filter(
            ApplicationComponent.id.in_(app_component_ids),
            ApplicationComponent.abacus_source.is_(True),
        ).all()
        app_components = {c.id: c for c in comps}

    # Build a set of element IDs whose linked app component is stale
    stale_element_ids = set()
    stale_app_name_map = {}
    stale_last_seen_map = {}
    for el in elements:
        comp = app_components.get(el.application_component_id)
        if comp and comp.last_sync_from_abacus and comp.last_sync_from_abacus < last_sync_at:
            stale_element_ids.add(el.id)
            stale_app_name_map[el.id] = comp.name
            stale_last_seen_map[el.id] = comp.last_sync_from_abacus.isoformat()

    # Build element name lookup for response
    el_name_map = {el.id: el.name for el in ArchiMateElement.query.filter(
        ArchiMateElement.id.in_(element_ids),
    ).all()}

    # Find relationships where source or target is stale
    rels = ArchiMateRelationship.query.filter(
        ArchiMateRelationship.source_id.in_(element_ids),
        ArchiMateRelationship.target_id.in_(element_ids),
    ).all()

    # Build set of relationship IDs already marked as architectural intent (suppress from review)
    intent_rel_ids = {
        sdr.relationship_id
        for sdr in SavedDiagramRelationship.query.filter_by(
            diagram_id=vp_id, is_architectural_intent=True,
        ).all()
    }

    stale_relationships = []
    for r in rels:
        if r.id in intent_rel_ids:
            continue
        stale_side_id = None
        if r.source_id in stale_element_ids:
            stale_side_id = r.source_id
        elif r.target_id in stale_element_ids:
            stale_side_id = r.target_id
        if stale_side_id is None:
            continue
        stale_relationships.append({
            "rel_id": r.id,
            "rel_type": r.type,
            "source_name": el_name_map.get(r.source_id, str(r.source_id)),
            "target_name": el_name_map.get(r.target_id, str(r.target_id)),
            "stale_app_name": stale_app_name_map.get(stale_side_id, "Unknown"),
            "last_seen": stale_last_seen_map.get(stale_side_id),
        })

    return jsonify({
        "stale_relationships": stale_relationships,
        "last_sync_at": last_sync_at.isoformat(),
        "checked_at": datetime.now(timezone.utc).isoformat(),
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>", methods=["PATCH"])
@login_required
def api_patch_saved_viewpoint(vp_id):
    """Partial update for a saved viewpoint.

    Supports:
        relationship_intent: { rel_id: int, is_architectural_intent: bool }
            Marks a stale relationship as kept-by-intent so it is excluded from
            future staleness review prompts.
    """
    from app.models.archimate_core import SavedDiagram, SavedDiagramRelationship

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return jsonify({"error": "Diagram not found"}), 404

    _check_solution_access(vp.solution_id)

    data = request.get_json(silent=True) or {}

    if "relationship_intent" in data:
        intent = data["relationship_intent"]
        rel_id = intent.get("rel_id")
        is_intent = bool(intent.get("is_architectural_intent", True))
        if not rel_id:
            return jsonify({"error": "relationship_intent.rel_id is required"}), 400

        sdr = SavedDiagramRelationship.query.filter_by(
            diagram_id=vp_id, relationship_id=int(rel_id),
        ).first()
        if sdr:
            sdr.is_architectural_intent = is_intent
        else:
            # Relationship not yet persisted in saved_diagram_relationships; create it
            sdr = SavedDiagramRelationship(
                diagram_id=vp_id,
                relationship_id=int(rel_id),
                is_architectural_intent=is_intent,
            )
            db.session.add(sdr)

        db.session.commit()
        return jsonify({"updated": True, "rel_id": rel_id, "is_architectural_intent": is_intent})

    return jsonify({"error": "No supported patch operation in request body"}), 400


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>", methods=["PUT"])
@login_required
def api_update_saved_viewpoint(vp_id):
    """Save/update a viewpoint — replaces element positions and relationship waypoints.

    JSON Body:
        name (str): Optional new name.
        description (str): Optional description.
        elements (list): Array of {element_id, x, y, width?, height?, rendering_mode?}.
        relationships (list): Array of {relationship_id, waypoints?, routing_style?}.
    """
    import json as _json

    from app.models.archimate_core import (
        SavedDiagram, SavedDiagramElement, SavedDiagramRelationship,
    )

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return jsonify({"error": "Diagram not found"}), 404

    # CMP-025: RBAC check
    _check_solution_access(vp.solution_id)

    data = request.get_json(silent=True) or {}

    if "name" in data:
        clean_name, _ = _validate_string(data["name"], "name", max_length=255)
        vp.name = clean_name or vp.name
    if "description" in data:
        clean_desc, _ = _validate_string(data["description"], "description", max_length=5000)
        vp.description = clean_desc

    # Materialize imported canvas items (string source ids) into real model rows
    # before the FK inserts below (same as the create route).
    element_id_map, relationship_id_map = _materialize_canvas_items(data)

    if "elements" in data:
        SavedDiagramElement.query.filter_by(diagram_id=vp.id).delete()
        for el_data in (data["elements"] or []):
            el_id = el_data.get("element_id")
            if not el_id:
                continue
            ve = SavedDiagramElement(
                diagram_id=vp.id,
                element_id=el_id,
                position_x=el_data.get("x", 0),
                position_y=el_data.get("y", 0),
                width=el_data.get("width", 180),
                height=el_data.get("height", 64),
                rendering_mode=el_data.get("rendering_mode", "black_box"),
            )
            db.session.add(ve)

    if "relationships" in data:
        SavedDiagramRelationship.query.filter_by(diagram_id=vp.id).delete()
        for rel_data in (data["relationships"] or []):
            rel_id = rel_data.get("relationship_id")
            if not rel_id:
                continue
            waypoints = rel_data.get("waypoints")
            vr = SavedDiagramRelationship(
                diagram_id=vp.id,
                relationship_id=rel_id,
                waypoints_json=_json.dumps(waypoints) if waypoints else None,
                routing_style=rel_data.get("routing_style", "manhattan"),
            )
            db.session.add(vr)

    db.session.commit()

    body = {
        "id": vp.id,
        "name": vp.name,
        "updated": True,
        "element_count": vp.positions.count(),
        "relationship_count": vp.rel_positions.count(),
    }
    if element_id_map or relationship_id_map:
        body["element_id_map"] = element_id_map
        body["relationship_id_map"] = relationship_id_map
    return jsonify(body)


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>", methods=["DELETE"])
@login_required
def api_delete_saved_viewpoint(vp_id):
    """Delete a saved diagram and its junction records (elements remain in catalog)."""
    from app.models.archimate_core import SavedDiagram

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return jsonify({"error": "Diagram not found"}), 404

    # CMP-025: RBAC check
    _check_solution_access(vp.solution_id)

    db.session.delete(vp)
    db.session.commit()

    return "", 204


# GAP-CMP-007: Submit viewpoint for ARB review
@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>/submit-review", methods=["POST"])
@login_required
def api_submit_viewpoint_review(vp_id):
    """Submit a viewpoint/diagram for ARB review."""
    from datetime import datetime as _dt
    from app.models.archimate_core import SavedDiagram

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return api_error("Viewpoint not found", 404)

    # CMP-025: RBAC check
    _check_solution_access(vp.solution_id)

    # Store review status in description metadata since SavedDiagram has no properties column
    # Use a simple convention: prepend review status marker
    try:
        from app.models.architecture_review_board import ARBAuditLog
        audit = ARBAuditLog(
            action="viewpoint_submitted_for_review",
            entity_type="SavedDiagram",
            entity_id=vp_id,
            details=json.dumps({
                "viewpoint_name": vp.name,
                "submitted_at": _dt.utcnow().isoformat(),
            }),
            user_id=current_user.id if hasattr(current_user, "id") else None,
        )
        db.session.add(audit)
        db.session.commit()
    except Exception:
        db.session.rollback()

    return api_success({"status": "submitted"})


# ── Element Search API (CMP-059: stable endpoint for composer element search) ──

@archimate_bp.route("/api/elements/search", methods=["GET"])
@login_required
def api_elements_search():
    """Search ArchiMate elements by name, type, and layer for the composer search panel.

    Query params:
        q        - name substring filter (case-insensitive)
        type     - exact element type filter (e.g. ApplicationComponent)
        layer    - layer filter (case-insensitive, e.g. application)
        limit    - max results (default 30, max 200)
        solution_id - restrict to elements linked to this solution
    """
    from app.models.archimate_core import ArchiMateElement

    query = ArchiMateElement.query

    elem_type = request.args.get("type", "").strip()
    if elem_type:
        query = query.filter(ArchiMateElement.type == elem_type)

    layer = request.args.get("layer", "").strip()
    if layer:
        query = query.filter(ArchiMateElement.layer.ilike(layer))

    search_q = request.args.get("q", "").strip()
    if search_q:
        query = query.filter(ArchiMateElement.name.ilike(f"%{search_q}%"))

    solution_id = request.args.get("solution_id", type=int)
    if solution_id:
        from app.models.solution_models import SolutionArchiMateElement
        linked_ids = db.session.query(SolutionArchiMateElement.element_id).filter_by(
            solution_id=solution_id
        ).subquery()
        query = query.filter(ArchiMateElement.id.in_(linked_ids))

    try:
        limit = min(int(request.args.get("limit", 30)), 200)
    except (ValueError, TypeError):
        limit = 30

    elements = query.order_by(ArchiMateElement.name).limit(limit).all()

    # ARCH-002: Include relationship + solution counts
    el_ids = [el.id for el in elements]
    rel_map = {}
    sol_map = {}
    if el_ids:
        try:
            from app.models.archimate_core import ArchiMateRelationship
            from sqlalchemy import func
            # Source counts
            src_rows = db.session.query(
                ArchiMateRelationship.source_id, func.count(ArchiMateRelationship.id)
            ).filter(ArchiMateRelationship.source_id.in_(el_ids)).group_by(ArchiMateRelationship.source_id).all()
            for eid, cnt in src_rows:
                rel_map[eid] = rel_map.get(eid, 0) + cnt
            # Target counts
            tgt_rows = db.session.query(
                ArchiMateRelationship.target_id, func.count(ArchiMateRelationship.id)
            ).filter(ArchiMateRelationship.target_id.in_(el_ids)).group_by(ArchiMateRelationship.target_id).all()
            for eid, cnt in tgt_rows:
                rel_map[eid] = rel_map.get(eid, 0) + cnt
        except Exception as exc:
            current_app.logger.debug("ARCH-002: relationship count query failed: %s", exc)
        try:
            from app.models.solution_archimate_element import SolutionArchiMateElement
            from sqlalchemy import func as fn
            sol_rows = db.session.query(
                SolutionArchiMateElement.element_id, fn.count(SolutionArchiMateElement.id)
            ).filter(SolutionArchiMateElement.element_id.in_(el_ids)).group_by(SolutionArchiMateElement.element_id).all()
            for eid, cnt in sol_rows:
                sol_map[eid] = cnt
        except Exception as exc:
            current_app.logger.debug("ARCH-002: solution count query failed: %s", exc)

    return jsonify({
        "data": [
            {
                "id": el.id,
                "name": el.name,
                "type": el.type,
                "layer": el.layer or "",
                "description": el.description or "",
                "relationship_count": rel_map.get(el.id, 0),
                "solution_count": sol_map.get(el.id, 0),
            }
            for el in elements
        ]
    })


# ── Element Detail API (for composer detail panel) ───────────────────────────

@archimate_bp.route("/api/elements/<int:element_id>/detail", methods=["GET"])
@login_required
def api_element_detail(element_id):
    """Return enriched element detail for the composer detail panel.

    Includes description, relationship count, viewpoint count, solution count,
    scope, and building block type.
    """
    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship

    el = db.session.get(ArchiMateElement, element_id)
    if not el:
        return jsonify({"error": "Element not found"}), 404

    rel_count = ArchiMateRelationship.query.filter(
        db.or_(
            ArchiMateRelationship.source_id == element_id,
            ArchiMateRelationship.target_id == element_id,
        )
    ).count()

    # tenant-filtered: scoped via parent FK (element_id)
    sol_count = db.session.execute(  # tenant-filtered: scoped via parent FK (element_id)
        db.text("SELECT COUNT(*) FROM solution_archimate_elements WHERE element_id = :eid"),  # tenant-filtered
        {"eid": element_id},
    ).scalar() or 0

    # REQ-CMP-001: Enrich with linked solutions, governance, capabilities
    linked_solutions = []
    try:
        sol_rows = db.session.execute(  # tenant-filtered: scoped via parent FK (element_id)
            db.text(  # tenant-filtered
                "SELECT sae.solution_id, s.name, s.adm_phase, sae.element_role "
                "FROM solution_archimate_elements sae "
                "JOIN solutions s ON s.id = sae.solution_id "
                "WHERE sae.element_id = :eid "
                "ORDER BY s.name "
                "LIMIT 20"
            ),
            {"eid": element_id},
        ).fetchall()
        for row in sol_rows:
            linked_solutions.append({
                "id": row[0], "name": row[1],
                "adm_phase": row[2] or "", "role": row[3] or "primary",
            })
    except Exception as exc:  # noqa: BLE001
        current_app.logger.debug("Solution link query failed for element %s: %s", element_id, exc)

    # Linked capabilities
    linked_capabilities = []
    try:
        cap_rows = db.session.execute(  # tenant-filtered: scoped via parent FK (element_id join)
            db.text(  # tenant-filtered
                "SELECT DISTINCT bc.id, bc.name, bc.level "
                "FROM business_capabilities bc "
                "JOIN capability_archimate_elements cae ON cae.capability_id = bc.id "
                "WHERE cae.archimate_element_id = :eid "
                "ORDER BY bc.name "
                "LIMIT 20"
            ),
            {"eid": element_id},
        ).fetchall()
        for row in cap_rows:
            linked_capabilities.append({"id": row[0], "name": row[1], "level": row[2]})
    except Exception as exc:  # noqa: BLE001
        current_app.logger.debug("Capability link query failed for element %s: %s", element_id, exc)

    # Connected elements (relationships with names)
    connected = []
    try:
        from app.models.archimate_core import ArchiMateRelationship as AMR
        rels = AMR.query.filter(
            db.or_(AMR.source_id == element_id, AMR.target_id == element_id)
        ).limit(30).all()
        for r in rels:
            other_id = r.target_id if r.source_id == element_id else r.source_id
            direction = "outgoing" if r.source_id == element_id else "incoming"
            other = db.session.get(ArchiMateElement, other_id)
            if other:
                connected.append({
                    "id": other.id, "name": other.name, "type": other.type,
                    "rel_type": r.type, "direction": direction,
                })
    except Exception as exc:  # noqa: BLE001
        current_app.logger.debug("Connected elements query failed for element %s: %s", element_id, exc)

    # GAP-CMP-008: Requirements linked via realization
    linked_requirements = []
    try:
        realization_rels = ArchiMateRelationship.query.filter_by(
            target_id=element_id, type="realization"
        ).all()
        for rel in realization_rels:
            source = db.session.get(ArchiMateElement, rel.source_id)
            if source and source.type == "Requirement":
                linked_requirements.append({
                    "id": source.id,
                    "name": source.name,
                    "type": source.type,
                })
    except Exception as exc:  # noqa: BLE001
        current_app.logger.debug("Requirements link query failed for element %s: %s", element_id, exc)

    # GAP-CMP-009: Custom properties (data classification, PII, etc.)
    cp = {}
    try:
        cp = getattr(el, "custom_properties", None) or {}
    except Exception:  # noqa: BLE001
        current_app.logger.debug("CMP-detail: custom_properties access failed for element %s", element_id)

    # GAP-INT-007: Interface metadata for ApplicationInterface elements
    interface_metadata = None
    if el.type == "ApplicationInterface":
        try:
            from app.models.integration_metadata import ApplicationInterfaceMetadata
            meta = ApplicationInterfaceMetadata.query.filter_by(archimate_element_id=element_id).first()
            if meta:
                interface_metadata = meta.to_dict()
        except Exception:  # noqa: BLE001
            current_app.logger.debug("CMP-detail: interface metadata lookup failed for element %s", element_id)

    return jsonify({
        "id": el.id,
        "name": el.name,
        "type": el.type,
        "layer": el.layer,
        "description": el.description or "",
        "scope": getattr(el, "scope", "") or "",
        "building_block_type": getattr(el, "building_block_type", "") or "",
        "relationship_count": rel_count,
        "solution_count": sol_count,
        "viewpoint_count": 0,
        "linked_solutions": linked_solutions,
        "linked_capabilities": linked_capabilities,
        "connected_elements": connected,
        "linked_requirements": linked_requirements,
        "custom_properties": cp,
        "interface_metadata": interface_metadata,
    })


# ── Element Usage API (for reuse detection) ──────────────────────────────────

@archimate_bp.route("/api/elements/<int:element_id>/usage", methods=["GET"])
@login_required
def api_element_usage(element_id):
    """Return usage counts for an element across the platform.

    Used by the composer reuse detection feature (CMP-009) to show
    how widely an element is referenced before creating a duplicate.
    """
    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship, SavedDiagramElement

    el = db.session.get(ArchiMateElement, element_id)
    if not el:
        return jsonify({"error": "Element not found"}), 404

    relationship_count = ArchiMateRelationship.query.filter(
        db.or_(
            ArchiMateRelationship.source_id == element_id,
            ArchiMateRelationship.target_id == element_id,
        )
    ).count()

    solution_count = db.session.execute(  # tenant-filtered: scoped via parent FK (element_id)
        db.text("SELECT COUNT(*) FROM solution_archimate_elements WHERE element_id = :eid"),  # tenant-filtered
        {"eid": element_id},
    ).scalar() or 0

    viewpoint_count = SavedDiagramElement.query.filter_by(element_id=element_id).count()

    return jsonify({
        "viewpoint_count": viewpoint_count,
        "solution_count": solution_count,
        "relationship_count": relationship_count,
    })


# ── Viewpoint Export API (CMP-014) ───────────────────────────────────────────

@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>/export", methods=["GET"])
@login_required
def api_export_saved_viewpoint(vp_id):
    """Export a saved viewpoint as ArchiMate Exchange Format XML.

    Query Parameters:
        format (str): Export format. Currently only 'archimate_exchange' is supported.

    Returns:
        application/xml with ArchiMate Open Exchange Format content.
    """
    fmt = request.args.get("format", "archimate_exchange")
    _supported = {"archimate_exchange", "mermaid", "lucid", "archi"}
    if fmt not in _supported:
        return jsonify({"error": f"Unsupported format: {fmt}"}), 400

    try:
        if fmt == "archimate_exchange":
            from app.services.archimate_export_service import export_saved_viewpoint
            body = export_saved_viewpoint(vp_id)
            mimetype, filename = "application/xml", f"viewpoint_{vp_id}.xml"
        elif fmt == "mermaid":
            from app.services.composer_export_formats import export_saved_viewpoint_mermaid
            body = export_saved_viewpoint_mermaid(vp_id)
            mimetype, filename = "text/plain; charset=utf-8", f"viewpoint_{vp_id}.mmd"
        elif fmt == "lucid":
            from app.services.composer_export_formats import export_saved_viewpoint_lucid
            body = export_saved_viewpoint_lucid(vp_id)  # bytes (.lucid ZIP)
            mimetype, filename = "application/octet-stream", f"viewpoint_{vp_id}.lucid"
        else:  # archi
            from app.services.composer_export_formats import export_saved_viewpoint_archi
            body = export_saved_viewpoint_archi(vp_id)
            mimetype, filename = "application/xml", f"viewpoint_{vp_id}.archimate"

        return Response(
            body,
            status=200,
            mimetype=mimetype,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404
    except Exception as exc:  # noqa: BLE001
        current_app.logger.error(f"[CMP-014] export_saved_viewpoint ({fmt}) failed: {exc}")
        return jsonify({"error": "Export failed"}), 500


@archimate_bp.route("/api/composer/export", methods=["POST"])
@login_required
def api_export_live_composer():
    """Export the CURRENT (possibly unsaved) composer canvas in an interchange
    format — so a user can import/edit/export without saving first. The client
    posts the live graph; the same renderers as the saved-viewpoint export run.

    Body: {format, viewpoint_name, elements:[{id,name,type,layer,x,y,w,h}],
           relationships:[{id,source_id,target_id,type,label}]}
    """
    data = request.get_json(silent=True) or {}
    fmt = data.get("format", "archimate_exchange")
    if fmt not in {"archimate_exchange", "mermaid", "lucid", "archi"}:
        return jsonify({"error": f"Unsupported format: {fmt}"}), 400

    elements = data.get("elements") or []
    relationships = data.get("relationships") or []
    if not isinstance(elements, list) or not isinstance(relationships, list):
        return jsonify({"error": "elements and relationships must be lists"}), 400
    if len(elements) > 5000 or len(relationships) > 10000:
        return jsonify({"error": "Diagram too large to export"}), 413
    if not elements:
        return jsonify({"error": "Nothing on the canvas to export"}), 400

    vp = {
        "viewpoint_name": (data.get("viewpoint_name") or "Composer Diagram")[:200],
        "phase_name": "Architecture",
        "elements": elements,
        "relationships": relationships,
    }
    try:
        from app.services.composer_export_formats import render_viewpoint

        body, mimetype, ext = render_viewpoint(vp, fmt)
        return Response(
            body,
            status=200,
            mimetype=mimetype,
            headers={"Content-Disposition": f"attachment; filename=diagram.{ext}"},
        )
    except Exception as exc:  # noqa: BLE001
        current_app.logger.error(f"[CMP-014] live composer export ({fmt}) failed: {exc}")
        return jsonify({"error": "Export failed"}), 500


# ── CMP-015 — Version Snapshots ──────────────────────────────────────────────

# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>/snapshots", methods=["POST"])
@login_required
def api_create_snapshot(vp_id):
    """Create a version snapshot of a saved viewpoint.

    Serializes the current viewpoint state (elements + relationships + positions)
    into a JSON blob and stores it as a named snapshot.

    JSON Body:
        name (str): Required snapshot name.
    """
    import json as _json

    from app.models.archimate_core import (
        ArchiMateRelationship, SavedDiagram,
    )
    from app.models.archimate_viewpoint import ArchimateViewpointSnapshot
    db.create_all()  # migration-exempt: creates archimate_viewpoint_snapshots if not yet present

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return jsonify({"error": "Diagram not found"}), 404

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400

    # Serialize full state
    positions = vp.positions.all()
    element_ids = [p.element_id for p in positions]

    from app.models.archimate_core import ArchiMateElement

    elements = []
    if element_ids:
        els = ArchiMateElement.query.filter(ArchiMateElement.id.in_(element_ids)).all()
        pos_map = {p.element_id: p for p in positions}
        for el in els:
            p = pos_map.get(el.id)
            elements.append({
                "id": el.id,
                "name": el.name,
                "type": el.type,
                "layer": el.layer,
                "x": p.position_x if p else 0,
                "y": p.position_y if p else 0,
                "width": p.width if p else 180,
                "height": p.height if p else 64,
                "rendering_mode": p.rendering_mode if p else "black_box",
            })

    rels = []
    if element_ids:
        rel_rows = ArchiMateRelationship.query.filter(
            ArchiMateRelationship.source_id.in_(element_ids),
            ArchiMateRelationship.target_id.in_(element_ids),
        ).all()
        rel_positions = vp.rel_positions.all()
        rel_pos_map = {rp.relationship_id: rp for rp in rel_positions}
        for r in rel_rows:
            rp = rel_pos_map.get(r.id)
            waypoints_raw = rp.waypoints_json if rp else None
            rels.append({
                "id": r.id,
                "source_id": r.source_id,
                "target_id": r.target_id,
                "type": r.type,
                "name": getattr(r, "name", None),
                "waypoints": _json.loads(waypoints_raw) if waypoints_raw else None,
                "routing_style": rp.routing_style if rp else "manhattan",
            })

    snapshot_data = {
        "viewpoint_name": vp.name,
        "viewpoint_type": vp.viewpoint_type,
        "solution_id": vp.solution_id,
        "elements": elements,
        "relationships": rels,
    }

    snapshot = ArchimateViewpointSnapshot(
        viewpoint_id=vp_id,
        name=name,
        snapshot_json=_json.dumps(snapshot_data),
    )
    db.session.add(snapshot)
    db.session.commit()

    return jsonify({
        "id": snapshot.id,
        "name": snapshot.name,
        "created_at": snapshot.created_at.isoformat() if snapshot.created_at else None,
    }), 201


@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>/snapshots", methods=["GET"])
@login_required
def api_list_snapshots(vp_id):
    """List all snapshots for a saved viewpoint, newest first."""
    from app.models.archimate_viewpoint import ArchimateViewpointSnapshot

    snapshots = (
        ArchimateViewpointSnapshot.query
        .filter_by(viewpoint_id=vp_id)
        .order_by(ArchimateViewpointSnapshot.created_at.desc())
        .all()
    )

    return jsonify({
        "snapshots": [
            {
                "id": s.id,
                "name": s.name,
                "created_at": s.created_at.isoformat() if s.created_at else None,
            }
            for s in snapshots
        ],
    })


@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>/snapshots/<int:sid>", methods=["GET"])
@login_required
def api_get_snapshot(vp_id, sid):
    """Get the full snapshot JSON for a specific snapshot."""
    import json as _json

    from app.models.archimate_viewpoint import ArchimateViewpointSnapshot

    snapshot = db.session.get(ArchimateViewpointSnapshot, sid)
    if not snapshot or snapshot.viewpoint_id != vp_id:
        return jsonify({"error": "Snapshot not found"}), 404

    return jsonify({
        "id": snapshot.id,
        "name": snapshot.name,
        "created_at": snapshot.created_at.isoformat() if snapshot.created_at else None,
        "data": _json.loads(snapshot.snapshot_json),
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/saved-viewpoints/<int:vp_id>/snapshots/<int:sid>/restore", methods=["POST"])
@login_required
def api_restore_snapshot(vp_id, sid):
    """Restore a snapshot — copy snapshot element/relationship data back to the active viewpoint.

    This replaces the current diagram positions and relationship waypoints
    with the data captured in the snapshot.
    """
    import json as _json

    from app.models.archimate_core import (
        SavedDiagram, SavedDiagramElement, SavedDiagramRelationship,
    )
    from app.models.archimate_viewpoint import ArchimateViewpointSnapshot

    snapshot = db.session.get(ArchimateViewpointSnapshot, sid)
    if not snapshot or snapshot.viewpoint_id != vp_id:
        return jsonify({"error": "Snapshot not found"}), 404

    vp = db.session.get(SavedDiagram, vp_id)
    if not vp:
        return jsonify({"error": "Diagram not found"}), 404

    snap_data = _json.loads(snapshot.snapshot_json)

    # Replace element positions
    SavedDiagramElement.query.filter_by(diagram_id=vp_id).delete()
    for el in (snap_data.get("elements") or []):
        ve = SavedDiagramElement(
            diagram_id=vp_id,
            element_id=el["id"],
            position_x=el.get("x", 0),
            position_y=el.get("y", 0),
            width=el.get("width", 180),
            height=el.get("height", 64),
            rendering_mode=el.get("rendering_mode", "black_box"),
        )
        db.session.add(ve)

    # Replace relationship waypoints
    SavedDiagramRelationship.query.filter_by(diagram_id=vp_id).delete()
    for rel in (snap_data.get("relationships") or []):
        waypoints = rel.get("waypoints")
        vr = SavedDiagramRelationship(
            diagram_id=vp_id,
            relationship_id=rel["id"],
            waypoints_json=_json.dumps(waypoints) if waypoints else None,
            routing_style=rel.get("routing_style", "manhattan"),
        )
        db.session.add(vr)

    db.session.commit()

    return jsonify({
        "restored": True,
        "snapshot_name": snapshot.name,
        "element_count": len(snap_data.get("elements") or []),
        "relationship_count": len(snap_data.get("relationships") or []),
    })


# ── CMP-017 — AI Phase-contextual Generation ─────────────────────────────

# ArchiMate element types per TOGAF ADM phase (used by mock generator)
_PHASE_ELEMENT_MAP = {
    "A": [
        ("Stakeholder", "motivation"),
        ("Driver", "motivation"),
        ("Goal", "motivation"),
        ("Principle", "motivation"),
    ],
    "B": [
        ("BusinessProcess", "business"),
        ("BusinessRole", "business"),
        ("BusinessService", "business"),
        ("BusinessActor", "business"),
        ("BusinessObject", "business"),
    ],
    "C": [
        ("ApplicationComponent", "application"),
        ("ApplicationService", "application"),
        ("ApplicationInterface", "application"),
        ("DataObject", "application"),
    ],
    "D": [
        ("Node", "technology"),
        ("SystemSoftware", "technology"),
        ("Device", "technology"),
        ("Artifact", "technology"),
        ("TechnologyService", "technology"),
    ],
}

# Default relationships per phase (mock)
_PHASE_REL_MAP = {
    "A": "influence",
    "B": "serving",
    "C": "serving",
    "D": "composition",
}

# All valid ArchiMate 3.2 element types (for validation)
_ALL_ARCHIMATE_TYPES = {
    "BusinessActor", "BusinessRole", "BusinessProcess", "BusinessFunction",
    "BusinessService", "BusinessObject", "BusinessEvent", "Contract", "Product",
    "ApplicationComponent", "ApplicationService", "ApplicationFunction",
    "ApplicationInterface", "ApplicationProcess", "DataObject",
    "Node", "Device", "SystemSoftware", "TechnologyService", "Artifact",
    "CommunicationNetwork", "Path", "TechnologyFunction", "TechnologyProcess",
    "TechnologyInterface", "TechnologyEvent",
    "Stakeholder", "Driver", "Goal", "Requirement", "Constraint", "Principle",
    "Assessment", "Value", "Meaning", "Outcome",
    "Capability", "Resource", "CourseOfAction", "ValueStream",
    "WorkPackage", "Deliverable", "Plateau", "Gap",
    "AndJunction", "OrJunction", "Grouping", "Location",
}


def _extract_nouns(description):
    """Extract key noun phrases from a natural-language description.

    Uses simple heuristic: split on common stop words and punctuation,
    keep phrases of 1-3 words that look like proper names or concepts.
    """
    import re

    stop_words = {
        "a", "an", "the", "and", "or", "but", "in", "on", "at", "to", "for",
        "of", "with", "by", "from", "is", "are", "was", "were", "be", "been",
        "being", "have", "has", "had", "do", "does", "did", "will", "would",
        "could", "should", "may", "might", "shall", "can", "need", "must",
        "that", "this", "these", "those", "it", "its", "they", "them", "their",
        "we", "our", "i", "my", "you", "your", "he", "she", "his", "her",
        "which", "who", "whom", "what", "where", "when", "how", "if", "then",
        "than", "so", "as", "not", "no", "all", "each", "every", "both",
        "such", "into", "through", "about", "between", "after", "before",
        "during", "without", "within", "also", "very", "just", "only", "new",
        "via", "using", "used", "use",
    }

    # Split on punctuation and conjunctions
    fragments = re.split(r"[,;:.\-\(\)\[\]\"\'\/\\&\n]+", description)
    nouns = []
    for frag in fragments:
        words = frag.strip().split()
        # Filter stop words from start and end, keep middle
        cleaned = [w for w in words if w.lower() not in stop_words]
        if cleaned:
            noun = " ".join(cleaned)
            if 1 <= len(noun) <= 60:
                nouns.append(noun.strip())
    return nouns[:8]  # Cap at 8 noun phrases


def _auto_detect_phase(description: str) -> str:
    """Detect TOGAF ADM phase from description keywords."""
    desc_lower = description.lower()
    if any(kw in desc_lower for kw in [
        "stakeholder", "vision", "goal", "driver", "principle", "motivation",
    ]):
        return "A"
    if any(kw in desc_lower for kw in [
        "business", "process", "role", "actor", "service", "organization",
    ]):
        return "B"
    if any(kw in desc_lower for kw in [
        "application", "system", "software", "data", "interface", "api",
        "integration",
    ]):
        return "C"
    if any(kw in desc_lower for kw in [
        "technology", "infrastructure", "server", "network", "device",
        "cloud", "deploy",
    ]):
        return "D"
    return "C"


def _mock_generate_elements(description: str, phase: str, existing_names: set) -> tuple:
    """Fallback: extract noun phrases and pair with phase-specific element types."""
    nouns = _extract_nouns(description)
    type_pool = _PHASE_ELEMENT_MAP[phase]

    elements = []
    for i, noun in enumerate(nouns):
        el_type, el_layer = type_pool[i % len(type_pool)]
        el_name = noun.title() if len(noun) > 2 else noun
        elements.append({
            "name": el_name,
            "type": el_type,
            "layer": el_layer,
            "valid": el_type in _ALL_ARCHIMATE_TYPES,
            "duplicate": el_name.lower() in existing_names,
        })

    if not elements:
        prefix = description[:30].strip().title()
        for el_type, el_layer in type_pool[:3]:
            label = el_type
            for strip_word in ("Business", "Application", "Technology"):
                label = label.replace(strip_word, "")
            el_name = (prefix + " " + label).strip()
            elements.append({
                "name": el_name,
                "type": el_type,
                "layer": el_layer,
                "valid": True,
                "duplicate": el_name.lower() in existing_names,
            })

    rel_type = _PHASE_REL_MAP.get(phase, "association")
    relationships = [
        {
            "source_name": elements[i]["name"],
            "target_name": elements[i + 1]["name"],
            "type": rel_type,
            "valid": True,
            "warning": "",
        }
        for i in range(len(elements) - 1)
    ]
    return elements, relationships


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@archimate_bp.route("/api/composer/generate", methods=["POST"])
@login_required
def api_composer_generate():
    """Generate ArchiMate elements and relationships from a natural-language description.

    Attempts LLM-powered generation via ArchiMateLLMService first; falls back to
    heuristic noun-extraction (CMP-017 mock) when no LLM provider is configured.

    JSON Body:
        description (str): Required natural-language architecture description.
        phase (str): Optional TOGAF phase (A/B/C/D). Auto-detects if empty.
        viewpoint_type (str): Optional viewpoint context passed to LLM as context.
        solution_id (int): Optional solution scope (used for dedup).

    Returns JSON with generated elements and relationships for review.
    """
    data = request.get_json(silent=True) or {}
    description = (data.get("description") or "").strip()
    if not description:
        return jsonify({"error": "description is required"}), 400

    phase = (data.get("phase") or "").strip().upper() or _auto_detect_phase(description)
    viewpoint_type = (data.get("viewpoint_type") or "").strip()
    solution_id = data.get("solution_id")

    if phase not in _PHASE_ELEMENT_MAP:
        return jsonify({"error": "phase must be one of: A, B, C, D"}), 400

    # Build dedup set from existing solution elements
    existing_names: set = set()
    if solution_id:
        try:
            rows = db.session.execute(  # tenant-filtered: scoped via parent FK (solution_id)
                db.text(  # tenant-filtered
                    "SELECT ae.name FROM archimate_elements ae "
                    "JOIN solution_archimate_elements sae ON sae.element_id = ae.id "
                    "WHERE sae.solution_id = :sid"
                ),
                {"sid": solution_id},
            ).fetchall()
            existing_names = {(r.name or "").lower() for r in rows}
        except Exception:  # noqa: BLE001
            current_app.logger.warning("CMP-036: dedup query failed, skipping duplicate check")

    # ── LLM-powered generation ──────────────────────────────────────────────
    llm_used = False
    elements: list = []
    relationships: list = []
    try:
        context = f"TOGAF Phase {phase}. Viewpoint: {viewpoint_type}." if viewpoint_type else f"TOGAF Phase {phase}."
        model_data = _run_archimate_llm_generation(
            requirements=description,
            context=context,
            target_layer="complete",
        )
        raw_elements = (model_data or {}).get("elements") or []
        raw_rels = (model_data or {}).get("relationships") or []
        if raw_elements:
            for el in raw_elements:
                el_name = (el.get("name") or "").strip()
                el_type = (el.get("type") or "").strip()
                el_layer = (el.get("layer") or "").strip().lower()
                if not el_name or not el_type:
                    continue
                elements.append({
                    "name": el_name,
                    "type": el_type,
                    "layer": el_layer,
                    "valid": el_type in _ALL_ARCHIMATE_TYPES,
                    "duplicate": el_name.lower() in existing_names,
                })
            for rel in raw_rels:
                raw_rel_type = (rel.get("type") or "association").lower()
                normalised_rel_type = _normalize_rel_type(raw_rel_type)
                is_valid = normalised_rel_type in ARCHIMATE_RELATIONSHIP_TYPES
                relationships.append({
                    "source_name": rel.get("source_name", ""),
                    "target_name": rel.get("target_name", ""),
                    "type": normalised_rel_type,
                    "description": rel.get("description", ""),
                    "valid": is_valid,
                    "warning": "" if is_valid else f"'{raw_rel_type}' is not a valid ArchiMate 3.2 relationship type",
                })
            llm_used = True
    except Exception:  # noqa: BLE001
        current_app.logger.info("CMP-036: LLM unavailable or failed, using mock generation")

    # ── Heuristic fallback ───────────────────────────────────────────────────
    if not elements:
        elements, relationships = _mock_generate_elements(description, phase, existing_names)

    return jsonify({
        "phase": phase,
        "elements": elements,
        "relationships": relationships,
        "llm_used": llm_used,
    })


# ── Context-Aware Generation (Enterprise Context Assembler) ─────────────────

@archimate_bp.route("/api/composer/generate-contextual", methods=["POST"])
@login_required
# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
def api_composer_generate_contextual():
    """Generate ArchiMate elements using full enterprise portfolio context.

    Unlike the basic /generate endpoint, this assembles real portfolio data
    (applications, ArchiMate elements, vendors, capabilities, principles)
    and injects it into the LLM prompt so generated architecture references
    actual enterprise entities instead of fabricating generic ones.

    JSON Body:
        description (str): Required natural-language architecture description.
        phase (str): Optional TOGAF phase (A/B/C/D). Auto-detects if empty.
        viewpoint_type (str): Optional viewpoint context.
        business_domain (str): Optional domain filter to narrow context.
        solution_id (int): Optional — includes entities linked to this solution.
        options (dict): Optional — {reference_existing: bool, include_gaps: bool}.

    Returns JSON with categorised elements (existing/new/possible_duplicate),
    validated relationships, identified gaps, and context stats.
    """
    from app.services.enterprise_context_assembler import EnterpriseContextAssembler

    data = request.get_json(silent=True) or {}
    description = (data.get("description") or "").strip()
    if not description:
        return jsonify({"error": "description is required"}), 400

    phase = (data.get("phase") or "").strip().upper() or _auto_detect_phase(description)
    viewpoint_type = (data.get("viewpoint_type") or "").strip()
    business_domain = (data.get("business_domain") or "").strip() or None
    solution_id = data.get("solution_id")
    options = data.get("options") or {}

    if phase not in _PHASE_ELEMENT_MAP:
        return jsonify({"error": "phase must be one of: A, B, C, D"}), 400

    assembler = EnterpriseContextAssembler()

    # 1. Assemble enterprise context
    try:
        ctx = assembler.assemble_context(
            description=description,
            phase=phase,
            business_domain=business_domain,
            solution_id=solution_id,
        )
    except Exception:
        current_app.logger.exception("Context assembly failed")
        ctx = None

    # 2. Generate with LLM using enriched context
    llm_used = False
    llm_response = {}

    if ctx:
        enriched_prompt = assembler.build_generation_prompt(ctx)

        try:
            llm_response = _run_archimate_llm_generation(
                requirements=enriched_prompt,
                context=f"TOGAF Phase {phase}. Viewpoint: {viewpoint_type}." if viewpoint_type else f"TOGAF Phase {phase}.",
                target_layer="complete",
            ) or {}
            if llm_response.get("elements"):
                llm_used = True
        except Exception:
            current_app.logger.info(
                "Contextual generation: LLM unavailable, falling back to basic generation"
            )

    # 3. Fallback to basic generation if LLM failed
    if not llm_response.get("elements"):
        existing_names: set = set()
        if solution_id:
            try:
                rows = db.session.execute(  # tenant-filtered: scoped via parent FK (solution_id)
                    db.text(  # tenant-filtered
                        "SELECT ae.name FROM archimate_elements ae "
                        "JOIN solution_archimate_elements sae ON sae.element_id = ae.id "
                        "WHERE sae.solution_id = :sid"
                    ),
                    {"sid": solution_id},
                ).fetchall()
                existing_names = {(r.name or "").lower() for r in rows}
            except Exception as exc:
                current_app.logger.debug("Failed to load existing element names for solution %s: %s", solution_id, exc)

        elements, relationships = _mock_generate_elements(description, phase, existing_names)
        return jsonify({
            "phase": phase,
            "elements": [dict(e, category="new") for e in elements],
            "relationships": relationships,
            "gaps": [],
            "llm_used": False,
            "context_stats": ctx.stats if ctx else {},
        })

    # 4. Post-process: validate, deduplicate, categorise
    if ctx:
        result = assembler.post_process(llm_response, ctx)
        return jsonify({
            "phase": phase,
            "elements": result.elements,
            "relationships": result.relationships,
            "gaps": result.gaps,
            "llm_used": llm_used,
            "context_stats": result.context_stats,
            "rationale": llm_response.get("rationale", ""),
        })

    # Shouldn't reach here, but return raw if no context
    return jsonify({
        "phase": phase,
        "elements": llm_response.get("elements", []),
        "relationships": llm_response.get("relationships", []),
        "gaps": [],
        "llm_used": llm_used,
        "context_stats": {},
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/context-preview", methods=["POST"])
@login_required
def api_composer_context_preview():
    """Quick context preview for the generation dialog.

    Returns counts of matching portfolio entities so the architect can see
    what enterprise data the AI will use before generating.

    JSON Body:
        description (str): Required description text.
        phase (str): Optional TOGAF phase.
        business_domain (str): Optional domain filter.

    Returns JSON with query_terms, counts per entity type, and domain.
    """
    from app.services.enterprise_context_assembler import EnterpriseContextAssembler

    data = request.get_json(silent=True) or {}
    description = (data.get("description") or "").strip()
    if not description:
        return jsonify({"error": "description is required"}), 400

    phase = (data.get("phase") or "").strip().upper() or "C"
    business_domain = (data.get("business_domain") or "").strip() or None

    assembler = EnterpriseContextAssembler()
    preview = assembler.get_context_preview(
        description=description,
        phase=phase,
        business_domain=business_domain,
    )
    return jsonify(preview)


# ── CMP-015 — Diagram Templates ─────────────────────────────────────────────

# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/templates", methods=["POST"])
@login_required
def api_create_template():
    """Create a diagram template from the current viewpoint.

    Templates store layout structure (element types + positions) without
    specific element IDs, so the template can be instantiated to create
    new diagrams with the same structure.

    JSON Body:
        name (str): Required template name.
        viewpoint_type (str): Optional viewpoint category.
        template_json (str/dict): Layout structure.
    """
    import json as _json

    from flask_login import current_user

    from app.models.archimate_viewpoint import ArchimateViewpointTemplate

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400

    template_json = data.get("template_json")
    if not template_json:
        return jsonify({"error": "template_json is required"}), 400

    if isinstance(template_json, dict) or isinstance(template_json, list):
        template_json = _json.dumps(template_json)

    template = ArchimateViewpointTemplate(
        name=name,
        viewpoint_type=data.get("viewpoint_type"),
        template_json=template_json,
        created_by=current_user.id if hasattr(current_user, "id") else None,
    )
    db.session.add(template)
    db.session.commit()

    return jsonify({
        "id": template.id,
        "name": template.name,
        "viewpoint_type": template.viewpoint_type,
        "created_at": template.created_at.isoformat() if template.created_at else None,
    }), 201


@archimate_bp.route("/api/templates", methods=["GET"])
@login_required
def api_list_templates():
    """List all diagram templates."""
    try:
        from app.models.archimate_viewpoint import ArchimateViewpointTemplate
        templates = (
            ArchimateViewpointTemplate.query
            .order_by(ArchimateViewpointTemplate.created_at.desc())
            .all()
        )
    except Exception:
        templates = []

    return jsonify({
        "templates": [
            {
                "id": t.id,
                "name": t.name,
                "viewpoint_type": t.viewpoint_type,
                "created_at": t.created_at.isoformat() if t.created_at else None,
            }
            for t in templates
        ],
    })


# ── CMP2-006 — Portfolio-generated diagram templates ─────────────────────────


@archimate_bp.route("/api/composer/portfolio-templates", methods=["GET"])
@login_required
def api_portfolio_templates():
    """Return diagram templates auto-generated from real portfolio data.

    Queries ApplicationComponent grouped by business_domain.  For each domain
    with 3 or more catalogued applications, a ready-to-load template definition
    is produced containing real element names, ArchiMate types, and suggested
    relationships (ServingRelationship between apps that share a domain).

    Query params:
        min_apps (int): Minimum applications per domain to generate a template
                        (default 3, min 2).

    Returns JSON:
        { "portfolio_templates": [ { id, name, description, domain,
          app_count, element_count, is_portfolio, elements, relationships } ] }
    """
    from collections import defaultdict

    from app.models.application_portfolio import ApplicationComponent

    min_apps = max(int(request.args.get("min_apps", 3)), 2)

    # Fetch all applications that have a non-empty business_domain
    try:
        apps = (
            ApplicationComponent.query
            .filter(
                ApplicationComponent.business_domain.isnot(None),
                ApplicationComponent.business_domain != "",
            )
            .order_by(ApplicationComponent.business_domain, ApplicationComponent.name)
            .all()
        )
    except Exception:  # noqa: BLE001
        apps = []

    # Group by business_domain
    domain_groups = defaultdict(list)
    for app in apps:
        domain = (app.business_domain or "").strip()
        if domain:
            domain_groups[domain].append(app)

    # ArchiMate type mapping based on application_type / deployment_model
    def _archimate_type(app_obj):
        app_type = (getattr(app_obj, "application_type", "") or "").lower()
        if app_type in ("saas", "cloud", "web"):
            return "ApplicationService"
        if app_type in ("mobile",):
            return "ApplicationComponent"
        deploy = (getattr(app_obj, "deployment_model", "") or "").lower()
        if deploy in ("saas", "cloud"):
            return "ApplicationService"
        return "ApplicationComponent"

    # Generate one template per qualifying domain
    templates = []
    for domain, domain_apps in sorted(domain_groups.items()):
        if len(domain_apps) < min_apps:
            continue

        # Cap at 12 apps per template to keep diagrams readable
        capped_apps = domain_apps[:12]

        elements = []
        spacing_x = 240
        spacing_y = 180
        cols = 3
        for idx, app_obj in enumerate(capped_apps):
            col = idx % cols
            row = idx // cols
            el_type = _archimate_type(app_obj)
            elements.append({
                "id": "portfolio_{}_{}".format(domain.lower().replace(" ", "_"), idx),
                "name": app_obj.name,
                "type": el_type,
                "layer": "application",
                "x": 40 + col * spacing_x,
                "y": 40 + row * spacing_y,
                "app_id": app_obj.id,
                "criticality": getattr(app_obj, "criticality", None) or getattr(app_obj, "business_criticality", None),
            })

        # Auto-detect relationships: apps in same domain get a CompositionRelationship
        # to a domain grouping node, plus ServingRelationships between adjacent apps
        relationships = []
        for idx in range(len(elements) - 1):
            relationships.append({
                "source": elements[idx]["id"],
                "target": elements[idx + 1]["id"],
                "type": "AssociationRelationship",
                "name": "same domain",
            })

        # Determine template style name
        domain_clean = domain.strip()
        app_count = len(capped_apps)
        total_in_domain = len(domain_apps)

        tpl = {
            "id": "portfolio-{}".format(domain.lower().replace(" ", "_").replace("/", "-")),
            "name": "{} Application Landscape".format(domain_clean),
            "description": "{} catalogued applications in the {} domain".format(
                total_in_domain, domain_clean
            ),
            "domain": domain_clean,
            "app_count": total_in_domain,
            "element_count": app_count,
            "is_portfolio": True,
            "elements": elements,
            "relationships": relationships,
        }
        templates.append(tpl)

    return jsonify({"portfolio_templates": templates})


# ── CMP-018 — Validation API ─────────────────────────────────────────────────

# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/validate", methods=["POST"])
@login_required
def api_composer_validate():
    """Validate a composer canvas for ArchiMate correctness.

    Runs three rule-based validation layers (no LLM calls):
      A. Relationship validity — checks against ArchiMate 3.2 matrix
      B. Phase completeness — checks required element types per ADM phase
      C. Naming quality — flags generic or duplicate-intent names

    JSON Body:
        elements (list): [{id, type, layer, name}]
        relationships (list): [{source_type, target_type, rel_type, source_name, target_name}]
        phase (str): ADM phase (e.g. 'A', 'B', 'C', 'D')
        viewpoint_type (str): Active viewpoint type

    Returns JSON with passed, warnings, errors arrays.
    """
    from app.services.archimate_validity_service import (
        ArchimateValidityService,
        _layer as validity_layer,
        _normalize_type,
    )

    data = request.get_json(silent=True) or {}
    elements = data.get("elements") or []
    relationships = data.get("relationships") or []
    phase = (data.get("phase") or "").strip().upper()

    passed = []
    warnings = []
    errors = []

    svc = ArchimateValidityService()

    # ── A. Relationship validity checks ──────────────────────────────────

    invalid_rels = []
    cross_layer_composition = []
    practitioner_warnings = []

    for rel in relationships:
        src_type = rel.get("source_type") or ""
        tgt_type = rel.get("target_type") or ""
        rel_type = rel.get("rel_type") or ""
        src_name = rel.get("source_name") or src_type
        tgt_name = rel.get("target_name") or tgt_type

        if not src_type or not tgt_type or not rel_type:
            continue

        if not svc.is_valid(src_type, tgt_type, rel_type):
            invalid_rels.append({
                "check": "invalid_relationship",
                "message": (
                    "Invalid " + rel_type + " from " + src_name
                    + " (" + src_type + ") to " + tgt_name
                    + " (" + tgt_type + ")"
                ),
                "element_ids": [],
                "fix_suggestion": (
                    "Check ArchiMate 3.2 validity matrix. Consider using "
                    "a different relationship type or adding an intermediary element."
                ),
            })

        src_layer = validity_layer(src_type)
        tgt_layer = validity_layer(tgt_type)
        if rel_type == "composition" and src_layer != tgt_layer:
            cross_layer_composition.append({
                "check": "cross_layer_composition",
                "message": (
                    "Composition from " + src_name + " (" + src_layer
                    + ") to " + tgt_name + " (" + tgt_layer
                    + ") crosses layers"
                ),
                "element_ids": [],
                "fix_suggestion": (
                    "Composition means ownership and must be within the same layer. "
                    "Use Aggregation or Realization instead."
                ),
            })

        pw = svc.get_practitioner_warnings(src_type, tgt_type, rel_type)
        for w in pw:
            practitioner_warnings.append({
                "check": "practitioner_warning",
                "message": w,
                "element_ids": [],
            })

    errors.extend(invalid_rels)
    errors.extend(cross_layer_composition)
    warnings.extend(practitioner_warnings)

    if not invalid_rels and not cross_layer_composition:
        passed.append({
            "check": "relationship_validity",
            "message": "All " + str(len(relationships)) + " relationships are valid per ArchiMate 3.2",
        })

    # ── A2. Orphan element check ─────────────────────────────────────────

    connected_names = set()
    for rel in relationships:
        connected_names.add(rel.get("source_name") or "")
        connected_names.add(rel.get("target_name") or "")

    orphan_ids = []
    orphan_names = []
    for el in elements:
        el_name = el.get("name") or ""
        el_type = el.get("type") or ""
        if el_type in ("AndJunction", "OrJunction", "Grouping", "Note"):
            continue
        if el_name not in connected_names:
            orphan_ids.append(str(el.get("id", "")))
            orphan_names.append(el_name or el_type)

    if orphan_names:
        warnings.append({
            "check": "orphan_elements",
            "message": (
                str(len(orphan_names)) + " orphan element(s) with no relationships: "
                + ", ".join(orphan_names[:5])
                + ("..." if len(orphan_names) > 5 else "")
            ),
            "element_ids": orphan_ids,
        })
    elif elements:
        passed.append({
            "check": "orphan_elements",
            "message": "All elements are connected by at least one relationship",
        })

    # ── B. Phase completeness checks ─────────────────────────────────────

    phase_requirements = {
        "A": {
            "required": ["Stakeholder", "Driver", "Goal"],
            "label": "Phase A (Architecture Vision)",
        },
        "B": {
            "required": ["BusinessProcess", "BusinessRole", "BusinessService"],
            "label": "Phase B (Business Architecture)",
        },
        "C": {
            "required": ["ApplicationComponent", "ApplicationService"],
            "label": "Phase C (Information Systems)",
        },
        "D": {
            "required": ["Node", "SystemSoftware", "Device"],
            "label": "Phase D (Technology Architecture)",
            "any_of": True,
        },
    }

    if phase and phase in phase_requirements:
        req = phase_requirements[phase]
        present_types = set()
        for el in elements:
            norm = _normalize_type(el.get("type") or "")
            present_types.add(norm)

        if req.get("any_of"):
            has_any = any(rt in present_types for rt in req["required"])
            if not has_any:
                warnings.append({
                    "check": "phase_completeness",
                    "message": (
                        req["label"] + ": needs at least one of "
                        + ", ".join(req["required"])
                    ),
                    "element_ids": [],
                })
            else:
                passed.append({
                    "check": "phase_completeness",
                    "message": req["label"] + ": required element types present",
                })
        else:
            missing = [rt for rt in req["required"] if rt not in present_types]
            if missing:
                warnings.append({
                    "check": "phase_completeness",
                    "message": (
                        req["label"] + ": missing required types: "
                        + ", ".join(missing)
                    ),
                    "element_ids": [],
                })
            else:
                passed.append({
                    "check": "phase_completeness",
                    "message": req["label"] + ": all required element types present",
                })

    # ── C. Naming quality checks ─────────────────────────────────────────

    generic_names = {
        "System", "Process", "Service", "Component", "Function",
        "Object", "Element", "Node", "Device", "Actor", "Role",
        "Interface", "Event", "Application", "Data", "Resource",
    }
    generic_flagged = []
    for el in elements:
        name = (el.get("name") or "").strip()
        if name in generic_names:
            generic_flagged.append({
                "id": str(el.get("id", "")),
                "name": name,
            })

    if generic_flagged:
        warnings.append({
            "check": "generic_names",
            "message": (
                str(len(generic_flagged)) + " element(s) have generic names: "
                + ", ".join(g["name"] for g in generic_flagged[:5])
                + ("..." if len(generic_flagged) > 5 else "")
            ),
            "element_ids": [g["id"] for g in generic_flagged],
        })
    elif elements:
        passed.append({
            "check": "generic_names",
            "message": "No generic element names detected",
        })

    # Duplicate-intent check (substring match)
    el_names_lower = []
    for el in elements:
        name = (el.get("name") or "").strip()
        if name:
            el_names_lower.append({
                "id": str(el.get("id", "")),
                "name": name,
                "lower": name.lower(),
            })

    duplicate_pairs = []
    for i in range(len(el_names_lower)):
        for j in range(i + 1, len(el_names_lower)):
            a = el_names_lower[i]
            b = el_names_lower[j]
            if len(a["lower"]) > 3 and len(b["lower"]) > 3:
                if a["lower"] in b["lower"] or b["lower"] in a["lower"]:
                    duplicate_pairs.append((a, b))

    if duplicate_pairs:
        pair_descriptions = []
        dup_ids = set()
        for a, b in duplicate_pairs[:3]:
            pair_descriptions.append('"' + a["name"] + '" / "' + b["name"] + '"')
            dup_ids.add(a["id"])
            dup_ids.add(b["id"])
        warnings.append({
            "check": "duplicate_intent",
            "message": (
                str(len(duplicate_pairs)) + " potential duplicate(s): "
                + "; ".join(pair_descriptions)
                + ("..." if len(duplicate_pairs) > 3 else "")
            ),
            "element_ids": list(dup_ids),
        })
    elif len(el_names_lower) > 1:
        passed.append({
            "check": "duplicate_intent",
            "message": "No duplicate-intent element names detected",
        })

    # ── Summary ──────────────────────────────────────────────────────────

    if not elements and not relationships:
        warnings.append({
            "check": "empty_canvas",
            "message": "Canvas is empty -- add elements and relationships to validate",
            "element_ids": [],
        })

    return jsonify({
        "passed": passed,
        "warnings": warnings,
        "errors": errors,
    })


# ── CMP-016 — AI Ambient Suggestion Mode ─────────────────────────────────────

# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/suggestions", methods=["POST"])
@login_required
def get_composer_suggestions():
    """Return rule-based suggestions for missing elements and relationships.

    Analyses the current canvas elements and relationships to suggest
    ArchiMate elements that are typically expected alongside the existing
    ones, and relationships that are missing between compatible pairs.

    JSON Body:
        elements (list): [{id, type, layer, name}]
        relationships (list): [{source_type, target_type, rel_type}]
        viewpoint_type (str): Optional viewpoint context.

    Returns:
        {missing_elements: [{type, reason, suggested_layer}],
         missing_relationships: [{source_id, target_id, suggested_type, reason}]}
    """
    data = request.get_json(silent=True) or {}
    elements = data.get("elements", [])
    relationships = data.get("relationships", [])

    if not elements:
        return jsonify({"missing_elements": [], "missing_relationships": []})

    # Build sets of element types present on the canvas
    type_set = set()
    for el in elements:
        el_type = el.get("type", "")
        if el_type:
            type_set.add(el_type)

    # ── Rule-based element suggestions ──────────────────────────────────
    missing_elements = []

    element_rules = [
        {
            "present": "ApplicationComponent",
            "missing": "ApplicationInterface",
            "reason": "ApplicationComponents typically expose ApplicationInterfaces",
            "layer": "application",
        },
        {
            "present": "ApplicationService",
            "missing": "ApplicationComponent",
            "reason": "ApplicationServices are usually realized by ApplicationComponents",
            "layer": "application",
        },
        {
            "present": "BusinessProcess",
            "missing": "BusinessRole",
            "reason": "BusinessProcesses are typically assigned to BusinessRoles",
            "layer": "business",
        },
        {
            "present": "Node",
            "missing": "SystemSoftware",
            "reason": "Nodes typically host SystemSoftware",
            "layer": "technology",
        },
        {
            "present": "BusinessService",
            "missing": "BusinessProcess",
            "reason": "BusinessServices are usually realized by BusinessProcesses",
            "layer": "business",
        },
        {
            "present": "ApplicationComponent",
            "missing": "ApplicationService",
            "reason": "ApplicationComponents typically expose ApplicationServices",
            "layer": "application",
        },
        {
            "present": "BusinessRole",
            "missing": "BusinessActor",
            "reason": "BusinessRoles are usually assigned to BusinessActors",
            "layer": "business",
        },
        {
            "present": "TechnologyService",
            "missing": "Node",
            "reason": "TechnologyServices are realized by Nodes or SystemSoftware",
            "layer": "technology",
        },
        {
            "present": "Goal",
            "missing": "Requirement",
            "reason": "Goals are typically refined into Requirements",
            "layer": "motivation",
        },
        {
            "present": "Stakeholder",
            "missing": "Driver",
            "reason": "Stakeholders typically have associated Drivers",
            "layer": "motivation",
        },
    ]

    seen_suggestions = set()
    for rule in element_rules:
        if rule["present"] in type_set and rule["missing"] not in type_set:
            if rule["missing"] not in seen_suggestions:
                missing_elements.append({
                    "type": rule["missing"],
                    "reason": rule["reason"],
                    "suggested_layer": rule["layer"],
                })
                seen_suggestions.add(rule["missing"])

    # ── Rule-based relationship suggestions ─────────────────────────────
    missing_relationships = []

    # Build existing relationship set (source_type → target_type → rel_type)
    existing_rels = set()
    for rel in relationships:
        key = (rel.get("source_type", ""), rel.get("target_type", ""), rel.get("rel_type", ""))
        existing_rels.add(key)

    # Build element lookup by type → list of {id, name}
    elements_by_type = {}
    for el in elements:
        el_type = el.get("type", "")
        if el_type:
            elements_by_type.setdefault(el_type, []).append(el)

    # Standard relationship patterns between element types
    rel_rules = [
        {"source": "ApplicationComponent", "target": "ApplicationInterface", "rel": "composition", "reason": "ApplicationComponent composes ApplicationInterface"},
        {"source": "ApplicationComponent", "target": "ApplicationService", "rel": "realization", "reason": "ApplicationComponent realizes ApplicationService"},
        {"source": "BusinessProcess", "target": "BusinessService", "rel": "realization", "reason": "BusinessProcess realizes BusinessService"},
        {"source": "BusinessRole", "target": "BusinessProcess", "rel": "assignment", "reason": "BusinessRole is assigned to BusinessProcess"},
        {"source": "Node", "target": "SystemSoftware", "rel": "composition", "reason": "Node hosts SystemSoftware"},
        {"source": "BusinessActor", "target": "BusinessRole", "rel": "assignment", "reason": "BusinessActor is assigned to BusinessRole"},
        {"source": "Stakeholder", "target": "Driver", "rel": "association", "reason": "Stakeholder has associated Driver"},
        {"source": "Goal", "target": "Requirement", "rel": "aggregation", "reason": "Goal is refined into Requirements"},
    ]

    rel_suggestion_count = 0
    max_rel_suggestions = 5

    for rule in rel_rules:
        if rel_suggestion_count >= max_rel_suggestions:
            break
        sources = elements_by_type.get(rule["source"], [])
        targets = elements_by_type.get(rule["target"], [])
        for src in sources:
            if rel_suggestion_count >= max_rel_suggestions:
                break
            for tgt in targets:
                key = (rule["source"], rule["target"], rule["rel"])
                if key not in existing_rels:
                    missing_relationships.append({
                        "source_id": src.get("id"),
                        "target_id": tgt.get("id"),
                        "suggested_type": rule["rel"],
                        "reason": rule["reason"],
                    })
                    existing_rels.add(key)
                    rel_suggestion_count += 1
                    break

    return jsonify({
        "missing_elements": missing_elements[:5],
        "missing_relationships": missing_relationships,
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/templates/<int:tid>/instantiate", methods=["POST"])
@login_required
def api_instantiate_template(tid):
    """Create a new saved viewpoint from a template.

    The template structure is used to create elements with the same types
    and positions. Each template slot becomes a placeholder element on the
    new diagram canvas.

    JSON Body:
        name (str): Required name for the new viewpoint.
        solution_id (int): Optional solution scope.
    """
    import json as _json

    from app.models.archimate_core import SavedDiagram, SavedDiagramElement
    from app.models.archimate_viewpoint import ArchimateViewpointTemplate

    template = db.session.get(ArchimateViewpointTemplate, tid)
    if not template:
        return jsonify({"error": "Template not found"}), 404

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400

    template_data = _json.loads(template.template_json)

    vp = SavedDiagram(
        name=name,
        viewpoint_type=template.viewpoint_type,
        solution_id=data.get("solution_id"),
        description=f"Created from template: {template.name}",
    )
    db.session.add(vp)
    db.session.flush()

    # Place template elements as diagram positions
    for slot in (template_data.get("elements") or []):
        el_id = slot.get("element_id")
        if not el_id:
            # Create a placeholder element from the template type
            from app.models.archimate_core import ArchiMateElement

            el_type = slot.get("element_type", "ApplicationComponent")
            el_name = slot.get("placeholder_name", "Unnamed")
            new_el = ArchiMateElement(
                name=el_name,
                type=el_type,
                layer=slot.get("element_layer", "application"),
                description="Created from template",
            )
            db.session.add(new_el)
            db.session.flush()
            el_id = new_el.id
        ve = SavedDiagramElement(
            diagram_id=vp.id,
            element_id=el_id,
            position_x=slot.get("x", 0),
            position_y=slot.get("y", 0),
            width=slot.get("width", 180),
            height=slot.get("height", 64),
            rendering_mode=slot.get("rendering_mode", "black_box"),
        )
        db.session.add(ve)

    db.session.commit()

    return jsonify({
        "id": vp.id,
        "name": vp.name,
        "viewpoint_type": vp.viewpoint_type,
    }), 201


# ── CMP-019 — AI diagram narration and impact explanation ────────────────────

# Relationship type plain-language explanations
_REL_EXPLANATIONS = {
    "composition": "is composed of",
    "aggregation": "aggregates",
    "assignment": "is assigned to",
    "realization": "realizes",
    "serving": "serves",
    "access": "accesses",
    "influence": "influences",
    "triggering": "triggers",
    "flow": "sends a flow to",
    "specialization": "is a specialization of",
    "association": "is associated with",
}

# Detailed descriptions per relationship type
_REL_DESCRIPTIONS = {
    "composition": "a strong ownership relationship where the target is an integral part of the source and cannot exist independently",
    "aggregation": "a grouping relationship where the target is part of the source but can exist independently",
    "assignment": "an allocation of responsibility -- the source carries out or executes the target",
    "realization": "the source implements or makes the target concrete",
    "serving": "the source provides functionality used by the target",
    "access": "the source reads, writes, or otherwise accesses the target data",
    "influence": "the source affects the implementation or achievement of the target",
    "triggering": "the source initiates or causes the target to start",
    "flow": "the source transfers information, goods, or value to the target",
    "specialization": "the source is a more specific variant of the target",
    "association": "a general, unspecified relationship between the two elements",
}

# Layer display names for narration
_LAYER_DISPLAY = {
    "business": "Business",
    "application": "Application",
    "technology": "Technology",
    "motivation": "Motivation",
    "strategy": "Strategy",
    "implementation": "Implementation & Migration",
    "physical": "Physical",
}

# Business-friendly type names (avoid ArchiMate jargon)
_BUSINESS_TYPE_NAMES = {
    "BusinessActor": "person or team",
    "BusinessRole": "role",
    "BusinessProcess": "business process",
    "BusinessFunction": "business function",
    "BusinessService": "business service",
    "BusinessObject": "business data",
    "BusinessEvent": "business event",
    "Contract": "contract",
    "Product": "product",
    "ApplicationComponent": "application",
    "ApplicationService": "application service",
    "ApplicationFunction": "application function",
    "ApplicationInterface": "interface",
    "ApplicationProcess": "application process",
    "DataObject": "data set",
    "Node": "server or platform",
    "Device": "device",
    "SystemSoftware": "system software",
    "TechnologyService": "infrastructure service",
    "Artifact": "artifact or file",
    "CommunicationNetwork": "network",
    "Stakeholder": "stakeholder",
    "Driver": "driver",
    "Goal": "goal",
    "Requirement": "requirement",
    "Constraint": "constraint",
    "Principle": "principle",
    "Capability": "capability",
    "Resource": "resource",
    "CourseOfAction": "course of action",
    "ValueStream": "value stream",
    "WorkPackage": "work package",
    "Deliverable": "deliverable",
    "Plateau": "plateau",
    "Gap": "gap",
}


def _friendly_type(el_type, audience):
    """Return element type name adjusted for audience."""
    if audience == "business":
        return _BUSINESS_TYPE_NAMES.get(el_type, el_type)
    return el_type


def _layer_from_type(el_type):
    """Guess the ArchiMate layer from an element type string."""
    t = (el_type or "").lower()
    if t.startswith("business") or t in ("contract", "product"):
        return "business"
    if t.startswith("application") or t == "dataobject":
        return "application"
    if t in ("node", "device", "systemsoftware", "technologyservice", "artifact", "communicationnetwork"):
        return "technology"
    if t in ("stakeholder", "driver", "goal", "requirement", "constraint", "principle"):
        return "motivation"
    if t in ("capability", "resource", "courseofaction", "valuestream"):
        return "strategy"
    if t in ("workpackage", "deliverable", "plateau", "gap"):
        return "implementation"
    return "other"


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/explain", methods=["POST"])
@login_required
def api_composer_explain():
    """Generate a rule-based narration of the current diagram.

    JSON Body:
        elements (list): [{name, type, layer}, ...]
        relationships (list): [{source_name, target_name, type}, ...]
        audience (str): 'technical' | 'business' | 'developer'

    Returns JSON with narration (markdown) and summary stats.
    """
    data = request.get_json(silent=True) or {}
    elements = data.get("elements") or []
    relationships = data.get("relationships") or []
    audience = data.get("audience", "technical")

    if audience not in ("technical", "business", "developer"):
        audience = "technical"

    if not elements:
        return jsonify({
            "narration": "The diagram is empty. Add elements to generate a narration.",
            "summary": {"element_count": 0, "relationship_count": 0, "layers": []},
        })

    # Count elements by layer
    layer_counts = {}
    for el in elements:
        layer = (el.get("layer") or _layer_from_type(el.get("type", ""))).lower()
        layer_counts[layer] = layer_counts.get(layer, 0) + 1

    layers_present = sorted(layer_counts.keys())
    el_count = len(elements)
    rel_count = len(relationships)

    # Build narration lines
    lines = []
    lines.append("## Diagram Overview\n")

    if audience == "business":
        lines.append(
            "This architecture diagram shows **{}** component{} "
            "connected by **{}** relationship{}.".format(
                el_count, "s" if el_count != 1 else "",
                rel_count, "s" if rel_count != 1 else "",
            )
        )
    elif audience == "developer":
        lines.append(
            "Architecture model with **{}** elements and **{}** relationships "
            "spanning {} layer{}.".format(
                el_count, rel_count,
                len(layers_present), "s" if len(layers_present) != 1 else "",
            )
        )
    else:
        lines.append(
            "ArchiMate diagram containing **{}** elements and **{}** relationships "
            "across {} layer{}: {}.".format(
                el_count, rel_count,
                len(layers_present), "s" if len(layers_present) != 1 else "",
                ", ".join(_LAYER_DISPLAY.get(l, l) for l in layers_present),
            )
        )

    lines.append("")

    # Layer breakdown
    if len(layers_present) > 1:
        lines.append("### Layers\n")
        for layer in layers_present:
            count = layer_counts[layer]
            display = _LAYER_DISPLAY.get(layer, layer.title())
            lines.append("- **{}**: {} element{}".format(display, count, "s" if count != 1 else ""))
        lines.append("")

    # Element listing
    lines.append("### Elements\n")
    for el in elements:
        name = el.get("name", "(unnamed)")
        el_type = el.get("type", "")
        friendly = _friendly_type(el_type, audience)
        if audience == "business":
            lines.append("- **{}** ({})".format(name, friendly))
        else:
            layer = (el.get("layer") or _layer_from_type(el_type)).lower()
            layer_label = _LAYER_DISPLAY.get(layer, layer.title())
            lines.append("- **{}** -- {} [{}]".format(name, friendly, layer_label))
    lines.append("")

    # Relationships
    if relationships:
        lines.append("### Relationships\n")
        for rel in relationships:
            src = rel.get("source_name", "?")
            tgt = rel.get("target_name", "?")
            rel_type = (rel.get("type") or "association").lower()
            verb = _REL_EXPLANATIONS.get(rel_type, rel_type)
            if audience == "business":
                lines.append("- {} {} {}".format(src, verb, tgt))
            elif audience == "developer":
                lines.append("- `{}` --[{}]--> `{}`".format(src, rel_type, tgt))
            else:
                lines.append("- **{}** {} **{}** ({})".format(src, verb, tgt, rel_type))
        lines.append("")

    # Cross-layer observations
    cross_layer_rels = []
    for rel in relationships:
        src_el = next((e for e in elements if e.get("name") == rel.get("source_name")), None)
        tgt_el = next((e for e in elements if e.get("name") == rel.get("target_name")), None)
        if src_el and tgt_el:
            src_layer = (src_el.get("layer") or _layer_from_type(src_el.get("type", ""))).lower()
            tgt_layer = (tgt_el.get("layer") or _layer_from_type(tgt_el.get("type", ""))).lower()
            if src_layer != tgt_layer:
                cross_layer_rels.append(rel)

    if cross_layer_rels:
        lines.append("### Cross-Layer Dependencies\n")
        if audience == "business":
            lines.append(
                "There {} **{}** connection{} "
                "that span across different architectural layers, showing how business needs "
                "connect to technical implementations.".format(
                    "is" if len(cross_layer_rels) == 1 else "are",
                    len(cross_layer_rels),
                    "s" if len(cross_layer_rels) != 1 else "",
                )
            )
        else:
            lines.append(
                "{} cross-layer relationship{} detected:".format(
                    len(cross_layer_rels), "s" if len(cross_layer_rels) != 1 else "",
                )
            )
            for rel in cross_layer_rels:
                src = rel.get("source_name", "?")
                tgt = rel.get("target_name", "?")
                r_type = (rel.get("type") or "association").lower()
                lines.append("- {} --[{}]--> {}".format(src, r_type, tgt))

    narration = "\n".join(lines)

    return jsonify({
        "narration": narration,
        "summary": {
            "element_count": el_count,
            "relationship_count": rel_count,
            "layers": layers_present,
        },
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/impact", methods=["POST"])
@login_required
def api_composer_impact():
    """Compute impact analysis by traversing relationships from a given element.

    JSON Body:
        element_id (int|str): ID of the element to analyse (matches canvas id).
        elements (list): [{id, name, type}, ...]
        relationships (list): [{source_id, target_id, type}, ...]

    Traverses up to 3 hops from the element. Returns affected elements and narrative.
    """
    data = request.get_json(silent=True) or {}
    element_id = data.get("element_id")
    elements = data.get("elements") or []
    relationships = data.get("relationships") or []

    if element_id is None:
        return jsonify({"error": "element_id is required"}), 400

    # Build lookup maps
    el_by_id = {}
    for el in elements:
        el_by_id[str(el.get("id", ""))] = el

    target_el = el_by_id.get(str(element_id))
    if not target_el:
        return jsonify({"error": "Element not found in provided elements list"}), 404

    # Build adjacency list (bidirectional -- impact flows both ways)
    adjacency = {}
    rel_lookup = {}
    for rel in relationships:
        src = str(rel.get("source_id", ""))
        tgt = str(rel.get("target_id", ""))
        r_type = (rel.get("type") or "association").lower()
        adjacency.setdefault(src, []).append(tgt)
        adjacency.setdefault(tgt, []).append(src)
        rel_lookup[(src, tgt)] = r_type
        rel_lookup[(tgt, src)] = r_type

    # BFS up to 3 hops
    visited = {str(element_id)}
    hop_details = []
    current_frontier = [str(element_id)]
    max_hops = 3

    for hop in range(1, max_hops + 1):
        next_frontier = []
        hop_elements = []
        for node_id in current_frontier:
            for neighbor_id in adjacency.get(node_id, []):
                if neighbor_id not in visited:
                    visited.add(neighbor_id)
                    next_frontier.append(neighbor_id)
                    neighbor_el = el_by_id.get(neighbor_id, {})
                    r_type_val = rel_lookup.get((node_id, neighbor_id), "association")
                    hop_elements.append({
                        "element_name": neighbor_el.get("name", "(unknown)"),
                        "element_id": neighbor_id,
                        "relationship": r_type_val,
                    })
        if hop_elements:
            hop_details.append({
                "hop": hop,
                "elements": [h["element_name"] for h in hop_elements],
                "element_ids": [h["element_id"] for h in hop_elements],
                "relationships": [h["relationship"] for h in hop_elements],
            })
        current_frontier = next_frontier
        if not current_frontier:
            break

    # Build impact narrative
    affected_ids = list(visited - {str(element_id)})
    el_name = target_el.get("name", "(unnamed)")

    if not affected_ids:
        narrative = (
            "**{}** has no direct or indirect connections in this diagram. "
            "Changes to it would have no impact on other elements.".format(el_name)
        )
    else:
        narrative_lines = ["## Impact Analysis: {}\n".format(el_name)]
        narrative_lines.append(
            "If **{}** is changed or removed, **{}** element{} "
            "could be affected across **{}** hop{}.\n".format(
                el_name, len(affected_ids),
                "s" if len(affected_ids) != 1 else "",
                len(hop_details),
                "s" if len(hop_details) != 1 else "",
            )
        )
        for hd in hop_details:
            hop_num = hd["hop"]
            narrative_lines.append("### Hop {} ({} impact)\n".format(
                hop_num, "direct" if hop_num == 1 else "indirect",
            ))
            for i, el_name_hop in enumerate(hd["elements"]):
                rel_val = hd["relationships"][i] if i < len(hd["relationships"]) else "association"
                verb = _REL_EXPLANATIONS.get(rel_val, rel_val)
                narrative_lines.append("- **{}** (via *{}*)".format(el_name_hop, verb))
            narrative_lines.append("")
        narrative = "\n".join(narrative_lines)

    return jsonify({
        "impact_narrative": narrative,
        "affected_element_ids": affected_ids,
        "hop_details": hop_details,
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/explain-relationship", methods=["POST"])
@login_required
def api_composer_explain_relationship():
    """Return a plain-language explanation of a single ArchiMate relationship.

    JSON Body:
        source_name (str): Name of the source element.
        target_name (str): Name of the target element.
        relationship_type (str): ArchiMate relationship type.
        source_type (str): ArchiMate type of the source element.
        target_type (str): ArchiMate type of the target element.

    Returns JSON with explanation string.
    """
    data = request.get_json(silent=True) or {}
    source_name = data.get("source_name", "(source)")
    target_name = data.get("target_name", "(target)")
    rel_type = (data.get("relationship_type") or "association").lower()
    source_type = data.get("source_type", "")
    target_type = data.get("target_type", "")

    verb = _REL_EXPLANATIONS.get(rel_type, rel_type)
    description = _REL_DESCRIPTIONS.get(rel_type, "a {} relationship".format(rel_type))

    explanation = "'{}' {} '{}' -- meaning {}.".format(source_name, verb, target_name, description)

    # Add type context if available
    if source_type and target_type:
        src_friendly = _BUSINESS_TYPE_NAMES.get(source_type, source_type)
        tgt_friendly = _BUSINESS_TYPE_NAMES.get(target_type, target_type)
        explanation += " Here, the {} ({}) {} the {} ({}).".format(
            src_friendly, source_type, verb, tgt_friendly, target_type,
        )

    return jsonify({"explanation": explanation})


# ── CMP-021 — Baseline-to-Target Delta and Plateau Generation ────────────────

# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/delta", methods=["POST"])
@login_required
def api_composer_delta():
    """Compare two saved diagrams and produce an element-level delta.

    JSON Body:
        baseline_viewpoint_id (int): Saved diagram ID for the baseline state.
        target_viewpoint_id (int): Saved diagram ID for the target state.

    Returns JSON with delta (added, retired, modified, unchanged) and summary counts.
    """
    from app.models.archimate_core import (
        ArchiMateElement,
        ArchiMateRelationship,
        SavedDiagram,
    )

    data = request.get_json(silent=True) or {}
    baseline_id = data.get("baseline_viewpoint_id")
    target_id = data.get("target_viewpoint_id")

    if not baseline_id or not target_id:
        return jsonify({"error": "baseline_viewpoint_id and target_viewpoint_id are required"}), 400

    baseline_vp = db.session.get(SavedDiagram, baseline_id)
    target_vp = db.session.get(SavedDiagram, target_id)

    if not baseline_vp:
        return jsonify({"error": f"Baseline diagram {baseline_id} not found"}), 404
    if not target_vp:
        return jsonify({"error": f"Target diagram {target_id} not found"}), 404

    # Gather element IDs from each diagram via junction table
    baseline_positions = baseline_vp.positions.all()
    target_positions = target_vp.positions.all()

    baseline_el_ids = set(p.element_id for p in baseline_positions)
    target_el_ids = set(p.element_id for p in target_positions)

    # Load all relevant elements in one query
    all_ids = baseline_el_ids | target_el_ids
    if all_ids:
        all_elements = {
            el.id: el
            for el in ArchiMateElement.query.filter(ArchiMateElement.id.in_(all_ids)).all()
        }
    else:
        all_elements = {}

    def _el_dict(el_id):
        el = all_elements.get(el_id)
        if not el:
            return {"element_id": el_id, "name": "(unknown)", "type": "Unknown"}
        return {"element_id": el.id, "name": el.name or "", "type": el.type or "Unknown", "layer": el.layer or ""}

    # Gather relationships for each diagram's element set
    def _get_rels_for_ids(el_ids):
        """Return dict mapping element_id -> set of relationship descriptors."""
        if not el_ids:
            return {}
        rels = ArchiMateRelationship.query.filter(
            ArchiMateRelationship.source_id.in_(el_ids),
            ArchiMateRelationship.target_id.in_(el_ids),
        ).all()
        el_rels = {}
        for r in rels:
            desc = f"{r.type}:{r.source_id}->{r.target_id}"
            el_rels.setdefault(r.source_id, set()).add(desc)
            el_rels.setdefault(r.target_id, set()).add(desc)
        return el_rels

    baseline_rels = _get_rels_for_ids(baseline_el_ids)
    target_rels = _get_rels_for_ids(target_el_ids)

    added = []
    retired = []
    modified = []
    unchanged = []

    # Elements in target but not baseline → added
    for eid in (target_el_ids - baseline_el_ids):
        added.append(_el_dict(eid))

    # Elements in baseline but not target → retired
    for eid in (baseline_el_ids - target_el_ids):
        retired.append(_el_dict(eid))

    # Elements in both → check relationship changes
    for eid in (baseline_el_ids & target_el_ids):
        b_rels = baseline_rels.get(eid, set())
        t_rels = target_rels.get(eid, set())
        if b_rels != t_rels:
            changes = []
            for r in (t_rels - b_rels):
                changes.append(f"relationship added: {r}")
            for r in (b_rels - t_rels):
                changes.append(f"relationship removed: {r}")
            entry = _el_dict(eid)
            entry["changes"] = changes
            modified.append(entry)
        else:
            unchanged.append(_el_dict(eid))

    return jsonify({
        "delta": {
            "added": added,
            "retired": retired,
            "modified": modified,
            "unchanged": unchanged,
        },
        "summary": {
            "added_count": len(added),
            "retired_count": len(retired),
            "modified_count": len(modified),
            "unchanged_count": len(unchanged),
        },
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/plateaus", methods=["POST"])
@login_required
def api_composer_plateaus():
    """Generate rule-based plateau suggestions from a delta result.

    JSON Body:
        delta (dict): The delta object returned by the /delta endpoint.

    Returns plateau suggestions grouped by ArchiMate layer with ordering rationale.
    """
    data = request.get_json(silent=True) or {}
    delta = data.get("delta")
    if not delta:
        return jsonify({"error": "delta is required"}), 400

    added = delta.get("added", [])
    modified = delta.get("modified", [])

    # Group added/modified elements by layer
    layer_groups = {}
    for el in added:
        layer = (el.get("layer") or "").lower()
        layer_groups.setdefault(layer, []).append(el.get("element_id"))
    for el in modified:
        layer = (el.get("layer") or "").lower()
        layer_groups.setdefault(layer, []).append(el.get("element_id"))

    # Layer ordering and naming
    LAYER_PLATEAU_CONFIG = {
        "technology": {"name": "Foundation Infrastructure", "rationale": "Deploy technology layer first", "order": 1},
        "physical": {"name": "Physical Infrastructure", "rationale": "Deploy physical infrastructure alongside technology", "order": 1},
        "application": {"name": "Application Modernization", "rationale": "Build on new infrastructure", "order": 2},
        "business": {"name": "Business Transformation", "rationale": "Enable new business capabilities on modernized applications", "order": 3},
        "strategy": {"name": "Strategic Alignment", "rationale": "Align strategy with implemented architecture", "order": 4},
        "motivation": {"name": "Motivation & Governance", "rationale": "Update governance and motivation elements", "order": 4},
        "implementation": {"name": "Migration Execution", "rationale": "Execute implementation and migration work packages", "order": 5},
    }

    plateaus = []
    for layer_key, el_ids in layer_groups.items():
        if not el_ids:
            continue
        config = LAYER_PLATEAU_CONFIG.get(layer_key, {
            "name": f"{layer_key.title()} Changes",
            "rationale": f"Changes in {layer_key} layer",
            "order": 6,
        })
        plateaus.append({
            "name": config["name"],
            "element_ids": el_ids,
            "rationale": config["rationale"],
            "order": config["order"],
            "layer": layer_key,
        })

    # Sort by order
    plateaus.sort(key=lambda p: p["order"])

    # Merge plateaus with same order (e.g., technology + physical)
    merged = {}
    for p in plateaus:
        key = p["order"]
        if key in merged:
            merged[key]["element_ids"].extend(p["element_ids"])
            if p["layer"] not in merged[key]["name"].lower():
                merged[key]["name"] += f" + {p['name']}"
        else:
            merged[key] = dict(p)
    plateaus = sorted(merged.values(), key=lambda p: p["order"])

    # Re-number sequentially
    for idx, p in enumerate(plateaus):
        p["order"] = idx + 1
        p.pop("layer", None)

    return jsonify({"plateaus": plateaus})


# ── CMP-020 — Enterprise intelligence overlays ─────────────────────────────

@archimate_bp.route("/api/composer/intelligence", methods=["GET"])
@login_required
def api_composer_intelligence():
    """Return enterprise intelligence enrichment for canvas elements.

    Query Parameters:
        element_ids (str): Comma-separated ArchiMate element IDs.

    Returns JSON with enrichment data per element (vendor, lifecycle, EOL, solution usage).
    """
    from datetime import datetime, timedelta

    from sqlalchemy import func

    from app.models.archimate_core import ArchiMateElement
    from app.models.application_portfolio import ApplicationComponent
    from app.models.business_capabilities import BusinessCapability
    from app.models.solution_archimate_element import SolutionArchiMateElement

    raw_ids = request.args.get("element_ids", "")
    if not raw_ids:
        return jsonify({"enrichment": {}})

    try:
        element_ids = [int(x.strip()) for x in raw_ids.split(",") if x.strip()]
    except (ValueError, TypeError):
        return jsonify({"error": "element_ids must be comma-separated integers"}), 400

    if not element_ids or len(element_ids) > 200:
        return jsonify({"enrichment": {}})

    # Fetch all requested ArchiMate elements in one query
    elements = db.session.query(ArchiMateElement).filter(
        ArchiMateElement.id.in_(element_ids)
    ).all()

    # Count solution usage for all element IDs in one query
    usage_counts = {}
    usage_rows = db.session.query(
        SolutionArchiMateElement.element_id,
        func.count(SolutionArchiMateElement.solution_id),
    ).filter(
        SolutionArchiMateElement.element_id.in_(element_ids)
    ).group_by(SolutionArchiMateElement.element_id).all()
    for row in usage_rows:
        usage_counts[row[0]] = row[1]

    # Types that map to ApplicationComponent
    app_types = {
        "ApplicationComponent", "ApplicationService", "ApplicationFunction",
        "ApplicationInterface", "ApplicationProcess",
    }
    # Types that map to VendorProduct (technology layer)
    tech_types = {"Node", "SystemSoftware", "Device", "TechnologyService"}
    # Types that map to BusinessCapability
    cap_types = {"Capability", "BusinessCapability"}

    # Collect names per category for batch queries
    app_names = []
    tech_names = []
    cap_names = []
    element_map = {}  # id -> element

    for el in elements:
        element_map[el.id] = el
        el_type = el.type or ""
        if el_type in app_types:
            app_names.append(el.name)
        elif el_type in tech_types:
            tech_names.append(el.name)
        elif el_type in cap_types:
            cap_names.append(el.name)

    # Batch query: ApplicationComponent by name (case-insensitive)
    app_lookup = {}
    if app_names:
        app_rows = db.session.query(ApplicationComponent).filter(
            func.lower(ApplicationComponent.name).in_([n.lower() for n in app_names])
        ).all()
        for ac in app_rows:
            app_lookup[ac.name.lower()] = ac

    # Batch query: VendorProduct by name
    vp_lookup = {}
    if tech_names:
        try:
            from app.models.vendor.vendor_organization import VendorProduct
            vp_rows = db.session.query(VendorProduct).options(
                joinedload(VendorProduct.vendor_organization)  # CMP-029: eager load to avoid N+1
            ).filter(
                func.lower(VendorProduct.name).in_([n.lower() for n in tech_names])
            ).all()
            for vp in vp_rows:
                vp_lookup[vp.name.lower()] = vp
        except Exception as exc:  # noqa: BLE001
            current_app.logger.debug("VendorProduct lookup failed (table may not exist): %s", exc)

    # Batch query: BusinessCapability by name
    cap_lookup = {}
    if cap_names:
        cap_rows = db.session.query(BusinessCapability).filter(
            func.lower(BusinessCapability.name).in_([n.lower() for n in cap_names])
        ).all()
        for bc in cap_rows:
            cap_lookup[bc.name.lower()] = bc

    eol_threshold = datetime.utcnow() + timedelta(days=548)  # ~18 months

    enrichment = {}
    for eid in element_ids:
        el = element_map.get(eid)
        if not el:
            continue

        info = {}
        signals = []
        el_type = el.type or ""
        name_lower = (el.name or "").lower()

        solution_count = usage_counts.get(eid, 0)
        info["solution_usage_count"] = solution_count
        if solution_count > 0:
            signals.append("solution_usage")

        if el_type in app_types:
            ac = app_lookup.get(name_lower)
            if ac:
                info["vendor_name"] = ac.vendor_name or None
                info["lifecycle_stage"] = ac.lifecycle_status or None
                info["app_criticality"] = ac.criticality or ac.business_criticality or None
                if ac.vendor_name:
                    signals.append("vendor_mapped")
                if ac.lifecycle_status in ("deprecated", "retired"):
                    signals.append("lifecycle_risk")

        elif el_type in tech_types:
            vp = vp_lookup.get(name_lower)
            if vp:
                vendor_name = None
                try:
                    if vp.vendor_organization:
                        vendor_name = vp.vendor_organization.name
                except Exception as exc:  # noqa: BLE001
                    current_app.logger.debug("Failed to read vendor_organization name: %s", exc)
                info["vendor_name"] = vendor_name
                eol_date = vp.end_of_life_date
                if eol_date and eol_date <= eol_threshold:
                    info["eol_warning"] = eol_date.strftime("%Y-%m-%d")
                    signals.append("eol_risk")
                else:
                    info["eol_warning"] = None
                if vendor_name:
                    signals.append("vendor_mapped")

        elif el_type in cap_types:
            bc = cap_lookup.get(name_lower)
            if bc:
                info["maturity_current"] = bc.current_maturity_level
                info["maturity_target"] = bc.target_maturity_level
                info["strategic_importance"] = bc.strategic_importance or None
                info["performance_score"] = bc.performance_score
                if bc.current_maturity_level and bc.target_maturity_level:
                    if bc.current_maturity_level < bc.target_maturity_level:
                        signals.append("maturity_gap")
                if bc.strategic_importance in ("critical", "high"):
                    signals.append("strategic")

        info["signals"] = signals
        if signals or info.get("solution_usage_count", 0) > 0:
            enrichment[str(eid)] = info

    return jsonify({"enrichment": enrichment})


# ── CMP-023: Pattern-based generation ─────────────────────────────────────────

import json as _json

BUILTIN_PATTERNS = [
    {
        "name": "API Gateway",
        "description": "API gateway fronting microservices",
        "elements": [
            {"role": "gateway", "type": "ApplicationComponent", "label": "{context} API Gateway"},
            {"role": "service_a", "type": "ApplicationService", "label": "{context} Service A"},
            {"role": "service_b", "type": "ApplicationService", "label": "{context} Service B"},
            {"role": "interface", "type": "ApplicationInterface", "label": "{context} REST API"},
        ],
        "relationships": [
            {"source_role": "gateway", "target_role": "service_a", "type": "serving"},
            {"source_role": "gateway", "target_role": "service_b", "type": "serving"},
            {"source_role": "interface", "target_role": "gateway", "type": "realization"},
        ],
    },
    {
        "name": "Event-Driven Microservices",
        "description": "Event-driven architecture with message broker",
        "elements": [
            {"role": "producer", "type": "ApplicationComponent", "label": "{context} Producer"},
            {"role": "broker", "type": "ApplicationComponent", "label": "{context} Message Broker"},
            {"role": "consumer", "type": "ApplicationComponent", "label": "{context} Consumer"},
            {"role": "event", "type": "DataObject", "label": "{context} Event"},
        ],
        "relationships": [
            {"source_role": "producer", "target_role": "broker", "type": "flow"},
            {"source_role": "broker", "target_role": "consumer", "type": "flow"},
            {"source_role": "producer", "target_role": "event", "type": "access"},
        ],
    },
    {
        "name": "CQRS",
        "description": "Command Query Responsibility Segregation",
        "elements": [
            {"role": "command", "type": "ApplicationService", "label": "{context} Command Service"},
            {"role": "query", "type": "ApplicationService", "label": "{context} Query Service"},
            {"role": "write_store", "type": "DataObject", "label": "{context} Write Store"},
            {"role": "read_store", "type": "DataObject", "label": "{context} Read Store"},
        ],
        "relationships": [
            {"source_role": "command", "target_role": "write_store", "type": "access"},
            {"source_role": "query", "target_role": "read_store", "type": "access"},
        ],
    },
    {
        "name": "Hub-and-Spoke Integration",
        "description": "Central integration hub connecting systems",
        "elements": [
            {"role": "hub", "type": "ApplicationComponent", "label": "{context} Integration Hub"},
            {"role": "system_a", "type": "ApplicationComponent", "label": "{context} System A"},
            {"role": "system_b", "type": "ApplicationComponent", "label": "{context} System B"},
            {"role": "system_c", "type": "ApplicationComponent", "label": "{context} System C"},
        ],
        "relationships": [
            {"source_role": "hub", "target_role": "system_a", "type": "serving"},
            {"source_role": "hub", "target_role": "system_b", "type": "serving"},
            {"source_role": "hub", "target_role": "system_c", "type": "serving"},
        ],
    },
    {
        "name": "Layered Application",
        "description": "Classic 3-tier application architecture",
        "elements": [
            {"role": "ui", "type": "ApplicationInterface", "label": "{context} UI"},
            {"role": "logic", "type": "ApplicationComponent", "label": "{context} Business Logic"},
            {"role": "data", "type": "DataObject", "label": "{context} Data Store"},
            {"role": "service", "type": "ApplicationService", "label": "{context} Service"},
        ],
        "relationships": [
            {"source_role": "service", "target_role": "logic", "type": "realization"},
            {"source_role": "ui", "target_role": "service", "type": "serving"},
            {"source_role": "logic", "target_role": "data", "type": "access"},
        ],
    },
    {
        "name": "ETL Pipeline",
        "description": "Extract-Transform-Load data pipeline",
        "elements": [
            {"role": "source", "type": "DataObject", "label": "{context} Source Data"},
            {"role": "extract", "type": "ApplicationFunction", "label": "{context} Extract"},
            {"role": "transform", "type": "ApplicationFunction", "label": "{context} Transform"},
            {"role": "load", "type": "ApplicationFunction", "label": "{context} Load"},
            {"role": "target", "type": "DataObject", "label": "{context} Target Data"},
        ],
        "relationships": [
            {"source_role": "extract", "target_role": "source", "type": "access"},
            {"source_role": "extract", "target_role": "transform", "type": "triggering"},
            {"source_role": "transform", "target_role": "load", "type": "triggering"},
            {"source_role": "load", "target_role": "target", "type": "access"},
        ],
    },
]


@archimate_bp.route("/api/patterns", methods=["GET"])
@login_required
def api_list_patterns():
    """List all architecture patterns (built-in + custom).

    Returns JSON list of pattern objects with id, name, description,
    is_builtin flag, and element/relationship counts.
    """
    results = []

    # Built-in patterns (negative IDs to distinguish from custom)
    for idx, pat in enumerate(BUILTIN_PATTERNS):
        results.append({
            "id": -(idx + 1),
            "name": pat["name"],
            "description": pat["description"],
            "is_builtin": True,
            "element_count": len(pat["elements"]),
            "relationship_count": len(pat["relationships"]),
        })

    # Custom patterns from database (model may not exist yet)
    try:
        from app.models.archimate_viewpoint import ArchimatePattern
        customs = ArchimatePattern.query.order_by(ArchimatePattern.created_at.desc()).all()
        for cp in customs:
            try:
                pdata = _json.loads(cp.pattern_json)
            except (ValueError, TypeError):
                pdata = {"elements": [], "relationships": []}
            results.append({
                "id": cp.id,
                "name": cp.name,
                "description": cp.description or "",
                "is_builtin": False,
                "element_count": len(pdata.get("elements", [])),
                "relationship_count": len(pdata.get("relationships", [])),
            })
    except Exception as exc:  # noqa: BLE001
        current_app.logger.debug("ComposerPattern query failed (table may not exist): %s", exc)

    return jsonify({"patterns": results})


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/patterns", methods=["POST"])
@login_required
def api_create_pattern():
    """Save a custom architecture pattern.

    JSON Body:
        name (str): Required — pattern name.
        description (str): Optional description.
        pattern_json (dict): Required — {elements: [...], relationships: [...]}.
    """
    from flask_login import current_user

    from app.models.archimate_viewpoint import ArchimatePattern

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400

    pattern_data = data.get("pattern_json")
    if not pattern_data or not isinstance(pattern_data, dict):
        return jsonify({"error": "pattern_json must be a dict with elements and relationships"}), 400

    elements = pattern_data.get("elements", [])
    if not elements:
        return jsonify({"error": "pattern_json must contain at least one element"}), 400

    try:
        pat = ArchimatePattern(
            name=name,
            description=(data.get("description") or "").strip() or None,
            pattern_json=_json.dumps(pattern_data),
            is_builtin=False,
            created_by=current_user.id if current_user and hasattr(current_user, "id") else None,
        )
        db.session.add(pat)
        db.session.commit()

        return jsonify({
            "id": pat.id,
            "name": pat.name,
            "description": pat.description,
            "is_builtin": False,
        }), 201
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        current_app.logger.error(f"[CMP-023] create pattern failed: {exc}")
        return jsonify({"error": str(exc)}), 500


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/patterns/<int:pattern_id>/instantiate", methods=["POST"])
@login_required
def api_instantiate_pattern(pattern_id):
    """Instantiate a pattern with a context string.

    Replaces {context} placeholders in element labels with the provided context.

    JSON Body:
        context (str): Required — context string (e.g., "Order Processing").

    Returns:
        elements: list of {role, type, label} with context applied.
        relationships: list of {source_role, target_role, type}.
    """
    from app.models.archimate_viewpoint import ArchimatePattern

    data = request.get_json(silent=True) or {}
    context = (data.get("context") or "").strip()
    if not context:
        return jsonify({"error": "context is required"}), 400

    # Resolve pattern — negative IDs are built-in
    if pattern_id < 0:
        idx = abs(pattern_id) - 1
        if idx < 0 or idx >= len(BUILTIN_PATTERNS):
            return jsonify({"error": "Built-in pattern not found"}), 404
        pattern_data = BUILTIN_PATTERNS[idx]
    else:
        try:
            pat = db.session.get(ArchimatePattern, pattern_id)
        except Exception:  # noqa: BLE001
            pat = None
        if not pat:
            return jsonify({"error": "Pattern not found"}), 404
        try:
            pattern_data = _json.loads(pat.pattern_json)
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid pattern data"}), 500

    # Apply context to element labels
    elements = []
    for el in pattern_data.get("elements", []):
        elements.append({
            "role": el["role"],
            "type": el["type"],
            "label": el["label"].replace("{context}", context),
        })

    relationships = []
    for rel in pattern_data.get("relationships", []):
        relationships.append({
            "source_role": rel["source_role"],
            "target_role": rel["target_role"],
            "type": rel["type"],
        })

    return jsonify({
        "pattern_name": pattern_data.get("name", ""),
        "context": context,
        "elements": elements,
        "relationships": relationships,
    })


# ── CMP-022 — AI document-to-model extraction ─────────────────────────────

# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/composer/extract", methods=["POST"])
@login_required
def api_composer_extract():
    """Rule-based extraction of ArchiMate elements from free-text.

    Accepts:
        text (str): The source text (architecture description, vendor proposal, etc.)
        target_phase (str): Optional TOGAF phase hint (A/B/C/D)
        viewpoint_type (str): Optional viewpoint scope

    Returns JSON with extracted elements (name, suggested_type, layer, confidence, source_quote)
    and inferred relationships.
    """
    import re

    payload = request.get_json(force=True) or {}
    text = (payload.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400

    target_phase = (payload.get("target_phase") or "").strip().upper()
    # viewpoint_type kept for future use
    # viewpoint_type = (payload.get("viewpoint_type") or "").strip()

    # ── Keyword → ArchiMate type mapping ──────────────────────────────
    KEYWORD_MAP = [
        # Application layer
        (r"\bapplication\b", "ApplicationComponent", "application", "high"),
        (r"\bsystem\b", "ApplicationComponent", "application", "high"),
        (r"\bplatform\b", "ApplicationComponent", "application", "high"),
        (r"\bsoftware\b", "ApplicationComponent", "application", "medium"),
        (r"\bservice\b", "ApplicationService", "application", "high"),
        (r"\bAPI\b", "ApplicationService", "application", "high"),
        (r"\bendpoint\b", "ApplicationService", "application", "medium"),
        (r"\bdatabase\b", "DataObject", "application", "high"),
        (r"\bdata\s*store\b", "DataObject", "application", "high"),
        (r"\brepository\b", "DataObject", "application", "medium"),
        # Technology layer
        (r"\bserver\b", "Node", "technology", "high"),
        (r"\bnode\b", "Node", "technology", "high"),
        (r"\binfrastructure\b", "Node", "technology", "medium"),
        # Business layer
        (r"\bprocess\b", "BusinessProcess", "business", "high"),
        (r"\bworkflow\b", "BusinessProcess", "business", "high"),
        (r"\bprocedure\b", "BusinessProcess", "business", "medium"),
        (r"\brole\b", "BusinessRole", "business", "high"),
        (r"\bactor\b", "BusinessActor", "business", "high"),
        (r"\bteam\b", "BusinessActor", "business", "medium"),
        (r"\bdepartment\b", "BusinessActor", "business", "medium"),
        # Motivation layer
        (r"\brequirement\b", "Requirement", "motivation", "high"),
        (r"\bneed\b", "Requirement", "motivation", "medium"),
        (r"\bmust\b", "Requirement", "motivation", "low"),
        (r"\bgoal\b", "Goal", "motivation", "high"),
        (r"\bobjective\b", "Goal", "motivation", "high"),
        (r"\btarget\b", "Goal", "motivation", "medium"),
        (r"\bstakeholder\b", "Stakeholder", "motivation", "high"),
        (r"\buser\b", "Stakeholder", "motivation", "medium"),
        (r"\bcustomer\b", "Stakeholder", "motivation", "medium"),
    ]

    # ── Phase-based bias: boost certain layers ─────────────────────────
    PHASE_LAYER_BIAS = {
        "A": ["motivation", "strategy"],
        "B": ["business"],
        "C": ["application"],
        "D": ["technology"],
    }

    # ── Split text into sentences ──────────────────────────────────────
    sentences = re.split(r'(?<=[.!?])\s+', text)
    if len(sentences) == 1 and len(text) > 200:
        sentences = re.split(r'[;\n]+', text)
    if not sentences:
        sentences = [text]

    elements = []
    seen_names = set()

    for sentence in sentences:
        sentence_stripped = sentence.strip()
        if not sentence_stripped:
            continue

        for pattern, el_type, layer, confidence in KEYWORD_MAP:
            matches = list(re.finditer(pattern, sentence_stripped, re.IGNORECASE))
            for match in matches:
                # Extract noun phrase around the keyword
                name = _extract_noun_phrase(sentence_stripped, match.start(), match.end())
                if not name or len(name) < 3:
                    continue

                # Deduplicate
                name_key = name.lower().strip()
                if name_key in seen_names:
                    continue
                seen_names.add(name_key)

                # Phase bias: boost confidence if layer matches phase
                final_confidence = confidence
                if target_phase and layer in PHASE_LAYER_BIAS.get(target_phase, []):
                    if confidence == "medium":
                        final_confidence = "high"
                    elif confidence == "low":
                        final_confidence = "medium"

                elements.append({
                    "name": name,
                    "suggested_type": el_type,
                    "layer": layer,
                    "confidence": final_confidence,
                    "source_quote": sentence_stripped[:200],
                })

    # ── Generate relationships between extracted elements ──────────────
    relationships = _infer_relationships(elements, sentences)

    return jsonify({
        "elements": elements,
        "relationships": relationships,
    })


def _extract_noun_phrase(sentence, kw_start, kw_end):
    """Extract a meaningful noun phrase around a keyword match position.

    Looks backwards and forwards from the keyword for capitalized words,
    determiners, and adjectives to build a sensible element name.
    """
    import re

    words = sentence.split()
    if not words:
        return ""

    # Find which word index contains the keyword
    char_pos = 0
    kw_word_idx = 0
    for i, w in enumerate(words):
        if char_pos <= kw_start < char_pos + len(w) + 1:
            kw_word_idx = i
            break
        char_pos += len(w) + 1

    # Expand left: grab preceding capitalized/adjective words
    start_idx = kw_word_idx
    stop_words = {"the", "a", "an", "is", "are", "was", "were", "and", "or", "to",
                  "in", "on", "at", "by", "for", "with", "that", "this", "it",
                  "from", "as", "of", "all", "each", "every", "will", "shall",
                  "should", "can", "could", "may", "might"}
    for j in range(kw_word_idx - 1, max(kw_word_idx - 5, -1), -1):
        w_clean = re.sub(r'[^\w]', '', words[j]).lower()
        if w_clean in stop_words or not w_clean:
            break
        if words[j][0].isupper() or w_clean.isalpha():
            start_idx = j
        else:
            break

    # Expand right: grab following capitalized words
    end_idx = kw_word_idx
    for j in range(kw_word_idx + 1, min(kw_word_idx + 5, len(words))):
        w_clean = re.sub(r'[^\w]', '', words[j]).lower()
        if w_clean in stop_words or not w_clean:
            break
        if words[j][0].isupper() or w_clean.isalpha():
            end_idx = j
        else:
            break

    phrase_words = words[start_idx:end_idx + 1]
    # Clean trailing punctuation
    phrase = " ".join(phrase_words)
    phrase = re.sub(r'[,;:.!?]+$', '', phrase).strip()

    # Title-case for consistency
    if phrase and not any(c.isupper() for c in phrase):
        phrase = phrase.title()

    return phrase


def _infer_relationships(elements, sentences):
    """Generate plausible relationships between extracted elements.

    Uses type compatibility and sentence proximity to suggest relationships.
    """
    relationships = []
    if len(elements) < 2:
        return relationships

    # Type-based relationship inference
    TYPE_REL_MAP = {
        ("ApplicationComponent", "DataObject"): "access",
        ("ApplicationComponent", "ApplicationService"): "realization",
        ("ApplicationService", "ApplicationService"): "flow",
        ("ApplicationComponent", "ApplicationComponent"): "flow",
        ("BusinessProcess", "ApplicationService"): "serving",
        ("BusinessProcess", "BusinessProcess"): "triggering",
        ("BusinessActor", "BusinessRole"): "assignment",
        ("BusinessRole", "BusinessProcess"): "assignment",
        ("Node", "ApplicationComponent"): "assignment",
        ("Node", "Node"): "association",
        ("Stakeholder", "Goal"): "association",
        ("Goal", "Requirement"): "realization",
        ("Requirement", "ApplicationComponent"): "realization",
        ("Requirement", "BusinessProcess"): "realization",
    }

    seen_rels = set()

    for i, el_a in enumerate(elements):
        for j, el_b in enumerate(elements):
            if i >= j:
                continue

            pair_key = (el_a["name"], el_b["name"])
            if pair_key in seen_rels:
                continue

            type_pair = (el_a["suggested_type"], el_b["suggested_type"])
            rev_pair = (el_b["suggested_type"], el_a["suggested_type"])

            rel_type = TYPE_REL_MAP.get(type_pair) or TYPE_REL_MAP.get(rev_pair)
            if not rel_type:
                # Check if they appear in the same sentence (proximity)
                same_sentence = False
                a_lower = el_a["name"].lower()
                b_lower = el_b["name"].lower()
                for sent in sentences:
                    sl = sent.lower()
                    if a_lower in sl and b_lower in sl:
                        same_sentence = True
                        break
                if same_sentence:
                    rel_type = "association"

            if rel_type:
                source_name = el_a["name"]
                target_name = el_b["name"]
                # Swap direction for certain types
                if rev_pair in TYPE_REL_MAP and type_pair not in TYPE_REL_MAP:
                    source_name, target_name = target_name, source_name

                seen_rels.add(pair_key)
                relationships.append({
                    "source_name": source_name,
                    "target_name": target_name,
                    "type": rel_type,
                    "confidence": "medium",
                })

    return relationships


# ── CMP-024: Comments API ──────────────────────────────────────────────────

@archimate_bp.route("/api/elements/<int:element_id>/comments", methods=["GET"])
@login_required
def api_list_element_comments(element_id):
    """List comments for an ArchiMate element, ordered by created_at."""
    try:
        from app.models.archimate_viewpoint import ArchimateElementComment
    except ImportError:
        # Comment storage isn't provisioned (no ArchimateElementComment model).
        # Degrade to "no comments" instead of 500ing the element view.
        return jsonify({"comments": []})

    viewpoint_id = request.args.get("viewpoint_id", type=int)
    query = ArchimateElementComment.query.options(
        joinedload(ArchimateElementComment.author)  # CMP-029: eager load to avoid N+1
    ).filter_by(element_id=element_id)
    if viewpoint_id:
        query = query.filter(
            db.or_(
                ArchimateElementComment.viewpoint_id == viewpoint_id,
                ArchimateElementComment.viewpoint_id.is_(None),
            )
        )
    comments = query.order_by(ArchimateElementComment.created_at.asc()).all()

    return jsonify({
        "comments": [
            {
                "id": c.id,
                "element_id": c.element_id,
                "viewpoint_id": c.viewpoint_id,
                "user_id": c.user_id,
                "user_name": c.author.name if c.author and hasattr(c.author, "name") else "User #" + str(c.user_id),
                "comment_text": c.comment_text,
                "created_at": c.created_at.isoformat() if c.created_at else None,
            }
            for c in comments
        ],
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/elements/<int:element_id>/comments", methods=["POST"])
@login_required
def api_create_element_comment(element_id):
    """Create a comment on an ArchiMate element."""
    from flask_login import current_user

    from app.models.archimate_viewpoint import ArchimateElementComment

    data = request.get_json(silent=True) or {}
    comment_text = (data.get("comment_text") or "").strip()
    if not comment_text:
        return jsonify({"error": "comment_text is required"}), 400

    comment = ArchimateElementComment(
        element_id=element_id,
        viewpoint_id=data.get("viewpoint_id"),
        user_id=current_user.id,
        comment_text=comment_text,
    )
    db.session.add(comment)
    db.session.commit()

    return jsonify({
        "id": comment.id,
        "element_id": comment.element_id,
        "viewpoint_id": comment.viewpoint_id,
        "user_id": comment.user_id,
        "user_name": current_user.name if hasattr(current_user, "name") else "User #" + str(current_user.id),
        "comment_text": comment.comment_text,
        "created_at": comment.created_at.isoformat() if comment.created_at else None,
    }), 201


# ── CMP-024: Audit log API ────────────────────────────────────────────────

@archimate_bp.route("/api/audit-log", methods=["GET"])
@login_required
def api_list_audit_log():
    """Get audit log entries, optionally filtered by viewpoint_id. Limit 100."""
    from app.models.archimate_viewpoint import ArchimateAuditLog

    viewpoint_id = request.args.get("viewpoint_id", type=int)
    query = ArchimateAuditLog.query.options(
        joinedload(ArchimateAuditLog.actor)  # CMP-029: eager load to avoid N+1
    )
    if viewpoint_id:
        query = query.filter(
            db.or_(
                ArchimateAuditLog.viewpoint_id == viewpoint_id,
                ArchimateAuditLog.viewpoint_id.is_(None),
            )
        )
    entries = query.order_by(ArchimateAuditLog.created_at.desc()).limit(100).all()

    return jsonify({
        "entries": [
            {
                "id": e.id,
                "viewpoint_id": e.viewpoint_id,
                "user_id": e.user_id,
                "user_name": e.actor.name if e.actor and hasattr(e.actor, "name") else "User #" + str(e.user_id),
                "action": e.action,
                "entity_type": e.entity_type,
                "entity_id": e.entity_id,
                "entity_name": e.entity_name,
                "old_value": e.old_value,
                "new_value": e.new_value,
                "created_at": e.created_at.isoformat() if e.created_at else None,
            }
            for e in entries
        ],
    })


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/audit-log", methods=["POST"])
@login_required
def api_create_audit_entry():
    """Log a composer action to the audit trail."""
    from flask_login import current_user

    from app.models.archimate_viewpoint import ArchimateAuditLog

    data = request.get_json(silent=True) or {}
    action = (data.get("action") or "").strip()
    if not action:
        return jsonify({"error": "action is required"}), 400

    entry = ArchimateAuditLog(
        viewpoint_id=data.get("viewpoint_id"),
        user_id=current_user.id,
        action=action,
        entity_type=data.get("entity_type"),
        entity_id=data.get("entity_id"),
        entity_name=data.get("entity_name"),
        old_value=data.get("old_value"),
        new_value=data.get("new_value"),
    )
    db.session.add(entry)
    db.session.commit()

    return jsonify({"id": entry.id, "action": entry.action}), 201


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/element-maturity", methods=["POST"])
@login_required
def api_element_maturity():
    """Resolve maturity for a batch of ArchiMate elements from existing domain data.

    Derives maturity from three sources (in priority order):
      1. Capability.current_maturity_level (1-5) via CapabilityArchiMateClassification
      2. ApplicationRationalizationScore.overall_health_score (0-100) via application_component_id
      3. ArchiMateElement.strategic_alignment_score (0-100) as fallback

    Request: {"element_ids": [1, 2, 3]}
    Response: {"maturity": {"1": {"level": 3, "label": "M3", "pct": 60, "source": "capability"}, ...}}
    """
    data = request.get_json(silent=True)
    if not data or not data.get("element_ids"):
        return jsonify({"maturity": {}}), 200

    element_ids = data["element_ids"]
    if not isinstance(element_ids, list) or len(element_ids) > 500:
        return jsonify({"error": "element_ids must be a list (max 500)"}), 400

    # Sanitize to ints
    try:
        element_ids = [int(eid) for eid in element_ids]
    except (ValueError, TypeError):
        return jsonify({"error": "element_ids must be integers"}), 400

    result = {}

    try:
        from app.models.archimate_core import ArchiMateElement

        elements = ArchiMateElement.query.filter(ArchiMateElement.id.in_(element_ids)).all()
        elem_map = {e.id: e for e in elements}

        # Source 1: Capability maturity via classification junction
        cap_maturity = {}
        try:
            from app.models.capability_archimate_mapping import CapabilityArchiMateClassification
            from app.models.business_capabilities import BusinessCapability

            mappings = (
                db.session.query(
                    CapabilityArchiMateClassification.archimate_element_id,
                    BusinessCapability.current_maturity_level,
                    BusinessCapability.target_maturity_level,
                )
                .join(BusinessCapability, BusinessCapability.id == CapabilityArchiMateClassification.capability_id)
                .filter(CapabilityArchiMateClassification.archimate_element_id.in_(element_ids))
                .all()
            )
            for eid, current, target in mappings:
                if current and eid not in cap_maturity:
                    cap_maturity[eid] = {"level": current, "target": target}
        except Exception as exc:
            current_app.logger.debug("Capability maturity lookup failed: %s", exc)

        # Source 2: Application health score via application_component_id
        app_health = {}
        app_linked_ids = [eid for eid in element_ids if elem_map.get(eid) and elem_map[eid].application_component_id]
        if app_linked_ids:
            try:
                from app.models.application_rationalization import ApplicationRationalizationScore

                scores = (
                    ApplicationRationalizationScore.query
                    .filter(ApplicationRationalizationScore.application_component_id.in_(
                        [elem_map[eid].application_component_id for eid in app_linked_ids]
                    ))
                    .all()
                )
                app_to_score = {}
                for s in scores:
                    if s.overall_health_score is not None:
                        app_to_score[s.application_component_id] = s.overall_health_score

                for eid in app_linked_ids:
                    app_id = elem_map[eid].application_component_id
                    if app_id in app_to_score:
                        app_health[eid] = app_to_score[app_id]
            except Exception as exc:
                current_app.logger.debug("Application health score lookup failed: %s", exc)

        # Build result: priority cascade
        for eid in element_ids:
            elem = elem_map.get(eid)
            if not elem:
                continue

            if eid in cap_maturity:
                level = cap_maturity[eid]["level"]
                pct = level * 20
                result[str(eid)] = {"level": level, "label": "M" + str(level), "pct": pct, "source": "capability"}
            elif eid in app_health:
                score = app_health[eid]
                level = max(1, min(5, int(score / 20) + (1 if score % 20 > 0 else 0)))
                result[str(eid)] = {"level": level, "label": "M" + str(level), "pct": int(score), "source": "health"}
            elif elem.strategic_alignment_score is not None:
                score = elem.strategic_alignment_score
                level = max(1, min(5, int(score / 20) + (1 if score % 20 > 0 else 0)))
                result[str(eid)] = {"level": level, "label": "M" + str(level), "pct": int(score), "source": "alignment"}

    except Exception as e:
        current_app.logger.error(f"Element maturity resolution failed: {e}", exc_info=True)

    return jsonify({"maturity": result})


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/elements/<int:element_id>/alignment-score", methods=["PUT"])
@login_required
def update_element_alignment_score(element_id):
    """Set the strategic_alignment_score on an ArchiMateElement (used by composer Set Maturity)."""
    from app.models.models import ArchiMateElement as AE

    elem = AE.query.get(element_id)
    if not elem:
        return jsonify({"error": "Element not found"}), 404

    data = request.get_json(silent=True) or {}
    score = data.get("score")
    if score is None or not isinstance(score, (int, float)) or score < 0 or score > 100:
        return jsonify({"error": "score must be a number 0-100"}), 400

    elem.strategic_alignment_score = float(score)
    db.session.commit()

    level = max(1, min(5, int(score / 20) + (1 if score % 20 > 0 else 0)))
    return jsonify({
        "id": elem.id,
        "strategic_alignment_score": elem.strategic_alignment_score,
        "maturity": {"level": level, "label": "M" + str(level), "pct": int(score), "source": "alignment"},
    })


# ── CMP-043: Custom element properties (tagged values) ──────────────────────

@archimate_bp.route("/api/elements/<int:element_id>/properties", methods=["GET"])
@login_required
def get_element_properties(element_id):
    """CMP-043: Get custom properties for an element."""
    import json
    from app.models.models import ArchiMateElement as AE

    el = AE.query.get(element_id)
    if not el:
        return jsonify({"error": "Element not found"}), 404
    try:
        props = json.loads(el.properties) if el.properties else {}
    except (json.JSONDecodeError, TypeError):
        props = {}
    return jsonify(props)


# CSRF: Protected via X-CSRFToken header sent by Platform.fetch
@login_required
@archimate_bp.route("/api/elements/<int:element_id>/properties", methods=["PUT"])
@login_required
def put_element_properties(element_id):
    """CMP-043: Replace custom properties for an element."""
    from app.models.models import ArchiMateElement as AE

    el = AE.query.get(element_id)
    if not el:
        return jsonify({"error": "Element not found"}), 404
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({"error": "Invalid JSON"}), 400
    if not isinstance(data, dict):
        return jsonify({"error": "Properties must be a JSON object"}), 400
    import json
    el.properties = json.dumps(data)
    db.session.commit()
    return jsonify(data)


# ── CMP-053: Landscape map view ──────────────────────────────────────────────

@archimate_bp.route("/api/landscape", methods=["GET"])
@login_required
def get_landscape_data():
    """CMP-053: Landscape map — grid view of element intersections with counts."""
    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship

    row_type = request.args.get("row_type", "Capability")
    col_type = request.args.get("col_type", "ApplicationComponent")

    rows = ArchiMateElement.query.filter_by(type=row_type).order_by(ArchiMateElement.name).all()
    cols = ArchiMateElement.query.filter_by(type=col_type).order_by(ArchiMateElement.name).all()

    row_ids = [r.id for r in rows]
    col_ids = [c.id for c in cols]

    if not row_ids or not col_ids:
        return jsonify({
            "rows": [{"id": r.id, "name": r.name} for r in rows],
            "columns": [{"id": c.id, "name": c.name} for c in cols],
            "cells": {},
        })

    # Query relationships in both directions
    rels = ArchiMateRelationship.query.filter(
        db.or_(
            db.and_(
                ArchiMateRelationship.source_id.in_(row_ids),
                ArchiMateRelationship.target_id.in_(col_ids),
            ),
            db.and_(
                ArchiMateRelationship.source_id.in_(col_ids),
                ArchiMateRelationship.target_id.in_(row_ids),
            ),
        )
    ).all()

    # Build cell data: "rowId_colId" -> {count, types[]}
    cells = {}
    row_set = set(row_ids)
    col_set = set(col_ids)
    for rel in rels:
        if rel.source_id in row_set and rel.target_id in col_set:
            r_id, c_id = rel.source_id, rel.target_id
        elif rel.source_id in col_set and rel.target_id in row_set:
            r_id, c_id = rel.target_id, rel.source_id
        else:
            continue
        key = f"{r_id}_{c_id}"
        if key not in cells:
            cells[key] = {"count": 0, "types": []}
        cells[key]["count"] += 1
        if rel.type not in cells[key]["types"]:
            cells[key]["types"].append(rel.type)

    return jsonify({
        "rows": [{"id": r.id, "name": r.name} for r in rows],
        "columns": [{"id": c.id, "name": c.name} for c in cols],
        "cells": cells,
    })



# ── CMP-041: Matrix cross-reference view ─────────────────────────────────────


@archimate_bp.route("/api/matrix", methods=["GET"])
@login_required
def get_matrix_data():
    """CMP-041: Matrix view data — cross-reference elements by type."""
    from app.models.models import ArchiMateElement, ArchiMateRelationship

    row_type = request.args.get("row_type", "")
    col_type = request.args.get("col_type", "")

    if not row_type or not col_type:
        return jsonify({"error": "row_type and col_type required"}), 400

    rows = ArchiMateElement.query.filter_by(type=row_type).order_by(ArchiMateElement.name).all()
    cols = ArchiMateElement.query.filter_by(type=col_type).order_by(ArchiMateElement.name).all()

    row_ids = [r.id for r in rows]
    col_ids = [c.id for c in cols]

    all_ids = row_ids + col_ids
    if not all_ids:
        return jsonify({
            "rows": [],
            "columns": [],
            "intersections": {},
        })

    # Get relationships between these elements (both directions)
    rels = ArchiMateRelationship.query.filter(
        ArchiMateRelationship.source_id.in_(all_ids),
        ArchiMateRelationship.target_id.in_(all_ids),
    ).all()

    # Build intersection map keyed by "rowId_colId"
    row_id_set = set(row_ids)
    col_id_set = set(col_ids)
    intersections = {}
    for rel in rels:
        src = rel.source_id
        tgt = rel.target_id
        rel_type = rel.type or "Association"
        # Forward: src is row, tgt is col
        if src in row_id_set and tgt in col_id_set:
            intersections[f"{src}_{tgt}"] = rel_type
        # Reverse: tgt is row, src is col
        if tgt in row_id_set and src in col_id_set:
            intersections[f"{tgt}_{src}"] = rel_type

    return jsonify({
        "rows": [{"id": r.id, "name": r.name, "type": r.type} for r in rows],
        "columns": [{"id": c.id, "name": c.name, "type": c.type} for c in cols],
        "intersections": intersections,
    })

# ── Active editor tracking (lightweight in-process store) ────────────────────
# Stores {diagram_id: {user_id: last_seen_ts}} — cleared on old entries each
# time a new heartbeat arrives. No DB required; acceptable for single-process.
import threading as _threading
import time as _time
import logging
logger = logging.getLogger(__name__)
_active_editors_lock = _threading.Lock()
_active_editors: dict = {}   # {diagram_id_str: {user_id_str: float_timestamp}}
_EDITOR_TIMEOUT_SECS = 30    # user considered gone after 30s without ping


@archimate_bp.route("/api/diagrams/<int:diagram_id>/active-editors", methods=["GET"])
@login_required
def api_diagram_active_editors(diagram_id):
    """Return list of users currently editing this diagram (heartbeat within 30s)."""
    now = _time.time()
    key = str(diagram_id)
    with _active_editors_lock:
        room = _active_editors.get(key, {})
        alive = {uid: ts for uid, ts in room.items() if now - ts < _EDITOR_TIMEOUT_SECS}
        _active_editors[key] = alive

    user_ids = [int(uid) for uid in alive]
    editors = []
    if user_ids:
        try:
            from app.models.user import User
            users = User.query.filter(User.id.in_(user_ids)).all()
            editors = [{"id": u.id, "name": u.full_name or u.email, "email": u.email} for u in users]
        except Exception:  # noqa: BLE001
            editors = [{"id": uid, "name": f"User {uid}", "email": ""} for uid in user_ids]

    return jsonify({"editors": editors})


@archimate_bp.route("/api/diagrams/<int:diagram_id>/editors/join", methods=["POST"])
@login_required
def api_diagram_editor_join(diagram_id):
    """Register the current user as actively editing diagram (call on open, heartbeat every 15s)."""
    key = str(diagram_id)
    uid = str(current_user.id)
    with _active_editors_lock:
        if key not in _active_editors:
            _active_editors[key] = {}
        _active_editors[key][uid] = _time.time()
    return jsonify({"ok": True})


@archimate_bp.route("/api/diagrams/<int:diagram_id>/editors/leave", methods=["POST"])
@login_required
def api_diagram_editor_leave(diagram_id):
    """Remove the current user from the active-editor list for this diagram."""
    key = str(diagram_id)
    uid = str(current_user.id)
    with _active_editors_lock:
        if key in _active_editors:
            _active_editors[key].pop(uid, None)
    return jsonify({"ok": True})


# ── Derived relationships (ArchiMate 3.2 structural derivation rules) ─────────
# Derivation rules per ArchiMate 3.2 section 5.7:
#   composition + *   = composition (parent derives through child)
#   aggregation + *   = aggregation
#   association is always derivable from any pair
# We implement a pragmatic single-hop rule: for every directed path A→B→C
# through any structural relationship, we emit A→C as a derived "association"
# unless a direct A→C already exists.

@archimate_bp.route("/api/composer/derived-relationships", methods=["GET"])
@login_required
def api_derived_relationships():
    """Compute one-hop derived relationships for the current diagram elements.

    Query param:
        element_ids  comma-separated ArchiMate element IDs visible on canvas.

    Returns list of derived relationships not already in the DB.
    """
    from app.models.models import ArchiMateRelationship, ArchiMateElement  # noqa: PLC0415

    id_param = request.args.get("element_ids", "")
    if not id_param:
        return jsonify({"derived": []})

    try:
        el_ids = [int(x) for x in id_param.split(",") if x.strip().isdigit()]
    except ValueError:
        return jsonify({"error": "invalid element_ids"}), 400

    if not el_ids:
        return jsonify({"derived": []})

    # Fetch all direct relationships among the canvas elements
    rels = ArchiMateRelationship.query.filter(
        ArchiMateRelationship.source_id.in_(el_ids),
        ArchiMateRelationship.target_id.in_(el_ids),
    ).all()

    # Build adjacency: source_id -> [(target_id, rel_type)]
    adj: dict = {}
    direct: set = set()
    for r in rels:
        adj.setdefault(r.source_id, []).append((r.target_id, r.type or "association"))
        direct.add((r.source_id, r.target_id))

    # Element metadata for display
    elements = {e.id: e for e in ArchiMateElement.query.filter(ArchiMateElement.id.in_(el_ids)).all()}

    # One-hop derivation
    STRUCTURAL = {"composition", "aggregation", "realization", "assignment", "serving"}
    derived = []
    seen = set()
    for a, hops_ab in adj.items():
        for b, type_ab in hops_ab:
            if b not in adj:
                continue
            for c, type_bc in adj[b]:
                if c == a:
                    continue  # skip self-loops
                if (a, c) in direct:
                    continue  # direct relationship already exists
                if (a, c) in seen:
                    continue
                seen.add((a, c))
                # Determine derived type using ArchiMate derivation rules
                if type_ab in STRUCTURAL and type_bc in STRUCTURAL:
                    derived_type = type_ab  # structural propagates
                else:
                    derived_type = "association"  # default fallback
                derived.append({
                    "source_id": a,
                    "target_id": c,
                    "via_id": b,
                    "type": derived_type,
                    "source_name": elements[a].name if a in elements else str(a),
                    "target_name": elements[c].name if c in elements else str(c),
                    "via_name": elements[b].name if b in elements else str(b),
                })

    return jsonify({"derived": derived})


# ── CSV Element Import ─────────────────────────────────────────────────────────
# Expected CSV columns (case-insensitive header row):
#   name, type, layer, description (description is optional)
# All rows create new ArchiMateElement records, returning their IDs so the
# caller can immediately place them on the canvas.

@archimate_bp.route("/api/import/csv", methods=["POST"])
@login_required
def api_import_csv():
    """Import ArchiMate elements from a CSV file upload.

    POST form-data:
        file: the CSV file (text/csv or .csv)
        solution_id: (optional) link created elements to this solution

    CSV must have a header row. Required columns: name, type.
    Optional columns: layer, description.
    Duplicate names (same name+type) are skipped — existing element ID returned.

    Returns:
        created: list of {id, name, type, layer} for newly created elements
        skipped: list of {name, type, reason} for skipped rows
        errors:  list of row-level error messages
    """
    import csv
    import io

    from app.models.models import ArchiMateElement  # noqa: PLC0415

    LAYER_DEFAULTS = {
        # Business layer types
        "businessactor": "business", "businessrole": "business",
        "businesscollaboration": "business", "businessinterface": "business",
        "businessprocess": "business", "businessfunction": "business",
        "businessinteraction": "business", "businessservice": "business",
        "businessobject": "business", "representation": "business",
        "product": "business", "contract": "business", "event": "business",
        # Application layer types
        "applicationcomponent": "application", "applicationcollaboration": "application",
        "applicationinterface": "application", "applicationfunction": "application",
        "applicationinteraction": "application", "applicationservice": "application",
        "applicationprocess": "application", "dataobject": "application",
        # Technology layer
        "node": "technology", "device": "technology", "systemsoftware": "technology",
        "technologycollaboration": "technology", "technologyinterface": "technology",
        "path": "technology", "communicationnetwork": "technology",
        "technologyfunction": "technology", "technologyprocess": "technology",
        "technologyinteraction": "technology", "technologyservice": "technology",
        "artifact": "technology",
        # Motivation layer
        "stakeholder": "motivation", "driver": "motivation", "assessment": "motivation",
        "goal": "motivation", "outcome": "motivation", "principle": "motivation",
        "requirement": "motivation", "constraint": "motivation", "meaning": "motivation",
        "value": "motivation",
        # Strategy layer
        "resource": "strategy", "capability": "strategy", "courseoaction": "strategy",
        "valuestream": "strategy",
        # Implementation layer
        "workpackage": "implementation", "deliverable": "implementation",
        "implementationevent": "implementation", "plateau": "implementation", "gap": "implementation",
    }

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded. POST with form-data field 'file'."}), 400

    solution_id = request.form.get("solution_id") or request.args.get("solution_id")

    try:
        raw = file.read().decode("utf-8-sig")  # handle BOM
        reader = csv.DictReader(io.StringIO(raw))
        rows = list(reader)
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"CSV parse error: {exc}"}), 400

    # Normalise header names to lowercase stripped
    if not rows:
        return jsonify({"error": "CSV is empty or has no data rows"}), 400

    headers = {k.lower().strip() for k in rows[0].keys()}
    if "name" not in headers or "type" not in headers:
        return jsonify({"error": "CSV must have 'name' and 'type' columns"}), 400

    created = []
    skipped = []
    errors = []

    for i, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        norm = {k.lower().strip(): (v or "").strip() for k, v in row.items()}
        name = norm.get("name", "").strip()
        el_type = norm.get("type", "").strip()
        if not name or not el_type:
            errors.append(f"Row {i}: missing name or type — skipped")
            continue

        # Derive layer from type if not provided
        layer = norm.get("layer", "").strip() or LAYER_DEFAULTS.get(el_type.lower().replace(" ", ""), "application")
        description = norm.get("description", "")

        # Check for existing element with same name+type to avoid duplicates
        existing = ArchiMateElement.query.filter_by(name=name, type=el_type).first()
        if existing:
            skipped.append({"name": name, "type": el_type, "reason": "already exists", "id": existing.id})
            continue

        try:
            el = ArchiMateElement(name=name, type=el_type, layer=layer, description=description or None)
            db.session.add(el)
            db.session.flush()  # get el.id without committing yet

            # Link to solution if requested
            if solution_id:
                try:
                    from app.models.solution_archimate_element import SolutionArchimateElement  # noqa: PLC0415
                    sae = SolutionArchimateElement(
                        solution_id=int(solution_id),
                        archimate_element_id=el.id,
                    )
                    db.session.add(sae)
                except Exception:  # noqa: BLE001  # fabricated-values-ok
                    pass  # solution link is best-effort

            created.append({"id": el.id, "name": el.name, "type": el.type, "layer": el.layer})
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            errors.append(f"Row {i} ({name}): {exc}")
            continue

    try:
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        return jsonify({"error": f"Database commit failed: {exc}"}), 500

    return jsonify({
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "summary": f"Imported {len(created)} elements, {len(skipped)} duplicates skipped, {len(errors)} errors",
    })


# ── REQ-ACD-001: CSV Upload → ArchiMate Elements (any type) ─────────────
# Accepts CSV/Excel with columns: name, type, description, [layer], [parent]
# Creates ArchiMateElement records with correct layer auto-assignment.

_TYPE_TO_LAYER = {
    "applicationcomponent": "Application", "applicationservice": "Application",
    "applicationinterface": "Application", "applicationfunction": "Application",
    "applicationevent": "Application", "dataobject": "Application",
    "businessprocess": "Business", "businessservice": "Business",
    "businessfunction": "Business", "businessobject": "Business",
    "businessactor": "Business", "businessrole": "Business",
    "businessevent": "Business", "businessinteraction": "Business",
    "contract": "Business", "product": "Business", "representation": "Business",
    "node": "Technology", "device": "Technology", "systemsoftware": "Technology",
    "artifact": "Technology", "communicationnetwork": "Technology",
    "technologyservice": "Technology", "technologyfunction": "Technology",
    "technologyinterface": "Technology", "technologyevent": "Technology",
    "path": "Technology",
    "resource": "Strategy", "capability": "Strategy",
    "valuestream": "Strategy", "courseofaction": "Strategy",
    "stakeholder": "Motivation", "driver": "Motivation", "assessment": "Motivation",
    "goal": "Motivation", "outcome": "Motivation", "principle": "Motivation",
    "requirement": "Motivation", "constraint": "Motivation",
    "meaning": "Motivation", "value": "Motivation",
    "workpackage": "Implementation", "deliverable": "Implementation",
    "plateau": "Implementation", "gap": "Implementation",
    "location": "Physical", "equipment": "Physical", "facility": "Physical",
    "distributionnetwork": "Physical", "material": "Physical",
    "grouping": "Other", "junction": "Other",
}


def _resolve_layer(element_type: str, explicit_layer: str = "") -> str:
    """Resolve ArchiMate layer from element type, with explicit override."""
    if explicit_layer and explicit_layer.strip():
        return explicit_layer.strip().capitalize()
    normalised = (element_type or "").replace(" ", "").replace("_", "").lower()
    return _TYPE_TO_LAYER.get(normalised, "Application")


@archimate_bp.route("/import", methods=["GET"])
@login_required
def import_page():
    """REQ-ACD-001: Render the unified ArchiMate import page."""
    return render_template("architecture/import_csv.html")


# Keep old URL as redirect for bookmarks
@archimate_bp.route("/import-csv", methods=["GET"])
@login_required
def import_csv_page():
    """Redirect old URL to unified import page."""
    from flask import redirect
    return redirect(url_for("archimate.import_page"), code=301)


@archimate_bp.route("/api/import-elements-csv", methods=["POST"])
@login_required
def api_import_elements_csv():
    """REQ-ACD-001: Import ArchiMate elements from CSV/Excel file.

    Accepts a file upload with columns: name, type, description, [layer], [parent].
    Creates ArchiMateElement records with correct layer auto-assignment.
    Returns summary with created/skipped/error counts and element IDs for
    immediate use in the Composer.
    """
    from app.models.archimate_core import ArchiMateElement

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    default_type = request.form.get("default_type", "").strip()
    filename = file.filename.lower()

    # Parse file
    rows = []
    try:
        if filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])))
            wb.close()
        elif filename.endswith(".csv"):
            import csv
            import io
            text = file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append({k.strip().lower(): (v or "").strip() for k, v in row.items()})
        else:
            return jsonify({"success": False, "error": "Unsupported file format. Use .csv or .xlsx"}), 400
    except Exception as parse_err:
        current_app.logger.error("CSV parse error: %s", parse_err)
        return jsonify({"success": False, "error": f"Could not parse file: {parse_err}"}), 400

    if not rows:
        return jsonify({"success": False, "error": "File contains no data rows"}), 400

    # Process rows
    created_ids = []
    skipped = []
    errors = []

    for idx, row in enumerate(rows, start=2):
        name = row.get("name", "").strip()
        if not name:
            errors.append({"row": idx, "error": "Missing name"})
            continue

        element_type = row.get("type", "").strip() or default_type
        if not element_type:
            errors.append({"row": idx, "name": name, "error": "No type specified and no default type set"})
            continue

        # Normalise type to PascalCase for storage
        type_clean = element_type.replace(" ", "").replace("_", "")
        # Capitalise first letter of each word-boundary segment
        if type_clean and type_clean[0].islower():
            type_clean = type_clean[0].upper() + type_clean[1:]

        layer = _resolve_layer(element_type, row.get("layer", ""))
        description = row.get("description", "").strip()

        # Check for duplicate by name + type
        existing = ArchiMateElement.query.filter_by(name=name, type=type_clean).first()
        if existing:
            skipped.append({"row": idx, "name": name, "type": type_clean, "existing_id": existing.id})
            continue

        try:
            elem = ArchiMateElement(
                name=name,
                type=type_clean,
                layer=layer,
                description=description or None,
                scope="enterprise",
            )
            db.session.add(elem)
            db.session.flush()
            created_ids.append({"id": elem.id, "name": name, "type": type_clean, "layer": layer})
        except Exception as row_err:
            db.session.rollback()
            errors.append({"row": idx, "name": name, "error": str(row_err)})

    if created_ids:
        db.session.commit()

    return jsonify({
        "success": True,
        "created": created_ids,
        "created_count": len(created_ids),
        "skipped": skipped,
        "skipped_count": len(skipped),
        "errors": errors,
        "error_count": len(errors),
        "total_rows": len(rows),
        "element_ids": [e["id"] for e in created_ids],
        "composer_url": f"/archimate/composer?elements={','.join(str(e['id']) for e in created_ids)}" if created_ids else None,
    })


@archimate_bp.route("/api/import-document", methods=["POST"])
@login_required
def api_import_document():
    """REQ-ACD-001: Unified document import — AI-powered extraction for PDF/DOCX/PPTX.

    Routes to the AI Chat document processing pipeline, extracts ArchiMate
    elements, creates them in the database, and returns element IDs for
    immediate use in the Composer.
    """
    import os

    from werkzeug.utils import secure_filename

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    filename = secure_filename(file.filename)
    if not filename:
        return jsonify({"success": False, "error": "Invalid filename"}), 400

    file_ext = os.path.splitext(filename)[1].lower()
    allowed = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".rtf", ".txt", ".md"}
    if file_ext not in allowed:
        return jsonify({"success": False, "error": f"File type '{file_ext}' not supported for AI analysis. Use: {', '.join(sorted(allowed))}"}), 400

    # Save file to disk for analysis
    upload_dir = os.path.join(
        current_app.config.get("UPLOAD_FOLDER", "uploads"), "import_documents"
    )
    os.makedirs(upload_dir, exist_ok=True)

    user_id = current_user.id if current_user.is_authenticated else 0
    from datetime import datetime as _dt
    ts = _dt.utcnow().strftime("%Y%m%d_%H%M%S")
    saved_filename = f"{user_id}_{ts}_{filename}"
    file_path = os.path.join(upload_dir, saved_filename)
    file.save(file_path)

    try:
        from app.services.archimate.document_analysis_service import DocumentAnalysisService
        from app.services.core.async_utils import get_or_create_event_loop

        analysis_service = DocumentAnalysisService()
        loop = get_or_create_event_loop()

        # Determine file type for routing
        if file_ext in {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".rtf"}:
            extracted_data, interactions = loop.run_until_complete(
                analysis_service._analyze_document(file_path, None, "architecture")
            )
        else:
            extracted_data, interaction = loop.run_until_complete(
                analysis_service._analyze_text_file(file_path, None, "architecture")
            )
            interactions = [interaction] if interaction else []

        # Normalize element types
        from app.services.archimate.element_type_normalizer import ElementTypeNormalizer
        normalizer = ElementTypeNormalizer()
        elements = normalizer.normalize_elements(extracted_data.get("elements", []))
        relationships = extracted_data.get("relationships", [])

        if not elements:
            metadata = extracted_data.get("metadata", {})
            return jsonify({
                "success": True,
                "created_count": 0,
                "skipped_count": 0,
                "error_count": 0,
                "errors": [],
                "element_ids": [],
                "message": metadata.get("error", "No ArchiMate elements could be extracted from this document."),
                "suggestion": metadata.get("suggestion", "Try a document with application names, system descriptions, or architecture content."),
            })

        # Create ArchiMate elements
        from app.models.archimate_core import ArchiMateElement

        created_ids = []
        skipped = []
        errors = []

        for elem in elements:
            name = (elem.get("name") or "").strip()
            if not name:
                continue

            el_type = elem.get("type", "ApplicationComponent")
            layer = _resolve_layer(el_type, elem.get("layer", ""))
            description = (elem.get("description") or "").strip()

            existing = ArchiMateElement.query.filter_by(name=name, type=el_type).first()
            if existing:
                skipped.append({"name": name, "type": el_type, "existing_id": existing.id})
                created_ids.append({"id": existing.id, "name": name, "type": el_type, "layer": layer})
                continue

            try:
                new_elem = ArchiMateElement(
                    name=name,
                    type=el_type,
                    layer=layer,
                    description=description or None,
                    scope="enterprise",
                )
                db.session.add(new_elem)
                db.session.flush()
                created_ids.append({"id": new_elem.id, "name": name, "type": el_type, "layer": layer})
            except Exception as row_err:
                db.session.rollback()
                errors.append({"name": name, "error": str(row_err)})

        if created_ids:
            db.session.commit()

        all_ids = [e["id"] for e in created_ids]
        return jsonify({
            "success": True,
            "created": created_ids,
            "created_count": len(created_ids) - len(skipped),
            "skipped_count": len(skipped),
            "errors": errors,
            "error_count": len(errors),
            "total_elements": len(elements),
            "element_ids": all_ids,
            "composer_url": f"/archimate/composer?elements={','.join(str(eid) for eid in all_ids)}" if all_ids else None,
        })

    except Exception as e:
        current_app.logger.error("Document analysis error: %s", e)
        db.session.rollback()
        return jsonify({"success": False, "error": f"Analysis failed: {e}"}), 500
    finally:
        # Cleanup temp file
        try:
            os.remove(file_path)
        except OSError:
            logger.exception("Failed to operation")
            pass


@archimate_bp.route("/api/create-diagram-from-elements", methods=["POST"])
@login_required
def api_create_diagram_from_elements():
    """Create a SavedDiagram from a list of element IDs with auto-layout.

    POST JSON: {"element_ids": [1, 2, 3], "name": "My Diagram"}
    Returns: SavedDiagram with composer_url for immediate viewing.
    """
    from app.models.archimate_core import (
        ArchiMateElement, SavedDiagram, SavedDiagramElement,
    )

    data = request.get_json(silent=True) or {}
    element_ids = data.get("element_ids", [])
    diagram_name = (data.get("name") or "Imported Elements Diagram").strip()

    if not element_ids:
        return jsonify({"success": False, "error": "No element IDs provided"}), 400

    # Fetch elements grouped by layer for layout
    elements = ArchiMateElement.query.filter(ArchiMateElement.id.in_(element_ids)).all()
    if not elements:
        return jsonify({"success": False, "error": "No elements found for given IDs"}), 404

    # Auto-layout: group by layer, arrange in grid
    layer_order = ["Motivation", "Strategy", "Business", "Application", "Technology", "Physical", "Implementation", "Other"]
    by_layer = {}
    for el in elements:
        layer = el.layer or "Other"
        by_layer.setdefault(layer, []).append(el)

    diagram = SavedDiagram(
        name=diagram_name,
        description=f"Auto-generated from {len(elements)} imported elements",
        created_by=current_user.id if current_user.is_authenticated else None,
    )
    db.session.add(diagram)
    db.session.flush()

    # Position elements in a grid, grouped by layer
    y_offset = 40
    cols_per_row = 4
    elem_w, elem_h = 180, 64
    gap_x, gap_y = 30, 20
    layer_gap = 40

    for layer_name in layer_order:
        layer_elements = by_layer.pop(layer_name, [])
        if not layer_elements:
            continue
        for idx, el in enumerate(layer_elements):
            col = idx % cols_per_row
            row = idx // cols_per_row
            x = 40 + col * (elem_w + gap_x)
            y = y_offset + row * (elem_h + gap_y)
            de = SavedDiagramElement(
                diagram_id=diagram.id,
                element_id=el.id,
                position_x=x,
                position_y=y,
                width=elem_w,
                height=elem_h,
                rendering_mode="black_box",
            )
            db.session.add(de)
        total_rows = (len(layer_elements) + cols_per_row - 1) // cols_per_row
        y_offset += total_rows * (elem_h + gap_y) + layer_gap

    # Handle any remaining layers not in layer_order
    for layer_name, layer_elements in by_layer.items():
        for idx, el in enumerate(layer_elements):
            col = idx % cols_per_row
            row = idx // cols_per_row
            x = 40 + col * (elem_w + gap_x)
            y = y_offset + row * (elem_h + gap_y)
            de = SavedDiagramElement(
                diagram_id=diagram.id,
                element_id=el.id,
                position_x=x,
                position_y=y,
                width=elem_w,
                height=elem_h,
                rendering_mode="black_box",
            )
            db.session.add(de)
        total_rows = (len(layer_elements) + cols_per_row - 1) // cols_per_row
        y_offset += total_rows * (elem_h + gap_y) + layer_gap

    db.session.commit()

    return jsonify({
        "success": True,
        "diagram_id": diagram.id,
        "diagram_name": diagram.name,
        "element_count": len(elements),
        "composer_url": f"/archimate/composer?viewpoint={diagram.id}",
    }), 201


# ── CMP2-004: Live metrics overlay — element-level portfolio metrics ─────────

@archimate_bp.route("/api/composer/element-metrics", methods=["GET"])
@login_required
def api_composer_element_metrics():
    """Return portfolio metrics (cost, lifecycle, risk, maturity) for canvas elements.

    Query Parameters:
        element_ids (str): Comma-separated ArchiMate element IDs.

    Returns JSON: {"metrics": {"<element_id>": {cost, lifecycle, risk, maturity}}}
    Each field is null when the data is not available.
    """
    from sqlalchemy import func

    from app.models.archimate_core import ArchiMateElement
    from app.models.application_portfolio import ApplicationComponent
    from app.models.business_capabilities import BusinessCapability
    from app.models.solution_archimate_element import SolutionArchiMateElement

    raw_ids = request.args.get("element_ids", "")
    if not raw_ids:
        return jsonify({"metrics": {}})

    try:
        element_ids = [int(x.strip()) for x in raw_ids.split(",") if x.strip()]
    except (ValueError, TypeError):
        return jsonify({"error": "element_ids must be comma-separated integers"}), 400

    if not element_ids or len(element_ids) > 500:
        return jsonify({"metrics": {}})

    elements = db.session.query(ArchiMateElement).filter(
        ArchiMateElement.id.in_(element_ids)
    ).all()
    elem_map = {e.id: e for e in elements}

    # Types that map to ApplicationComponent
    app_types = {
        "ApplicationComponent", "ApplicationService", "ApplicationFunction",
        "ApplicationInterface", "ApplicationProcess",
    }
    cap_types = {"Capability", "BusinessCapability"}

    # Collect names for batch lookup
    app_names = []
    cap_names = []
    for el in elements:
        el_type = el.type or ""
        if el_type in app_types:
            app_names.append(el.name)
        elif el_type in cap_types:
            cap_names.append(el.name)

    # Batch: ApplicationComponent by name (case-insensitive)
    app_lookup = {}
    if app_names:
        app_rows = db.session.query(ApplicationComponent).filter(
            func.lower(ApplicationComponent.name).in_([n.lower() for n in app_names])
        ).all()
        for ac in app_rows:
            app_lookup[ac.name.lower()] = ac

    # Batch: BusinessCapability by name
    cap_lookup = {}
    if cap_names:
        cap_rows = db.session.query(BusinessCapability).filter(
            func.lower(BusinessCapability.name).in_([n.lower() for n in cap_names])
        ).all()
        for bc in cap_rows:
            cap_lookup[bc.name.lower()] = bc

    # Capability maturity via classification junction (same pattern as element-maturity)
    cap_maturity = {}
    try:
        from app.models.capability_archimate_mapping import CapabilityArchiMateClassification

        mappings = (
            db.session.query(
                CapabilityArchiMateClassification.archimate_element_id,
                BusinessCapability.current_maturity_level,
                BusinessCapability.target_maturity_level,
            )
            .join(BusinessCapability, BusinessCapability.id == CapabilityArchiMateClassification.capability_id)
            .filter(CapabilityArchiMateClassification.archimate_element_id.in_(element_ids))
            .all()
        )
        for eid, current, target in mappings:
            if current and eid not in cap_maturity:
                cap_maturity[eid] = {"level": current, "target": target}
    except Exception as exc:
        current_app.logger.debug("CMP2-004: capability maturity lookup failed: %s", exc)

    metrics = {}
    for eid in element_ids:
        el = elem_map.get(eid)
        if not el:
            continue

        entry = {"cost": None, "lifecycle": None, "risk": None, "maturity": None}
        el_type = el.type or ""
        name_lower = (el.name or "").lower()

        if el_type in app_types:
            ac = app_lookup.get(name_lower)
            if ac:
                # Cost: total_cost_of_ownership from ApplicationComponent
                if ac.total_cost_of_ownership is not None:
                    entry["cost"] = round(ac.total_cost_of_ownership, 0)
                # Lifecycle: lifecycle_status
                entry["lifecycle"] = ac.lifecycle_status or None
                # Risk: vendor_risk field
                entry["risk"] = getattr(ac, "vendor_risk", None) or None

        elif el_type in cap_types:
            bc = cap_lookup.get(name_lower)
            if bc:
                entry["maturity"] = bc.current_maturity_level

        # Override maturity from classification junction if available
        if eid in cap_maturity:
            entry["maturity"] = cap_maturity[eid]["level"]

        # Fallback maturity from strategic_alignment_score
        if entry["maturity"] is None and el.strategic_alignment_score is not None:
            entry["maturity"] = max(1, min(5, int(el.strategic_alignment_score / 20) + (1 if el.strategic_alignment_score % 20 > 0 else 0)))

        metrics[str(eid)] = entry

    return jsonify({"metrics": metrics})


# ── GAP-CMP-006: OEF XML Import ─────────────────────────────────────────
# Imports ArchiMate elements and relationships from Open Exchange Format XML.
# Frontend calls this from composer_persistence.js → importOef().

# Valid element types (mirrors ArchiMateExchangeService.ALL_ELEMENT_TYPES)
_OEF_VALID_TYPES = {
    # Motivation
    "Stakeholder", "Driver", "Assessment", "Goal", "Outcome",
    "Principle", "Requirement", "Constraint", "Meaning", "Value",
    # Strategy
    "Resource", "Capability", "CourseOfAction", "ValueStream",
    # Business
    "BusinessActor", "BusinessRole", "BusinessCollaboration", "BusinessInterface",
    "BusinessProcess", "BusinessFunction", "BusinessInteraction", "BusinessEvent",
    "BusinessService", "BusinessObject", "Contract", "Representation", "Product",
    # Application
    "ApplicationComponent", "ApplicationCollaboration", "ApplicationInterface",
    "ApplicationFunction", "ApplicationInteraction", "ApplicationProcess",
    "ApplicationEvent", "ApplicationService", "DataObject",
    # Technology
    "Node", "Device", "SystemSoftware", "TechnologyCollaboration",
    "TechnologyInterface", "Path", "CommunicationNetwork", "TechnologyFunction",
    "TechnologyProcess", "TechnologyInteraction", "TechnologyEvent",
    "TechnologyService", "Artifact",
    # Physical
    "Equipment", "Facility", "DistributionNetwork", "Material",
    # Implementation
    "WorkPackage", "Deliverable", "ImplementationEvent", "Plateau", "Gap",
    # Composite
    "Grouping", "Location",
}


@archimate_bp.route("/api/import/oef", methods=["POST"])
@login_required
def api_import_oef():
    """Import ArchiMate elements and relationships from Open Exchange Format XML.

    Accepts multipart/form-data with a 'file' field containing the XML.
    Parses elements and relationships, creates/links to existing catalog entries.

    Returns JSON with:
        elements: list of imported elements with {id, name, type, layer, status}
        relationships: list of imported relationships
        stats: {elements, relationships, elements_created, elements_linked}
        warnings: list of warning messages
    """
    import xml.etree.ElementTree as ET  # noqa: PLC0415, N813

    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship  # noqa: PLC0415

    # Namespace constants (match archimate_exchange_service.py)
    ARCHIMATE_NS = "http://www.opengroup.org/xsd/archimate/3.0/"
    XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded. POST with form-data field 'file'."}), 400

    if not file.filename or not file.filename.lower().endswith(".xml"):
        return jsonify({"error": "File must be an XML file (.xml)"}), 400

    # Read and parse XML
    try:
        raw = file.read()
        if len(raw) > 10 * 1024 * 1024:
            return jsonify({"error": "File exceeds 10MB limit"}), 400
        root = ET.fromstring(raw)  # noqa: S314
    except ET.ParseError as exc:
        return jsonify({"error": f"XML parse error: {exc}"}), 400
    except Exception as exc:  # noqa: BLE001
        return jsonify({"error": f"Failed to read file: {exc}"}), 400

    # Register namespaces for lookup
    ns = {"am": ARCHIMATE_NS}

    # ── Parse elements ──────────────────────────────────────────────
    xml_elements = root.findall(".//am:element", ns)
    if not xml_elements:
        # Try without namespace (some exporters omit it)
        xml_elements = root.findall(".//element")
    # Also try direct children under <elements> container
    if not xml_elements:
        elements_container = root.find("am:elements", ns) or root.find("elements")
        if elements_container is not None:
            xml_elements = list(elements_container)

    elements_out = []
    warnings = []
    # Map XML identifier → db ID for relationship resolution
    xml_id_to_db_id = {}
    elements_created = 0
    elements_linked = 0

    for elem in xml_elements:
        # Extract type from xsi:type attribute
        elem_type = elem.get(f"{{{XSI_NS}}}type")
        if not elem_type:
            tag = elem.tag
            if "}" in tag:
                elem_type = tag.split("}")[1]
            else:
                elem_type = tag

        # Strip namespace prefix if present (e.g. "archimate:BusinessProcess")
        if ":" in elem_type:
            elem_type = elem_type.split(":")[-1]

        # Remove trailing "Type" suffix some exporters add
        if elem_type.endswith("Type") and elem_type != "DataObject":
            elem_type = elem_type[:-4]

        # Validate type is known
        if elem_type not in _OEF_VALID_TYPES:
            # Try re-joined variant (strip spaces then check)
            cleaned = _re.sub(r"\s+", "", elem_type)
            if cleaned in _OEF_VALID_TYPES:
                elem_type = cleaned
            else:
                warnings.append(f"Unknown element type '{elem_type}', defaulting to Grouping")
                elem_type = "Grouping"

        # Extract name
        name_el = elem.find(f"{{{ARCHIMATE_NS}}}name")
        if name_el is None:
            name_el = elem.find("name")
        name = (name_el.text if name_el is not None and name_el.text else
                elem.get("name", "Unnamed Element"))

        # Extract documentation
        doc_el = elem.find(f"{{{ARCHIMATE_NS}}}documentation")
        if doc_el is None:
            doc_el = elem.find("documentation")
        description = doc_el.text if doc_el is not None and doc_el.text else None

        # Determine layer
        layer = _resolve_layer(elem_type)

        # XML identifier for relationship mapping
        xml_id = elem.get("identifier", "")

        # De-duplicate: check for existing element with same name + type
        existing = ArchiMateElement.query.filter_by(name=name, type=elem_type).first()
        if existing:
            xml_id_to_db_id[xml_id] = existing.id
            elements_linked += 1
            elements_out.append({
                "id": existing.id,
                "name": existing.name,
                "type": existing.type,
                "layer": (existing.layer or layer).lower(),
                "status": "linked",
            })
        else:
            try:
                new_el = ArchiMateElement(
                    name=name,
                    type=elem_type,
                    layer=layer,
                    description=description,
                )
                db.session.add(new_el)
                db.session.flush()
                xml_id_to_db_id[xml_id] = new_el.id
                elements_created += 1
                elements_out.append({
                    "id": new_el.id,
                    "name": new_el.name,
                    "type": new_el.type,
                    "layer": (new_el.layer or layer).lower(),
                    "status": "created",
                })
            except Exception as exc:  # noqa: BLE001
                db.session.rollback()
                warnings.append(f"Failed to create element '{name}': {exc}")
                continue

    # ── Parse relationships ─────────────────────────────────────────
    xml_rels = root.findall(".//am:relationship", ns)
    if not xml_rels:
        xml_rels = root.findall(".//relationship")
    if not xml_rels:
        rels_container = root.find("am:relationships", ns) or root.find("relationships")
        if rels_container is not None:
            xml_rels = list(rels_container)

    relationships_out = []
    rels_created = 0

    for rel in xml_rels:
        # Extract relationship type
        rel_type_raw = rel.get(f"{{{XSI_NS}}}type")
        if not rel_type_raw:
            tag = rel.tag
            if "}" in tag:
                rel_type_raw = tag.split("}")[1]
            else:
                rel_type_raw = tag

        # Strip namespace prefix
        if ":" in rel_type_raw:
            rel_type_raw = rel_type_raw.split(":")[-1]

        # Normalize using the existing function
        rel_type = _normalize_rel_type(rel_type_raw)

        # Get source and target XML identifiers
        source_xml_id = rel.get("source", "")
        target_xml_id = rel.get("target", "")

        source_db_id = xml_id_to_db_id.get(source_xml_id)
        target_db_id = xml_id_to_db_id.get(target_xml_id)

        if not source_db_id or not target_db_id:
            warnings.append(
                f"Relationship {rel_type}: could not resolve "
                f"source ({source_xml_id}) or target ({target_xml_id})"
            )
            continue

        # Check for existing relationship to avoid duplicates
        existing_rel = ArchiMateRelationship.query.filter_by(
            source_id=source_db_id,
            target_id=target_db_id,
            type=rel_type,
        ).first()

        if existing_rel:
            relationships_out.append({
                "id": existing_rel.id,
                "source_id": existing_rel.source_id,
                "target_id": existing_rel.target_id,
                "type": rel_type,
                "status": "linked",
            })
            continue

        try:
            new_rel = ArchiMateRelationship(
                type=rel_type,
                source_id=source_db_id,
                target_id=target_db_id,
                created_by_id=current_user.id if hasattr(current_user, "id") else None,
            )
            db.session.add(new_rel)
            db.session.flush()
            rels_created += 1
            relationships_out.append({
                "id": new_rel.id,
                "source_id": new_rel.source_id,
                "target_id": new_rel.target_id,
                "type": rel_type,
                "status": "created",
            })
        except Exception as exc:  # noqa: BLE001
            db.session.rollback()
            warnings.append(f"Failed to create relationship {rel_type}: {exc}")
            continue

    # ── Commit ──────────────────────────────────────────────────────
    try:
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        return jsonify({"error": f"Database commit failed: {exc}"}), 500

    return jsonify({
        "elements": elements_out,
        "relationships": relationships_out,
        "stats": {
            "elements": len(elements_out),
            "relationships": len(relationships_out),
            "elements_created": elements_created,
            "elements_linked": elements_linked,
            "relationships_created": rels_created,
        },
        "warnings": warnings,
    })


# ── GOV-01: Cross-solution dependency tracking ──────────────────────────


@archimate_bp.route("/api/element/<int:element_id>/dependent-solutions", methods=["GET"])
@login_required
def api_element_dependent_solutions(element_id):
    """Return all solutions that reference a given ArchiMate element.

    This enables cross-solution dependency analysis: if two solutions share
    the same element, a change to that element impacts both.
    """
    from app.models.archimate_core import ArchiMateElement
    from app.models.solution_archimate_element import SolutionArchiMateElement
    from app.models.solution_models import Solution

    element = db.session.get(ArchiMateElement, element_id)
    if not element:
        return api_error("Element not found", 404)

    links = (
        db.session.query(SolutionArchiMateElement)
        .filter(SolutionArchiMateElement.element_id == element_id)
        .all()
    )

    solutions = []
    for link in links:
        sol = db.session.get(Solution, link.solution_id)
        if sol:
            solutions.append({
                "id": sol.id,
                "name": getattr(sol, "name", None) or getattr(sol, "title", "Untitled"),
                "status": getattr(sol, "status", None),
                "governance_status": getattr(sol, "governance_status", None),
                "element_role": link.element_role,
                "linked_at": link.created_at.isoformat() if link.created_at else None,
            })

    return jsonify({
        "element": {
            "id": element.id,
            "name": element.name,
            "type": element.type,
            "layer": element.layer,
        },
        "solutions": solutions,
        "total": len(solutions),
    })


@archimate_bp.route("/api/shared-elements", methods=["GET"])
@login_required
def api_shared_elements():
    """Return ArchiMate elements referenced by 2+ solutions.

    This is the core governance query: elements shared across solutions
    represent architectural coupling. Changes to shared elements have
    cross-solution impact.

    Query Parameters:
        min_solutions (int): Minimum solution count to include (default 2).
        limit (int): Max results to return (default 20, max 100).
    """
    from app.models.archimate_core import ArchiMateElement
    from app.models.solution_archimate_element import SolutionArchiMateElement

    min_solutions = request.args.get("min_solutions", 2, type=int)
    if min_solutions < 2:
        min_solutions = 2
    limit = min(request.args.get("limit", 20, type=int), 100)

    # Subquery: element_id → distinct solution count
    shared_q = (
        db.session.query(
            SolutionArchiMateElement.element_id,
            db.func.count(db.func.distinct(SolutionArchiMateElement.solution_id)).label("solution_count"),
        )
        .group_by(SolutionArchiMateElement.element_id)
        .having(db.func.count(db.func.distinct(SolutionArchiMateElement.solution_id)) >= min_solutions)
        .subquery()
    )

    rows = (
        db.session.query(ArchiMateElement, shared_q.c.solution_count)
        .join(shared_q, ArchiMateElement.id == shared_q.c.element_id)
        .order_by(shared_q.c.solution_count.desc())
        .limit(limit)
        .all()
    )

    elements = []
    for elem, sol_count in rows:
        elements.append({
            "id": elem.id,
            "name": elem.name,
            "type": elem.type,
            "layer": elem.layer,
            "solution_count": sol_count,
        })

    return jsonify({
        "shared_elements": elements,
        "total": len(elements),
        "min_solutions": min_solutions,
    })


# ── GOV-01: Cross-Solution Dependency Graph ──────────────────────────────────


@archimate_bp.route("/api/impact-analysis/<int:element_id>", methods=["GET"])
@login_required
def get_impact_analysis(element_id):
    """Return impact analysis for a given ArchiMate element.

    Shows which solutions directly and transitively depend on this element,
    enabling change-impact assessment before modifying shared architecture
    components.

    Response shape:
        element            — id, name, type
        direct_dependents  — solutions that directly reference this element
        transitive_dependents — solutions reachable via ArchiMate relationships
        risk_score         — 0.0–1.0, proportional to total dependent solutions
    """
    from app.models.archimate_core import ArchiMateElement, ArchiMateRelationship
    from app.models.solution_archimate_element import SolutionArchiMateElement
    from app.models.solution_models import Solution

    element = db.session.get(ArchiMateElement, element_id)
    if not element:
        return jsonify({"error": "Element not found"}), 404

    # ── Direct dependents ────────────────────────────────────────────────────
    direct_links = (
        db.session.query(SolutionArchiMateElement, Solution)
        .join(Solution, SolutionArchiMateElement.solution_id == Solution.id)
        .filter(SolutionArchiMateElement.element_id == element_id)
        .all()
    )

    direct_dependents = []
    direct_solution_ids = set()
    for link, sol in direct_links:
        direct_dependents.append({
            "solution_id": sol.id,
            "solution_name": sol.name,
            "role": link.element_role,
        })
        direct_solution_ids.add(sol.id)

    # ── Transitive dependents ────────────────────────────────────────────────
    # Find elements connected to this element via ArchiMate relationships,
    # then find solutions that use THOSE connected elements.
    connected_element_ids = (
        db.session.query(ArchiMateRelationship.target_id)
        .filter(ArchiMateRelationship.source_id == element_id)
        .union(
            db.session.query(ArchiMateRelationship.source_id)
            .filter(ArchiMateRelationship.target_id == element_id)
        )
        .subquery()
    )

    transitive_rows = (
        db.session.query(
            SolutionArchiMateElement.solution_id,
            SolutionArchiMateElement.element_id.label("via_element_id"),
            ArchiMateElement.name.label("via_element_name"),
            Solution.name.label("solution_name"),
        )
        .join(ArchiMateElement, SolutionArchiMateElement.element_id == ArchiMateElement.id)
        .join(Solution, SolutionArchiMateElement.solution_id == Solution.id)
        .filter(SolutionArchiMateElement.element_id.in_(connected_element_ids))
        .filter(~SolutionArchiMateElement.solution_id.in_(list(direct_solution_ids)))
        .all()
    )

    seen_transitive = set()
    transitive_dependents = []
    for row in transitive_rows:
        key = (row.solution_id, row.via_element_id)
        if key not in seen_transitive:
            seen_transitive.add(key)
            transitive_dependents.append({
                "solution_id": row.solution_id,
                "solution_name": row.solution_name,
                "via_element_id": row.via_element_id,
                "via_element_name": row.via_element_name,
            })

    total_dependents = len(direct_solution_ids) + len({r["solution_id"] for r in transitive_dependents})
    risk_score = round(min(1.0, total_dependents * 0.1), 2)

    return jsonify({
        "element": {
            "id": element.id,
            "name": element.name,
            "type": element.type,
        },
        "direct_dependents": direct_dependents,
        "transitive_dependents": transitive_dependents,
        "risk_score": risk_score,
    })
