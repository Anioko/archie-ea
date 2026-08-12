"""Bulk import of a capability model from CSV or Excel.

The gap this closes: Archie could import an *application inventory* and could
not import a capability model. A business architect arriving with an existing
Level 1–3 model — which is every business architect, since the model predates
the tool — had no path but re-keying it one capability at a time.

Three properties make this usable on a real model rather than a demo:

**It plans before it writes.** ``plan()`` reads the file, resolves it against
what the organisation already has, and returns exactly what would be created,
updated, left alone and rejected — without touching the database. The UI shows
that, and ``apply()`` re-plans from the same rows rather than trusting anything
the browser sends back.

**It is idempotent.** Rows are matched on ``code`` when the file supplies one
and on name-within-parent otherwise, so importing the same file twice produces
the same model rather than a second copy of it. That matters more than it
sounds: an import you cannot safely repeat is an import nobody dares run on
production data.

**A rejected row does not cost you the file.** Errors are per row, carry the
1-based line number the spreadsheet shows, and never abort the rows around
them. The alternative — one bad maturity value rolling back 400 good
capabilities — is what makes people give up and re-key.

Hierarchy is resolved in two passes: every capability is inserted first, then
parents are linked, so a file may reference a parent that appears below its
child. Cycles are detected in the plan and rejected there, because a cycle
committed to the database is a tree that no longer terminates.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional, Tuple

from app import db

MAX_ROWS = 5000
MAX_BYTES = 10 * 1024 * 1024

STRATEGIC_IMPORTANCE = {"critical", "high", "medium", "low"}

# Header aliases. Business architects export from Excel, Sparx, LeanIX and
# APQC, and none of them agree on a spelling; rejecting a file over "Capability
# Name" vs "name" would be the tool's problem presented as the user's.
COLUMN_ALIASES: Dict[str, str] = {
    "code": "code",
    "capability code": "code",
    "id": "code",
    "capability id": "code",
    "ref": "code",
    "name": "name",
    "capability": "name",
    "capability name": "name",
    "title": "name",
    "description": "description",
    "definition": "description",
    "summary": "description",
    "level": "level",
    "tier": "level",
    "capability level": "level",
    "parent": "parent",
    "parent code": "parent",
    "parent id": "parent",
    "parent capability": "parent",
    "parent name": "parent",
    "domain": "business_domain",
    "business domain": "business_domain",
    "capability domain": "business_domain",
    "category": "category",
    "type": "category",
    "capability type": "category",
    "business owner": "business_owner",
    "owner": "business_owner",
    "accountable": "business_owner",
    "it owner": "it_owner",
    "technology owner": "it_owner",
    "strategic importance": "strategic_importance",
    "importance": "strategic_importance",
    "criticality": "strategic_importance",
    "current maturity": "current_maturity",
    "current maturity level": "current_maturity",
    "as-is maturity": "current_maturity",
    "target maturity": "target_maturity",
    "target maturity level": "target_maturity",
    "to-be maturity": "target_maturity",
    "notes": "notes",
    "assessment notes": "notes",
}

TEMPLATE_HEADERS = [
    "code",
    "name",
    "description",
    "level",
    "parent",
    "business_domain",
    "category",
    "business_owner",
    "strategic_importance",
    "current_maturity",
    "target_maturity",
]

TEMPLATE_ROWS = [
    ["CAP-01", "Supply Chain", "Plan, source, make and deliver", "1", "", "Operations", "Operations", "COO", "critical", "", ""],
    ["CAP-01-01", "Demand Planning", "Forecast demand across channels", "2", "CAP-01", "Operations", "Operations", "Head of Planning", "high", "2", "4"],
    ["CAP-01-02", "Inventory Management", "Hold the right stock in the right place", "2", "CAP-01", "Operations", "Operations", "Head of Logistics", "high", "3", "4"],
    ["CAP-02", "Customer Management", "Win, serve and retain customers", "1", "", "Commercial", "Customer Management", "CCO", "critical", "", ""],
]


@dataclass
class RowError:
    line: int
    column: Optional[str]
    message: str

    def as_dict(self) -> Dict[str, Any]:
        return {"line": self.line, "column": self.column, "message": self.message}


@dataclass
class PlannedRow:
    line: int
    action: str  # "create" | "update" | "unchanged"
    values: Dict[str, Any]
    existing_id: Optional[int] = None
    changes: Dict[str, Tuple[Any, Any]] = field(default_factory=dict)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "line": self.line,
            "action": self.action,
            "code": self.values.get("code"),
            "name": self.values.get("name"),
            "level": self.values.get("level"),
            "parent": self.values.get("parent"),
            "changes": {k: {"from": v[0], "to": v[1]} for k, v in self.changes.items()},
        }


@dataclass
class ImportPlan:
    rows: List[PlannedRow] = field(default_factory=list)
    errors: List[RowError] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    columns_seen: List[str] = field(default_factory=list)
    columns_ignored: List[str] = field(default_factory=list)

    @property
    def creates(self) -> int:
        return sum(1 for r in self.rows if r.action == "create")

    @property
    def updates(self) -> int:
        return sum(1 for r in self.rows if r.action == "update")

    @property
    def unchanged(self) -> int:
        return sum(1 for r in self.rows if r.action == "unchanged")

    def as_dict(self) -> Dict[str, Any]:
        return {
            "creates": self.creates,
            "updates": self.updates,
            "unchanged": self.unchanged,
            "errors": [e.as_dict() for e in self.errors],
            "warnings": self.warnings,
            "columns_seen": self.columns_seen,
            "columns_ignored": self.columns_ignored,
            "rows": [r.as_dict() for r in self.rows],
        }


class CapabilityImportError(Exception):
    """The file could not be read at all — as distinct from rows being rejected."""


class CapabilityImportService:
    """Parse, plan and apply a capability-model import."""

    # ------------------------------------------------------------------
    # Reading the file
    # ------------------------------------------------------------------

    @staticmethod
    def template_csv() -> str:
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        writer.writerow(TEMPLATE_HEADERS)
        writer.writerows(TEMPLATE_ROWS)
        return buf.getvalue()

    @classmethod
    def read(cls, stream: io.BufferedIOBase, filename: str) -> Tuple[List[Dict[str, str]], List[str]]:
        """Return (raw rows keyed by canonical column, headers seen).

        Raises CapabilityImportError when the file itself is unusable; a file
        that parses but contains bad values is the plan's problem, not this
        method's.
        """
        data = stream.read(MAX_BYTES + 1)
        if not data:
            raise CapabilityImportError("The file is empty.")
        if len(data) > MAX_BYTES:
            raise CapabilityImportError(
                f"The file is larger than {MAX_BYTES // (1024 * 1024)} MB."
            )

        lower = (filename or "").lower()
        if lower.endswith((".xlsx", ".xlsm")):
            rows, headers = cls._read_excel(data)
        elif lower.endswith(".csv") or lower.endswith(".txt"):
            rows, headers = cls._read_csv(data)
        else:
            raise CapabilityImportError(
                "Only .csv and .xlsx files are accepted. Export your model as "
                "either and try again."
            )

        if not headers:
            raise CapabilityImportError("The first row must be a header row.")
        return rows, headers

    @staticmethod
    def _decode(data: bytes) -> str:
        # Excel writes UTF-8 with a BOM, and a fair number of European exports
        # are cp1252. Failing the whole import on a stray accent is not
        # acceptable when the alternative is one substitution.
        for encoding in ("utf-8-sig", "utf-8", "cp1252"):
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    @classmethod
    def _read_csv(cls, data: bytes) -> Tuple[List[Dict[str, str]], List[str]]:
        text_data = cls._decode(data)
        sample = text_data[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel  # a single-column file sniffs as nothing
        reader = csv.reader(io.StringIO(text_data), dialect)
        rows = list(reader)
        if not rows:
            raise CapabilityImportError("The file has no rows.")
        headers = [h.strip() for h in rows[0]]
        return cls._to_dicts(headers, rows[1:]), headers

    @classmethod
    def _read_excel(cls, data: bytes) -> Tuple[List[Dict[str, str]], List[str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - openpyxl is a pinned dependency
            raise CapabilityImportError(
                "Excel support is unavailable on this deployment; export as CSV."
            ) from exc

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = wb.active
        raw = []
        for row in sheet.iter_rows(values_only=True):
            raw.append(["" if v is None else str(v).strip() for v in row])
            if len(raw) > MAX_ROWS + 1:
                break
        wb.close()
        if not raw:
            raise CapabilityImportError("The first sheet has no rows.")
        headers = [h.strip() for h in raw[0]]
        return cls._to_dicts(headers, raw[1:]), headers

    @staticmethod
    def _to_dicts(headers: List[str], rows: Iterable[List[str]]) -> List[Dict[str, str]]:
        canonical = [COLUMN_ALIASES.get(h.strip().lower().replace("_", " ")) for h in headers]
        out = []
        for offset, values in enumerate(rows):
            if not any((v or "").strip() for v in values):
                continue  # a blank spacer row is not an error
            record: Dict[str, str] = {"__line__": str(offset + 2)}
            for idx, key in enumerate(canonical):
                if key is None:
                    continue
                value = values[idx] if idx < len(values) else ""
                record[key] = (value or "").strip()
            out.append(record)
        return out

    # ------------------------------------------------------------------
    # Planning
    # ------------------------------------------------------------------

    @classmethod
    def plan(cls, raw_rows: List[Dict[str, str]], headers: List[str]) -> ImportPlan:
        """Resolve the file against the current organisation. Writes nothing."""
        from app.models.business_capabilities import BusinessCapability

        plan = ImportPlan(columns_seen=list(headers))
        recognised = {COLUMN_ALIASES.get(h.strip().lower().replace("_", " ")) for h in headers}
        plan.columns_ignored = [
            h
            for h in headers
            if COLUMN_ALIASES.get(h.strip().lower().replace("_", " ")) is None and h.strip()
        ]
        if "name" not in recognised:
            plan.errors.append(
                RowError(1, None, "No 'name' column found. A capability needs a name.")
            )
            return plan

        if len(raw_rows) > MAX_ROWS:
            plan.errors.append(
                RowError(
                    1, None,
                    f"The file has {len(raw_rows)} rows; the limit is {MAX_ROWS}. "
                    "Split it by domain and import each part.",
                )
            )
            return plan

        # BusinessCapability is a TenantMixin model: this query is already
        # scoped to the caller's organisation by do_orm_execute.
        existing = BusinessCapability.query.all()
        by_code = {c.code.strip().lower(): c for c in existing if c.code}
        by_name = {}
        for c in existing:
            by_name.setdefault(c.name.strip().lower(), c)

        cleaned: List[Tuple[int, Dict[str, Any]]] = []
        seen_codes: Dict[str, int] = {}
        seen_names: Dict[str, int] = {}

        for record in raw_rows:
            line = int(record.get("__line__", "0"))
            values, errors = cls._clean(record, line)
            if errors:
                plan.errors.extend(errors)
                continue

            code_key = (values.get("code") or "").strip().lower()
            name_key = values["name"].strip().lower()

            if code_key:
                if code_key in seen_codes:
                    plan.errors.append(
                        RowError(line, "code",
                                 f"Code {values['code']!r} is already used on line {seen_codes[code_key]}.")
                    )
                    continue
                seen_codes[code_key] = line
            else:
                # Without a code, name is the identity — so it has to be unique
                # in the file, or two rows silently become one capability.
                if name_key in seen_names:
                    plan.errors.append(
                        RowError(line, "name",
                                 f"{values['name']!r} appears on line {seen_names[name_key]} too. "
                                 "Give the rows distinct codes if they are different capabilities.")
                    )
                    continue
                seen_names[name_key] = line

            cleaned.append((line, values))

        # Parent resolution: a parent may be a code or a name, and may appear
        # anywhere in the file — including below its child.
        file_codes = {v.get("code", "").strip().lower(): ln for ln, v in cleaned if v.get("code")}
        file_names = {v["name"].strip().lower(): ln for ln, v in cleaned}
        parent_of: Dict[int, Optional[int]] = {}

        for line, values in cleaned:
            parent_ref = (values.get("parent") or "").strip()
            if not parent_ref:
                parent_of[line] = None
                continue
            key = parent_ref.lower()
            if key in file_codes:
                parent_of[line] = file_codes[key]
            elif key in file_names:
                parent_of[line] = file_names[key]
            elif key in by_code or key in by_name:
                parent_of[line] = None  # resolved against the database at apply time
                values["__db_parent__"] = key
            else:
                plan.errors.append(
                    RowError(line, "parent",
                             f"Parent {parent_ref!r} is not in this file and does not exist yet. "
                             "Add the parent row, or import the parent level first.")
                )
                parent_of[line] = None
                values["__unresolved_parent__"] = True

        # Cycle detection, across the file's own edges AND through the database.
        #
        # Walking only the file's edges was not enough. A parent that resolves
        # against an existing capability set parent_of[line] = None, so the walk
        # stopped at the first step and reported nothing — while apply() went on
        # to write that same reference. Concretely: the database holds A (root)
        # and B (child of A); import one row `code=A, parent=B`; it plans as a
        # clean update and commits A → B → A. The hierarchy view and the report
        # builder were both hardened against unbounded recursion in this branch,
        # which is the tell that a cycle is reachable; every other consumer of
        # parent_capability_id is not.
        #
        # `existing_parent` walks upward through what is already stored, with
        # rows this import is about to re-parent overridden by their new parent.
        by_id = {c.id: c for c in existing}
        row_for_capability = {}
        for line, values in cleaned:
            code_key = (values.get("code") or "").strip().lower()
            match = by_code.get(code_key) if code_key else by_name.get(values["name"].strip().lower())
            if match is not None:
                row_for_capability[match.id] = line

        def resolves_to_capability(line_no):
            """The stored capability a row's parent reference points at, if any."""
            values = dict(cleaned_by_line[line_no])
            key = (values.get("parent") or "").strip().lower()
            if not key:
                return None
            return by_code.get(key) or by_name.get(key)

        cleaned_by_line = {ln: v for ln, v in cleaned}

        for line, values in cleaned:
            if values.get("__unresolved_parent__"):
                continue

            seen_lines, seen_ids = set(), set()
            cursor_line = parent_of.get(line)
            cursor_cap = resolves_to_capability(line)
            cycle = False

            # Follow file edges first.
            while cursor_line is not None:
                if cursor_line in seen_lines or cursor_line == line:
                    cycle = True
                    break
                seen_lines.add(cursor_line)
                cursor_cap = resolves_to_capability(cursor_line)
                cursor_line = parent_of.get(cursor_line)

            # Then continue through stored ancestry, honouring any re-parenting
            # this same import performs on the way up.
            while not cycle and cursor_cap is not None:
                if cursor_cap.id in seen_ids:
                    cycle = True
                    break
                seen_ids.add(cursor_cap.id)

                overriding_line = row_for_capability.get(cursor_cap.id)
                if overriding_line is not None and overriding_line != line:
                    if overriding_line in seen_lines:
                        cycle = True
                        break
                    seen_lines.add(overriding_line)
                    cursor_cap = resolves_to_capability(overriding_line)
                    continue
                if overriding_line == line:
                    cycle = True
                    break
                cursor_cap = by_id.get(cursor_cap.parent_capability_id)

            if cycle:
                plan.errors.append(
                    RowError(line, "parent",
                             f"{values['name']!r} is its own ancestor. A capability tree "
                             "cannot contain a cycle.")
                )
                values["__cycle__"] = True

        for line, values in cleaned:
            if values.get("__unresolved_parent__") or values.get("__cycle__"):
                continue

            code_key = (values.get("code") or "").strip().lower()
            match = by_code.get(code_key) if code_key else by_name.get(values["name"].strip().lower())

            if match is None:
                plan.rows.append(PlannedRow(line=line, action="create", values=values))
                continue

            changes = {}
            for field_name in (
                "name", "description", "level", "business_domain", "category",
                "business_owner", "it_owner", "strategic_importance",
            ):
                if field_name not in values:
                    continue
                new = values[field_name]
                if new in (None, ""):
                    continue  # a blank cell means "leave it alone", not "erase it"
                old = getattr(match, field_name, None)
                if old != new:
                    changes[field_name] = (old, new)

            for src, dest in (("current_maturity", "current_maturity_level"),
                              ("target_maturity", "target_maturity_level")):
                if values.get(src) is None:
                    continue
                old = getattr(match, dest, None)
                if old != values[src]:
                    changes[dest] = (old, values[src])

            plan.rows.append(
                PlannedRow(
                    line=line,
                    action="update" if changes else "unchanged",
                    values=values,
                    existing_id=match.id,
                    changes=changes,
                )
            )

        if plan.columns_ignored:
            plan.warnings.append(
                "Ignored column(s): " + ", ".join(plan.columns_ignored)
                + ". Nothing in them will be imported."
            )

        return plan

    @staticmethod
    def _clean(record: Dict[str, str], line: int) -> Tuple[Dict[str, Any], List[RowError]]:
        errors: List[RowError] = []
        values: Dict[str, Any] = {}

        name = (record.get("name") or "").strip()
        if not name:
            errors.append(RowError(line, "name", "Name is required."))
        elif len(name) > 256:
            errors.append(RowError(line, "name", "Name is longer than 256 characters."))
        else:
            values["name"] = name

        code = (record.get("code") or "").strip()
        if code:
            if len(code) > 50:
                errors.append(RowError(line, "code", "Code is longer than 50 characters."))
            elif not re.match(r"^[\w.\-/ ]+$", code):
                errors.append(
                    RowError(line, "code", f"Code {code!r} contains characters that are not letters, "
                                           "digits, dot, dash, slash or space.")
                )
            else:
                values["code"] = code

        level_raw = (record.get("level") or "").strip()
        if level_raw:
            try:
                level = int(float(level_raw))
            except ValueError:
                errors.append(RowError(line, "level", f"Level {level_raw!r} is not a number."))
            else:
                if not 1 <= level <= 5:
                    errors.append(RowError(line, "level", f"Level {level} is outside 1–5."))
                else:
                    values["level"] = level

        for src in ("current_maturity", "target_maturity"):
            raw = (record.get(src) or "").strip()
            if not raw:
                continue
            try:
                maturity = int(float(raw))
            except ValueError:
                errors.append(RowError(line, src, f"{raw!r} is not a maturity level."))
                continue
            if not 1 <= maturity <= 5:
                errors.append(RowError(line, src, f"Maturity {maturity} is outside 1–5."))
                continue
            values[src] = maturity

        importance = (record.get("strategic_importance") or "").strip().lower()
        if importance:
            if importance not in STRATEGIC_IMPORTANCE:
                errors.append(
                    RowError(line, "strategic_importance",
                             f"{importance!r} is not one of: "
                             + ", ".join(sorted(STRATEGIC_IMPORTANCE)))
                )
            else:
                values["strategic_importance"] = importance

        for src in ("description", "business_domain", "category", "business_owner",
                    "it_owner", "parent", "notes"):
            raw = (record.get(src) or "").strip()
            if raw:
                values[src] = raw

        return values, errors

    # ------------------------------------------------------------------
    # Applying
    # ------------------------------------------------------------------

    @classmethod
    def apply(cls, plan: ImportPlan) -> Dict[str, Any]:
        """Commit a plan. Returns a summary; raises nothing a caller must catch.

        Everything commits or nothing does. A partially imported capability
        model is worse than none: the operator cannot tell which half is real,
        and re-running is the only recovery — which is exactly what a
        half-written parent link makes unsafe.
        """
        from app.models.business_capabilities import BusinessCapability

        actionable = [r for r in plan.rows if r.action in ("create", "update")]
        if not actionable:
            return {"created": 0, "updated": 0, "unchanged": plan.unchanged, "archimate_elements": 0}

        existing = BusinessCapability.query.all()
        by_code = {c.code.strip().lower(): c for c in existing if c.code}
        by_name = {}
        for c in existing:
            by_name.setdefault(c.name.strip().lower(), c)

        created, updated = 0, 0
        by_line: Dict[int, BusinessCapability] = {}

        try:
            # Pass 1 — rows, without parents.
            for row in actionable:
                values = row.values
                if row.action == "create":
                    cap = BusinessCapability(
                        name=values["name"],
                        code=values.get("code") or None,
                        description=values.get("description"),
                        level=values.get("level", 1),
                        business_domain=values.get("business_domain"),
                        category=values.get("category"),
                        business_owner=values.get("business_owner"),
                        it_owner=values.get("it_owner"),
                        strategic_importance=values.get("strategic_importance"),
                    )
                    db.session.add(cap)
                    created += 1
                else:
                    cap = db.session.get(BusinessCapability, row.existing_id)
                    if cap is None:  # deleted between plan and apply
                        continue
                    for field_name, (_old, new) in row.changes.items():
                        setattr(cap, field_name, new)
                    updated += 1

                is_create = row.action == "create"
                cls._apply_maturity(cap, values, is_create=is_create)
                db.session.flush()
                if is_create:
                    # The column defaults (current 1, target 3) are applied by
                    # SQLAlchemy at INSERT even when the attribute is explicitly
                    # None, so a level the file never mentioned comes back as a
                    # number. Null it after the flush, where the assignment
                    # becomes a plain UPDATE and the default cannot re-apply.
                    if values.get("current_maturity") is None:
                        cap.current_maturity_level = None
                    if values.get("target_maturity") is None:
                        cap.target_maturity_level = None
                    if cap.current_maturity_level is None or cap.target_maturity_level is None:
                        cap.maturity_gap = None
                    db.session.flush()
                by_line[row.line] = cap

            # Pass 2 — parents, now that every row has an id.
            for row in actionable:
                cap = by_line.get(row.line)
                if cap is None:
                    continue
                parent_ref = (row.values.get("parent") or "").strip().lower()
                if not parent_ref:
                    continue
                parent = None
                for candidate in actionable:
                    other = by_line.get(candidate.line)
                    if other is None:
                        continue
                    if (other.code or "").strip().lower() == parent_ref or \
                       other.name.strip().lower() == parent_ref:
                        parent = other
                        break
                if parent is None:
                    parent = by_code.get(parent_ref) or by_name.get(parent_ref)
                if parent is not None and parent.id != cap.id:
                    cap.parent_capability_id = parent.id

            elements = cls._sync_archimate([by_line[r.line] for r in actionable if r.line in by_line])
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise

        return {
            "created": created,
            "updated": updated,
            "unchanged": plan.unchanged,
            "archimate_elements": elements,
        }

    @staticmethod
    def _apply_maturity(cap, values: Dict[str, Any], is_create: bool = False) -> None:
        """Set maturity, and stamp the assessment date on the same write.

        ``current_maturity_level`` defaults to 1 and ``target_maturity_level``
        defaults to 3, so neither is ever NULL on a row this importer creates.
        Every reader that has to tell "assessed at level 1" from "never
        assessed" keys off ``maturity_assessment_date`` — which means stamping
        it publishes *both* columns as measurements, including the one the file
        never mentioned.

        Concretely: import ``code,name,current maturity`` with no target column,
        and every capability is reported in the PDF, the Word file, the workbook
        and the frameworks table as having a target of 3 and an average target
        of 3.0 that nobody entered, with a maturity_gap computed from it.

        So on a create the unsupplied column is set to NULL explicitly, beating
        the default. On an update it is left alone — a blank cell means "leave
        this alone" everywhere else in this importer, and the existing value is
        the tenant's data rather than something invented here.
        """
        from datetime import datetime

        current = values.get("current_maturity")
        target = values.get("target_maturity")

        if current is not None:
            cap.current_maturity_level = current
        elif is_create:
            cap.current_maturity_level = None

        if target is not None:
            cap.target_maturity_level = target
        elif is_create:
            cap.target_maturity_level = None

        if values.get("notes"):
            cap.maturity_assessment_notes = values["notes"]

        if current is None and target is None:
            return

        cap.maturity_assessment_date = datetime.utcnow()
        # Only a gap between two real numbers. Deriving it from a default would
        # put an invented number in the "Gap" column of a board paper.
        if cap.current_maturity_level is not None and cap.target_maturity_level is not None:
            cap.maturity_gap = cap.target_maturity_level - cap.current_maturity_level
        else:
            cap.maturity_gap = None

    @staticmethod
    def _sync_archimate(capabilities: List[Any]) -> int:
        """Mirror the capability hierarchy onto the ArchiMate elements.

        ArchiMate is the backbone of this repository rather than a view of it,
        so a capability that exists only in ``business_capability`` is invisible
        to the composer, the traceability matrix and every viewpoint.

        The *elements* are created by the ``before_insert`` listener on
        ``BusinessCapability``, so this does not create them — creating a second
        one would orphan the first. What the listener does not do, and what a
        decomposition needs, is set ``parent_id``: without it every imported
        element sits at the root and the composer shows a flat list where the
        map shows a tree.

        Returns the number of elements whose parent this linked.
        """
        try:
            from app.models.archimate_core import ArchiMateElement
            from app.models.business_capabilities import BusinessCapability
        except ImportError:
            return 0

        linked = 0
        for cap in capabilities:
            element_id = getattr(cap, "archimate_element_id", None)
            parent_id = getattr(cap, "parent_capability_id", None)
            if not element_id or not parent_id:
                continue
            parent = db.session.get(BusinessCapability, parent_id)
            if parent is None or not getattr(parent, "archimate_element_id", None):
                continue
            element = db.session.get(ArchiMateElement, element_id)
            if element is not None and element.parent_id != parent.archimate_element_id:
                element.parent_id = parent.archimate_element_id
                linked += 1
        return linked
