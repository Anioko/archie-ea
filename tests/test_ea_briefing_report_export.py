"""The EA briefing as a deliverable — PDF, Word, Excel, shareable page.

``EnterpriseBriefingService`` already computes the findings and persists them;
what it could not do was hand somebody the result. A briefing whose only form is
a web page gets screenshotted into a slide, which loses the ``action_url`` —
the link that says where each finding can be verified, and the reason the
artefact is worth circulating at all.

The properties these tests exist to protect:

* **The export reads, it does not recompute.** The findings and counts come
  back exactly as they were persisted, so the document and the page it was
  exported from say the same thing. An export that re-ran the gatherers would
  quietly produce a different briefing from the one on screen.
* **A count that was never written is an em dash, not a zero.**
  ``finding_count`` is nullable; a briefing with no count recorded is not a
  briefing that found nothing.
* **One builder feeds all four renderers.**
"""

from __future__ import annotations

import io
import uuid

import pytest

from app.modules.solutions_strategic.v2.services.briefing_report_service import (
    BriefingReportError,
    BriefingReportService,
)

FINDINGS = [
    {
        "category": "portfolio",
        "severity": "info",
        "title": "Portfolio: 120 applications under management",
        "detail": "4 applications are in the sunset pipeline.",
        "evidence": "Application lifecycle distribution",
        "action_label": "Open applications",
        "action_url": "/applications/",
    },
    {
        "category": "drift",
        "severity": "high",
        "title": "Drift detected on Core Banking Replacement",
        "detail": "Scope changed after the last gate.",
        "evidence": "Snapshot 12",
        "action_label": "Open programme cockpit",
        "action_url": "/solutions/programmes/4",
    },
]


@pytest.fixture
def client(app):
    previous = app.config.get("LOGIN_DISABLED", False)
    app.config["LOGIN_DISABLED"] = True
    try:
        yield app.test_client()
    finally:
        app.config["LOGIN_DISABLED"] = previous


def _briefing_without_counts(db_session):
    """A briefing row whose counts are NULL.

    ``finding_count`` and ``flagged_count`` carry a Python-side ``default=0``,
    so they cannot be inserted as NULL — they have to be nulled afterwards. The
    state is reachable all the same (a row written outside the ORM, or a column
    added to an existing table by ``reconcile-schema``, arrives NULL), and it is
    the state the em-dash rule exists for.
    """
    briefing = _briefing(db_session)
    briefing.finding_count = None
    briefing.flagged_count = None
    db_session.flush()
    return briefing


def _briefing(db_session, **kw):
    from app.models.strategic import EnterpriseBriefing

    defaults = {
        "headline": "Drift detected on Core Banking Replacement",
        "summary": "This briefing surfaces 2 findings (1 needing attention).",
        "source": "manual",
        "findings": FINDINGS,
        "finding_count": 2,
        "flagged_count": 1,
    }
    defaults.update(kw)
    briefing = EnterpriseBriefing(**defaults)
    db_session.add(briefing)
    db_session.flush()
    return briefing


# ---------------------------------------------------------------------------
# What the builder gathers
# ---------------------------------------------------------------------------


def test_the_findings_come_back_as_persisted_ranked_by_severity(db_session):
    briefing = _briefing(db_session)
    report = BriefingReportService.build(briefing_id=briefing.id)

    assert report is not None
    assert [f["severity"] for f in report["findings"]] == ["high", "info"], (
        "the reader needs what needs attention first"
    )
    assert report["findings"][0]["action_url"] == "/solutions/programmes/4", (
        "the verification link is the point of the artefact"
    )
    assert report["totals"]["findings_listed"] == 2


def test_the_counts_are_the_stored_ones_not_recomputed(db_session):
    """An export that recomputed would disagree with the page it came from."""
    briefing = _briefing(db_session, finding_count=7, flagged_count=3)
    report = BriefingReportService.build(briefing_id=briefing.id)

    assert report["briefing"]["finding_count"] == 7
    assert report["briefing"]["flagged_count"] == 3
    assert report["totals"]["findings_listed"] == 2


def test_a_count_that_was_never_written_stays_absent(db_session):
    briefing = _briefing_without_counts(db_session)
    report = BriefingReportService.build(briefing_id=briefing.id)

    assert report["briefing"]["finding_count"] is None, (
        "a briefing with no count recorded is not a briefing that found nothing"
    )
    assert report["briefing"]["flagged_count"] is None


def test_findings_are_bucketed_by_area(db_session):
    briefing = _briefing(db_session)
    report = BriefingReportService.build(briefing_id=briefing.id)

    by_name = {row["name"]: row for row in report["categories"]}
    assert by_name["drift"] == {"name": "drift", "count": 1, "flagged": 1}
    assert by_name["portfolio"] == {"name": "portfolio", "count": 1, "flagged": 0}
    assert report["totals"]["categories"] == 2


def test_a_briefing_with_no_findings_recorded_is_not_an_error(db_session):
    briefing = _briefing(db_session, findings=[], finding_count=0, flagged_count=0)
    report = BriefingReportService.build(briefing_id=briefing.id)

    assert report["findings"] == []
    assert report["categories"] == []
    assert report["briefing"]["finding_count"] == 0


def test_a_missing_briefing_builds_to_none(db_session):
    assert BriefingReportService.build(briefing_id=987654321) is None


# ---------------------------------------------------------------------------
# The renderers
# ---------------------------------------------------------------------------


def test_the_workbook_carries_the_verification_link_for_each_finding(db_session):
    from openpyxl import load_workbook

    briefing = _briefing(db_session)
    payload = BriefingReportService.render_xlsx(
        BriefingReportService.build(briefing_id=briefing.id)
    )

    workbook = load_workbook(io.BytesIO(payload))
    assert {"Findings", "Summary", "Areas", "History"} <= set(workbook.sheetnames)

    sheet = workbook["Findings"]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index("Where to verify") + 1
    links = {sheet.cell(row=row, column=column).value for row in (2, 3)}
    assert links == {"/solutions/programmes/4", "/applications/"}


def test_the_workbook_leaves_an_unwritten_count_cell_empty(db_session):
    from openpyxl import load_workbook

    briefing = _briefing_without_counts(db_session)
    payload = BriefingReportService.render_xlsx(
        BriefingReportService.build(briefing_id=briefing.id)
    )

    workbook = load_workbook(io.BytesIO(payload))
    summary = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(values_only=True)
        if row[0]
    }
    assert summary["Findings"] in (None, "")
    assert summary["Needing attention"] in (None, "")


def test_the_word_document_names_the_findings_and_their_evidence(db_session):
    from docx import Document

    briefing = _briefing(db_session)
    payload = BriefingReportService.render_docx(
        BriefingReportService.build(briefing_id=briefing.id)
    )

    document = Document(io.BytesIO(payload))
    text = "\n".join(p.text for p in document.paragraphs)
    assert "Drift detected on Core Banking Replacement" in text
    assert "Scope changed after the last gate." in text
    assert "evidence: Snapshot 12" in text
    assert "/solutions/programmes/4" in text


def test_the_html_report_states_when_a_briefing_found_nothing(app, db_session):
    briefing = _briefing(db_session, findings=[], finding_count=0, flagged_count=0)
    report = BriefingReportService.build(briefing_id=briefing.id)
    with app.test_request_context("/"):
        html = BriefingReportService.render_html(report)

    assert "This briefing recorded no findings" in html
    assert "not a failure to look" in html


# ---------------------------------------------------------------------------
# The route
# ---------------------------------------------------------------------------


@pytest.fixture
def seeded_briefing(db_session):
    return _briefing(db_session, source=f"test-{uuid.uuid4().hex[:6]}").id


@pytest.mark.parametrize("fmt", ["html", "xlsx", "docx"])
def test_each_format_downloads(client, seeded_briefing, fmt):
    resp = client.get(f"/solutions/briefings/report.{fmt}?briefing_id={seeded_briefing}")
    assert resp.status_code == 200, resp.data[:300]
    assert len(resp.data) > 0
    if fmt != "html":
        assert "attachment" in resp.headers.get("Content-Disposition", "")


def test_the_html_response_declares_one_charset(client, seeded_briefing):
    resp = client.get(f"/solutions/briefings/report.html?briefing_id={seeded_briefing}")
    assert resp.headers["Content-Type"].count("charset") == 1


def test_an_unsupported_format_is_refused_with_the_list(client, seeded_briefing):
    resp = client.get(f"/solutions/briefings/report.rtf?briefing_id={seeded_briefing}")
    assert resp.status_code == 400
    body = resp.get_json()
    assert "rtf" in body["error"]
    assert "pdf" in body["error"]


def test_a_briefing_that_does_not_exist_is_a_404(client):
    resp = client.get("/solutions/briefings/report.html?briefing_id=987654321")
    assert resp.status_code == 404
    assert "briefing" in resp.get_json()["error"].lower()


def test_a_format_this_deployment_cannot_render_says_so(
    client, seeded_briefing, monkeypatch
):
    """WeasyPrint binds native libraries that are absent on Windows and slim images."""
    from app.modules.solutions_strategic.v2.services import briefing_report_service as mod

    def _unavailable(_report):
        raise BriefingReportError("PDF rendering is unavailable on this deployment.")

    monkeypatch.setattr(
        mod.BriefingReportService, "render_pdf", staticmethod(_unavailable)
    )

    resp = client.get(f"/solutions/briefings/report.pdf?briefing_id={seeded_briefing}")
    assert resp.status_code == 503
    assert "unavailable" in resp.get_json()["error"].lower()


def test_the_report_requires_a_login(app, seeded_briefing):
    resp = app.test_client().get(
        f"/solutions/briefings/report.xlsx?briefing_id={seeded_briefing}"
    )
    assert resp.status_code in (302, 401)
