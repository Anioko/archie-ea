"""The capability model as a deliverable — PDF, Word, Excel, shareable page.

Archie could export a Solution Architecture Document and a solution deck, and
nothing a *business* architect produces: the capability map offered CSV, JSON
and a picture, none of which is a deliverable. A business architect's output is
a document that goes to an executive committee.

The property these tests exist to protect is that **the report never invents a
measurement**. ``current_maturity_level`` carries a column default of 1 and is
therefore never NULL, so a capability nobody has assessed is indistinguishable
from one deliberately scored 1 unless the renderer keys off
``maturity_assessment_date``. A "0" or a "1.0" in a maturity column of a board
paper is a number an executive acts on.

One builder feeds all four renderers, so the PDF, the Word file and the
workbook cannot disagree about the same capability count in the same meeting.
"""

from __future__ import annotations

import io
import uuid

import pytest

from app.modules.capabilities.services.capability_report_service import (
    CapabilityReportError,
    CapabilityReportService,
)


@pytest.fixture
def client(app):
    previous = app.config.get("LOGIN_DISABLED", False)
    app.config["LOGIN_DISABLED"] = True
    try:
        yield app.test_client()
    finally:
        app.config["LOGIN_DISABLED"] = previous


def _capability(db_session, name, **kw):
    from app.models.business_capabilities import BusinessCapability

    cap = BusinessCapability(name=name, **kw)
    db_session.add(cap)
    db_session.flush()
    return cap


# ---------------------------------------------------------------------------
# What the builder gathers
# ---------------------------------------------------------------------------


def test_an_unassessed_model_reports_no_average_rather_than_zero(
    db_session, make_org, tenant_ctx
):
    org = make_org(f"report-unassessed-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        _capability(db_session, "Supply Chain", level=1, business_domain="Operations")
        _capability(db_session, "Customer Management", level=1, business_domain="Commercial")
        report = CapabilityReportService.build()

    assert report["totals"]["capabilities"] == 2
    assert report["totals"]["assessed"] == 0
    assert report["totals"]["avg_current"] is None, (
        "current_maturity_level defaults to 1, so averaging it over unassessed "
        "rows puts a maturity score nobody measured into a board paper"
    )
    assert all(row["current"] is None for row in report["flat"])


def test_averages_cover_only_the_assessed(db_session, make_org, tenant_ctx):
    from datetime import datetime

    org = make_org(f"report-assessed-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        _capability(
            db_session, "Assessed A", level=1, business_domain="Operations",
            current_maturity_level=2, target_maturity_level=4,
            maturity_assessment_date=datetime.utcnow(),
        )
        _capability(
            db_session, "Assessed B", level=1, business_domain="Operations",
            current_maturity_level=4, target_maturity_level=4,
            maturity_assessment_date=datetime.utcnow(),
        )
        _capability(db_session, "Never Assessed", level=1, business_domain="Operations")
        report = CapabilityReportService.build()

    totals = report["totals"]
    assert (totals["capabilities"], totals["assessed"], totals["unassessed"]) == (3, 2, 1)
    assert totals["avg_current"] == pytest.approx(3.0)
    assert totals["avg_target"] == pytest.approx(4.0)

    operations = next(d for d in report["domains"] if d["name"] == "Operations")
    assert operations["count"] == 3
    assert operations["assessed"] == 2
    assert operations["avg_current"] == pytest.approx(3.0)


def test_a_parentless_capability_below_level_one_is_still_in_the_report(
    db_session, make_org, tenant_ctx
):
    """The same root rule as the tree view — a report that drops rows is worse than none."""
    org = make_org(f"report-orphan-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        _capability(db_session, "Orphaned Sub-Capability", level=3)
        report = CapabilityReportService.build()

    assert [row["name"] for row in report["flat"]] == ["Orphaned Sub-Capability"]


def test_the_hierarchy_is_flattened_with_its_depth(db_session, make_org, tenant_ctx):
    org = make_org(f"report-depth-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        parent = _capability(db_session, "Supply Chain", level=1)
        _capability(db_session, "Demand Planning", level=2, parent_capability_id=parent.id)
        report = CapabilityReportService.build()

    by_name = {row["name"]: row for row in report["flat"]}
    assert by_name["Supply Chain"]["depth"] == 0
    assert by_name["Demand Planning"]["depth"] == 1


# ---------------------------------------------------------------------------
# The renderers
# ---------------------------------------------------------------------------


def test_the_workbook_leaves_an_unassessed_maturity_cell_empty(
    db_session, make_org, tenant_ctx
):
    from openpyxl import load_workbook

    org = make_org(f"report-xlsx-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        _capability(db_session, "Unassessed Capability", level=1, business_domain="Ops")
        payload = CapabilityReportService.render_xlsx(CapabilityReportService.build())

    workbook = load_workbook(io.BytesIO(payload))
    assert {"Capabilities", "Summary", "Domains"} <= set(workbook.sheetnames)

    sheet = workbook["Capabilities"]
    headers = [c.value for c in sheet[1]]
    current_col = headers.index("Current maturity") + 1
    assert sheet.cell(row=2, column=current_col).value in (None, ""), (
        "a 0 in a maturity column reads as a measured score of zero"
    )

    summary = {row[0]: row[1] for row in workbook["Summary"].iter_rows(values_only=True) if row[0]}
    assert summary["Average current maturity"] == "not assessed"


def test_the_word_document_renders_with_content(db_session, make_org, tenant_ctx):
    from docx import Document

    org = make_org(f"report-docx-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        _capability(db_session, "Supply Chain Reporting", level=1, business_domain="Ops")
        payload = CapabilityReportService.render_docx(CapabilityReportService.build())

    document = Document(io.BytesIO(payload))
    text = "\n".join(p.text for p in document.paragraphs)
    assert "Supply Chain Reporting" in text
    assert "Capability Model" in text


def test_the_html_report_shows_an_em_dash_for_an_unassessed_capability(
    app, db_session, make_org, tenant_ctx
):
    org = make_org(f"report-html-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        _capability(db_session, "Unassessed Capability", level=1)
        with app.test_request_context("/"):
            html = CapabilityReportService.render_html(CapabilityReportService.build())

    assert "Unassessed Capability" in html
    assert "—" in html
    assert "No capability in this model carries a maturity assessment" in html


# ---------------------------------------------------------------------------
# The route
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("fmt", ["html", "xlsx", "docx"])
def test_each_format_downloads(app, client, fmt):
    resp = client.get(f"/capability-map/report.{fmt}")
    assert resp.status_code == 200, resp.data[:300]
    assert len(resp.data) > 0
    if fmt != "html":
        assert "attachment" in resp.headers.get("Content-Disposition", "")


def test_the_html_response_declares_one_charset(app, client):
    resp = client.get("/capability-map/report.html")
    assert resp.headers["Content-Type"].count("charset") == 1


def test_an_unsupported_format_is_refused_with_the_list(app, client):
    resp = client.get("/capability-map/report.rtf")
    assert resp.status_code == 400
    body = resp.get_json()
    assert "rtf" in body["error"]
    assert "xlsx" in body["error"]


def test_a_format_this_deployment_cannot_render_says_so(app, client, monkeypatch):
    """WeasyPrint binds native libraries that are absent on Windows and slim images.

    The answer has to name the reason and leave the other three formats
    working — not 500, and not a zero-byte file the user opens and puzzles over.
    """
    from app.modules.capabilities.services import capability_report_service as mod

    def _unavailable(_report):
        raise CapabilityReportError("PDF rendering is unavailable on this deployment.")

    monkeypatch.setattr(
        mod.CapabilityReportService, "render_pdf", staticmethod(_unavailable)
    )

    resp = client.get("/capability-map/report.pdf")
    assert resp.status_code == 503
    assert "unavailable" in resp.get_json()["error"].lower()


def test_the_report_requires_a_login(app):
    resp = app.test_client().get("/capability-map/report.xlsx")
    assert resp.status_code in (302, 401)
