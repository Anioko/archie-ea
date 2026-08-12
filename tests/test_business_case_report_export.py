"""A business case as a deliverable — PDF, Word, Excel, shareable page.

The business case is the one business-architecture artefact that already *is* a
document, and until now it existed only as a web form. What reached an
investment board was whatever somebody retyped into their own template.

The properties these tests exist to protect:

* **The report never invents a figure.** Every financial column on
  ``BusinessCase`` is nullable. A case with no capex entered must show an em
  dash and an empty workbook cell, never 0 — a zero capex is a claim that the
  change is free, and somebody approves against it.
* **Rendering never writes.** ``aggregate_financials`` will happily
  pre-populate blank fields from the linked capability / initiative / solution.
  Exporting a document must not mutate the document, so the builder calls it
  with ``apply_missing=False`` and shows the linked figures as a labelled
  cross-check instead.
* **A section nobody wrote says so**, rather than appearing as an empty heading
  that reads as "there was nothing to say here".
"""

from __future__ import annotations

import io
import uuid
from decimal import Decimal

import pytest

from app.modules.business_case.report_service import (
    BusinessCaseReportError,
    BusinessCaseReportService,
)


@pytest.fixture
def client(app):
    previous = app.config.get("LOGIN_DISABLED", False)
    app.config["LOGIN_DISABLED"] = True
    try:
        yield app.test_client()
    finally:
        app.config["LOGIN_DISABLED"] = previous


def _case(db_session, title, **kw):
    from app.models.business_case import BusinessCase

    case = BusinessCase(title=title, **kw)
    db_session.add(case)
    db_session.flush()
    return case


# ---------------------------------------------------------------------------
# What the builder gathers
# ---------------------------------------------------------------------------


def test_an_empty_case_reports_no_financials_rather_than_zero(
    db_session, make_org, tenant_ctx
):
    org = make_org(f"bc-empty-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        case = _case(db_session, "Replace the billing platform")
        report = BusinessCaseReportService.build(case.id)

    assert report is not None
    assert all(row["value"] is None for row in report["financials"]), (
        "a zero capex is a claim that the change is free, and an investment "
        "board approves against it"
    )
    assert report["totals"]["financials_entered"] == 0


def test_entered_financials_are_carried_through_as_numbers(
    db_session, make_org, tenant_ctx
):
    org = make_org(f"bc-money-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        case = _case(
            db_session,
            "Consolidate the CRM estate",
            capex=Decimal("1250000.00"),
            roi_percentage=Decimal("18.50"),
            payback_months=30,
        )
        report = BusinessCaseReportService.build(case.id)

    by_key = {row["key"]: row for row in report["financials"]}
    assert by_key["capex"]["value"] == pytest.approx(1250000.0)
    assert by_key["roi_percentage"]["value"] == pytest.approx(18.5)
    assert by_key["payback_months"]["value"] == pytest.approx(30.0)
    assert by_key["opex_annual"]["value"] is None
    assert report["totals"]["financials_entered"] == 3


def test_an_unwritten_narrative_section_is_none_not_an_empty_string(
    db_session, make_org, tenant_ctx
):
    org = make_org(f"bc-narrative-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        case = _case(
            db_session,
            "Retire the mainframe",
            problem_statement="Support ends in 2028.",
        )
        report = BusinessCaseReportService.build(case.id)

    by_key = {row["key"]: row for row in report["narrative"]}
    assert by_key["problem_statement"]["text"] == "Support ends in 2028."
    assert by_key["key_risks"]["text"] is None
    assert report["totals"]["sections_written"] == 1
    assert report["totals"]["sections"] == 5


def test_building_the_report_does_not_write_to_the_case(
    db_session, make_org, tenant_ctx
):
    """aggregate_financials(apply_missing=True) would fill capex in from the link.

    Exporting a document must not change the document. A reader who exports a
    draft twice and gets two different capex figures cannot trust either.
    """
    from app.models.solution_models import Solution

    org = make_org(f"bc-readonly-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        solution = Solution(name=f"Linked Solution {uuid.uuid4().hex[:6]}",
                            estimated_cost=Decimal("500000.00"))
        db_session.add(solution)
        db_session.flush()

        case = _case(db_session, "Linked case", solution_id=solution.id)
        report = BusinessCaseReportService.build(case.id)

        assert case.capex is None, "the export pre-populated a field it was only reading"

    by_key = {row["key"]: row for row in report["financials"]}
    assert by_key["capex"]["value"] is None
    assert report["links"][2] == {"label": "Solution", "name": solution.name}


def test_a_divergence_from_a_linked_record_is_reported_as_a_cross_check(
    db_session, make_org, tenant_ctx
):
    from app.models.solution_models import Solution

    org = make_org(f"bc-crosscheck-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        solution = Solution(name=f"Cross Solution {uuid.uuid4().hex[:6]}",
                            estimated_cost=Decimal("500000.00"))
        db_session.add(solution)
        db_session.flush()
        case = _case(
            db_session, "Diverging case",
            solution_id=solution.id, capex=Decimal("600000.00"),
        )
        report = BusinessCaseReportService.build(case.id)

    by_label = {row["label"]: row for row in report["cross_checks"]}
    assert "Capital expenditure" in by_label
    assert by_label["Capital expenditure"]["entered"] == pytest.approx(600000.0)
    assert by_label["Capital expenditure"]["linked"] == pytest.approx(500000.0)
    assert by_label["Capital expenditure"]["difference"] == pytest.approx(100000.0)


def test_a_missing_business_case_builds_to_none(db_session, make_org, tenant_ctx):
    org = make_org(f"bc-missing-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        assert BusinessCaseReportService.build(987654321) is None


# ---------------------------------------------------------------------------
# The renderers
# ---------------------------------------------------------------------------


def test_the_workbook_leaves_an_unentered_cost_cell_empty(
    db_session, make_org, tenant_ctx
):
    from openpyxl import load_workbook

    org = make_org(f"bc-xlsx-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        case = _case(db_session, "Unpriced case")
        payload = BusinessCaseReportService.render_xlsx(
            BusinessCaseReportService.build(case.id)
        )

    workbook = load_workbook(io.BytesIO(payload))
    assert {"Summary", "Financials", "Narrative", "Cross-check"} <= set(
        workbook.sheetnames
    )

    financials = {
        row[0]: row[1]
        for row in workbook["Financials"].iter_rows(min_row=2, values_only=True)
    }
    assert financials["Capital expenditure"] in (None, ""), (
        "a 0 in a cost column reads as a measured figure and gets summed into a "
        "portfolio total"
    )
    assert financials["Return on investment"] in (None, "")


def test_the_word_document_marks_a_section_nobody_wrote(
    db_session, make_org, tenant_ctx
):
    from docx import Document

    org = make_org(f"bc-docx-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        case = _case(
            db_session, "Partially written case",
            problem_statement="The estate carries four overlapping CRMs.",
        )
        payload = BusinessCaseReportService.render_docx(
            BusinessCaseReportService.build(case.id)
        )

    document = Document(io.BytesIO(payload))
    text = "\n".join(p.text for p in document.paragraphs)
    assert "Partially written case" in text
    assert "The estate carries four overlapping CRMs." in text
    assert "Not documented. This section has not been written yet." in text


def test_the_html_report_shows_an_em_dash_for_an_unentered_figure(
    app, db_session, make_org, tenant_ctx
):
    org = make_org(f"bc-html-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        case = _case(db_session, "Unpriced HTML case")
        report = BusinessCaseReportService.build(case.id)
        with app.test_request_context("/"):
            html = BusinessCaseReportService.render_html(report)

    assert "Unpriced HTML case" in html
    assert "—" in html
    assert "It is not a zero" in html
    assert "Not documented" in html


# ---------------------------------------------------------------------------
# The route
# ---------------------------------------------------------------------------


@pytest.fixture
def seeded_case(db_session, make_org, tenant_ctx):
    org = make_org(f"bc-route-{uuid.uuid4().hex[:6]}")
    with tenant_ctx(org.id):
        case = _case(
            db_session, "Routed case",
            problem_statement="Something needs fixing.",
            capex=Decimal("10000.00"),
        )
    return case.id


@pytest.mark.parametrize("fmt", ["html", "xlsx", "docx"])
def test_each_format_downloads(client, seeded_case, fmt):
    resp = client.get(f"/business-case/{seeded_case}/report.{fmt}")
    assert resp.status_code == 200, resp.data[:300]
    assert len(resp.data) > 0
    if fmt != "html":
        assert "attachment" in resp.headers.get("Content-Disposition", "")


def test_the_html_response_declares_one_charset(client, seeded_case):
    resp = client.get(f"/business-case/{seeded_case}/report.html")
    assert resp.headers["Content-Type"].count("charset") == 1


def test_an_unsupported_format_is_refused_with_the_list(client, seeded_case):
    resp = client.get(f"/business-case/{seeded_case}/report.rtf")
    assert resp.status_code == 400
    body = resp.get_json()
    assert "rtf" in body["error"]
    assert "docx" in body["error"]


def test_a_business_case_that_does_not_exist_is_a_404(client):
    resp = client.get("/business-case/987654321/report.html")
    assert resp.status_code == 404
    assert "not found" in resp.get_json()["error"].lower()


def test_a_format_this_deployment_cannot_render_says_so(
    client, seeded_case, monkeypatch
):
    """WeasyPrint binds native libraries that are absent on Windows and slim images."""
    from app.modules.business_case import report_service as mod

    def _unavailable(_report):
        raise BusinessCaseReportError("PDF rendering is unavailable on this deployment.")

    monkeypatch.setattr(
        mod.BusinessCaseReportService, "render_pdf", staticmethod(_unavailable)
    )

    resp = client.get(f"/business-case/{seeded_case}/report.pdf")
    assert resp.status_code == 503
    assert "unavailable" in resp.get_json()["error"].lower()


def test_the_report_requires_a_login(app, seeded_case):
    resp = app.test_client().get(f"/business-case/{seeded_case}/report.xlsx")
    assert resp.status_code in (302, 401)
